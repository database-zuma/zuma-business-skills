#!/usr/bin/env python3
"""
Coretax Faktur Generator — Generate DJP Coretax-ready XLSX from Register Penjualan.

Reads monthly Register Penjualan workbook, filters NON-RETAIL invoices,
and outputs Faktur + DetailFaktur sheets in DJP Coretax import template format.

Usage:
    python3 coretax_faktur.py "/path/to/Register_Penjualan_Januari_Tax_2026.xlsx"
    python3 coretax_faktur.py "/path/to/Register.xlsx" --entity DDD --output /outbox/
    python3 coretax_faktur.py "/path/to/Register.xlsx" --all-nonretail
    python3 coretax_faktur.py "/path/to/Register.xlsx" --dry-run
"""

from __future__ import annotations

import argparse
import os
import re
import sys
from collections import OrderedDict
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print(
        "ERROR: openpyxl is required. Install with: pip3 install openpyxl",
        file=sys.stderr,
    )
    sys.exit(1)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

ENTITY_NPWP = {
    "DDD": "0803371400624000",
    "MBB": "0997550074617000",
    "UBB": "0628502825617000",
    "LJBB": "",  # TODO: add when available
}

# Month sheet name mapping (Register uses 3-letter abbreviations)
MONTH_SHEETS = {
    "JAN": "JAN",
    "FEB": "FEB",
    "MAR": "MAR",
    "APR": "APR",
    "MAY": "MAY",
    "MEI": "MEI",
    "JUN": "JUN",
    "JUL": "JUL",
    "AUG": "AUG",
    "AGS": "AGS",
    "SEP": "SEP",
    "OKT": "OKT",
    "OCT": "OCT",
    "NOV": "NOV",
    "DEC": "DEC",
    "DES": "DES",
}

# Default customer filter (Bu Aulia's pattern: MBB only)
DEFAULT_CUSTOMERS = ["MAKMUR BESAR BERSAMA"]

# Non-retail Jenis Pelanggan values
NON_RETAIL_TYPES = {"WHOLESALE", "CONSIGNMENT"}


# JAN sheet column indices (1-based, header at row 2)
class JanCol:
    ID_PELANGGAN = 2
    TANGGAL = 3
    NAMA_PELANGGAN = 4
    NOMOR = 7  # Invoice number (Nomor #)
    KODE = 8  # Product code (Kode #)
    NAMA_BARANG = 9
    KUANTITAS = 10
    SATUAN = 11
    HARGA_AT = 12  # Harga @ (price per unit, inc tax)
    DISKON = 15
    HARGA_INCLUDE = 16
    TOTAL_HARGA = 17
    DEPARTEMEN = 19
    CABANG = 20
    JENIS_PELANGGAN = 21


# Master Pelanggan column indices (1-based, header at row 1)
class MasterCol:
    ID = 1
    NAMA = 2
    NPWP_15 = 4
    NPWP_16 = 5
    NITKU = 6
    ALAMAT = 7


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------


class LineItem:
    """A single line item from the JAN sheet."""

    __slots__ = (
        "invoice",
        "date",
        "customer_name",
        "customer_id",
        "product_code",
        "product_name",
        "base_name",
        "qty",
        "harga_at",
        "diskon",
        "harga_include",
        "total_harga",
        "jenis_pelanggan",
    )

    def __init__(self, row_data: Dict[str, Any]):
        self.invoice = row_data["invoice"]  # type: str
        self.date = row_data["date"]  # type: datetime
        self.customer_name = row_data["customer_name"]  # type: str
        self.customer_id = row_data["customer_id"]  # type: str
        self.product_code = row_data["product_code"]  # type: str
        self.product_name = row_data["product_name"]  # type: str
        self.base_name = row_data["base_name"]  # type: str
        self.qty = row_data["qty"]  # type: float
        self.harga_at = row_data["harga_at"]  # type: float
        self.diskon = row_data["diskon"]  # type: float
        self.harga_include = row_data["harga_include"]  # type: float
        self.total_harga = row_data["total_harga"]  # type: float
        self.jenis_pelanggan = row_data["jenis_pelanggan"]  # type: str


class CustomerInfo:
    """Customer master data."""

    __slots__ = ("id", "name", "npwp_16", "nitku", "alamat")

    def __init__(self, id_: str, name: str, npwp_16: str, nitku: str, alamat: str):
        self.id = id_
        self.name = name
        self.npwp_16 = npwp_16
        self.nitku = nitku
        self.alamat = alamat


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def strip_article_name(raw_name: str) -> str:
    """Strip size/color from article name, add SANDAL prefix.

    'BABY BOYS TOY STORY 1, 21/22, BROWN' -> 'SANDAL BABY BOYS TOY STORY 1'
    """
    base = raw_name.split(",")[0].strip()
    return "SANDAL " + base


def safe_float(val: Any) -> float:
    """Convert cell value to float, defaulting to 0."""
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def safe_str(val: Any) -> str:
    """Convert cell value to string, defaulting to empty."""
    if val is None:
        return ""
    return str(val).strip()


def detect_month_sheet(wb: openpyxl.Workbook) -> Optional[str]:
    """Auto-detect the month data sheet from workbook."""
    names_upper = {n.upper(): n for n in wb.sheetnames}
    for abbrev in MONTH_SHEETS:
        if abbrev in names_upper:
            return names_upper[abbrev]
    # Fallback: look for sheet with >1000 rows that's not Master/Rekap/etc
    skip = {
        "MASTER HARGA",
        "MASTER PELANGGAN",
        "REKAP",
        "INPUT CORETAX",
        "DIGUNGGUNG",
        "REKON CORETAX",
        "SALES FREETAX",
    }
    for name in wb.sheetnames:
        if name.upper() in skip:
            continue
        ws = wb[name]
        if ws.max_row and ws.max_row > 1000:
            return name
    return None


def find_header_row(
    ws: openpyxl.worksheet.worksheet.Worksheet, marker: str = "Nomor #"
) -> int:
    """Find the row containing column headers by looking for a marker string."""
    for row in range(1, min(10, ws.max_row + 1)):
        for col in range(1, min(35, ws.max_column + 1)):
            val = ws.cell(row, col).value
            if val and marker in str(val):
                return row
    return 2  # default fallback


def build_column_map(
    ws: openpyxl.worksheet.worksheet.Worksheet, header_row: int
) -> Dict[str, int]:
    """Build a flexible column name -> index map from the header row."""
    col_map = {}
    for col in range(1, min(35, ws.max_column + 1)):
        val = ws.cell(header_row, col).value
        if val:
            col_map[str(val).strip()] = col
    return col_map


# ---------------------------------------------------------------------------
# Core: Read Register
# ---------------------------------------------------------------------------


def read_master_pelanggan(wb: openpyxl.Workbook) -> Dict[str, CustomerInfo]:
    """Read Master Pelanggan sheet into lookup dict keyed by customer name (uppercased)."""
    ws = wb["Master Pelanggan"]
    customers = {}  # type: Dict[str, CustomerInfo]

    for row in range(2, ws.max_row + 1):
        cid = safe_str(ws.cell(row, MasterCol.ID).value)
        name = safe_str(ws.cell(row, MasterCol.NAMA).value)
        npwp_16 = safe_str(ws.cell(row, MasterCol.NPWP_16).value)
        nitku = safe_str(ws.cell(row, MasterCol.NITKU).value)
        alamat = safe_str(ws.cell(row, MasterCol.ALAMAT).value)

        if not name:
            continue

        info = CustomerInfo(cid, name, npwp_16, nitku, alamat)
        customers[name.upper()] = info
        # Also index by ID for fallback lookup
        if cid:
            customers[cid.upper()] = info

    return customers


def read_line_items(
    wb: openpyxl.Workbook,
    month_sheet: str,
    customer_filter: Optional[List[str]] = None,
    all_nonretail: bool = False,
) -> List[LineItem]:
    """Read line items from the month data sheet, filtered by customer or jenis pelanggan."""
    ws = wb[month_sheet]
    header_row = find_header_row(ws)

    # Build column map for flexibility
    col_map = build_column_map(ws, header_row)

    # Determine column indices (fallback to hardcoded if header names differ)
    def get_col(names: List[str], default: int) -> int:
        for n in names:
            if n in col_map:
                return col_map[n]
        return default

    col_inv = get_col(["Nomor #"], JanCol.NOMOR)
    col_date = get_col(["Tanggal"], JanCol.TANGGAL)
    col_cust_name = get_col(["Nama Pelanggan"], JanCol.NAMA_PELANGGAN)
    col_cust_id = get_col(["Id Pelanggan"], JanCol.ID_PELANGGAN)
    col_code = get_col(["Kode #"], JanCol.KODE)
    col_name = get_col(["Nama Barang"], JanCol.NAMA_BARANG)
    col_qty = get_col(["Kuantitas"], JanCol.KUANTITAS)
    col_harga = get_col(["Harga @", "Harga@"], JanCol.HARGA_AT)
    col_diskon = get_col(["diskon"], JanCol.DISKON)
    col_harga_inc = get_col(["Harga Include"], JanCol.HARGA_INCLUDE)
    col_total = get_col(["Total Harga"], JanCol.TOTAL_HARGA)
    col_jenis = get_col(["Jenis Pelanggan"], JanCol.JENIS_PELANGGAN)

    # Prepare filter sets
    filter_names = None  # type: Optional[set]
    if customer_filter and not all_nonretail:
        filter_names = {n.upper() for n in customer_filter}

    items = []  # type: List[LineItem]
    data_start = header_row + 1

    for row in range(data_start, ws.max_row + 1):
        inv = safe_str(ws.cell(row, col_inv).value)
        if not inv or not inv.startswith("INV/"):
            continue

        jenis = safe_str(ws.cell(row, col_jenis).value)
        cust_name = safe_str(ws.cell(row, col_cust_name).value)

        # Filter by jenis pelanggan
        if jenis.upper() in ("RETAIL",):
            continue
        if jenis.upper() == "RETUR PPNK":
            continue

        # Filter by customer name
        if filter_names and cust_name.upper() not in filter_names:
            # Try partial match
            matched = False
            for fn in filter_names:
                if fn in cust_name.upper() or cust_name.upper() in fn:
                    matched = True
                    break
            if not matched:
                continue

        # Parse date
        raw_date = ws.cell(row, col_date).value
        if isinstance(raw_date, datetime):
            date = raw_date
        elif isinstance(raw_date, str):
            try:
                date = datetime.strptime(raw_date, "%Y-%m-%d %H:%M:%S")
            except ValueError:
                try:
                    date = datetime.strptime(raw_date, "%Y-%m-%d")
                except ValueError:
                    date = datetime.now()
        else:
            date = datetime.now()

        raw_name = safe_str(ws.cell(row, col_name).value)
        base_name = strip_article_name(raw_name) if raw_name else ""

        item = LineItem(
            {
                "invoice": inv,
                "date": date,
                "customer_name": cust_name,
                "customer_id": safe_str(ws.cell(row, col_cust_id).value),
                "product_code": safe_str(ws.cell(row, col_code).value),
                "product_name": raw_name,
                "base_name": base_name,
                "qty": safe_float(ws.cell(row, col_qty).value),
                "harga_at": safe_float(ws.cell(row, col_harga).value),
                "diskon": safe_float(ws.cell(row, col_diskon).value),
                "harga_include": safe_float(ws.cell(row, col_harga_inc).value),
                "total_harga": safe_float(ws.cell(row, col_total).value),
                "jenis_pelanggan": jenis,
            }
        )
        items.append(item)

    return items


# ---------------------------------------------------------------------------
# Core: Aggregate & Generate
# ---------------------------------------------------------------------------


def aggregate_invoices(
    items: List[LineItem],
    customers: Dict[str, CustomerInfo],
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    """Aggregate line items into Faktur rows and DetailFaktur rows.

    Returns: (faktur_rows, detail_rows)
    """
    # Group by invoice number
    invoice_groups = OrderedDict()  # type: OrderedDict[str, List[LineItem]]
    for item in items:
        if item.invoice not in invoice_groups:
            invoice_groups[item.invoice] = []
        invoice_groups[item.invoice].append(item)

    # Sort by invoice number (extract numeric suffix)
    def sort_key(inv_entry):
        inv_num, inv_items = inv_entry
        # Extract trailing number from e.g. INV/DDD/2026/I/017 -> 17
        match = re.search(r'(\d+)$', inv_num)
        num = int(match.group(1)) if match else 0
        return num

    sorted_groups = OrderedDict(sorted(invoice_groups.items(), key=sort_key))
    faktur_rows = []  # type: List[Dict[str, Any]]
    detail_rows = []  # type: List[Dict[str, Any]]

    for baris_idx, (inv_num, inv_items) in enumerate(sorted_groups.items(), start=1):
        first = inv_items[0]

        # Lookup customer from Master Pelanggan
        cust = _lookup_customer(first, customers)

        # Faktur header row
        faktur = {
            "baris": baris_idx,
            "tanggal": first.date.strftime("%d/%m/%Y"),
            "jenis_faktur": "Normal",
            "kode_transaksi": "04",
            "keterangan_tambahan": "",
            "dokumen_pendukung": "",
            "period_dok": "",
            "referensi": inv_num,
            "cap_fasilitas": "",
            "id_tku_penjual": "",  # filled per entity
            "npwp_pembeli": cust.npwp_16 if cust else "",
            "jenis_id_pembeli": "TIN",
            "negara_pembeli": "IDN",
            "nomor_dok_pembeli": "-",
            "nama_pembeli": cust.name if cust else first.customer_name,
            "alamat_pembeli": cust.alamat if cust else "",
            "email_pembeli": "-",
            "id_tku_pembeli": cust.nitku if cust else "",
        }
        faktur_rows.append(faktur)

        # Aggregate detail by base article name
        article_groups = OrderedDict()  # type: OrderedDict[str, List[LineItem]]
        for item in inv_items:
            key = item.base_name
            if key not in article_groups:
                article_groups[key] = []
            article_groups[key].append(item)

        for art_name, art_items in article_groups.items():
            total_qty = sum(i.qty for i in art_items)
            total_harga_sum = sum(i.total_harga for i in art_items)
            total_diskon = sum(i.diskon for i in art_items)

            # Tax math (Kode Transaksi 04: DPP Nilai Lain)
            # Harga Satuan = weighted avg Harga@ / 1.11
            if total_qty > 0:
                avg_harga_at = total_harga_sum / total_qty
                harga_satuan = avg_harga_at / 1.11
            else:
                harga_satuan = 0.0

            dpp = harga_satuan * total_qty
            dpp_nilai_lain = (11.0 / 12.0) * dpp
            ppn = 0.12 * dpp_nilai_lain

            detail = {
                "baris": baris_idx,
                "barang_jasa": "A",
                "kode_barang": "640000",
                "nama_barang": art_name,
                "satuan": "UM.0019",
                "harga_satuan": harga_satuan,
                "qty": int(total_qty),
                "total_diskon": total_diskon,
                "dpp": dpp,
                "dpp_nilai_lain": dpp_nilai_lain,
                "tarif_ppn": 12,
                "ppn": ppn,
                "tarif_ppnbm": 0,
                "ppnbm": 0,
            }
            detail_rows.append(detail)

    return faktur_rows, detail_rows


def _lookup_customer(
    item: LineItem, customers: Dict[str, CustomerInfo]
) -> Optional[CustomerInfo]:
    """Lookup customer info from Master Pelanggan, trying multiple keys."""
    # Try exact name match
    key = item.customer_name.upper()
    if key in customers:
        return customers[key]

    # Try stripping CV/PT prefix
    stripped = re.sub(r"^(CV\.?\s*|PT\.?\s*)", "", key, flags=re.IGNORECASE).strip()
    if stripped in customers:
        return customers[stripped]

    # Try customer ID
    if item.customer_id and item.customer_id.upper() in customers:
        return customers[item.customer_id.upper()]

    # Try partial match
    for cust_key, cust_info in customers.items():
        if stripped and stripped in cust_key:
            return cust_info
        if cust_key in key:
            return cust_info

    return None


# ---------------------------------------------------------------------------
# Core: Write XLSX
# ---------------------------------------------------------------------------


def write_coretax_xlsx(
    faktur_rows: List[Dict[str, Any]],
    detail_rows: List[Dict[str, Any]],
    entity: str,
    output_path: str,
) -> str:
    """Write Coretax DJP template XLSX with Faktur + DetailFaktur sheets."""
    npwp = ENTITY_NPWP.get(entity, "")
    id_tku = npwp + "000000" if npwp else ""

    wb = openpyxl.Workbook()

    # --- Faktur sheet ---
    ws_f = wb.active
    ws_f.title = "Faktur"

    # Row 1: NPWP Penjual header
    ws_f.cell(1, 1, "NPWP Penjual")
    ws_f.cell(1, 3, npwp)

    # Row 2: blank
    # Row 3: column headers
    faktur_headers = [
        "Baris",
        "Tanggal Faktur",
        "Jenis Faktur",
        "Kode Transaksi",
        "Keterangan Tambahan",
        "Dokumen Pendukung",
        "Period Dok Pendukung",
        "Referensi",
        "Cap Fasilitas",
        "ID TKU Penjual",
        "NPWP/NIK Pembeli",
        "Jenis ID Pembeli",
        "Negara Pembeli",
        "Nomor Dokumen Pembeli",
        "Nama Pembeli",
        "Alamat Pembeli",
        "Email Pembeli",
        "ID TKU Pembeli",
    ]
    for col_idx, header in enumerate(faktur_headers, start=1):
        ws_f.cell(3, col_idx, header)

    # Row 4+: data
    for row_idx, frow in enumerate(faktur_rows, start=4):
        ws_f.cell(row_idx, 1, frow["baris"])
        ws_f.cell(row_idx, 2, frow["tanggal"])
        ws_f.cell(row_idx, 3, frow["jenis_faktur"])
        ws_f.cell(row_idx, 4, frow["kode_transaksi"])
        ws_f.cell(row_idx, 5, frow["keterangan_tambahan"])
        ws_f.cell(row_idx, 6, frow["dokumen_pendukung"])
        ws_f.cell(row_idx, 7, frow["period_dok"])
        ws_f.cell(row_idx, 8, frow["referensi"])
        ws_f.cell(row_idx, 9, frow["cap_fasilitas"])
        ws_f.cell(row_idx, 10, id_tku)
        ws_f.cell(row_idx, 11, frow["npwp_pembeli"])
        ws_f.cell(row_idx, 12, frow["jenis_id_pembeli"])
        ws_f.cell(row_idx, 13, frow["negara_pembeli"])
        ws_f.cell(row_idx, 14, frow["nomor_dok_pembeli"])
        ws_f.cell(row_idx, 15, frow["nama_pembeli"])
        ws_f.cell(row_idx, 16, frow["alamat_pembeli"])
        ws_f.cell(row_idx, 17, frow["email_pembeli"])
        ws_f.cell(row_idx, 18, frow["id_tku_pembeli"])

    # --- DetailFaktur sheet ---
    ws_d = wb.create_sheet("DetailFaktur")

    detail_headers = [
        "Baris",
        "Barang/Jasa",
        "Kode Barang Jasa",
        "Nama Barang/Jasa",
        "Nama Satuan Ukur",
        "Harga Satuan",
        "Jumlah Barang Jasa",
        "Total Diskon",
        "DPP",
        "DPP Nilai Lain",
        "Tarif PPN",
        "PPN",
        "Tarif PPnBM",
        "PPnBM",
    ]
    for col_idx, header in enumerate(detail_headers, start=1):
        ws_d.cell(1, col_idx, header)

    # Data rows
    for row_idx, drow in enumerate(detail_rows, start=2):
        ws_d.cell(row_idx, 1, drow["baris"])
        ws_d.cell(row_idx, 2, drow["barang_jasa"])
        ws_d.cell(row_idx, 3, drow["kode_barang"])
        ws_d.cell(row_idx, 4, drow["nama_barang"])
        ws_d.cell(row_idx, 5, drow["satuan"])
        ws_d.cell(row_idx, 6, drow["harga_satuan"])
        ws_d.cell(row_idx, 7, drow["qty"])
        ws_d.cell(row_idx, 8, drow["total_diskon"])
        ws_d.cell(row_idx, 9, drow["dpp"])
        ws_d.cell(row_idx, 10, drow["dpp_nilai_lain"])
        ws_d.cell(row_idx, 11, drow["tarif_ppn"])
        ws_d.cell(row_idx, 12, drow["ppn"])
        ws_d.cell(row_idx, 13, drow["tarif_ppnbm"])
        ws_d.cell(row_idx, 14, drow["ppnbm"])

    # Save
    wb.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def main():
    parser = argparse.ArgumentParser(
        description="Generate Coretax DJP XLSX from Register Penjualan"
    )
    parser.add_argument("register", help="Path to Register Penjualan XLSX")
    parser.add_argument(
        "--entity",
        default="DDD",
        choices=list(ENTITY_NPWP.keys()),
        help="Selling entity (default: DDD)",
    )
    parser.add_argument(
        "--output", default=None, help="Output directory (default: same as input)"
    )
    parser.add_argument(
        "--customers",
        nargs="+",
        default=None,
        help="Customer name filter (default: MAKMUR BESAR BERSAMA)",
    )
    parser.add_argument(
        "--all-nonretail", action="store_true", help="Include ALL non-retail customers"
    )
    parser.add_argument(
        "--month-sheet",
        default=None,
        help="Explicit month sheet name (default: auto-detect)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse and show summary without generating output",
    )
    args = parser.parse_args()

    register_path = os.path.expanduser(args.register)
    if not os.path.isfile(register_path):
        print(f"ERROR: File not found: {register_path}", file=sys.stderr)
        sys.exit(1)

    # Load workbook
    print(f"Loading: {register_path}")
    wb = openpyxl.load_workbook(register_path, data_only=True)
    print(f"Sheets: {wb.sheetnames}")

    # Detect month sheet
    month_sheet = args.month_sheet or detect_month_sheet(wb)
    if not month_sheet or month_sheet not in wb.sheetnames:
        print(
            f"ERROR: Could not detect month sheet. Available: {wb.sheetnames}",
            file=sys.stderr,
        )
        print("Use --month-sheet to specify explicitly.", file=sys.stderr)
        sys.exit(1)
    print(f"Month sheet: {month_sheet}")

    # Read Master Pelanggan
    if "Master Pelanggan" not in wb.sheetnames:
        print("ERROR: 'Master Pelanggan' sheet not found in workbook.", file=sys.stderr)
        sys.exit(1)
    customers = read_master_pelanggan(wb)
    print(f"Master Pelanggan: {len(customers)} entries loaded")

    # Determine customer filter
    customer_filter = args.customers or DEFAULT_CUSTOMERS

    # Read line items
    print(f"Reading line items from '{month_sheet}'...")
    if args.all_nonretail:
        print("  Filter: ALL non-retail customers")
    else:
        print(f"  Filter: {customer_filter}")

    items = read_line_items(
        wb,
        month_sheet,
        customer_filter=customer_filter,
        all_nonretail=args.all_nonretail,
    )
    print(f"  Found: {len(items)} line items")

    if not items:
        print(
            "WARNING: No matching line items found. Check customer filter.",
            file=sys.stderr,
        )
        wb.close()
        sys.exit(0)

    # Aggregate
    faktur_rows, detail_rows = aggregate_invoices(items, customers)
    print(f"\nAggregated:")
    print(f"  Faktur rows (invoices): {len(faktur_rows)}")
    print(f"  DetailFaktur rows (articles): {len(detail_rows)}")

    # Summary
    total_dpp_nl = sum(d["dpp_nilai_lain"] for d in detail_rows)
    total_ppn = sum(d["ppn"] for d in detail_rows)
    print(f"  Total DPP Nilai Lain: Rp {total_dpp_nl:,.2f}")
    print(f"  Total PPN 12%: Rp {total_ppn:,.2f}")

    if args.dry_run:
        print("\n[DRY RUN] No output file generated.")
        # Show first 5 faktur rows
        print("\nSample Faktur rows:")
        for f in faktur_rows[:5]:
            print(
                f"  Baris {f['baris']}: {f['tanggal']} | {f['referensi']} | {f['nama_pembeli']}"
            )
        print("\nSample DetailFaktur rows:")
        for d in detail_rows[:5]:
            print(
                f"  Baris {d['baris']}: {d['nama_barang']} | qty={d['qty']} | DPP NL={d['dpp_nilai_lain']:,.2f}"
            )
        wb.close()
        sys.exit(0)

    # Generate output filename
    # Extract month/year from first invoice date
    first_date = items[0].date
    month_name = first_date.strftime("%B")
    year = first_date.year

    output_dir = args.output or os.path.dirname(register_path)
    output_dir = os.path.expanduser(output_dir)
    os.makedirs(output_dir, exist_ok=True)

    output_filename = f"Coretax_Faktur_{args.entity}_{month_name}_{year}.xlsx"
    output_path = os.path.join(output_dir, output_filename)

    # Write
    write_coretax_xlsx(faktur_rows, detail_rows, args.entity, output_path)
    print(f"\n✅ Output: {output_path}")
    print(f"   Faktur: {len(faktur_rows)} invoices")
    print(f"   DetailFaktur: {len(detail_rows)} line items")

    wb.close()


if __name__ == "__main__":
    main()
