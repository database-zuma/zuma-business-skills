"""
Step 2: Planogram Visualization — Zuma Royal Plaza
Reads PLANOGRAM_Royal_Plaza.xlsx (Step 1 output) and generates:
  Option A: Excel floor plan (VISUAL_PLANOGRAM_Royal_Plaza.xlsx)
  Option B: ASCII floor plan  (VISUAL_PLANOGRAM_Royal_Plaza.txt)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import re
from datetime import datetime

# ============================================================
# CONFIG
# ============================================================
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_FILE = os.path.join(SCRIPT_DIR, "..", "PLANOGRAM_Royal_Plaza.xlsx")
EXCEL_OUTPUT = os.path.join(SCRIPT_DIR, "VISUAL_PLANOGRAM_Royal_Plaza.xlsx")
ASCII_OUTPUT = os.path.join(SCRIPT_DIR, "VISUAL_PLANOGRAM_Royal_Plaza.txt")

STORE_NAME = "Zuma Royal Plaza"
DATE_STR = datetime.now().strftime("%d %b %Y")

# Series -> hex color (without #, for openpyxl)
SERIES_COLORS = {
    "ZORRO": "EAD1DC",
    "DALLAS": "E6B8AF",
    "SLIDE": "FFF2CC",
    "SLIDE BASIC": "FFF2CC",
    "SLIDE MAX": "FFF2CC",
    "SLIDE PUFFY": "FFF2CC",
    "FLO": "F9CB9C",
    "PUFFY": "A4C2F4",
    "ELSA": "CCCCCC",
    "CLASSIC": "FFF2CC",
    "BLACKSERIES": "D9D9D9",
    "BLACK SERIES": "D9D9D9",
    "STRIPE": "FFE0B2",
    "ONYX": "CFD8DC",
    "ONYX Z": "CFD8DC",
    "MERCI": "D9D2E9",
    "STITCH": "B4E7CE",
    "WEDGES": "D7CCC8",
    "WBB": "80DEEA",
    "VELCRO": "E8F5E9",
    "VELCRO BRICKS": "A5D6A7",
    "VELCRO STITCH": "B4E7CE",
    "POOH": "FFE082",
    "LOTSO": "F48FB1",
    "BOYS COLLAB": "CE93D8",
    "GIRLS COLLAB": "F8BBD0",
    "DISNEY": "FFAB91",
    "DISNEY MINNIE": "FFAB91",
    "DISNEY MICKEY": "FFAB91",
    "BATMAN": "90A4AE",
    "MICKEY": "FFCC80",
    "MICKEY & FRIENDS": "FFCC80",
    "BABY CLASSIC": "80DEEA",
    "BABY KARAKTER": "B7E1CD",
    "BABY VELCRO": "E06666",
    "KIDS CLASSIC": "B6D7A8",
    "KIDS KARAKTER": "B6D7A8",
    "TOY STORY": "81D4FA",
    "BOYS TOY STORY": "81D4FA",
    "(KOSONG)": "F5F5F5",
}

# Series abbreviation for ASCII (max 6 chars)
SERIES_ABBREV = {
    "CLASSIC": "CLS",
    "BLACKSERIES": "BLK",
    "BLACK SERIES": "BLK",
    "STRIPE": "STP",
    "ZORRO": "ZRO",
    "DALLAS": "DAL",
    "SLIDE": "SLD",
    "SLIDE BASIC": "SLB",
    "SLIDE MAX": "SLM",
    "SLIDE PUFFY": "SLP",
    "FLO": "FLO",
    "PUFFY": "PFY",
    "ELSA": "ELS",
    "MERCI": "MRC",
    "STITCH": "STH",
    "WEDGES": "WDG",
    "ONYX": "ONX",
    "ONYX Z": "ONX",
    "WBB": "WBB",
    "VELCRO": "VLC",
    "VELCRO BRICKS": "VLB",
    "VELCRO STITCH": "VLS",
    "POOH": "POO",
    "LOTSO": "LOT",
    "BOYS COLLAB": "BCL",
    "GIRLS COLLAB": "GCL",
    "DISNEY": "DSN",
    "DISNEY MINNIE": "DMN",
    "DISNEY MICKEY": "DMK",
    "BATMAN": "BAT",
    "MICKEY": "MCK",
    "MICKEY & FRIENDS": "M&F",
    "TOY STORY": "TOY",
    "BOYS TOY STORY": "BTS",
    "(KOSONG)": "---",
    "AVAILABLE": "---",
}


def get_series_color_hex(series_name):
    """Get color hex (no #) for a series. Fallback to auto-generated pastel."""
    s = series_name.upper().strip()
    if s in SERIES_COLORS:
        return SERIES_COLORS[s]
    # Try partial match
    for key, color in SERIES_COLORS.items():
        if key in s or s in key:
            return color
    # Auto-generate pastel from hash
    import hashlib

    h = int(hashlib.md5(s.encode()).hexdigest()[:6], 16)
    r = 180 + (h % 75)
    g = 180 + ((h >> 8) % 75)
    b = 180 + ((h >> 16) % 75)
    return f"{r:02x}{g:02x}{b:02x}"


def get_series_abbrev(series_name):
    """Get 3-char abbreviation for ASCII view."""
    s = series_name.upper().strip()
    if s in SERIES_ABBREV:
        return SERIES_ABBREV[s]
    # Try partial match
    for key, abbr in SERIES_ABBREV.items():
        if key in s or s in key:
            return abbr
    # Fallback: first 3 chars
    return s[:3] if len(s) >= 3 else s


# ============================================================
# PARSE PLANOGRAM XLSX
# ============================================================
def parse_planogram_xlsx(filepath):
    """
    Parse PLANOGRAM_Royal_Plaza.xlsx -> dict with backwall data + rak baby.
    Returns:
      {
        "backwalls": {
          "BW-1": {"gender_type": ..., "hooks": ..., "hpa": ..., "slots": ...,
                    "articles": [{"name": ..., "series": ..., "tier": ..., "avg": ...}, ...]},
          ...
        },
        "rak_baby": [{"name": ..., "series": ..., "tier": ..., "avg": ...}, ...],
        "summary": {"total_slots": ..., "total_used": ..., "util_pct": ...}
      }
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    data = {"backwalls": {}, "rak_baby": [], "summary": {}}

    for sheet_name in wb.sheetnames:
        if sheet_name.startswith("BW-"):
            bw_id = sheet_name.split(" ")[0]  # "BW-1", "BW-2", etc.
            bw_data = parse_backwall_sheet(wb[sheet_name], bw_id, sheet_name)
            data["backwalls"][bw_id] = bw_data
        elif sheet_name == "Rak Baby":
            data["rak_baby"] = parse_rak_baby_sheet(wb[sheet_name])

    # Summary from the summary sheet
    ws_sum = wb["Summary Report"] if "Summary Report" in wb.sheetnames else None
    if ws_sum:
        total_slots = 0
        total_used = 0
        for r in range(3, ws_sum.max_row + 1):
            label = str(ws_sum.cell(row=r, column=1).value or "")
            val = str(ws_sum.cell(row=r, column=3).value or "")
            if "hooks/slots available" in label:
                total_slots = int(val) if val.isdigit() else 0
            elif "hooks/slots used" in label:
                total_used = int(val) if val.isdigit() else 0
        util_pct = round(total_used / total_slots * 100, 1) if total_slots > 0 else 0
        data["summary"] = {
            "total_slots": total_slots,
            "total_used": total_used,
            "util_pct": util_pct,
        }
    else:
        # Compute from parsed data
        total_slots = sum(bw["slots"] for bw in data["backwalls"].values()) + 2
        total_used = sum(
            len(bw["articles"]) for bw in data["backwalls"].values()
        ) + len(data["rak_baby"])
        data["summary"] = {
            "total_slots": total_slots,
            "total_used": total_used,
            "util_pct": round(total_used / total_slots * 100, 1)
            if total_slots > 0
            else 0,
        }

    wb.close()
    return data


def parse_backwall_sheet(ws, bw_id, sheet_name):
    """Parse a single BW-N sheet."""
    # Extract gender_type from sheet name: "BW-1 Men Jepit" -> "Men Jepit"
    gender_type = " ".join(sheet_name.split(" ")[1:])

    # Row 1: title — extract hooks count
    title = str(ws.cell(row=1, column=1).value or "")
    hooks_match = re.search(r"(\d+)\s*Hooks", title)
    hooks = int(hooks_match.group(1)) if hooks_match else 0

    # Row 2: mode description — extract hooks_per_article
    mode_desc = str(ws.cell(row=2, column=1).value or "")
    hpa_match = re.search(r"(\d+)\s*hook\s*=\s*1\s*artikel", mode_desc)
    hpa = int(hpa_match.group(1)) if hpa_match else 2
    slots = hooks // hpa if hpa > 0 else 0

    # Parse articles by stepping through columns with stride = hpa
    articles = []
    col = 2  # articles start at column 2
    slot_idx = 0
    while slot_idx < slots:
        series = str(ws.cell(row=5, column=col).value or "").strip()
        article = str(ws.cell(row=6, column=col).value or "").strip()
        tier_str = str(ws.cell(row=7, column=col).value or "").strip()

        # Parse tier: "T1 | Avg: 6.3" -> tier=1, avg=6.3
        tier = ""
        avg = 0.0
        tier_match = re.match(r"T(\d+)", tier_str)
        if tier_match:
            tier = tier_match.group(1)
        avg_match = re.search(r"Avg:\s*([\d.]+)", tier_str)
        if avg_match:
            avg = float(avg_match.group(1))

        if article and article != "None" and article.upper() != "AVAILABLE":
            articles.append(
                {
                    "name": article,
                    "series": series if series and series != "(Kosong)" else "UNKNOWN",
                    "tier": tier,
                    "avg": avg,
                }
            )
        elif article.upper() == "AVAILABLE":
            articles.append(
                {
                    "name": "AVAILABLE",
                    "series": "(KOSONG)",
                    "tier": "",
                    "avg": 0,
                }
            )

        col += hpa
        slot_idx += 1

    return {
        "gender_type": gender_type,
        "hooks": hooks,
        "hpa": hpa,
        "slots": slots,
        "articles": articles,
    }


def parse_rak_baby_sheet(ws):
    """Parse Rak Baby sheet."""
    articles = []
    for r in range(5, ws.max_row + 1):
        layer = ws.cell(row=r, column=1).value
        article = ws.cell(row=r, column=2).value
        tier_raw = str(ws.cell(row=r, column=3).value or "")
        avg = ws.cell(row=r, column=4).value

        if layer and article and str(layer).startswith("Layer"):
            tier_match = re.match(r"T?(\d+)", tier_raw)
            tier = tier_match.group(1) if tier_match else ""
            articles.append(
                {
                    "name": str(article),
                    "series": extract_series_from_article(str(article)),
                    "tier": tier,
                    "avg": float(avg) if avg else 0,
                }
            )
    return articles


def extract_series_from_article(article_name):
    """Extract series name from article name like 'BABY WBB 3' -> 'WBB'."""
    name = article_name.upper().strip()
    # Remove gender prefix
    for prefix in ["MEN ", "LADIES ", "BABY ", "BOYS ", "GIRLS ", "JUNIOR "]:
        if name.startswith(prefix):
            name = name[len(prefix) :]
            break
    # The series is usually the first word(s) before the last number
    # e.g., "WBB 3" -> "WBB", "VELCRO STITCH 1" -> "VELCRO STITCH"
    parts = name.split()
    # Find last numeric part
    last_num_idx = -1
    for i in range(len(parts) - 1, -1, -1):
        if parts[i].isdigit():
            last_num_idx = i
            break
    if last_num_idx > 0:
        return " ".join(parts[:last_num_idx])
    return name


# ============================================================
# ROYAL PLAZA LAYOUT CONFIG
# ============================================================
# Grid coordinates for Excel visual (col, row) — 1-based
# Each display unit gets a rectangular zone in the grid

ROYAL_PLAZA_LAYOUT = {
    "store_name": "Zuma Royal Plaza",
    "grid_width": 90,  # total columns (wider for readable article names)
    "grid_height": 50,  # total rows
    "title_row": 1,
    "display_units": [
        # BW-3 Men Fashion — top-left horizontal
        # 18 articles × 2 cols each = 36 cols
        {
            "bw_id": "BW-3",
            "label": "BW-3 MEN FASHION",
            "dimension": "8X7",
            "orientation": "horizontal",
            "grid_col": 7,
            "grid_row": 5,  # series row (article row = 6, tier row = 7)
            "label_row": 3,
            "dim_row": 4,  # row for dimension label
        },
        # BW-2 Ladies Fashion — top-right horizontal
        # 18 articles × 2 cols each = 36 cols
        {
            "bw_id": "BW-2",
            "label": "BW-2 LADIES FASHION",
            "dimension": "8X7",
            "orientation": "horizontal",
            "grid_col": 47,
            "grid_row": 5,
            "label_row": 3,
            "dim_row": 4,
        },
        # BW-1 Men Jepit — left vertical
        # 4 cols wide (series, article×2, tier), up to 24 rows
        {
            "bw_id": "BW-1",
            "label": "BW-1 CLASSIC & BLACKSERIES",
            "dimension": "7X7",
            "orientation": "vertical",
            "grid_col": 2,
            "grid_row": 14,
            "label_row": 10,
            "dim_row": 11,
        },
        # BW-4 Baby & Kids — right vertical
        # 4 cols wide, up to 26 rows
        {
            "bw_id": "BW-4",
            "label": "BW-4 BABY & KIDS",
            "dimension": "7X7",
            "orientation": "vertical",
            "grid_col": 84,
            "grid_row": 14,
            "label_row": 10,
            "dim_row": 11,
        },
        # Rak Baby — center
        {
            "bw_id": "RAK_BABY",
            "label": "RAK BABY",
            "type": "rak_baby",
            "grid_col": 28,
            "grid_row": 24,
            "label_row": 23,
        },
        # KASIR — center-right
        {
            "bw_id": "KASIR",
            "label": "KASIR",
            "type": "landmark",
            "grid_col": 58,
            "grid_row": 18,
            "width": 6,
            "height": 3,
        },
        # AIRMOVE — lower center
        {
            "bw_id": "AIRMOVE",
            "label": "AIRMOVE",
            "type": "landmark",
            "grid_col": 36,
            "grid_row": 34,
            "width": 8,
            "height": 3,
        },
    ],
    "entrance": {
        "grid_col": 35,
        "grid_row": 46,
        "label": "ENTRANCE",
    },
}


# ============================================================
# OPTION A: EXCEL VISUAL
# ============================================================
def generate_excel_visual(planogram_data, layout, output_path):
    """Generate an Excel floor plan with colored grid cells showing article names."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Layout Visual"

    # Wider columns for readable article names
    for c in range(1, layout["grid_width"] + 5):
        ws.column_dimensions[get_column_letter(c)].width = 8
    for r in range(1, layout["grid_height"] + 10):
        ws.row_dimensions[r].height = 20

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    thick_border = Border(
        left=Side(style="medium"),
        right=Side(style="medium"),
        top=Side(style="medium"),
        bottom=Side(style="medium"),
    )

    # Title
    title_col = 20
    ws.merge_cells(
        start_row=1, start_column=title_col, end_row=1, end_column=title_col + 20
    )
    title_cell = ws.cell(
        row=1, column=title_col, value=f"LAYOUT {layout['store_name'].upper()}"
    )
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center")

    # Date subtitle
    ws.merge_cells(
        start_row=2, start_column=title_col, end_row=2, end_column=title_col + 20
    )
    date_cell = ws.cell(
        row=2,
        column=title_col,
        value=f"Planogram Period: Feb 2025 - Jan 2026 | Generated: {DATE_STR}",
    )
    date_cell.font = Font(size=9, italic=True)
    date_cell.alignment = Alignment(horizontal="center")

    # Draw each display unit
    for unit in layout["display_units"]:
        unit_type = unit.get("type", "backwall")

        if unit_type == "landmark":
            _excel_draw_landmark(ws, unit, thin_border, thick_border)
        elif unit_type == "rak_baby":
            _excel_draw_rak_baby(
                ws, unit, planogram_data["rak_baby"], thin_border, thick_border
            )
        else:
            bw_id = unit["bw_id"]
            bw_data = planogram_data["backwalls"].get(bw_id)
            if bw_data:
                _excel_draw_backwall(ws, unit, bw_data, thin_border, thick_border)

    # Entrance
    ent = layout["entrance"]
    ws.merge_cells(
        start_row=ent["grid_row"],
        start_column=ent["grid_col"],
        end_row=ent["grid_row"],
        end_column=ent["grid_col"] + 14,
    )
    ent_cell = ws.cell(
        row=ent["grid_row"], column=ent["grid_col"], value="=== ENTRANCE ==="
    )
    ent_cell.font = Font(bold=True, size=12)
    ent_cell.alignment = Alignment(horizontal="center")

    # Legend
    _excel_draw_legend(ws, planogram_data, layout)

    # Summary stats at bottom
    summary = planogram_data["summary"]
    sum_row = layout["grid_height"] + 2
    ws.cell(
        row=sum_row,
        column=2,
        value=f"Total: {summary['total_used']}/{summary['total_slots']} slots used ({summary['util_pct']}%)",
    )
    ws.cell(row=sum_row, column=2).font = Font(bold=True, size=10)

    wb.save(output_path)
    print(f"  Excel visual saved: {output_path}")


def _apply_fill_border(ws, row, col, fill, border):
    """Apply fill and border to a cell."""
    c = ws.cell(row=row, column=col)
    c.fill = fill
    c.border = border


def _excel_draw_backwall(ws, unit, bw_data, thin_border, thick_border):
    """Draw a single backwall on the Excel grid with article names, series, tier & dimension."""
    articles = bw_data["articles"]
    orientation = unit.get("orientation", "horizontal")
    start_col = unit["grid_col"]
    start_row = unit["grid_row"]
    label_row = unit.get("label_row", start_row - 2)
    dim_row = unit.get("dim_row", start_row - 1)
    dimension = unit.get("dimension", "")

    n = len(articles)
    if n == 0:
        return

    COLS_PER_ART = 2  # each article gets 2 merged columns

    if orientation == "horizontal":
        total_cols = n * COLS_PER_ART

        # Label row (backwall name)
        ws.merge_cells(
            start_row=label_row,
            start_column=start_col,
            end_row=label_row,
            end_column=start_col + total_cols - 1,
        )
        label_cell = ws.cell(row=label_row, column=start_col, value=unit["label"])
        label_cell.font = Font(bold=True, size=10)
        label_cell.alignment = Alignment(horizontal="center")

        # Dimension row
        ws.merge_cells(
            start_row=dim_row,
            start_column=start_col,
            end_row=dim_row,
            end_column=start_col + total_cols - 1,
        )
        dim_cell = ws.cell(
            row=dim_row,
            column=start_col,
            value=f"[ {dimension} ]  —  {n} articles  —  {bw_data['hooks']} hooks  —  {bw_data['hpa']} hooks/artikel",
        )
        dim_cell.font = Font(bold=True, size=9, color="333333")
        dim_cell.alignment = Alignment(horizontal="center")

        # 3 data rows per article: Series, Article Name, Tier & Avg
        for i, art in enumerate(articles):
            col = start_col + i * COLS_PER_ART
            series = art["series"]
            color_hex = get_series_color_hex(series)
            fill = PatternFill(
                start_color=color_hex, end_color=color_hex, fill_type="solid"
            )

            # Row 1: Series name
            ws.merge_cells(
                start_row=start_row,
                start_column=col,
                end_row=start_row,
                end_column=col + COLS_PER_ART - 1,
            )
            cell_s = ws.cell(row=start_row, column=col, value=series)
            cell_s.fill = fill
            cell_s.border = thin_border
            cell_s.alignment = Alignment(horizontal="center", vertical="center")
            cell_s.font = Font(size=8, italic=True)
            for mc in range(col, col + COLS_PER_ART):
                _apply_fill_border(ws, start_row, mc, fill, thin_border)

            # Row 2: Article name (FULL — the key fix)
            ws.merge_cells(
                start_row=start_row + 1,
                start_column=col,
                end_row=start_row + 1,
                end_column=col + COLS_PER_ART - 1,
            )
            cell_a = ws.cell(row=start_row + 1, column=col, value=art["name"])
            cell_a.fill = fill
            cell_a.border = thin_border
            cell_a.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell_a.font = Font(size=8, bold=True)
            for mc in range(col, col + COLS_PER_ART):
                _apply_fill_border(ws, start_row + 1, mc, fill, thin_border)

            # Row 3: Tier & Avg
            tier_label = f"T{art['tier']} | Avg: {art['avg']}" if art["tier"] else ""
            ws.merge_cells(
                start_row=start_row + 2,
                start_column=col,
                end_row=start_row + 2,
                end_column=col + COLS_PER_ART - 1,
            )
            cell_t = ws.cell(row=start_row + 2, column=col, value=tier_label)
            cell_t.fill = fill
            cell_t.border = thin_border
            cell_t.alignment = Alignment(horizontal="center", vertical="center")
            cell_t.font = Font(size=7)
            for mc in range(col, col + COLS_PER_ART):
                _apply_fill_border(ws, start_row + 2, mc, fill, thin_border)

    elif orientation == "vertical":
        VCOLS = 4  # series, article(×2 merged), tier

        # Label row
        ws.merge_cells(
            start_row=label_row,
            start_column=start_col,
            end_row=label_row,
            end_column=start_col + VCOLS - 1,
        )
        label_cell = ws.cell(row=label_row, column=start_col, value=unit["label"])
        label_cell.font = Font(bold=True, size=9)
        label_cell.alignment = Alignment(horizontal="center")

        # Dimension row
        ws.merge_cells(
            start_row=dim_row,
            start_column=start_col,
            end_row=dim_row,
            end_column=start_col + VCOLS - 1,
        )
        dim_cell = ws.cell(
            row=dim_row, column=start_col, value=f"[ {dimension} ] {n} articles"
        )
        dim_cell.font = Font(bold=True, size=8, color="333333")
        dim_cell.alignment = Alignment(horizontal="center")

        # Header row — write all cells BEFORE merging
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True, size=7)
        hr = start_row - 1
        for hi, hdr in enumerate(["Series", "Article", "", "Tier"]):
            hc = ws.cell(row=hr, column=start_col + hi, value=hdr)
            hc.fill = header_fill
            hc.font = header_font
            hc.border = thin_border
            hc.alignment = Alignment(horizontal="center")
        # Now merge article header across 2 cols
        ws.merge_cells(
            start_row=hr,
            start_column=start_col + 1,
            end_row=hr,
            end_column=start_col + 2,
        )

        # Articles top-to-bottom
        for i, art in enumerate(articles):
            row = start_row + i
            series = art["series"]
            color_hex = get_series_color_hex(series)
            fill = PatternFill(
                start_color=color_hex, end_color=color_hex, fill_type="solid"
            )

            # Col 1: Series
            cell = ws.cell(row=row, column=start_col, value=series)
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(size=7)

            # Col 2-3 merged: Article name (FULL)
            ws.merge_cells(
                start_row=row,
                start_column=start_col + 1,
                end_row=row,
                end_column=start_col + 2,
            )
            cell2 = ws.cell(row=row, column=start_col + 1, value=art["name"])
            cell2.fill = fill
            cell2.border = thin_border
            cell2.alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True
            )
            cell2.font = Font(size=8, bold=True)
            _apply_fill_border(ws, row, start_col + 2, fill, thin_border)

            # Col 4: Tier & Avg
            tier_label = f"T{art['tier']}|{art['avg']}" if art["tier"] else ""
            cell3 = ws.cell(row=row, column=start_col + 3, value=tier_label)
            cell3.fill = fill
            cell3.border = thin_border
            cell3.alignment = Alignment(horizontal="center", vertical="center")
            cell3.font = Font(size=7)


def _excel_draw_rak_baby(ws, unit, rak_articles, thin_border, thick_border):
    """Draw Rak Baby on Excel grid."""
    start_col = unit["grid_col"]
    start_row = unit["grid_row"]
    label_row = unit.get("label_row", start_row - 1)

    width = 6  # merged cols for readable text

    # Label
    ws.merge_cells(
        start_row=label_row,
        start_column=start_col,
        end_row=label_row,
        end_column=start_col + width - 1,
    )
    label_cell = ws.cell(row=label_row, column=start_col, value=unit["label"])
    label_cell.font = Font(bold=True, size=10)
    label_cell.alignment = Alignment(horizontal="center")

    for i, art in enumerate(rak_articles):
        row = start_row + i
        series = art["series"]
        color_hex = get_series_color_hex(series)
        fill = PatternFill(
            start_color=color_hex, end_color=color_hex, fill_type="solid"
        )

        ws.merge_cells(
            start_row=row,
            start_column=start_col,
            end_row=row,
            end_column=start_col + width - 1,
        )
        cell = ws.cell(
            row=row,
            column=start_col,
            value=f"Layer {i + 1}: {art['name']} ({series}) T{art['tier']}|Avg:{art['avg']}",
        )
        cell.fill = fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(size=9, bold=True)
        for mc in range(start_col, start_col + width):
            _apply_fill_border(ws, row, mc, fill, thin_border)


def _excel_draw_landmark(ws, unit, thin_border, thick_border):
    """Draw a landmark (KASIR, AIRMOVE, etc.) on Excel grid."""
    start_col = unit["grid_col"]
    start_row = unit["grid_row"]
    width = unit.get("width", 6)
    height = unit.get("height", 3)

    ws.merge_cells(
        start_row=start_row,
        start_column=start_col,
        end_row=start_row + height - 1,
        end_column=start_col + width - 1,
    )
    cell = ws.cell(row=start_row, column=start_col, value=unit["label"])
    cell.font = Font(bold=True, size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thick_border

    # Apply border to all cells in the merged area
    for r in range(start_row, start_row + height):
        for c in range(start_col, start_col + width):
            ws.cell(row=r, column=c).border = thick_border


def _excel_draw_legend(ws, planogram_data, layout):
    """Draw a legend section below the floor plan."""
    legend_row = layout["grid_height"] - 5
    legend_col = 2

    ws.cell(row=legend_row, column=legend_col, value="LEGEND (Series Colors):")
    ws.cell(row=legend_row, column=legend_col).font = Font(bold=True, size=10)

    # Collect unique series from all backwalls
    all_series = set()
    for bw_data in planogram_data["backwalls"].values():
        for art in bw_data["articles"]:
            if art["series"] and art["series"] != "(KOSONG)":
                all_series.add(art["series"])
    for art in planogram_data["rak_baby"]:
        if art["series"]:
            all_series.add(art["series"])

    sorted_series = sorted(all_series)
    row_offset = 1
    items_per_row = 6

    for i, series in enumerate(sorted_series):
        r = legend_row + row_offset + (i // items_per_row)
        c = legend_col + (i % items_per_row) * 4

        color_hex = get_series_color_hex(series)
        fill = PatternFill(
            start_color=color_hex, end_color=color_hex, fill_type="solid"
        )

        # Color swatch + name in 2 merged cols
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c + 1)
        cell = ws.cell(row=r, column=c, value=series)
        cell.fill = fill
        cell.font = Font(size=8, bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        _apply_fill_border(
            ws,
            r,
            c + 1,
            fill,
            Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            ),
        )


# ============================================================
# OPTION B: ASCII VISUAL
# ============================================================
def generate_ascii_visual(planogram_data, layout, output_path):
    """Generate an ASCII art floor plan."""
    lines = []
    W = 115  # total width

    # Header
    lines.append("")
    lines.append(center_text(f"LAYOUT {layout['store_name'].upper()}", W))
    lines.append(center_text("=" * 40, W))
    lines.append(
        center_text(f"Planogram Feb 2025 - Jan 2026 | Generated: {DATE_STR}", W)
    )
    lines.append("")

    # ─── TOP WALL: BW-3 Men Fashion + BW-2 Ladies Fashion ───
    bw3 = planogram_data["backwalls"].get("BW-3")
    bw2 = planogram_data["backwalls"].get("BW-2")

    bw3_block = (
        _ascii_horizontal_block("BW-3 MEN FASHION (8X7)", bw3, 50) if bw3 else []
    )
    bw2_block = (
        _ascii_horizontal_block("BW-2 LADIES FASHION (8X7)", bw2, 50) if bw2 else []
    )

    # Side by side with gap
    max_lines_top = max(len(bw3_block), len(bw2_block))
    for i in range(max_lines_top):
        left = bw3_block[i] if i < len(bw3_block) else " " * 50
        right = bw2_block[i] if i < len(bw2_block) else " " * 50
        lines.append(f"  {left}  {right}")

    lines.append("")

    # ─── MIDDLE: BW-1 (left) | Center (Rak Baby, Kasir) | BW-4 (right) ───
    bw1 = planogram_data["backwalls"].get("BW-1")
    bw4 = planogram_data["backwalls"].get("BW-4")
    rak = planogram_data["rak_baby"]

    bw1_block = _ascii_vertical_block("BW-1 MEN JEPIT (7X7)", bw1, 18) if bw1 else []
    bw4_block = _ascii_vertical_block("BW-4 BABY & KIDS (7X7)", bw4, 18) if bw4 else []

    # Center content
    center_block = _ascii_center_block(rak, 65)

    # Combine: left | center | right
    max_lines_mid = max(len(bw1_block), len(bw4_block), len(center_block))
    for i in range(max_lines_mid):
        left = bw1_block[i] if i < len(bw1_block) else " " * 18
        mid = center_block[i] if i < len(center_block) else " " * 65
        right = bw4_block[i] if i < len(bw4_block) else " " * 18
        lines.append(f"{left} {mid} {right}")

    lines.append("")

    # ─── ENTRANCE ───
    lines.append(center_text("=== ENTRANCE ===", W))
    lines.append("")

    # ─── LEGEND ───
    lines.append("  " + "-" * 80)
    lines.append("  LEGEND (Series Abbreviations):")
    lines.append("")

    # Collect all series used
    all_series = set()
    for bw_data in planogram_data["backwalls"].values():
        for art in bw_data["articles"]:
            if art["series"] and art["series"] != "(KOSONG)":
                all_series.add(art["series"])
    for art in planogram_data["rak_baby"]:
        if art["series"]:
            all_series.add(art["series"])

    sorted_series = sorted(all_series)
    legend_line = "  "
    for i, series in enumerate(sorted_series):
        abbr = get_series_abbrev(series)
        entry = f"  {abbr} = {series}"
        if len(legend_line) + len(entry) > W - 5:
            lines.append(legend_line)
            legend_line = "  "
        legend_line += entry
    if legend_line.strip():
        lines.append(legend_line)

    lines.append("")

    # ─── SUMMARY ───
    summary = planogram_data["summary"]
    lines.append("  " + "-" * 80)
    lines.append(
        f"  SUMMARY: {summary['total_used']}/{summary['total_slots']} slots used "
        f"({summary['util_pct']}% utilization)"
    )

    bw_summaries = []
    for bw_id in ["BW-1", "BW-3", "BW-2", "BW-4"]:
        bw = planogram_data["backwalls"].get(bw_id)
        if bw:
            used = len([a for a in bw["articles"] if a["name"] != "AVAILABLE"])
            bw_summaries.append(f"{bw_id} {bw['gender_type']}: {used}/{bw['slots']}")
    lines.append(f"  {' | '.join(bw_summaries)}")
    lines.append(f"  Rak Baby: {len(rak)}/2 layers")
    lines.append("  " + "-" * 80)
    lines.append("")

    output = "\n".join(lines)

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(output)

    print(f"  ASCII visual saved: {output_path}")
    return output


def center_text(text, width):
    """Center text within given width."""
    padding = max(0, (width - len(text)) // 2)
    return " " * padding + text


def _ascii_horizontal_block(title, bw_data, width):
    """
    Draw a horizontal backwall as ASCII block.
    Articles displayed left-to-right in rows of cells.
    """
    articles = bw_data["articles"]
    n = len(articles)
    if n == 0:
        return [f"  {title}", "  (empty)"]

    cell_w = 6  # chars per cell including border
    cols_per_row = min(n, (width - 2) // cell_w)
    if cols_per_row < 1:
        cols_per_row = 1

    lines = []
    lines.append(f"  {title}")

    # Process articles in chunks of cols_per_row
    for chunk_start in range(0, n, cols_per_row):
        chunk = articles[chunk_start : chunk_start + cols_per_row]
        actual_cols = len(chunk)

        # Top border
        top = "  +" + "+".join(["-----"] * actual_cols) + "+"
        lines.append(top)

        # Series row
        series_row = "  |"
        for art in chunk:
            abbr = get_series_abbrev(art["series"])
            series_row += f"{abbr:^5}|"
        lines.append(series_row)

        # Tier row
        tier_row = "  |"
        for art in chunk:
            t = f"T{art['tier']}" if art["tier"] else "   "
            tier_row += f"{t:^5}|"
        lines.append(tier_row)

        # Bottom border
        bot = "  +" + "+".join(["-----"] * actual_cols) + "+"
        lines.append(bot)

    return lines


def _ascii_vertical_block(title, bw_data, width):
    """
    Draw a vertical backwall as ASCII block.
    Articles displayed top-to-bottom, one per row.
    """
    articles = bw_data["articles"]
    n = len(articles)

    inner_w = width - 4  # inside the box borders
    lines = []

    # Title
    lines.append(f"+{'-' * (width - 2)}+")
    lines.append(f"|{title[: width - 2]:^{width - 2}}|")
    lines.append(f"+{'-' * (width - 2)}+")

    for art in articles:
        abbr = get_series_abbrev(art["series"])
        tier = f"T{art['tier']}" if art["tier"] else "  "
        # Format: "ABR T1" or "--- " for available
        if art["name"] == "AVAILABLE":
            label = "  (empty)   "
        else:
            label = f" {abbr} {tier}"
        lines.append(f"|{label:<{width - 2}}|")

    lines.append(f"+{'-' * (width - 2)}+")

    return lines


def _ascii_center_block(rak_articles, width):
    """
    Draw center area: Rak Baby, KASIR, AIRMOVE.
    Returns list of lines, each exactly `width` chars.
    """
    lines = []

    # Pad top to align with vertical backwalls
    for _ in range(2):
        lines.append(" " * width)

    # KASIR (right side of center)
    kasir_line = " " * 40 + "+--------+"
    lines.append(kasir_line.ljust(width))
    kasir_line2 = " " * 40 + "| KASIR  |"
    lines.append(kasir_line2.ljust(width))
    kasir_line3 = " " * 40 + "+--------+"
    lines.append(kasir_line3.ljust(width))

    lines.append(" " * width)
    lines.append(" " * width)

    # RAK BABY
    rak_w = 35
    rak_pad = 10
    rak_header = " " * rak_pad + "+" + "-" * (rak_w - 2) + "+"
    lines.append(rak_header.ljust(width))
    rak_title_text = "RAK BABY"
    rak_title = " " * rak_pad + "|" + f"{rak_title_text:^{rak_w - 2}}" + "|"
    lines.append(rak_title.ljust(width))

    rak_sep = " " * rak_pad + "+" + "-" * (rak_w - 2) + "+"
    lines.append(rak_sep.ljust(width))

    for i, art in enumerate(rak_articles):
        abbr = get_series_abbrev(art["series"])
        label = f"Layer {i + 1}: {art['name'][:18]} ({abbr})"
        rak_line = " " * rak_pad + "|" + f"{label:^{rak_w - 2}}" + "|"
        lines.append(rak_line.ljust(width))

    rak_bottom = " " * rak_pad + "+" + "-" * (rak_w - 2) + "+"
    lines.append(rak_bottom.ljust(width))

    # Spacer
    lines.append(" " * width)
    lines.append(" " * width)

    # AIRMOVE
    am_w = 25
    am_pad = 15
    am_top = " " * am_pad + "+" + "-" * (am_w - 2) + "+"
    lines.append(am_top.ljust(width))
    am_title_text = "AIRMOVE (Display)"
    am_title = " " * am_pad + "|" + f"{am_title_text:^{am_w - 2}}" + "|"
    lines.append(am_title.ljust(width))
    am_bot = " " * am_pad + "+" + "-" * (am_w - 2) + "+"
    lines.append(am_bot.ljust(width))

    # Pad bottom to match vertical wall height
    while len(lines) < 30:
        lines.append(" " * width)

    return lines


# ============================================================
# MAIN
# ============================================================
if __name__ == "__main__":
    print("=" * 60)
    print(f"PLANOGRAM VISUALIZATION: {STORE_NAME}")
    print("=" * 60)

    print(f"\n[1/3] Parsing planogram XLSX: {INPUT_FILE}")
    if not os.path.exists(INPUT_FILE):
        print(f"ERROR: Input file not found: {INPUT_FILE}")
        exit(1)

    data = parse_planogram_xlsx(INPUT_FILE)

    # Print parsed summary
    for bw_id, bw in data["backwalls"].items():
        used = len([a for a in bw["articles"] if a["name"] != "AVAILABLE"])
        print(f"  {bw_id} ({bw['gender_type']}): {used}/{bw['slots']} articles")
    print(f"  Rak Baby: {len(data['rak_baby'])} articles")
    print(
        f"  Summary: {data['summary']['total_used']}/{data['summary']['total_slots']} "
        f"({data['summary']['util_pct']}%)"
    )

    print(f"\n[2/3] Generating Excel visual...")
    generate_excel_visual(data, ROYAL_PLAZA_LAYOUT, EXCEL_OUTPUT)

    print(f"\n[3/3] Generating ASCII visual...")
    ascii_output = generate_ascii_visual(data, ROYAL_PLAZA_LAYOUT, ASCII_OUTPUT)

    print(f"\n{'=' * 60}")
    print("DONE!")
    print(f"  Excel: {EXCEL_OUTPUT}")
    print(f"  ASCII: {ASCII_OUTPUT}")
    print(f"{'=' * 60}")

    # Print the ASCII output for verification
    print("\n--- ASCII OUTPUT PREVIEW ---")
    print(ascii_output)
