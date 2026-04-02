#!/usr/bin/env python3
"""
Generate SOPB (Surat Order Pengiriman Barang) xlsx in Accurate import format.
Requires user input: SOPB number per entity + tanggal diminta.

Usage:
    python3 generate_sopb.py --store ROYAL --entity DDD \
        --sopb-number "SOPB/DDD/WHS/2026/IV/001" \
        --tanggal-diminta 2026-04-04 \
        [--date 2026-03-31]

    # Multiple entities:
    python3 generate_sopb.py --store ROYAL --entity DDD,LJBB \
        --sopb-number "SOPB/DDD/WHS/2026/IV/001,SOPB/LJBB/WHS/2026/IV/001" \
        --tanggal-diminta 2026-04-04

Format: Accurate "Permintaan Barang" import template
  Row 1: Header-level column names
  Row 2: Item-level column names
  Row 3: HEADER row (SOPB number, keterangan, "Kirim Barang")
  Row 4+: ITEM rows (kode_besar, nama, qty pairs, tanggal)
"""
import argparse
import os
import sys
from datetime import date, datetime

try:
    import psycopg2
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    os.system("pip install psycopg2-binary openpyxl 2>/dev/null")
    import psycopg2
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

DB_CONFIG = {
    "host": "76.13.194.120",
    "port": 5432,
    "dbname": "openclaw_ops",
    "user": "openclaw_app",
    "password": "Zuma-0psCl4w-2026!",
}

STORE_SHORT = {
    "ZUMA GALAXY MALL": "GM",
    "Zuma PTC": "PTC",
    "Zuma Tunjungan Plaza 3": "TP",
    "Zuma Royal Plaza": "ROYAL",
    "Zuma Lippo Sidoarjo": "SDA",
    "Zuma City Of Tomorrow Mall": "CITO",
    "Zuma Lippo Batu": "BATU",
    "Zuma Mall Olympic Garden": "MOG",
    "Zuma MATOS": "MATOS",
    "Zuma Icon Mall Gresik": "ICON",
    "Zuma Sunrise Mall Mojokerto": "SUNRISE",
}
SHORT_TO_FULL = {v: k for k, v in STORE_SHORT.items()}

# Accurate "Permintaan Barang" column headers
HEADER_LEVEL = [
    '', 'No Permintaan', 'Tgl Permintaan', 'Keterangan', 'Nama Cabang', 'Tipe Permintaan', 'Nama Gudang',
    *[f'Kustom Karakter {i+1}' for i in range(10)],
    'Kustom Angka 1', *[f'Kustom Angka {i+2}' for i in range(9)],
    'Kustom Tanggal 1', 'Kustom Tanggal 2',
]

ITEM_LEVEL = [
    '', 'Kode Barang', 'Nama Barang', 'Kuantitas', 'Satuan', 'Tgl Diminta', 'Catatan Barang',
    'Nama Dept Barang', 'No Proyek Barang',
    *[f'Kustom Karakter {i+1}' for i in range(15)],
    'Kustom Angka1',  # Accurate quirk: no space before 1
    *[f'Kustom Angka {i+2}' for i in range(9)],
    'Kustom Tanggal 1', 'Kustom Tanggal 2',
    *[f'Kategori Keuangan {i+1}' for i in range(10)],
    'Harga Estimasi',
]

HEADER_BG = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
HEADER_ROW_BG = PatternFill(start_color="93C47D", end_color="93C47D", fill_type="solid")
ITEM_BG = PatternFill(start_color="CFE2F3", end_color="CFE2F3", fill_type="solid")

OUTPUT_DIR = os.path.expanduser(
    "~/.paperclip/instances/default/workspaces/b1b9ff7c-0b4b-44cd-a012-21e77d02edc2/outbox"
)


def expand_kode_kecil_to_sizes(conn, kode_kecil, boxes):
    """Expand kode_kecil (box) to kode_besar per size using portal.kodemix.
    Returns list of (kode_besar, nama_variant, qty_pairs) tuples.
    qty_pairs = count_by_assortment × boxes
    """
    with conn.cursor() as cur:
        cur.execute("""
            SELECT DISTINCT kode_besar, nama_variant, COALESCE(count_by_assortment::int, 0) AS assortment
            FROM portal.kodemix
            WHERE UPPER(kode) = UPPER(%s)
              AND count_by_assortment IS NOT NULL AND count_by_assortment != ''
            ORDER BY kode_besar;
        """, (kode_kecil,))
        rows = cur.fetchall()

    if not rows:
        # Fallback: return single row with boxes * 12
        return [(kode_kecil, kode_kecil, boxes * 12)]

    return [(r[0], r[1], r[2] * boxes) for r in rows if r[2] * boxes > 0]


def fetch_ro_skus(conn, analysis_date, store_name, entity):
    """Fetch RO articles and expand to size-level using portal.kodemix (DB fallback)."""
    with conn.cursor() as cur:
        cur.execute("""
            SELECT kode_kecil, article_name, recomms_ro
            FROM public.pcp_ro_weekly_analysis
            WHERE analysis_date = %s AND store_name = %s AND ro_type = 'RO_BOX'
            ORDER BY kode_kecil;
        """, (analysis_date, store_name))
        articles = cur.fetchall()

    results = []
    for kode_kecil, article_name, boxes in articles:
        expanded = expand_kode_kecil_to_sizes(conn, kode_kecil, boxes)
        for kode_besar, nama_variant, qty_pairs in expanded:
            results.append((kode_kecil, article_name, boxes, kode_besar, nama_variant, qty_pairs))
    return results


def fetch_ro_skus_from_gsheet(conn, gsheet_id, entity, store_code=None):
    """Read qty from GSheet (Picking List or ROBOX), expand to size-level.
    Auto-detects format: tries Picking List first, falls back to ROBOX.
    """
    sys.path.insert(0, os.path.dirname(__file__))
    from read_gsheet import read_picking_list_actual, read_robox_actual_ro

    data = {}
    # Try Picking List format first (if store_code provided)
    if store_code:
        try:
            data = read_picking_list_actual(gsheet_id, store_code)
            if data:
                print(f"  Read {len(data)} articles from Picking List GSheet")
        except Exception:
            pass

    # Fallback to ROBOX format
    if not data:
        try:
            data = read_robox_actual_ro(gsheet_id)
            if data:
                print(f"  Read {len(data)} articles from ROBOX GSheet")
        except Exception:
            pass

    if not data:
        return []

    results = []
    for kode_kecil, info in data.items():
        boxes = info["qty"]
        if boxes <= 0:
            continue
        expanded = expand_kode_kecil_to_sizes(conn, kode_kecil, boxes)
        for kode_besar, nama_variant, qty_pairs in expanded:
            results.append((kode_kecil, info["artikel"], boxes, kode_besar, nama_variant, qty_pairs))
    return results


def format_date_ddmmyyyy(date_str):
    d = datetime.strptime(date_str, "%Y-%m-%d")
    return d.strftime("%d/%m/%Y")


def build_sopb_xlsx(store_short, store_full, entity, sopb_number, tanggal_diminta, skus, output_path):
    """Generate SOPB xlsx in Accurate import format."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{store_short}_{entity.upper()}"

    max_cols = max(len(HEADER_LEVEL), len(ITEM_LEVEL))

    # Row 1: Header-level column names
    for ci, h in enumerate(HEADER_LEVEL):
        cell = ws.cell(row=1, column=ci + 1, value=h)
        cell.font = Font(bold=True, size=10)
        cell.fill = HEADER_BG

    # Row 2: Item-level column names
    for ci, h in enumerate(ITEM_LEVEL):
        cell = ws.cell(row=2, column=ci + 1, value=h)
        cell.font = Font(bold=True, size=10)
        cell.fill = HEADER_BG

    # Calculate total pairs — skus already expanded to size-level, column 5 = qty_pairs
    total_pairs = sum(sku[5] for sku in skus)
    keterangan = f"Protol kirim ke {store_full} {total_pairs} pairs"
    formatted_date = format_date_ddmmyyyy(tanggal_diminta)

    # Row 3: HEADER data row
    header_data = [''] * max_cols
    header_data[0] = 'HEADER'
    header_data[1] = sopb_number
    header_data[2] = ''  # Tgl Permintaan (Accurate defaults)
    header_data[3] = keterangan
    header_data[4] = ''  # Nama Cabang
    header_data[5] = 'Kirim Barang'
    header_data[6] = ''  # Nama Gudang
    for ci, v in enumerate(header_data):
        cell = ws.cell(row=3, column=ci + 1, value=v)
        cell.font = Font(bold=True, size=10)
        cell.fill = HEADER_ROW_BG

    # Row 4+: ITEM rows — already expanded to size-level
    for ri, sku in enumerate(skus, 4):
        # sku: kode_kecil(0), article_name(1), boxes(2), kode_besar(3), nama_variant(4), qty_pairs(5)
        item_data = [''] * len(ITEM_LEVEL)
        item_data[0] = 'ITEM'
        item_data[1] = sku[3]             # kode_besar (size-level code)
        item_data[2] = sku[4]             # nama_variant (with size + color)
        item_data[3] = sku[5]             # qty_pairs (count_by_assortment × boxes)
        item_data[4] = ''                 # Satuan
        item_data[5] = formatted_date
        item_data[6] = ''                 # Catatan

        for ci, v in enumerate(item_data):
            cell = ws.cell(row=ri, column=ci + 1, value=v)
            cell.fill = ITEM_BG

    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18

    wb.save(output_path)
    return output_path


def main():
    parser = argparse.ArgumentParser(description="Generate SOPB xlsx (Accurate import format)")
    parser.add_argument("--store", required=True, help="Store code (e.g. ROYAL)")
    parser.add_argument("--entity", required=True, help="Entity: DDD,LJBB,MBB,UBB (comma-sep for multiple)")
    parser.add_argument("--sopb-number", required=True, help="SOPB number(s), comma-sep matching entities")
    parser.add_argument("--tanggal-diminta", required=True, help="Tanggal diminta (YYYY-MM-DD)")
    parser.add_argument("--date", default=str(date.today()), help="Analysis date (YYYY-MM-DD)")
    parser.add_argument("--gsheet-id", default=None, help="ROBOX GSheet ID to read Actual RO from")
    args = parser.parse_args()

    store_code = args.store.upper()
    store_full = SHORT_TO_FULL.get(store_code)
    if not store_full:
        print(f"Unknown store: {store_code}. Valid: {', '.join(sorted(SHORT_TO_FULL.keys()))}")
        sys.exit(1)

    entities = [e.strip().upper() for e in args.entity.split(",")]
    sopb_numbers = [s.strip() for s in args.sopb_number.split(",")]

    if len(entities) != len(sopb_numbers):
        print(f"Mismatch: {len(entities)} entities but {len(sopb_numbers)} SOPB numbers")
        sys.exit(1)

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    conn = psycopg2.connect(**DB_CONFIG)
    generated = []

    try:
        for entity, sopb_num in zip(entities, sopb_numbers):
            if entity not in ('DDD', 'LJBB', 'MBB', 'UBB'):
                print(f"Invalid entity: {entity}. Must be DDD/LJBB/MBB/UBB")
                continue

            if args.gsheet_id:
                print(f"  Reading from GSheet: {args.gsheet_id[:20]}...")
                skus = fetch_ro_skus_from_gsheet(conn, args.gsheet_id, entity, store_code=store_code)
            else:
                skus = fetch_ro_skus(conn, args.date, store_full, entity)
            if not skus:
                print(f"  {store_code}/{entity}: no RO data, skipping")
                continue

            filename = f"SOPB-{args.date.replace('-','')}-{store_code}-{entity}.xlsx"
            filepath = os.path.join(OUTPUT_DIR, filename)
            build_sopb_xlsx(store_code, store_full, entity, sopb_num, args.tanggal_diminta, skus, filepath)
            generated.append({"entity": entity, "file": filepath, "articles": len(skus)})
            print(f"  {store_code}/{entity}: {len(skus)} articles → {filename}")

        if generated:
            print(f"\nGenerated {len(generated)} SOPB file(s).")
            print(f"GENERATED_FILES={json.dumps([g['file'] for g in generated])}")
        else:
            print("No SOPB files generated.")

    finally:
        conn.close()


# Need json for output
import json

if __name__ == "__main__":
    main()
