#!/usr/bin/env python3
"""
Generate weekly RO analysis xlsx — one file per store.
Naming: ROBOX-YYYYMMDD-{STORE}-{SEQ:05d}.xlsx
Sequence tracked in public.pcp_ro_file_tracker table.

Usage:
    python3 generate_ro.py [--date YYYY-MM-DD] [--store STORE_SHORT]

    # Generate for all stores needing RO:
    python3 generate_ro.py

    # Generate for specific store:
    python3 generate_ro.py --store GM
"""
import argparse
import json
import os
import sys
from datetime import date

try:
    import psycopg2
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    os.system("pip install psycopg2-binary openpyxl 2>/dev/null")
    import psycopg2
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

DB_CONFIG = {
    "host": "76.13.194.120",
    "port": 5432,
    "dbname": "openclaw_ops",
    "user": "openclaw_app",
    "password": "Zuma-0psCl4w-2026!",
}

STORE_SHORT = {
    "ZUMA GALAXY MALL": "GM",
    "Zuma PTC": "PTC",
    "Zuma Tunjungan Plaza 3": "TP",
    "Zuma Royal Plaza": "ROYAL",
    "Zuma Lippo Sidoarjo": "SDA",
    "Zuma City Of Tomorrow Mall": "CITO",
    "Zuma Lippo Batu": "BATU",
    "Zuma Mall Olympic Garden": "MOG",
    "Zuma MATOS": "MATOS",
    "Zuma Icon Mall Gresik": "ICON",
    "Zuma Sunrise Mall Mojokerto": "SUNRISE",
}
SHORT_TO_FULL = {v: k for k, v in STORE_SHORT.items()}

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
RO_BOX_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
NO_STOCK_FILL = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
OUTPUT_DIR = os.path.expanduser("~/.paperclip/instances/default/workspaces/b1b9ff7c-0b4b-44cd-a012-21e77d02edc2/outbox")


def ensure_tracker_table(conn):
    """Create pcp_ro_file_tracker if not exists."""
    with conn.cursor() as cur:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS public.pcp_ro_file_tracker (
                id SERIAL PRIMARY KEY,
                seq INTEGER NOT NULL,
                store_short TEXT NOT NULL,
                store_name TEXT NOT NULL,
                analysis_date DATE NOT NULL,
                filename TEXT NOT NULL,
                filepath TEXT NOT NULL,
                gdrive_link TEXT,
                sent_at TIMESTAMPTZ,
                created_at TIMESTAMPTZ DEFAULT NOW()
            );
            CREATE INDEX IF NOT EXISTS idx_ro_tracker_seq ON public.pcp_ro_file_tracker(seq);
        """)
        conn.commit()


def get_next_seq(conn):
    """Get next global sequence number."""
    with conn.cursor() as cur:
        cur.execute("SELECT COALESCE(MAX(seq), 0) + 1 FROM public.pcp_ro_file_tracker;")
        return cur.fetchone()[0]


def record_file(conn, seq, store_short, store_name, analysis_date, filename, filepath):
    """Record generated file in tracker."""
    with conn.cursor() as cur:
        cur.execute("""
            INSERT INTO public.pcp_ro_file_tracker (seq, store_short, store_name, analysis_date, filename, filepath)
            VALUES (%s, %s, %s, %s, %s, %s)
            RETURNING id;
        """, (seq, store_short, store_name, analysis_date, filename, filepath))
        conn.commit()
        return cur.fetchone()[0]


def refresh_analysis(conn):
    """Run the daily analysis function."""
    with conn.cursor() as cur:
        cur.execute("SELECT * FROM public.refresh_pcp_ro_weekly();")
        row = cur.fetchone()
        conn.commit()
        return {"stores_analyzed": row[0], "articles_needing_ro": row[1]}


def fetch_stores_needing_ro(conn, analysis_date):
    """Get list of stores that have at least one RO_BOX recommendation."""
    with conn.cursor() as cur:
        cur.execute("""
            SELECT store_name,
                   SUM(CASE WHEN ro_type = 'RO_BOX' THEN 1 ELSE 0 END) AS ro_box,
                   SUM(recomms_ro) AS total_boxes
            FROM public.pcp_ro_weekly_analysis
            WHERE analysis_date = %s AND ro_type = 'RO_BOX'
            GROUP BY store_name
            HAVING SUM(CASE WHEN ro_type = 'RO_BOX' THEN 1 ELSE 0 END) > 0
            ORDER BY store_name;
        """, (analysis_date,))
        return cur.fetchall()


def fetch_store_data(conn, analysis_date, store_name):
    """Fetch RO data for a single store."""
    with conn.cursor() as cur:
        cur.execute("""
            SELECT kode_kecil, article_name, tier,
                   stock_whs, stock_ljbb, stock_total,
                   stock_onhand, planogram_box, recomms_ro
            FROM public.pcp_ro_weekly_analysis
            WHERE analysis_date = %s AND store_name = %s AND ro_type = 'RO_BOX'
            ORDER BY tier, kode_kecil;
        """, (analysis_date, store_name))
        return cur.fetchall()


def build_store_xlsx(store_name, store_short, rows, analysis_date, seq, output_path):
    """Build xlsx for a single store."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = store_short

    # Title
    headers = ["Kode Kecil", "Artikel", "Tier", "Stock WHS", "Stock LJBB",
               "Stock Total", "On-Hand", "Planogram", "Recomms RO", "Actual RO"]
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=f"RO Analysis — {store_name} — {analysis_date}")
    title_cell.font = Font(name="Calibri", size=12, bold=True)
    title_cell.alignment = Alignment(horizontal="center")

    # Info row
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    info_cell = ws.cell(row=2, column=1, value=f"File #{seq:05d} | Branch: Jatim | Generated: {date.today()}")
    info_cell.font = Font(name="Calibri", size=9, italic=True, color="666666")
    info_cell.alignment = Alignment(horizontal="center")

    # Headers
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    # Data — all rows are RO_BOX only
    ri = 5
    for row in rows:
        # kode_kecil(0), article_name(1), tier(2), stock_whs(3), stock_ljbb(4),
        # stock_total(5), stock_onhand(6), planogram_box(7), recomms_ro(8)
        # kode_kecil(0), article_name(1), tier(2), stock_whs(3), stock_ljbb(4),
        # stock_total(5), stock_onhand(6), planogram_box(7), recomms_ro(8)
        vals = [row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8], row[8]]
        #                                                                              ↑ Actual RO = copy of Recomms RO
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center" if ci >= 3 else "left")
            cell.fill = RO_BOX_FILL
            if ci in (4, 5, 6) and v == 0:
                cell.fill = NO_STOCK_FILL
            if ci == 10:  # Actual RO column — editable by user
                cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        ri += 1

    # Summary row
    ri += 1
    ws.cell(row=ri, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=ri, column=9, value=sum(r[8] for r in rows)).font = Font(bold=True)
    ws.cell(row=ri, column=9).alignment = Alignment(horizontal="center")
    ws.cell(row=ri, column=10, value=sum(r[8] for r in rows)).font = Font(bold=True)
    ws.cell(row=ri, column=10).alignment = Alignment(horizontal="center")

    # Column widths
    widths = [12, 38, 6, 12, 12, 12, 12, 12, 12, 12]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    wb.save(output_path)
    return output_path


def main():
    parser = argparse.ArgumentParser(description="Generate weekly RO xlsx per store")
    parser.add_argument("--date", type=str, default=str(date.today()))
    parser.add_argument("--store", type=str, default=None, help="Single store short code (e.g. GM, PTC)")
    parser.add_argument("--no-refresh", action="store_true", help="Skip refreshing analysis")
    args = parser.parse_args()

    analysis_date = args.date
    date_compact = analysis_date.replace("-", "")
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    conn = psycopg2.connect(**DB_CONFIG)

    try:
        ensure_tracker_table(conn)

        if not args.no_refresh:
            summary = refresh_analysis(conn)
            print(f"Analysis refreshed: {summary['stores_analyzed']} stores, {summary['articles_needing_ro']} articles needing RO")

        # Determine which stores to process
        if args.store:
            store_full = SHORT_TO_FULL.get(args.store.upper())
            if not store_full:
                print(f"Unknown store: {args.store}. Valid: {', '.join(sorted(SHORT_TO_FULL.keys()))}")
                sys.exit(1)
            stores_to_process = [(store_full,)]
        else:
            stores_to_process = fetch_stores_needing_ro(conn, analysis_date)

        if not stores_to_process:
            print(f"No stores need RO for {analysis_date}.")
            print("GENERATED_FILES=[]")
            sys.exit(0)

        generated = []
        for store_row in stores_to_process:
            store_name = store_row[0]
            store_short = STORE_SHORT.get(store_name, store_name[:10])
            rows = fetch_store_data(conn, analysis_date, store_name)

            if not rows:
                continue

            seq = get_next_seq(conn)
            filename = f"ROBOX-{date_compact}-{store_short}-{seq:05d}.xlsx"
            filepath = os.path.join(OUTPUT_DIR, filename)

            build_store_xlsx(store_name, store_short, rows, analysis_date, seq, filepath)
            record_file(conn, seq, store_short, store_name, analysis_date, filename, filepath)

            generated.append({
                "seq": seq,
                "store": store_short,
                "store_name": store_name,
                "filename": filename,
                "filepath": filepath,
                "articles": len(rows),
            })
            print(f"  [{seq:05d}] {store_short}: {len(rows)} articles → {filename}")

        print(f"\nGenerated {len(generated)} files.")
        print(f"GENERATED_FILES={json.dumps([g['filepath'] for g in generated])}")

    finally:
        conn.close()


if __name__ == "__main__":
    main()
