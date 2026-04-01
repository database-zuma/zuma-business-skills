#!/usr/bin/env python3
"""
Generate Picking List xlsx from approved RO data.
Input: approved GSheet or ro_daily_analysis table for specific stores.

Usage:
    python3 generate_picking_list.py --stores GM,PTC,ROYAL [--date 2026-03-31]
    python3 generate_picking_list.py --stores ROYAL

Format matches existing Picking List template:
  - 1 sheet per store
  - Header: PICKING LIST DELIVERY, Tanggal, Destination, Picker, Delivered by
  - Section 1: Supplies (Inner Box, Shopping Bag, etc) — optional
  - Section 2: Articles (kode_kecil, article_name, REQ qty by box, ACTUAL blank)
"""
import argparse
import json
import os
import sys
from datetime import date, datetime

try:
    import psycopg2
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    os.system("pip install psycopg2-binary openpyxl 2>/dev/null")
    import psycopg2
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

DB_CONFIG = {
    "host": "76.13.194.120",
    "port": 5432,
    "dbname": "openclaw_ops",
    "user": "openclaw_app",
    "password": "Zuma-0psCl4w-2026!",
}

STORE_SHORT = {
    "ZUMA GALAXY MALL": "GM",
    "Zuma PTC": "PTC",
    "Zuma Tunjungan Plaza 3": "TP",
    "Zuma Royal Plaza": "ROYAL",
    "Zuma Lippo Sidoarjo": "SDA",
    "Zuma City Of Tomorrow Mall": "CITO",
    "Zuma Lippo Batu": "BATU",
    "Zuma Mall Olympic Garden": "MOG",
    "Zuma MATOS": "MATOS",
    "Zuma Icon Mall Gresik": "ICON",
    "Zuma Sunrise Mall Mojokerto": "SUNRISE",
}
SHORT_TO_FULL = {v: k for k, v in STORE_SHORT.items()}

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=12, bold=True)
LABEL_FONT = Font(name="Calibri", size=10, bold=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

OUTPUT_DIR = os.path.expanduser(
    "~/.paperclip/instances/default/workspaces/b1b9ff7c-0b4b-44cd-a012-21e77d02edc2/outbox"
)


def fetch_ro_articles(conn, analysis_date, store_name):
    """Fetch RO_BOX articles for a store from ro_daily_analysis (fallback)."""
    with conn.cursor() as cur:
        cur.execute("""
            SELECT kode_kecil, article_name, tier, recomms_ro
            FROM public.ro_daily_analysis
            WHERE analysis_date = %s AND store_name = %s AND ro_type = 'RO_BOX'
            ORDER BY tier, kode_kecil;
        """, (analysis_date, store_name))
        return cur.fetchall()


def fetch_ro_from_gsheet(gsheet_id):
    """Read Actual RO column from ROBOX GSheet. Returns [(kode, artikel, '', qty), ...]."""
    sys.path.insert(0, os.path.dirname(__file__))
    from read_gsheet import read_robox_actual_ro
    data = read_robox_actual_ro(gsheet_id)
    return [(kode, info["artikel"], "", info["qty"]) for kode, info in data.items()]


def build_picking_list(stores_data, analysis_date, output_path):
    """Build picking list xlsx with 1 sheet per store."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    today_str = datetime.now().strftime("%d/%m/%Y")
    date_compact = analysis_date.replace("-", "")
    date_ddm = f"{date_compact[6:8]}{date_compact[4:6]}"

    for store_short, (store_full, articles) in stores_data.items():
        sheet_name = f"{store_short} {date_ddm}"
        ws = wb.create_sheet(title=sheet_name[:31])

        # --- Header block ---
        ws.merge_cells("A1:F1")
        ws["A1"] = "PICKING LIST DELIVERY"
        ws["A1"].font = TITLE_FONT
        ws["A1"].alignment = Alignment(horizontal="center")

        ws["A2"] = "Tanggal "
        ws["A2"].font = LABEL_FONT
        ws["B2"] = f": {today_str}"
        ws["D2"] = "Picker"
        ws["D2"].font = LABEL_FONT
        ws["E2"] = ": ___________"

        ws["A3"] = "Destination"
        ws["A3"].font = LABEL_FONT
        ws["B3"] = f": {store_short}"
        ws["D3"] = "Delivered by "
        ws["D3"].font = LABEL_FONT
        ws["E3"] = ": ___________"

        # --- Section 1: Supplies (empty rows for manual fill) ---
        row = 5
        supply_headers = ["NO", "CODE", "ARTIKEL", "ASSORTMEN", "REQ", "ACTUAL"]
        for ci, h in enumerate(supply_headers, 1):
            cell = ws.cell(row=row, column=ci, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

        # 3 blank supply rows
        for i in range(1, 4):
            ws.cell(row=row + i, column=1, value=i).border = THIN_BORDER
            for ci in range(2, 7):
                ws.cell(row=row + i, column=ci).border = THIN_BORDER

        # --- Section 2: Articles ---
        row = row + 5
        art_headers = ["No", "ARTIKEL", "NAMA ARTIKEL", "", "REQ", "ACTUAL"]
        for ci, h in enumerate(art_headers, 1):
            cell = ws.cell(row=row, column=ci, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

        for idx, art in enumerate(articles, 1):
            r = row + idx
            # kode_kecil(0), article_name(1), tier(2), recomms_ro(3)
            ws.cell(row=r, column=1, value=idx).border = THIN_BORDER
            ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")

            ws.cell(row=r, column=2, value=art[0]).border = THIN_BORDER  # kode_kecil

            ws.cell(row=r, column=3, value=art[1]).border = THIN_BORDER  # article_name

            ws.cell(row=r, column=4).border = THIN_BORDER  # empty

            ws.cell(row=r, column=5, value=art[3]).border = THIN_BORDER  # REQ (recomms_ro = boxes)
            ws.cell(row=r, column=5).alignment = Alignment(horizontal="center")

            ws.cell(row=r, column=6).border = THIN_BORDER  # ACTUAL (blank for picker to fill)
            ws.cell(row=r, column=6).alignment = Alignment(horizontal="center")

        # Total row
        total_row = row + len(articles) + 1
        ws.cell(row=total_row, column=1, value="TOTAL").font = LABEL_FONT
        ws.cell(row=total_row, column=5, value=sum(a[3] for a in articles)).font = LABEL_FONT
        ws.cell(row=total_row, column=5).alignment = Alignment(horizontal="center")
        ws.cell(row=total_row, column=5).border = THIN_BORDER
        ws.cell(row=total_row, column=6, value=0).alignment = Alignment(horizontal="center")
        ws.cell(row=total_row, column=6).border = THIN_BORDER

        # Signature block
        sig_row = total_row + 2
        sig_font = Font(name="Calibri", size=10)
        sig_bold = Font(name="Calibri", size=10, bold=True)

        ws.cell(row=sig_row, column=1, value="Picker").font = sig_bold
        ws.cell(row=sig_row, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=sig_row, column=3, value="Approved").font = sig_bold
        ws.cell(row=sig_row, column=3).alignment = Alignment(horizontal="center")
        ws.cell(row=sig_row, column=5, value="Adm. Warehouse").font = sig_bold
        ws.merge_cells(start_row=sig_row, start_column=5, end_row=sig_row, end_column=6)
        ws.cell(row=sig_row, column=5).alignment = Alignment(horizontal="center")

        # Dotted signature lines
        sig_line_row = sig_row + 2
        dots = "................................."
        ws.cell(row=sig_line_row, column=1, value=dots).font = sig_font
        ws.cell(row=sig_line_row, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=sig_line_row, column=3, value=dots).font = sig_font
        ws.cell(row=sig_line_row, column=3).alignment = Alignment(horizontal="center")
        ws.cell(row=sig_line_row, column=5, value=dots).font = sig_font
        ws.merge_cells(start_row=sig_line_row, start_column=5, end_row=sig_line_row, end_column=6)
        ws.cell(row=sig_line_row, column=5).alignment = Alignment(horizontal="center")

        # Note: *Diisi jika assortmen tidak sesuai
        note_row = sig_line_row + 2
        note_font = Font(name="Calibri", size=9, italic=True)
        ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=4)
        ws.cell(row=note_row, column=1, value="*Diisi jika assortmen tidak sesuai").font = note_font

        # Column widths
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 38
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 8
        ws.column_dimensions["F"].width = 8

    wb.save(output_path)
    return output_path


def main():
    parser = argparse.ArgumentParser(description="Generate Picking List for approved stores")
    parser.add_argument("--stores", required=True, help="Comma-separated store codes: GM,PTC,ROYAL")
    parser.add_argument("--date", default=str(date.today()), help="Analysis date (YYYY-MM-DD)")
    parser.add_argument("--gsheet-id", default=None, help="GSheet ID/URL of ROBOX to read Actual RO from (chain from Flow 1)")
    args = parser.parse_args()

    store_codes = [s.strip().upper() for s in args.stores.split(",")]
    analysis_date = args.date
    date_compact = analysis_date.replace("-", "")

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    conn = psycopg2.connect(**DB_CONFIG)
    stores_data = {}

    try:
        for code in store_codes:
            full = SHORT_TO_FULL.get(code)
            if not full:
                print(f"Unknown store: {code}. Valid: {', '.join(sorted(SHORT_TO_FULL.keys()))}")
                continue

            if args.gsheet_id:
                # Read from ROBOX GSheet (Actual RO column) — chain from Flow 1
                print(f"  {code}: reading from GSheet {args.gsheet_id[:20]}...")
                articles = fetch_ro_from_gsheet(args.gsheet_id)
            else:
                # Fallback: read from DB
                articles = fetch_ro_articles(conn, analysis_date, full)

            if not articles:
                print(f"  {code}: no articles, skipping")
                continue
            stores_data[code] = (full, articles)
            print(f"  {code}: {len(articles)} articles")

        if not stores_data:
            print("No stores with RO data. Nothing to generate.")
            sys.exit(0)

        filename = f"PickingList-{date_compact}-{'+'.join(stores_data.keys())}.xlsx"
        filepath = os.path.join(OUTPUT_DIR, filename)
        build_picking_list(stores_data, analysis_date, filepath)
        print(f"\nGenerated: {filepath}")
        print(f"OUTPUT_FILE={filepath}")

    finally:
        conn.close()


if __name__ == "__main__":
    main()
