#!/usr/bin/env python3
"""
SPG Achievement Recap Generator — Paperclip Agent
Queries iSeller sales + store targets, calculates per-SPG metrics,
generates leaderboard Excel with achievement %.
"""
import argparse
import os
import sys
from datetime import datetime
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import psycopg2

THIN = Border(left=Side(style="thin"), right=Side(style="thin"),
              top=Side(style="thin"), bottom=Side(style="thin"))
CELL_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
HEADER_FILL = PatternFill(start_color="002A3A", end_color="002A3A", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=9)

MONTH_NAMES = {1:"Januari",2:"Februari",3:"Maret",4:"April",5:"Mei",6:"Juni",
               7:"Juli",8:"Agustus",9:"September",10:"Oktober",11:"November",12:"Desember"}
MONTH_COLS = {1:"jan",2:"feb",3:"mar",4:"apr",5:"may",6:"jun",
              7:"jul",8:"aug",9:"sep",10:"oct",11:"nov",12:"dec"}

# Area → branch mapping for filtering
AREA_BRANCHES = {
    "jatim": ["jawa timur","jatim"],
    "jakarta": ["jakarta","jabodetabek"],
    "bali": ["bali"],
    "lombok": ["lombok","ntb"],
    "sumatera": ["sumatera","pekanbaru","medan"],
    "sulawesi": ["sulawesi","manado","makassar"],
    "batam": ["batam","kepri"],
}

def get_db():
    return psycopg2.connect(
        host=os.environ.get("PGHOST","76.13.194.120"),
        port=os.environ.get("PGPORT","5432"),
        dbname=os.environ.get("PGDATABASE","openclaw_ops"),
        user=os.environ.get("PGUSER","openclaw_app"),
        password=os.environ.get("PGPASSWORD",""),
    )

def count_workdays(year, month):
    """Count Mon-Sat in a month."""
    import calendar
    cal = calendar.monthcalendar(year, month)
    count = 0
    for week in cal:
        for i, day in enumerate(week):
            if day != 0 and i < 6:  # Mon=0 to Sat=5
                count += 1
    return count

def pull_spg_sales(area, month, year):
    """Pull SPG sales data from core.iseller for the given month."""
    conn = get_db()
    cur = conn.cursor()

    date_start = f"{year}-{month:02d}-01"
    if month == 12:
        date_end = f"{year+1}-01-01"
    else:
        date_end = f"{year}-{month+1:02d}-01"

    area_filter = ""
    if area.upper() != "ALL":
        # branch column uses exact names: Jatim, Jakarta, Bali, Lombok, etc.
        area_filter = f"AND LOWER(TRIM(branch)) = LOWER(TRIM('{area}'))"

    cur.execute(f"""
        SELECT TRIM(kasir) as spg,
               TRIM(toko) as toko,
               COALESCE(TRIM(branch), '') as branch,
               COALESCE(TRIM(area), '') as area,
               COUNT(DISTINCT nomor_pesanan) as trx,
               SUM(jumlah::numeric) as qty,
               SUM(subtotal_per_order::numeric) as sales
        FROM core.iseller
        WHERE tanggal_pesanan::date >= '{date_start}'
          AND tanggal_pesanan::date < '{date_end}'
          AND kasir IS NOT NULL AND TRIM(kasir) != ''
          AND status_pembayaran = 'paid'
          {area_filter}
        GROUP BY TRIM(kasir), TRIM(toko), TRIM(branch), TRIM(area)
        ORDER BY sales DESC
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()
    print(f"Pulled {len(rows)} SPG-store combinations")
    return rows

def pull_targets(year, month):
    """Pull store monthly targets."""
    conn = get_db()
    cur = conn.cursor()
    month_col = MONTH_COLS.get(month, "jan")
    cur.execute(f"""
        SELECT TRIM(store_name_norm) as store, TRIM(branch) as branch,
               COALESCE({month_col}::numeric, 0) as target
        FROM portal.store_monthly_target
        WHERE year = %s AND COALESCE({month_col}::numeric, 0) > 0
    """, (year,))
    rows = cur.fetchall()
    cur.close()
    conn.close()
    targets = {}
    for store, branch, target in rows:
        targets[store.lower() if store else ""] = {"branch": branch or "", "target": float(target)}
    print(f"Loaded {len(targets)} store targets for {month_col} {year}")
    return targets

def pull_spg_count_per_store(area, month, year):
    """Count distinct SPGs per store for target splitting."""
    conn = get_db()
    cur = conn.cursor()
    date_start = f"{year}-{month:02d}-01"
    date_end = f"{year}-{month+1:02d}-01" if month < 12 else f"{year+1}-01-01"

    cur.execute(f"""
        SELECT TRIM(toko) as toko, COUNT(DISTINCT TRIM(kasir)) as spg_count
        FROM core.iseller
        WHERE tanggal_pesanan::date >= '{date_start}' AND tanggal_pesanan::date < '{date_end}'
          AND kasir IS NOT NULL AND TRIM(kasir) != '' AND status_pembayaran = 'paid'
        GROUP BY TRIM(toko)
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return {toko.lower(): int(cnt) for toko, cnt in rows}

def match_store_to_target(toko, targets):
    """Fuzzy match toko name to target store."""
    toko_lower = toko.lower().strip()
    # Exact
    if toko_lower in targets: return targets[toko_lower]
    # Substring
    for key, val in targets.items():
        if key in toko_lower or toko_lower in key:
            return val
    # Strip "zuma " and compare
    stripped = toko_lower.replace("zuma ", "").replace("pameran ", "").strip()
    for key, val in targets.items():
        ks = key.replace("zuma ", "").strip()
        if ks == stripped or ks in stripped or stripped in ks:
            return val
    # Word overlap: if most words match, consider it a match
    toko_words = set(stripped.split()) - {"mall", "the", "di", "and"}
    for key, val in targets.items():
        ks = key.replace("zuma ", "").strip()
        key_words = set(ks.split()) - {"mall", "the", "di", "and"}
        overlap = toko_words & key_words
        if len(overlap) >= max(1, len(key_words) - 1):  # allow 1 word mismatch
            return val
    # Xchange/exchange normalization
    normalized = stripped.replace("xchange", "exchange").replace("x change", "exchange")
    for key, val in targets.items():
        ks = key.replace("zuma ", "").strip()
        if ks == normalized or ks in normalized or normalized in ks:
            return val
    return None

def generate_excel(spg_data, targets, spg_counts, area, month, year, output_path):
    """Generate SPG Achievement Recap Excel."""
    wb = openpyxl.Workbook()
    month_name = MONTH_NAMES.get(month, str(month))
    workdays = count_workdays(year, month)

    # ── Aggregate per SPG ──
    spg_agg = defaultdict(lambda: {"toko": "", "branch": "", "area": "", "trx": 0, "qty": 0, "sales": 0, "stores": set()})
    for spg, toko, branch, area_val, trx, qty, sales in spg_data:
        s = spg_agg[spg]
        s["trx"] += int(trx)
        s["qty"] += float(qty)
        s["sales"] += float(sales)
        s["stores"].add(toko)
        if not s["toko"] or float(sales) > 0:
            s["toko"] = toko
            s["branch"] = branch
            s["area"] = area_val

    # Find home store (store with most sales)
    for spg, data in spg_agg.items():
        # Re-check: find the store with highest sales for this SPG
        store_sales = defaultdict(float)
        for s, t, b, a, trx, qty, sales in spg_data:
            if s == spg:
                store_sales[t] += float(sales)
        if store_sales:
            data["toko"] = max(store_sales, key=store_sales.get)

    # Sort by sales desc
    ranked = sorted(spg_agg.items(), key=lambda x: -x[1]["sales"])

    # ── Sheet 1: SPG Leaderboard ──
    ws = wb.active
    ws.title = "SPG Leaderboard"

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 12
    for c in ['E','F','G','H','I','J','K','L','M','N']:
        ws.column_dimensions[c].width = 15

    ws.cell(row=1, column=1, value=f"SPG ACHIEVEMENT RECAP — {area.upper()} — {month_name} {year}").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"Generated by SPG-Recap-Agent — {datetime.now().strftime('%d %B %Y %H:%M')}").font = Font(size=9, italic=True, color="666666")
    ws.cell(row=3, column=1, value=f"Hari kerja (Sen-Sab): {workdays} hari").font = Font(size=9, color="888888")

    headers = ["Rank", "Nama SPG", "Toko (Home Store)", "Area",
               "Sales (Rp)", "Qty (prs)", "Trx",
               "ATV (Rp)", "ATU (prs)", "ARP (Rp)",
               "Target Toko", "Target SPG", "Achievement %"]
    for ci, h in enumerate(headers):
        cell = ws.cell(row=5, column=ci+1, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CELL_ALIGN
        cell.border = THIN

    r = 6
    total_sales = total_qty = total_trx = 0
    for rank, (spg, data) in enumerate(ranked, 1):
        sales = data["sales"]
        qty = data["qty"]
        trx = data["trx"]
        atv = sales / trx if trx > 0 else 0
        atu = qty / trx if trx > 0 else 0
        arp = sales / qty if qty > 0 else 0

        # Target
        target_info = match_store_to_target(data["toko"], targets)
        target_toko = target_info["target"] if target_info else 0
        spg_count = spg_counts.get(data["toko"].lower(), 1)
        target_spg = target_toko / spg_count if spg_count > 0 else 0
        achievement = (sales / target_spg * 100) if target_spg > 0 else 0

        ws.cell(row=r, column=1, value=rank).alignment = CELL_ALIGN; ws.cell(row=r, column=1).border = THIN
        ws.cell(row=r, column=2, value=spg).border = THIN; ws.cell(row=r, column=2).font = Font(size=9)
        ws.cell(row=r, column=3, value=data["toko"]).border = THIN; ws.cell(row=r, column=3).font = Font(size=9)
        ws.cell(row=r, column=4, value=data["area"] or data["branch"]).border = THIN; ws.cell(row=r, column=4).font = Font(size=9)

        for ci, val in enumerate([sales, qty, trx, atv, atu, arp, target_toko, target_spg], 5):
            cell = ws.cell(row=r, column=ci, value=round(val))
            cell.alignment = CELL_ALIGN; cell.border = THIN
            cell.number_format = '#,##0' if ci != 6 else '#,##0'

        # Achievement with color
        cell = ws.cell(row=r, column=13, value=round(achievement, 1))
        cell.alignment = CELL_ALIGN; cell.border = THIN
        cell.number_format = '0.0"%"'
        if achievement >= 100:
            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            cell.font = Font(bold=True, color="006100")
        elif achievement >= 80:
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        else:
            cell.fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
            cell.font = Font(color="CC0000")

        total_sales += sales; total_qty += qty; total_trx += trx
        r += 1

    # Totals row
    r += 1
    GT_FILL = PatternFill(start_color="002A3A", end_color="002A3A", fill_type="solid")
    GT_FONT = Font(bold=True, size=10, color="FFFFFF")
    for cc in range(1, 14):
        ws.cell(row=r, column=cc).border = THIN
        ws.cell(row=r, column=cc).fill = GT_FILL
    ws.cell(row=r, column=1, value="TOTAL").font = GT_FONT
    ws.cell(row=r, column=2, value=f"{len(ranked)} SPG").font = GT_FONT
    ws.cell(row=r, column=5, value=round(total_sales)).font = GT_FONT; ws.cell(row=r, column=5).number_format = '#,##0'
    ws.cell(row=r, column=6, value=round(total_qty)).font = GT_FONT
    ws.cell(row=r, column=7, value=total_trx).font = GT_FONT

    # ── Sheet 2: Double Store ──
    ws2 = wb.create_sheet("Double Store")
    ws2.cell(row=1, column=1, value=f"DOUBLE STORE — {area.upper()} — {month_name} {year}").font = Font(bold=True, size=14)
    ws2.cell(row=2, column=1, value="SPG yang transaksi di >1 toko (selain home store)").font = Font(size=9, italic=True, color="666666")

    ws2.column_dimensions['A'].width = 5
    ws2.column_dimensions['B'].width = 28
    ws2.column_dimensions['C'].width = 25
    ws2.column_dimensions['D'].width = 25
    ws2.column_dimensions['E'].width = 15
    ws2.column_dimensions['F'].width = 10

    ds_headers = ["No", "Nama SPG", "Home Store", "Other Store", "Sales (Rp)", "Trx"]
    for ci, h in enumerate(ds_headers):
        cell = ws2.cell(row=4, column=ci+1, value=h)
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL; cell.alignment = CELL_ALIGN; cell.border = THIN

    r2 = 5
    no = 1
    for spg, data in ranked:
        if len(data["stores"]) > 1:
            home = data["toko"]
            for s, t, b, a, trx, qty, sales in spg_data:
                if s == spg and t != home:
                    ws2.cell(row=r2, column=1, value=no).alignment = CELL_ALIGN; ws2.cell(row=r2, column=1).border = THIN
                    ws2.cell(row=r2, column=2, value=spg).border = THIN
                    ws2.cell(row=r2, column=3, value=home).border = THIN
                    ws2.cell(row=r2, column=4, value=t).border = THIN
                    ws2.cell(row=r2, column=5, value=round(float(sales))).border = THIN; ws2.cell(row=r2, column=5).number_format = '#,##0'
                    ws2.cell(row=r2, column=6, value=int(trx)).alignment = CELL_ALIGN; ws2.cell(row=r2, column=6).border = THIN
                    no += 1; r2 += 1

    if r2 == 5:
        ws2.cell(row=5, column=1, value="Tidak ada SPG double store bulan ini.").font = Font(italic=True, color="888888")

    wb.save(output_path)
    print(f"Excel saved: {output_path}")
    print(f"Total: {len(ranked)} SPG, {round(total_sales):,} sales, {round(total_qty):,} pairs, {total_trx:,} trx")
    return len(ranked)

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--area", default="ALL")
    parser.add_argument("--month", type=int, default=datetime.now().month)
    parser.add_argument("--year", type=int, default=datetime.now().year)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    print(f"=== SPG Recap: {args.area} — {MONTH_NAMES.get(args.month)} {args.year} ===")

    spg_data = pull_spg_sales(args.area, args.month, args.year)
    if not spg_data:
        print("ERROR: No SPG sales data found")
        sys.exit(1)

    targets = pull_targets(args.year, args.month)
    spg_counts = pull_spg_count_per_store(args.area, args.month, args.year)

    generate_excel(spg_data, targets, spg_counts, args.area, args.month, args.year, args.output)
    print("=== Done ===")

if __name__ == "__main__":
    main()
