#!/usr/bin/env python3
"""
=============================================================================
  RO REQUEST GENERATOR — Zuma Royal Plaza

  Generates Restock (RO Box / RO Protol) + Surplus Pull List
  based on planogram targets vs actual store stock.

  Data sources:
    - Planogram: RO Input Jatim.xlsx → Planogram sheet (Royal Plaza rows)
    - Stock:     openclaw_ops DB (core.stock_with_product)
    - Sales:     openclaw_ops DB (core.sales_with_product)
    - Warehouse: Warehouse Pusat Box (DDD+LJBB), Warehouse Pusat Protol (DDD)

  Output: RO_REQUEST_Royal_Plaza.xlsx

  Business Rules (from SKILL.md — zuma-distribution-flow):
    - RO Protol: when <50% sizes empty → send individual pairs from Gudang Protol
    - RO Box:    when >=50% sizes empty → send full box (12 pairs) from Gudang Box
    - Surplus:   when tier actual% > tier ideal% → pull lowest-selling articles
    - Royal Plaza: STORAGE = 0 → every RO Box creates surplus risk

  ⚠ TO METRIC NOTE:
    There is ambiguity between two TO definitions used in Zuma:
      (a) TO = stock_coverage = current_stock / monthly_sales (months of stock remaining)
          → High TO = slow mover, Low TO = fast seller
      (b) TO = turnover_rate = monthly_sales / current_stock (sales velocity)
          → Low TO = slow mover, High TO = fast seller
    The SKILL.md says "TO terendah = dead stock" which matches definition (b).
    The user described "TO 4.5 = article can hold 4.5 months" which matches definition (a).

    RESOLUTION: This script outputs BOTH metrics:
      - stock_coverage_months: stock / monthly_sales (user's definition)
      - turnover_rate: monthly_sales / stock (skill doc's definition)
    For surplus pull: sorts by avg_monthly_sales ASC (slowest sellers first).
    This is UNAMBIGUOUS regardless of which TO definition is adopted.

    → ASK TEAM: Which TO label does the Allocation Planner use? Then standardize.

  ⚠ IDEAL TIER CAPACITY % NOTE:
    No official per-store ideal tier % exists yet. This script derives the ideal
    from the planogram itself (count of articles per tier on display).
    → ASK TEAM: Define real per-store tier capacity targets.

=============================================================================
"""

import os
import sys
import io
from datetime import datetime, date
from collections import defaultdict

# Fix Windows console encoding for Unicode arrows/symbols
if sys.stdout.encoding != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

import psycopg2
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =============================================================================
# CONFIG
# =============================================================================

STORE_NAME = "Zuma Royal Plaza"
STORE_DB_PATTERN = (
    "zuma royal plaza"  # For ILIKE matching in DB (exact, not 'royal surf')
)
STORAGE_CAPACITY = 0  # Royal Plaza has NO storage

# Planogram source
PLANOGRAM_FILE = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "..", "RO Input Jatim.xlsx"
)
PLANOGRAM_SHEET = "Planogram"

# Output
OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "RO_REQUEST_Royal_Plaza.xlsx")

# DB Connection
DB_CONFIG = {
    "host": "76.13.194.120",
    "port": 5432,
    "dbname": "openclaw_ops",
    "user": "openclaw_app",
    "password": "Zuma-0psCl4w-2026!",
}

# Planogram Excel column indices (0-based)
COL_STORE = 0
COL_GENDER = 1
COL_SERIES = 2
COL_ARTICLE = 3
COL_TIER = 4
COL_KODE_MIX = 5
COL_SIZE_START = 6  # Column G
COL_SIZE_END = 33  # Column AH (inclusive)
COL_AVG_SALES_3M_PAIRS = 34
COL_AVG_SALES_3M_BOX = 35
COL_PCT_SALES_MIX = 36
COL_REKO_PAIRS = 37
COL_REKO_BOX = 38
COL_AVG_SALES_WEEK = 39
COL_TOTAL_PLANO = 40
COL_TIER_WEIGHT = 41
COL_SALES_WEIGHT = 42
COL_COMBINED_WEIGHT = 43

# Size column headers (will be read from Excel row 1)
SIZE_COL_HEADERS = {}  # Populated at runtime: {col_index: "size_label"}

# Surplus tiers to check (T1, T2, T3 only — T4/T5 excluded, T8 protected)
SURPLUS_CHECK_TIERS = [1, 2, 3]
SURPLUS_EXCLUDE_TIERS = [4, 5, 8]

# RO threshold
RO_BOX_THRESHOLD = 0.50  # >=50% sizes empty → RO Box

# Excel styling
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL_DARK = PatternFill(
    start_color="1B2A4A", end_color="1B2A4A", fill_type="solid"
)
HEADER_FILL_GREEN = PatternFill(
    start_color="0D6B3E", end_color="0D6B3E", fill_type="solid"
)
HEADER_FILL_RED = PatternFill(
    start_color="8B1A1A", end_color="8B1A1A", fill_type="solid"
)
HEADER_FILL_ORANGE = PatternFill(
    start_color="B35900", end_color="B35900", fill_type="solid"
)
HEADER_FILL_PURPLE = PatternFill(
    start_color="4A1B6D", end_color="4A1B6D", fill_type="solid"
)
HEADER_FILL_TEAL = PatternFill(
    start_color="0E4D5C", end_color="0E4D5C", fill_type="solid"
)
FILL_LIGHT_GREEN = PatternFill(
    start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"
)
FILL_LIGHT_RED = PatternFill(
    start_color="FFEBEE", end_color="FFEBEE", fill_type="solid"
)
FILL_LIGHT_YELLOW = PatternFill(
    start_color="FFF8E1", end_color="FFF8E1", fill_type="solid"
)
FILL_LIGHT_ORANGE = PatternFill(
    start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"
)
FILL_LIGHT_PURPLE = PatternFill(
    start_color="F3E5F5", end_color="F3E5F5", fill_type="solid"
)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


# =============================================================================
# 1. READ PLANOGRAM
# =============================================================================


def read_planogram():
    """
    Read planogram targets from RO Input Jatim.xlsx → Planogram sheet.
    Returns dict of displayed articles for Royal Plaza:
    {
        kode_mix: {
            'gender': str, 'series': str, 'article': str, 'tier': int,
            'kode_mix': str,
            'sizes': {size_label: target_pairs, ...},
            'total_planogram': float,
            'avg_sales_3m_pairs': float,
            'avg_sales_3m_box': float,
            'pct_sales_mix': float,
            'avg_sales_week': float,
        }
    }
    """
    global SIZE_COL_HEADERS

    print(f"[1/8] Reading planogram from: {PLANOGRAM_FILE}")
    wb = openpyxl.load_workbook(PLANOGRAM_FILE, read_only=True, data_only=True)
    ws = wb[PLANOGRAM_SHEET]

    articles = {}
    row_count = 0

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            # Read size column headers
            for col_idx in range(COL_SIZE_START, COL_SIZE_END + 1):
                label = str(row[col_idx]) if row[col_idx] is not None else ""
                SIZE_COL_HEADERS[col_idx] = label
            continue

        store = row[COL_STORE]
        if not store or "royal plaza" not in str(store).lower():
            continue

        kode_mix = row[COL_KODE_MIX]
        total_plano = row[COL_TOTAL_PLANO]

        if not kode_mix:
            continue

        # Only articles that are ON DISPLAY (Total Planogram > 0)
        if total_plano is None or total_plano <= 0:
            continue

        # Extract size targets
        sizes = {}
        for col_idx in range(COL_SIZE_START, COL_SIZE_END + 1):
            val = row[col_idx]
            if val is not None and val > 0:
                size_label = SIZE_COL_HEADERS[col_idx]
                sizes[size_label] = float(val)

        tier_raw = row[COL_TIER]
        tier = int(tier_raw) if tier_raw is not None else 0

        articles[kode_mix] = {
            "gender": row[COL_GENDER] or "",
            "series": row[COL_SERIES] or "",
            "article": row[COL_ARTICLE] or "",
            "tier": tier,
            "kode_mix": kode_mix,
            "sizes": sizes,
            "total_planogram": float(total_plano),
            "avg_sales_3m_pairs": float(row[COL_AVG_SALES_3M_PAIRS] or 0),
            "avg_sales_3m_box": float(row[COL_AVG_SALES_3M_BOX] or 0),
            "pct_sales_mix": float(row[COL_PCT_SALES_MIX] or 0),
            "avg_sales_week": float(row[COL_AVG_SALES_WEEK] or 0),
        }
        row_count += 1

    wb.close()
    print(f"       → {row_count} Royal Plaza articles on display loaded")

    # Print tier breakdown
    tier_counts = defaultdict(int)
    for a in articles.values():
        tier_counts[a["tier"]] += 1
    for t in sorted(tier_counts):
        print(f"         Tier {t}: {tier_counts[t]} articles")

    return articles


# =============================================================================
# 2. QUERY DATABASE
# =============================================================================


def query_db():
    """
    Pull all necessary data from openclaw_ops:
    - Royal Plaza actual stock per article per size
    - Warehouse Pusat Box stock (DDD + LJBB)
    - Warehouse Pusat Protol stock (DDD only)
    - Sales last 3 months at Royal Plaza (for TO / coverage calc)
    - Off-planogram stock (articles at store but not on planogram)
    """
    print(f"[2/8] Querying database...")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # --- 2a. Royal Plaza actual stock per kode_mix per size ---
    print("       → Royal Plaza actual stock...")
    cur.execute(
        """
        SELECT kode_mix, ukuran, SUM(quantity) as qty
        FROM core.stock_with_product
        WHERE LOWER(nama_gudang) = %s
          AND kode_mix IS NOT NULL
        GROUP BY kode_mix, ukuran
        ORDER BY kode_mix, ukuran
    """,
        (STORE_DB_PATTERN,),
    )

    store_stock = defaultdict(lambda: defaultdict(float))
    for kode_mix, ukuran, qty in cur.fetchall():
        store_stock[kode_mix][str(ukuran)] = float(qty) if qty else 0.0

    print(f"         {len(store_stock)} articles with stock data")

    # --- 2b. Warehouse Pusat Box stock (DDD + LJBB) per kode_mix per size ---
    print("       → Warehouse Pusat BOX stock (DDD + LJBB)...")
    cur.execute("""
        SELECT kode_mix, ukuran, SUM(quantity) as qty
        FROM core.stock_with_product
        WHERE LOWER(nama_gudang) = 'warehouse pusat'
          AND source_entity IN ('DDD', 'LJBB')
          AND kode_mix IS NOT NULL
        GROUP BY kode_mix, ukuran
        ORDER BY kode_mix, ukuran
    """)

    wh_box_stock = defaultdict(lambda: defaultdict(float))
    for kode_mix, ukuran, qty in cur.fetchall():
        wh_box_stock[kode_mix][str(ukuran)] = float(qty) if qty else 0.0

    print(f"         {len(wh_box_stock)} articles in WH Pusat Box")

    # --- 2c. Warehouse Pusat Protol stock (DDD only) per kode_mix per size ---
    print("       → Warehouse Pusat PROTOL stock (DDD only)...")
    cur.execute("""
        SELECT kode_mix, ukuran, SUM(quantity) as qty
        FROM core.stock_with_product
        WHERE LOWER(nama_gudang) = 'warehouse pusat protol'
          AND source_entity = 'DDD'
          AND kode_mix IS NOT NULL
        GROUP BY kode_mix, ukuran
        ORDER BY kode_mix, ukuran
    """)

    wh_protol_stock = defaultdict(lambda: defaultdict(float))
    for kode_mix, ukuran, qty in cur.fetchall():
        wh_protol_stock[kode_mix][str(ukuran)] = float(qty) if qty else 0.0

    print(f"         {len(wh_protol_stock)} articles in WH Pusat Protol")

    # --- 2d. Sales last 3 months at Royal Plaza (for TO calculation) ---
    print("       → Sales last 3 months at Royal Plaza...")
    cur.execute(
        """
        SELECT kode_mix, SUM(quantity) as total_pairs
        FROM core.sales_with_product
        WHERE LOWER(matched_store_name) = %s
          AND transaction_date >= CURRENT_DATE - INTERVAL '3 months'
          AND kode_mix IS NOT NULL
        GROUP BY kode_mix
        ORDER BY kode_mix
    """,
        (STORE_DB_PATTERN,),
    )

    sales_3m = {}
    for kode_mix, total_pairs in cur.fetchall():
        sales_3m[kode_mix] = float(total_pairs) if total_pairs else 0.0

    print(f"         {len(sales_3m)} articles with sales data")

    # --- 2e. All articles at Royal Plaza (for off-planogram detection) ---
    print("       → Full article list at Royal Plaza (for off-planogram check)...")
    cur.execute(
        """
        SELECT kode_mix, article, gender, series, tier, ukuran, SUM(quantity) as qty
        FROM core.stock_with_product
        WHERE LOWER(nama_gudang) = %s
          AND kode_mix IS NOT NULL
          AND quantity > 0
        GROUP BY kode_mix, article, gender, series, tier, ukuran
        ORDER BY kode_mix, ukuran
    """,
        (STORE_DB_PATTERN,),
    )

    all_store_articles = defaultdict(
        lambda: {"article": "", "gender": "", "series": "", "tier": 0, "sizes": {}}
    )
    for kode_mix, article, gender, series, tier, ukuran, qty in cur.fetchall():
        rec = all_store_articles[kode_mix]
        rec["article"] = article or ""
        rec["gender"] = gender or ""
        rec["series"] = series or ""
        rec["tier"] = int(tier) if tier else 0
        rec["sizes"][str(ukuran)] = float(qty) if qty else 0.0

    print(f"         {len(all_store_articles)} articles with stock > 0 at store")

    # --- 2f. Snapshot date ---
    cur.execute(
        """
        SELECT MAX(snapshot_date) FROM core.stock_with_product
        WHERE LOWER(nama_gudang) = %s
    """,
        (STORE_DB_PATTERN,),
    )
    snapshot_date = cur.fetchone()[0]
    print(f"         Stock snapshot date: {snapshot_date}")

    cur.close()
    conn.close()

    return {
        "store_stock": dict(store_stock),
        "wh_box_stock": dict(wh_box_stock),
        "wh_protol_stock": dict(wh_protol_stock),
        "sales_3m": sales_3m,
        "all_store_articles": dict(all_store_articles),
        "snapshot_date": snapshot_date,
    }


# =============================================================================
# 3. GAP ANALYSIS
# =============================================================================


def calculate_gaps(planogram, db_data):
    """
    For each displayed article:
    - Compare planogram target per size vs actual stock
    - Calculate gap per size
    - Calculate % size kosong (sizes with 0 stock / total assortment)
    - Decide RO type: Box (>=50%) or Protol (<50%)

    Returns list of article analysis dicts.
    """
    print(f"[3/8] Calculating gaps...")

    store_stock = db_data["store_stock"]
    results = []

    for kode_mix, plano in planogram.items():
        target_sizes = plano["sizes"]  # {size_label: target_pairs}
        actual_sizes = store_stock.get(kode_mix, {})

        total_assortment = len(target_sizes)
        if total_assortment == 0:
            continue

        sizes_detail = []
        sizes_zero = 0
        total_gap = 0
        total_actual = 0
        total_target = 0

        for size_label, target in sorted(target_sizes.items(), key=lambda x: x[0]):
            actual = actual_sizes.get(size_label, 0)
            gap = max(0, target - actual)
            is_zero = actual <= 0

            sizes_detail.append(
                {
                    "size": size_label,
                    "target": target,
                    "actual": actual,
                    "gap": gap,
                    "is_zero": is_zero,
                }
            )

            if is_zero:
                sizes_zero += 1
            total_gap += gap
            total_actual += actual
            total_target += target

        pct_size_kosong = sizes_zero / total_assortment if total_assortment > 0 else 0

        # RO type decision
        if total_gap <= 0:
            ro_type = "NO_RESTOCK"
        elif pct_size_kosong >= RO_BOX_THRESHOLD:
            ro_type = "RO_BOX"
        else:
            ro_type = "RO_PROTOL"

        results.append(
            {
                **plano,
                "sizes_detail": sizes_detail,
                "total_assortment": total_assortment,
                "sizes_zero": sizes_zero,
                "pct_size_kosong": pct_size_kosong,
                "total_gap": total_gap,
                "total_actual": total_actual,
                "total_target": total_target,
                "ro_type": ro_type,
            }
        )

    # Stats
    ro_box = sum(1 for r in results if r["ro_type"] == "RO_BOX")
    ro_protol = sum(1 for r in results if r["ro_type"] == "RO_PROTOL")
    no_restock = sum(1 for r in results if r["ro_type"] == "NO_RESTOCK")
    print(
        f"       → RO Box: {ro_box}, RO Protol: {ro_protol}, No restock: {no_restock}"
    )

    return results


# =============================================================================
# 4. RO DECISION + WAREHOUSE AVAILABILITY CHECK
# =============================================================================


def generate_ro_decisions(gap_results, db_data):
    """
    For each article needing restock:
    - RO_PROTOL: check WH Pusat Protol availability per size
    - RO_BOX: check WH Pusat Box availability (full box)

    Adds warehouse availability info to each result.
    Also handles RO Protol → Box fallback when protol unavailable.
    """
    print(f"[4/8] Checking warehouse availability...")

    wh_box = db_data["wh_box_stock"]
    wh_protol = db_data["wh_protol_stock"]

    ro_protol_list = []
    ro_box_list = []
    no_restock_list = []

    for art in gap_results:
        kode_mix = art["kode_mix"]
        ro_type = art["ro_type"]

        if ro_type == "NO_RESTOCK":
            no_restock_list.append(art)
            continue

        if ro_type == "RO_PROTOL":
            # Check protol availability per size
            protol_available = wh_protol.get(kode_mix, {})
            protol_can_fill = []
            protol_cannot_fill = []

            for sd in art["sizes_detail"]:
                if sd["gap"] <= 0:
                    continue
                avail = protol_available.get(sd["size"], 0)
                can_fill = min(sd["gap"], avail)
                remaining = sd["gap"] - can_fill

                sd["protol_available"] = avail
                sd["protol_fill"] = can_fill
                sd["protol_remaining"] = remaining

                if can_fill > 0:
                    protol_can_fill.append(sd)
                if remaining > 0:
                    protol_cannot_fill.append(sd)

            art["protol_can_fill"] = protol_can_fill
            art["protol_cannot_fill"] = protol_cannot_fill

            # If protol can't fill everything, check if fallback to box needed
            if protol_cannot_fill:
                # Fallback: check box availability
                box_avail = wh_box.get(kode_mix, {})
                total_box_pairs = sum(box_avail.values())
                art["fallback_box_available"] = total_box_pairs > 0
                art["fallback_box_total_pairs"] = total_box_pairs

                if total_box_pairs > 0 and STORAGE_CAPACITY == 0:
                    # ⚠ Royal Plaza has NO storage — flag surplus pre-plan needed
                    sizes_that_exist = [
                        sd
                        for sd in art["sizes_detail"]
                        if sd["actual"] > 0 and sd["gap"] <= 0
                    ]
                    art["surplus_preplan_sizes"] = sizes_that_exist
                    art["surplus_preplan_warning"] = (
                        f"⚠ STORAGE=0: Box will add surplus in "
                        f"{len(sizes_that_exist)} sizes already stocked. "
                        f"Pre-plan redistribution BEFORE approving."
                    )
            else:
                art["fallback_box_available"] = False
                art["fallback_box_total_pairs"] = 0

            ro_protol_list.append(art)

        elif ro_type == "RO_BOX":
            # Check box availability (DDD + LJBB)
            box_avail = wh_box.get(kode_mix, {})
            total_box_pairs = sum(box_avail.values())

            art["box_available"] = total_box_pairs > 0
            art["box_total_pairs"] = total_box_pairs
            art["box_sizes"] = dict(box_avail)

            if total_box_pairs > 0 and STORAGE_CAPACITY == 0:
                # Even for RO Box (>=50% empty), flag surplus for sizes that DO have stock
                sizes_with_stock = [
                    sd for sd in art["sizes_detail"] if sd["actual"] > 0
                ]
                art["surplus_preplan_sizes"] = sizes_with_stock
                if sizes_with_stock:
                    art["surplus_preplan_warning"] = (
                        f"⚠ STORAGE=0: Box will add surplus in "
                        f"{len(sizes_with_stock)} sizes still stocked. "
                        f"Pre-plan redistribution."
                    )
                else:
                    art["surplus_preplan_warning"] = ""
            else:
                art["surplus_preplan_warning"] = ""

            ro_box_list.append(art)

    print(f"       → RO Protol requests: {len(ro_protol_list)}")
    print(f"       → RO Box requests: {len(ro_box_list)}")
    print(f"       → No restock needed: {len(no_restock_list)}")

    return ro_protol_list, ro_box_list, no_restock_list


# =============================================================================
# 5. TO / COVERAGE CALCULATION
# =============================================================================


def calculate_to(gap_results, db_data):
    """
    Calculate Turnover / Stock Coverage for each article.

    stock_coverage_months = current_store_stock / avg_monthly_sales
      → High = slow mover (user definition: "TO 4.5 = lasts 4.5 months")

    For surplus: we sort by avg_monthly_sales ASC (slowest first = pull first).
    """
    print(f"[5/8] Calculating TO / stock coverage...")

    sales_3m = db_data["sales_3m"]

    for art in gap_results:
        kode_mix = art["kode_mix"]
        total_pairs_3m = sales_3m.get(kode_mix, 0)
        avg_monthly_sales = total_pairs_3m / 3.0 if total_pairs_3m > 0 else 0

        current_stock = art["total_actual"]

        # Stock coverage (months) — user's definition
        if avg_monthly_sales > 0:
            stock_coverage_months = current_stock / avg_monthly_sales
        else:
            stock_coverage_months = float("inf") if current_stock > 0 else 0

        # Turnover rate — skill doc's definition
        if current_stock > 0:
            turnover_rate = avg_monthly_sales / current_stock
        else:
            turnover_rate = float("inf") if avg_monthly_sales > 0 else 0

        art["total_pairs_3m"] = total_pairs_3m
        art["avg_monthly_sales"] = avg_monthly_sales
        art["stock_coverage_months"] = stock_coverage_months
        art["turnover_rate"] = turnover_rate

    print(f"       → TO calculated for {len(gap_results)} articles")


# =============================================================================
# 6. SURPLUS CALCULATION (Tier-Based)
# =============================================================================


def calculate_surplus(planogram, gap_results, db_data):
    """
    Surplus = when a tier is OVER-CAPACITY at the store.

    Steps:
    1. Derive ideal tier % from planogram (articles on display per tier)
    2. Calculate actual tier distribution from current stock
    3. For over-capacity tiers (T1, T2, T3 only):
       - Rank articles by avg_monthly_sales ASC (slowest first)
       - Mark bottom N articles as surplus candidates
    4. T4/T5: excluded (clearance)
    5. T8: excluded (3-month protection)

    Also detects off-planogram stock (articles at store not on planogram).
    """
    print(f"[6/8] Calculating surplus...")

    all_store_articles = db_data["all_store_articles"]

    # --- Step 1: Ideal tier distribution from planogram ---
    ideal_tier_counts = defaultdict(int)
    for art in planogram.values():
        ideal_tier_counts[art["tier"]] += 1

    total_planogram_articles = sum(ideal_tier_counts.values())

    ideal_tier_pct = {}
    for tier, count in ideal_tier_counts.items():
        ideal_tier_pct[tier] = (
            count / total_planogram_articles if total_planogram_articles > 0 else 0
        )

    print(
        f"       Ideal tier distribution (from planogram, {total_planogram_articles} articles):"
    )
    for t in sorted(ideal_tier_pct):
        print(
            f"         Tier {t}: {ideal_tier_counts[t]} articles ({ideal_tier_pct[t] * 100:.1f}%)"
        )

    # --- Step 2: Actual tier distribution at store ---
    # Build a lookup of gap_results by kode_mix for TO data
    gap_lookup = {art["kode_mix"]: art for art in gap_results}

    # Count articles per tier that have stock at the store
    actual_tier_articles = defaultdict(list)  # {tier: [article_info, ...]}

    # First: articles on planogram that have stock
    store_stock = db_data["store_stock"]
    for art in gap_results:
        if art["total_actual"] > 0:
            # Get per-size stock from DB for surplus pull detail
            size_stock = store_stock.get(art["kode_mix"], {})
            actual_tier_articles[art["tier"]].append(
                {
                    "kode_mix": art["kode_mix"],
                    "article": art["article"],
                    "gender": art["gender"],
                    "series": art["series"],
                    "tier": art["tier"],
                    "total_stock": art["total_actual"],
                    "sizes": {s: q for s, q in size_stock.items() if q > 0},
                    "avg_monthly_sales": art.get("avg_monthly_sales", 0),
                    "stock_coverage_months": art.get("stock_coverage_months", 0),
                    "on_planogram": True,
                }
            )

    # Second: off-planogram articles with stock
    off_planogram = []
    for kode_mix, info in all_store_articles.items():
        if kode_mix not in planogram:
            total_stock = sum(info["sizes"].values())
            if total_stock > 0:
                sales_3m = db_data["sales_3m"].get(kode_mix, 0)
                avg_monthly = sales_3m / 3.0 if sales_3m > 0 else 0
                coverage = (
                    total_stock / avg_monthly
                    if avg_monthly > 0
                    else (float("inf") if total_stock > 0 else 0)
                )

                art_info = {
                    "kode_mix": kode_mix,
                    "article": info["article"],
                    "gender": info["gender"],
                    "series": info["series"],
                    "tier": info["tier"],
                    "total_stock": total_stock,
                    "sizes": info["sizes"],
                    "avg_monthly_sales": avg_monthly,
                    "stock_coverage_months": coverage,
                    "on_planogram": False,
                }
                off_planogram.append(art_info)
                actual_tier_articles[info["tier"]].append(art_info)

    # Actual counts
    actual_tier_counts = {t: len(arts) for t, arts in actual_tier_articles.items()}
    total_actual_articles = sum(actual_tier_counts.values())

    actual_tier_pct = {}
    for tier, count in actual_tier_counts.items():
        actual_tier_pct[tier] = (
            count / total_actual_articles if total_actual_articles > 0 else 0
        )

    print(
        f"\n       Actual tier distribution (articles with stock, {total_actual_articles} total):"
    )
    for t in sorted(actual_tier_counts):
        ideal = ideal_tier_counts.get(t, 0)
        actual = actual_tier_counts[t]
        diff = actual - ideal
        status = "OVER" if diff > 0 else ("UNDER" if diff < 0 else "OK")
        check = "✓ check" if t in SURPLUS_CHECK_TIERS else "✗ skip"
        print(
            f"         Tier {t}: {actual} actual vs {ideal} ideal "
            f"({'+' if diff > 0 else ''}{diff}) [{status}] {check}"
        )

    # --- Step 3: Surplus candidates for over-capacity tiers ---
    surplus_pull_list = []

    for tier in SURPLUS_CHECK_TIERS:
        ideal_count = ideal_tier_counts.get(tier, 0)
        actual_count = actual_tier_counts.get(tier, 0)

        if actual_count <= ideal_count:
            continue  # Under or at capacity — no surplus

        over_by = actual_count - ideal_count
        tier_articles = actual_tier_articles[tier]

        # Sort by avg_monthly_sales ASC (slowest sellers first = surplus priority)
        tier_articles_sorted = sorted(
            tier_articles, key=lambda x: x["avg_monthly_sales"]
        )

        # Pull the slowest N articles
        for art in tier_articles_sorted[:over_by]:
            surplus_pull_list.append(
                {
                    **art,
                    "surplus_reason": f"Tier {tier} over-capacity by {over_by} articles",
                    "pull_priority": "HIGH"
                    if art["avg_monthly_sales"] == 0
                    else "MEDIUM",
                }
            )

    print(f"\n       → Surplus candidates: {len(surplus_pull_list)}")
    print(f"       → Off-planogram articles with stock: {len(off_planogram)}")

    return {
        "surplus_pull_list": surplus_pull_list,
        "off_planogram": off_planogram,
        "ideal_tier_counts": dict(ideal_tier_counts),
        "ideal_tier_pct": dict(ideal_tier_pct),
        "actual_tier_counts": dict(actual_tier_counts),
        "actual_tier_pct": dict(actual_tier_pct),
        "total_planogram_articles": total_planogram_articles,
        "total_actual_articles": total_actual_articles,
    }


# =============================================================================
# 7. EXCEL OUTPUT
# =============================================================================


def style_header_row(ws, row_num, max_col, fill=HEADER_FILL_DARK):
    """Apply header styling to a row."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = THIN_BORDER


def auto_width(ws, min_width=10, max_width=35):
    """Auto-adjust column widths."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(
            min_width, min(max_len + 2, max_width)
        )


def write_excel(
    ro_protol_list,
    ro_box_list,
    no_restock_list,
    surplus_data,
    gap_results,
    planogram,
    db_data,
):
    """
    Write the RO Request Excel workbook — official form/document format.

    This is a weekly handover document from Area Supervisor to Warehouse Supervisor.
    Sheets:
      1. RO Request       — Cover page with summary, instructions, signatures
      2. Daftar RO Protol — Clean picking list for WH Protol
      3. Daftar RO Box    — Clean picking list for WH Box (omitted if none)
      4. Daftar Surplus    — Clean pull list (size-level) for store staff
      5. Reference         — Tier analysis + full article status (internal use)
    """
    print(f"\n[7/8] Writing Excel output: {OUTPUT_FILE}")

    wb = openpyxl.Workbook()
    snapshot = db_data.get("snapshot_date", date.today())
    today = datetime.now()
    has_ro_box = len(ro_box_list) > 0

    # --- Extra styles for official form ---
    TITLE_FONT = Font(bold=True, size=20, color="1B2A4A")
    STORE_FONT = Font(bold=True, size=15, color="0D6B3E")
    SECTION_TITLE_FONT = Font(bold=True, size=13, color="1B2A4A")
    LABEL_FONT = Font(bold=True, size=11)
    VALUE_FONT = Font(size=11)
    SMALL_FONT = Font(size=10, color="666666")
    SIGN_FONT = Font(size=10)
    TOTAL_FONT = Font(bold=True, size=11, color="1B2A4A")
    FILL_ZEBRA = PatternFill(
        start_color="F0F4F8", end_color="F0F4F8", fill_type="solid"
    )
    FILL_LIGHT_GRAY = PatternFill(
        start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"
    )

    def write_form_header(ws, row, title, subtitle, total_cols):
        """Write a consistent form header on each list sheet."""
        ws.merge_cells(
            start_row=row, start_column=1, end_row=row, end_column=total_cols
        )
        ws.cell(row=row, column=1, value=title).font = Font(
            bold=True, size=16, color="1B2A4A"
        )
        row += 1
        ws.merge_cells(
            start_row=row, start_column=1, end_row=row, end_column=total_cols
        )
        ws.cell(row=row, column=1, value=STORE_NAME).font = Font(
            bold=True, size=12, color="0D6B3E"
        )
        row += 1
        ws.merge_cells(
            start_row=row, start_column=1, end_row=row, end_column=total_cols
        )
        ws.cell(
            row=row,
            column=1,
            value=f"Week of {today.strftime('%d %B %Y')}  |  Stock Snapshot: {snapshot}",
        ).font = SMALL_FONT
        row += 1
        ws.merge_cells(
            start_row=row, start_column=1, end_row=row, end_column=total_cols
        )
        ws.cell(row=row, column=1, value=subtitle).font = Font(size=11, color="333333")
        row += 2
        return row

    # --- Calculate totals ---
    total_protol_pairs = 0
    for art in ro_protol_list:
        for sd in art["sizes_detail"]:
            if sd["gap"] > 0:
                total_protol_pairs += sd.get("protol_fill", 0)

    total_surplus_pairs = sum(
        a["total_stock"] for a in surplus_data["surplus_pull_list"]
    )

    # =========================================================================
    # SHEET 1: RO REQUEST — Cover Page / Official Form
    # =========================================================================
    ws = wb.active
    ws.title = "RO Request"
    ws.sheet_properties.tabColor = "1B2A4A"

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 22

    row = 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1, value="WEEKLY RO REQUEST").font = TITLE_FONT
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1, value=STORE_NAME).font = STORE_FONT
    row += 2

    # Info block
    info = [
        ("Week of:", today.strftime("%d %B %Y")),
        ("Stock Snapshot:", str(snapshot)),
        ("Storage Capacity:", f"{STORAGE_CAPACITY} boxes"),
    ]
    for label, val in info:
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        ws.cell(row=row, column=2, value=val).font = VALUE_FONT
        row += 1

    row += 1

    # From / To
    ws.cell(row=row, column=1, value="From:").font = LABEL_FONT
    ws.cell(row=row, column=2, value="Area Supervisor").font = VALUE_FONT
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
    ws.cell(row=row, column=3, value="___________________________").font = SIGN_FONT
    row += 1
    ws.cell(row=row, column=1, value="To:").font = LABEL_FONT
    ws.cell(row=row, column=2, value="Warehouse Supervisor").font = VALUE_FONT
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
    ws.cell(row=row, column=3, value="___________________________").font = SIGN_FONT
    row += 2

    # --- REQUEST SUMMARY TABLE ---
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1, value="REQUEST SUMMARY").font = SECTION_TITLE_FONT
    row += 1

    sum_headers = ["Type", "Articles", "Total", "Source / Destination", "See Sheet"]
    for ci, h in enumerate(sum_headers, 1):
        ws.cell(row=row, column=ci, value=h)
    style_header_row(ws, row, len(sum_headers), HEADER_FILL_DARK)
    row += 1

    summary_rows = [
        (
            "RO PROTOL",
            len(ro_protol_list),
            f"{int(total_protol_pairs)} pairs",
            "FROM: WH Pusat Protol",
            "Daftar RO Protol",
        ),
        (
            "RO BOX",
            len(ro_box_list) if has_ro_box else 0,
            f"{len(ro_box_list)} boxes" if has_ro_box else "NONE this week",
            "FROM: WH Pusat Box" if has_ro_box else "-",
            "Daftar RO Box" if has_ro_box else "-",
        ),
        (
            "SURPLUS PULL",
            len(surplus_data["surplus_pull_list"]),
            f"{int(total_surplus_pairs)} pairs",
            "TO: WH Pusat Protol",
            "Daftar Surplus",
        ),
    ]
    for label, count, total, source, sheet in summary_rows:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=count)
        ws.cell(row=row, column=3, value=total)
        ws.cell(row=row, column=4, value=source)
        ws.cell(row=row, column=5, value=sheet)
        for ci in range(1, 6):
            ws.cell(row=row, column=ci).border = THIN_BORDER
        row += 1

    row += 2

    # --- INSTRUCTIONS ---
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1, value="INSTRUCTIONS").font = SECTION_TITLE_FONT
    row += 1
    instructions = [
        "1. RO Protol: Pick individual pairs from Warehouse Pusat Protol per article list.",
        "2. RO Box: Pick full boxes (1 box = 12 pairs) from Warehouse Pusat Box per article list.",
        "3. Surplus: Pull listed items from store display, return to Warehouse Pusat Protol.",
        "4. Restock + Surplus happen SAME DAY: deliver RO items IN, pull surplus items OUT.",
        "5. Priority: RO Protol first. RO Box only when >=50% sizes are empty.",
    ]
    for inst in instructions:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.cell(row=row, column=1, value=inst).font = Font(size=10)
        row += 1

    row += 2

    # --- SIGNATURES ---
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1, value="SIGNATURES").font = SECTION_TITLE_FONT
    row += 1

    sign_cols = ["", "Prepared by", "Approved by", "Received by"]
    for ci, h in enumerate(sign_cols, 1):
        ws.cell(row=row, column=ci, value=h).font = LABEL_FONT
        ws.cell(row=row, column=ci).alignment = Alignment(horizontal="center")
    row += 1
    for field in ["Name:", "Date:", "Signature:"]:
        ws.cell(row=row, column=1, value=field).font = SIGN_FONT
        for ci in range(2, 5):
            ws.cell(row=row, column=ci, value="___________________").font = SIGN_FONT
            ws.cell(row=row, column=ci).alignment = Alignment(horizontal="center")
        row += 1

    # =========================================================================
    # SHEET 2: DAFTAR RO PROTOL — Clean Picking List
    # =========================================================================
    ws2 = wb.create_sheet("Daftar RO Protol")
    ws2.sheet_properties.tabColor = "0D6B3E"

    headers2 = [
        "No",
        "Article",
        "Kode Mix",
        "Tier",
        "Sizes Needed (size:qty)",
        "Total Pairs",
    ]
    total_cols2 = len(headers2)

    row = 1
    row = write_form_header(
        ws2,
        row,
        "DAFTAR RO PROTOL",
        f"Source: Warehouse Pusat Protol (DDD)  |  Total: {len(ro_protol_list)} articles, {int(total_protol_pairs)} pairs",
        total_cols2,
    )

    for ci, h in enumerate(headers2, 1):
        ws2.cell(row=row, column=ci, value=h)
    style_header_row(ws2, row, total_cols2, HEADER_FILL_GREEN)
    row += 1

    num = 1
    for art in sorted(ro_protol_list, key=lambda x: (x["tier"], x["article"])):
        sizes_needed = []
        total_gap = 0
        for sd in art["sizes_detail"]:
            if sd["gap"] > 0:
                sizes_needed.append(f"{sd['size']}:{int(sd['gap'])}")
                total_gap += sd.get("protol_fill", 0)
        if not sizes_needed:
            continue

        ws2.cell(row=row, column=1, value=num)
        ws2.cell(row=row, column=2, value=art["article"])
        ws2.cell(row=row, column=3, value=art["kode_mix"])
        ws2.cell(row=row, column=4, value=art["tier"])
        ws2.cell(row=row, column=5, value=", ".join(sizes_needed))
        ws2.cell(row=row, column=6, value=int(total_gap))

        if num % 2 == 0:
            for ci in range(1, total_cols2 + 1):
                ws2.cell(row=row, column=ci).fill = FILL_ZEBRA
        for ci in range(1, total_cols2 + 1):
            ws2.cell(row=row, column=ci).border = THIN_BORDER
        num += 1
        row += 1

    # Total row
    for ci in range(1, total_cols2 + 1):
        ws2.cell(row=row, column=ci).border = THIN_BORDER
        ws2.cell(row=row, column=ci).fill = FILL_LIGHT_GRAY
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws2.cell(row=row, column=1, value="TOTAL PAIRS").font = TOTAL_FONT
    ws2.cell(row=row, column=6, value=int(total_protol_pairs)).font = TOTAL_FONT

    auto_width(ws2)

    # =========================================================================
    # SHEET 3: DAFTAR RO BOX — Clean Picking List
    # =========================================================================
    if has_ro_box:
        ws3 = wb.create_sheet("Daftar RO Box")
        ws3.sheet_properties.tabColor = "8B1A1A"

        headers3 = [
            "No",
            "Article",
            "Kode Mix",
            "Tier",
            "Box Qty",
            "WH Available",
        ]
        total_cols3 = len(headers3)

        row = 1
        row = write_form_header(
            ws3,
            row,
            "DAFTAR RO BOX",
            f"Source: Warehouse Pusat Box (DDD + LJBB)  |  Total: {len(ro_box_list)} boxes",
            total_cols3,
        )

        for ci, h in enumerate(headers3, 1):
            ws3.cell(row=row, column=ci, value=h)
        style_header_row(ws3, row, total_cols3, HEADER_FILL_RED)
        row += 1

        num = 1
        for art in sorted(ro_box_list, key=lambda x: (x["tier"], x["article"])):
            ws3.cell(row=row, column=1, value=num)
            ws3.cell(row=row, column=2, value=art["article"])
            ws3.cell(row=row, column=3, value=art["kode_mix"])
            ws3.cell(row=row, column=4, value=art["tier"])
            ws3.cell(row=row, column=5, value=1)
            ws3.cell(
                row=row,
                column=6,
                value="YES" if art.get("box_available") else "NO",
            )

            if not art.get("box_available"):
                for ci in range(1, total_cols3 + 1):
                    ws3.cell(row=row, column=ci).fill = FILL_LIGHT_RED
            elif num % 2 == 0:
                for ci in range(1, total_cols3 + 1):
                    ws3.cell(row=row, column=ci).fill = FILL_ZEBRA
            for ci in range(1, total_cols3 + 1):
                ws3.cell(row=row, column=ci).border = THIN_BORDER
            num += 1
            row += 1

        # Total row
        for ci in range(1, total_cols3 + 1):
            ws3.cell(row=row, column=ci).border = THIN_BORDER
            ws3.cell(row=row, column=ci).fill = FILL_LIGHT_GRAY
        ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws3.cell(row=row, column=1, value="TOTAL BOXES").font = TOTAL_FONT
        ws3.cell(row=row, column=5, value=len(ro_box_list)).font = TOTAL_FONT

        auto_width(ws3)

    # =========================================================================
    # SHEET 4: DAFTAR SURPLUS — Size-Level Pull List
    # =========================================================================
    ws4 = wb.create_sheet("Daftar Surplus")
    ws4.sheet_properties.tabColor = "B35900"

    headers4 = ["No", "Article", "Kode Mix", "Size", "Pairs to Pull"]
    total_cols4 = len(headers4)

    row = 1
    row = write_form_header(
        ws4,
        row,
        "DAFTAR SURPLUS PULL",
        f"Destination: Warehouse Pusat Protol  |  Total: {len(surplus_data['surplus_pull_list'])} articles, {int(total_surplus_pairs)} pairs",
        total_cols4,
    )

    for ci, h in enumerate(headers4, 1):
        ws4.cell(row=row, column=ci, value=h)
    style_header_row(ws4, row, total_cols4, HEADER_FILL_ORANGE)
    row += 1

    num = 1
    running_total = 0
    for art in surplus_data["surplus_pull_list"]:
        sizes = art.get("sizes", {})
        if not sizes:
            ws4.cell(row=row, column=1, value=num)
            ws4.cell(row=row, column=2, value=art["article"])
            ws4.cell(row=row, column=3, value=art["kode_mix"])
            ws4.cell(row=row, column=4, value="(all)")
            ws4.cell(row=row, column=5, value=int(art["total_stock"]))
            running_total += int(art["total_stock"])
            if num % 2 == 0:
                for ci in range(1, total_cols4 + 1):
                    ws4.cell(row=row, column=ci).fill = FILL_ZEBRA
            for ci in range(1, total_cols4 + 1):
                ws4.cell(row=row, column=ci).border = THIN_BORDER
            num += 1
            row += 1
        else:
            for size_label in sorted(sizes.keys()):
                qty = sizes[size_label]
                if qty <= 0:
                    continue
                ws4.cell(row=row, column=1, value=num)
                ws4.cell(row=row, column=2, value=art["article"])
                ws4.cell(row=row, column=3, value=art["kode_mix"])
                ws4.cell(row=row, column=4, value=size_label)
                ws4.cell(row=row, column=5, value=int(qty))
                running_total += int(qty)
                if num % 2 == 0:
                    for ci in range(1, total_cols4 + 1):
                        ws4.cell(row=row, column=ci).fill = FILL_ZEBRA
                for ci in range(1, total_cols4 + 1):
                    ws4.cell(row=row, column=ci).border = THIN_BORDER
                num += 1
                row += 1

    # Total row
    for ci in range(1, total_cols4 + 1):
        ws4.cell(row=row, column=ci).border = THIN_BORDER
        ws4.cell(row=row, column=ci).fill = FILL_LIGHT_GRAY
    ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws4.cell(row=row, column=1, value="TOTAL PAIRS TO PULL").font = TOTAL_FONT
    ws4.cell(row=row, column=5, value=int(running_total)).font = TOTAL_FONT

    auto_width(ws4)

    # =========================================================================
    # SHEET 5: REFERENCE — Tier Analysis + Full Status + Off-Planogram
    # =========================================================================
    ws5 = wb.create_sheet("Reference")
    ws5.sheet_properties.tabColor = "4A1B6D"

    row = 1
    ws5.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
    ws5.cell(row=row, column=1, value="REFERENCE DATA (Internal Use)").font = Font(
        bold=True, size=14, color="4A1B6D"
    )
    row += 2

    # --- Tier Capacity Analysis ---
    ws5.cell(row=row, column=1, value="TIER CAPACITY ANALYSIS").font = Font(
        bold=True, size=12
    )
    row += 1
    tier_headers = [
        "Tier",
        "Ideal (Planogram)",
        "Ideal %",
        "Actual (Stock)",
        "Actual %",
        "Diff",
        "Status",
    ]
    for ci, h in enumerate(tier_headers, 1):
        ws5.cell(row=row, column=ci, value=h)
    style_header_row(ws5, row, len(tier_headers), HEADER_FILL_DARK)
    row += 1

    all_tiers = sorted(
        set(
            list(surplus_data["ideal_tier_counts"].keys())
            + list(surplus_data["actual_tier_counts"].keys())
        )
    )
    for tier in all_tiers:
        ideal = surplus_data["ideal_tier_counts"].get(tier, 0)
        actual = surplus_data["actual_tier_counts"].get(tier, 0)
        ideal_p = surplus_data["ideal_tier_pct"].get(tier, 0)
        actual_p = surplus_data["actual_tier_pct"].get(tier, 0)
        diff = actual - ideal
        status = "OVER" if diff > 0 else ("UNDER" if diff < 0 else "OK")

        ws5.cell(row=row, column=1, value=f"Tier {tier}")
        ws5.cell(row=row, column=2, value=ideal)
        ws5.cell(row=row, column=3, value=f"{ideal_p * 100:.1f}%")
        ws5.cell(row=row, column=4, value=actual)
        ws5.cell(row=row, column=5, value=f"{actual_p * 100:.1f}%")
        ws5.cell(row=row, column=6, value=f"{'+' if diff > 0 else ''}{diff}")
        ws5.cell(row=row, column=7, value=status)

        fill = None
        if diff > 0 and tier in SURPLUS_CHECK_TIERS:
            fill = FILL_LIGHT_RED
        elif diff < 0:
            fill = FILL_LIGHT_YELLOW
        elif diff == 0:
            fill = FILL_LIGHT_GREEN
        if fill:
            for ci in range(1, 8):
                ws5.cell(row=row, column=ci).fill = fill
        for ci in range(1, 8):
            ws5.cell(row=row, column=ci).border = THIN_BORDER
        row += 1

    row += 2

    # --- Full Article Status ---
    ws5.cell(row=row, column=1, value="FULL ARTICLE STATUS").font = Font(
        bold=True, size=12
    )
    row += 1
    status_headers = [
        "Article",
        "Kode Mix",
        "Gender",
        "Series",
        "Tier",
        "Target",
        "Actual",
        "Gap",
        "% Kosong",
        "RO Type",
        "Avg Monthly Sales",
        "Stock Coverage",
    ]
    for ci, h in enumerate(status_headers, 1):
        ws5.cell(row=row, column=ci, value=h)
    style_header_row(ws5, row, len(status_headers), HEADER_FILL_PURPLE)
    row += 1

    for art in sorted(gap_results, key=lambda x: (x["tier"], x["article"])):
        ws5.cell(row=row, column=1, value=art["article"])
        ws5.cell(row=row, column=2, value=art["kode_mix"])
        ws5.cell(row=row, column=3, value=art["gender"])
        ws5.cell(row=row, column=4, value=art["series"])
        ws5.cell(row=row, column=5, value=art["tier"])
        ws5.cell(row=row, column=6, value=art["total_target"])
        ws5.cell(row=row, column=7, value=art["total_actual"])
        ws5.cell(row=row, column=8, value=art["total_gap"])
        ws5.cell(row=row, column=9, value=f"{art['pct_size_kosong'] * 100:.0f}%")
        ws5.cell(row=row, column=10, value=art["ro_type"])
        ws5.cell(row=row, column=11, value=round(art.get("avg_monthly_sales", 0), 1))
        cov = art.get("stock_coverage_months", 0)
        ws5.cell(
            row=row,
            column=12,
            value="dead stock" if cov == float("inf") else round(cov, 1),
        )

        if art["ro_type"] == "RO_BOX":
            fill = FILL_LIGHT_RED
        elif art["ro_type"] == "RO_PROTOL":
            fill = FILL_LIGHT_YELLOW
        else:
            fill = FILL_LIGHT_GREEN
        for ci in range(1, len(status_headers) + 1):
            ws5.cell(row=row, column=ci).fill = fill
            ws5.cell(row=row, column=ci).border = THIN_BORDER
        row += 1

    row += 2

    # --- Off-Planogram Stock ---
    ws5.cell(row=row, column=1, value="OFF-PLANOGRAM STOCK").font = Font(
        bold=True, size=12
    )
    row += 1
    off_headers = [
        "Article",
        "Kode Mix",
        "Gender",
        "Series",
        "Tier",
        "Stock (pairs)",
        "Sizes",
        "Recommendation",
    ]
    for ci, h in enumerate(off_headers, 1):
        ws5.cell(row=row, column=ci, value=h)
    style_header_row(ws5, row, len(off_headers), HEADER_FILL_TEAL)
    row += 1

    for art in sorted(
        surplus_data["off_planogram"],
        key=lambda x: (-x["total_stock"], x["article"]),
    ):
        sizes_with_stock = [
            f"{s}({int(q)})" for s, q in sorted(art["sizes"].items()) if q > 0
        ]
        tier = art["tier"]
        sales = art["avg_monthly_sales"]
        if tier in (4, 5):
            reco = "CLEARANCE"
        elif tier == 8:
            reco = "T8 PROTECTION"
        elif sales == 0:
            reco = "DEAD STOCK - pull"
        else:
            reco = "SURPLUS - pull"

        ws5.cell(row=row, column=1, value=art["article"])
        ws5.cell(row=row, column=2, value=art["kode_mix"])
        ws5.cell(row=row, column=3, value=art["gender"])
        ws5.cell(row=row, column=4, value=art["series"])
        ws5.cell(row=row, column=5, value=tier)
        ws5.cell(row=row, column=6, value=art["total_stock"])
        ws5.cell(row=row, column=7, value=", ".join(sizes_with_stock))
        ws5.cell(row=row, column=8, value=reco)
        for ci in range(1, len(off_headers) + 1):
            ws5.cell(row=row, column=ci).border = THIN_BORDER
        row += 1

    auto_width(ws5)

    # =========================================================================
    # SAVE
    # =========================================================================
    wb.save(OUTPUT_FILE)
    sheet_names = [s.title for s in wb.worksheets]
    print(f"\n       Saved: {OUTPUT_FILE}")
    print(f"       Sheets: {sheet_names}")
    if not has_ro_box:
        print("       Note: RO Box sheet omitted (no box requests this week)")


# =============================================================================
# 8. MAIN
# =============================================================================


def main():
    print("=" * 70)
    print(f"  RO REQUEST GENERATOR — {STORE_NAME}")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 70)
    print()

    # Step 1: Read planogram
    planogram = read_planogram()
    if not planogram:
        print("ERROR: No planogram data found for Royal Plaza.")
        sys.exit(1)

    # Step 2: Query DB
    db_data = query_db()

    # Step 3: Gap analysis
    gap_results = calculate_gaps(planogram, db_data)

    # Step 4: RO decisions + warehouse check
    ro_protol_list, ro_box_list, no_restock_list = generate_ro_decisions(
        gap_results, db_data
    )

    # Step 5: TO / Coverage calculation
    calculate_to(gap_results, db_data)

    # Step 6: Surplus calculation
    surplus_data = calculate_surplus(planogram, gap_results, db_data)

    # Step 7: Excel output
    write_excel(
        ro_protol_list,
        ro_box_list,
        no_restock_list,
        surplus_data,
        gap_results,
        planogram,
        db_data,
    )

    # Final summary
    print()
    print("=" * 70)
    print("  SUMMARY")
    print("=" * 70)
    print(f"  Planogram articles:    {len(planogram)}")
    print(f"  RO Protol requests:    {len(ro_protol_list)}")
    print(f"  RO Box requests:       {len(ro_box_list)}")
    print(f"  No restock needed:     {len(no_restock_list)}")
    print(f"  Surplus pull list:     {len(surplus_data['surplus_pull_list'])}")
    print(f"  Off-planogram stock:   {len(surplus_data['off_planogram'])}")
    print(f"  Storage capacity:      {STORAGE_CAPACITY} boxes")
    print(f"  Output: {OUTPUT_FILE}")
    print("=" * 70)


if __name__ == "__main__":
    main()
