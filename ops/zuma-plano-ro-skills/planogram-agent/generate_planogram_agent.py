#!/usr/bin/env python3
"""
Planogram Agent Generator — Paperclip
Reads pre-planogram data from Store Layout GSheet (cached xlsx),
generates v4 visual layout + summary, writes to portal.planogram_paperclip.

Usage:
  python3 generate_planogram_agent.py \
    --store-code ROYAL \
    --store-name "Zuma Royal Plaza" \
    --store-db ROYAL \
    --area Jatim \
    --output /path/to/output.xlsx
"""

import argparse
import json
import math
import os
import sys
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import psycopg2

# ============================================================
# CONSTANTS
# ============================================================
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
STORE_LAYOUT_XLSX = "/tmp/store_layout_jatim.xlsx"
DATA_OPTION_XLSX = "/tmp/data_option_jatim.xlsx"

# GSheet IDs for re-download if cache missing
STORE_LAYOUT_GSHEET = "1rQCC93t6f7HY1Dnoliplu2Mg61i-nnAupErn_U0fPhw"
DATA_OPTION_GSHEET = "1FweMqoNikHB7F9efH0xv5rOtpt4mxnuxqzbPqb41eCo"

SERIES_COLORS = {
    "ZORRO": "EAD1DC", "DALLAS": "E6B8AF", "SLIDE": "FFF2CC",
    "FLO": "F9CB9C", "PUFFY": "A4C2F4", "ELSA": "CCCCCC",
    "CLASSIC": "FFF2CC", "CLASSICEARTH": "F0E6C0", "CLASSIC EARTH": "F0E6C0",
    "BLACKSERIES": "D9D9D9", "BLACK SERIES": "D9D9D9", "STRIPE": "FFE0B2",
    "ONYX": "CFD8DC", "STITCH": "B4E7CE", "WEDGES": "D7CCC8",
    "LUCA": "FFCCBC", "AIRMOVE": "E0E0E0", "LUNA": "D5A6BD",
    "DISNEY": "FFAB91", "BATMAN": "90A4AE", "MICKEY": "FFCC80",
    "MICKEY & FRIENDS": "FFCC80", "POOH": "FFE082", "TOY STORY": "81D4FA",
    "WBB": "80DEEA", "VELCRO": "E8F5E9", "MERCI": "D9D2E9",
    "SPIDER-MAN": "CE93D8", "LOTSO": "F48FB1", "(KOSONG)": "F5F5F5",
}

# Display rules
HOOKS_PER_ARTICLE = {"jepit": 2, "fashion": 3}
PAIRS_PER_HOOK = {"jepit": 6, "fashion": 4}
PAIRS_PER_BOX = 12
NHOOKS = 7  # Standard backwall hook rows

# Series → type mapping
JEPIT_SERIES = {"CLASSIC", "CLASSICEARTH", "CLASSIC EARTH", "BLACKSERIES", "BLACK SERIES", "STRIPE", "SJ-CLASSIC"}
FASHION_SERIES = {"ZORRO", "DALLAS", "SLIDE", "FLO", "PUFFY", "ELSA", "ONYX", "STITCH", "MERCI", "LUNA", "LUCA", "WEDGES"}

THIN = Border(left=Side(style="thin"), right=Side(style="thin"),
              top=Side(style="thin"), bottom=Side(style="thin"))
CELL_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

def get_color(series):
    s = series.upper().strip()
    if s in SERIES_COLORS: return SERIES_COLORS[s]
    for k, v in SERIES_COLORS.items():
        if k in s or s in k: return v
    import hashlib
    h = int(hashlib.md5(s.encode()).hexdigest()[:6], 16)
    return f"{180+(h%75):02x}{180+((h>>8)%75):02x}{180+((h>>16)%75):02x}"

def get_display_type(series):
    s = series.upper().strip()
    if s in JEPIT_SERIES: return "jepit"
    return "fashion"

# ============================================================
# DATA LOADING
# ============================================================
def ensure_cache(xlsx_path, gsheet_id):
    """Download GSheet as xlsx if not cached."""
    if os.path.exists(xlsx_path):
        age_hours = (datetime.now().timestamp() - os.path.getmtime(xlsx_path)) / 3600
        if age_hours < 24:
            return  # Fresh enough
    print(f"Downloading GSheet {gsheet_id}...")
    import subprocess
    subprocess.run([
        "curl", "-sL",
        f"https://docs.google.com/spreadsheets/d/{gsheet_id}/export?format=xlsx",
        "-o", xlsx_path
    ], check=True)
    print(f"Downloaded: {xlsx_path}")

def load_store_planogram_data(store_name, store_db, area, total_display_pairs=0):
    """Calculate fresh planogram from sales history — recalculated every run."""
    conn = psycopg2.connect(
        host=os.environ.get("PGHOST", "76.13.194.120"),
        port=os.environ.get("PGPORT", "5432"),
        dbname=os.environ.get("PGDATABASE", "openclaw_ops"),
        user=os.environ.get("PGUSER", "openclaw_app"),
        password=os.environ.get("PGPASSWORD", ""),
    )
    cur = conn.cursor()

    # Single query: pull all articles sold at this store (12mo history + 3mo avg)
    # Bridge kode_mix → kode_kecil via portal.kodemix
    cur.execute("""
        WITH sales_12mo AS (
            SELECT kode_mix,
                   MAX(article) as article, MAX(gender) as gender,
                   MAX(series) as series,
                   COALESCE(MAX(NULLIF(TRIM(tier),'')),'3') as tier,
                   SUM(quantity) as total_qty,
                   ROUND(SUM(quantity)::numeric / 12, 2) as avg_monthly
            FROM core.sales_with_product
            WHERE UPPER(TRIM(matched_store_name)) = UPPER(TRIM(%s))
              AND transaction_date >= (CURRENT_DATE - INTERVAL '12 months')
              AND quantity > 0 AND kode_mix IS NOT NULL
              AND (is_intercompany IS NULL OR is_intercompany = FALSE)
            GROUP BY kode_mix
            HAVING SUM(quantity) > 0
        ),
        sales_3mo AS (
            SELECT kode_mix, ROUND(SUM(quantity)::numeric / 3, 2) as avg_3mo
            FROM core.sales_with_product
            WHERE UPPER(TRIM(matched_store_name)) = UPPER(TRIM(%s))
              AND transaction_date >= (CURRENT_DATE - INTERVAL '3 months')
              AND quantity > 0 AND kode_mix IS NOT NULL
              AND (is_intercompany IS NULL OR is_intercompany = FALSE)
            GROUP BY kode_mix
        ),
        total_3mo AS (
            SELECT SUM(quantity)::numeric as total FROM core.sales_with_product
            WHERE UPPER(TRIM(matched_store_name)) = UPPER(TRIM(%s))
              AND transaction_date >= (CURRENT_DATE - INTERVAL '3 months')
              AND quantity > 0 AND kode_mix IS NOT NULL
              AND (is_intercompany IS NULL OR is_intercompany = FALSE)
        ),
        kode_bridge AS (
            SELECT DISTINCT ON (kode_mix) kode_mix, kode as kode_kecil
            FROM portal.kodemix WHERE kode IS NOT NULL AND kode_mix IS NOT NULL
            ORDER BY kode_mix, no_urut
        )
        SELECT s.kode_mix, kb.kode_kecil, s.article, s.gender, s.series, s.tier,
               COALESCE(s3.avg_3mo, s.avg_monthly) as avg_sales,
               ROUND(COALESCE(s3.avg_3mo, s.avg_monthly) / NULLIF((SELECT total FROM total_3mo), 0), 6) as sales_mix,
               s.total_qty
        FROM sales_12mo s
        LEFT JOIN sales_3mo s3 ON s3.kode_mix = s.kode_mix
        LEFT JOIN kode_bridge kb ON kb.kode_mix = s.kode_mix
        ORDER BY COALESCE(s3.avg_3mo, s.avg_monthly) DESC
    """, (store_db, store_db, store_db))

    rows = cur.fetchall()
    cur.close()
    conn.close()

    if not rows:
        print(f"WARNING: No sales data for store '{store_db}'")
        return []

    # Build article list with proportional reko_box
    total_avg = sum(float(r[6]) for r in rows if r[6])
    articles = []
    for row in rows:
        kode_mix, kode_kecil, article, gender, series, tier, avg_sales, sales_mix, total_qty = row
        kode_mix = str(kode_mix or "").strip()
        if not kode_mix: continue

        avg_f = float(avg_sales) if avg_sales else 0
        mix_f = float(sales_mix) if sales_mix else 0

        # Reko box proportional to display capacity
        if total_display_pairs > 0 and total_avg > 0:
            reko_pairs = mix_f * total_display_pairs
            reko_box = max(1, round(reko_pairs / PAIRS_PER_BOX))
        else:
            reko_pairs = avg_f
            reko_box = max(1, round(avg_f / PAIRS_PER_BOX))

        articles.append({
            "gender": str(gender or "").strip(),
            "series": str(series or "").strip(),
            "article": str(article or kode_mix).strip(),
            "tier": str(tier or "3").strip(),
            "article_mix": kode_kecil.strip() if kode_kecil else kode_mix,
            "reko_box": reko_box,
            "reko_pairs": reko_pairs,
            "avg_sales": avg_f,
            "sales_mix": mix_f,
            "sizes": {},
            "tipe": get_display_type(str(series or "").strip()),
        })

    print(f"Calculated {len(articles)} articles from fresh sales data (12mo history, 3mo avg)")
    return articles

def load_backwall_config(store_code, area):
    """Load backwall/display config from Data Option GSheet."""
    ensure_cache(DATA_OPTION_XLSX, DATA_OPTION_GSHEET)
    wb = openpyxl.load_workbook(DATA_OPTION_XLSX, data_only=True)

    # Try "Jatim New" first (has gender/kategori), then "Jatim"
    sheet_name = f"{area} New" if f"{area} New" in wb.sheetnames else area
    if sheet_name not in wb.sheetnames:
        print(f"WARNING: Sheet '{sheet_name}' not found in Data Option. Using defaults.")
        wb.close()
        return get_default_config(store_code)

    ws = wb[sheet_name]

    # Find store column (stores may be in row 1 or row 2)
    store_col = None
    for search_row in [1, 2, 3]:
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=search_row, column=c).value
            if val and store_code.upper() in str(val).upper():
                store_col = c
                break
        if store_col:
            break

    if not store_col:
        print(f"WARNING: Store '{store_code}' not found in Data Option sheet '{sheet_name}'")
        wb.close()
        return get_default_config(store_code)

    # Read display components (rows 3+)
    config = {"backwalls": [], "gondolas": [], "rak_baby": 0, "basket": 0, "meja": 0, "gudang": 0}
    bw_idx = 0
    for r in range(3, ws.max_row + 1):
        label = str(ws.cell(row=r, column=1).value or "").strip().lower()
        val = ws.cell(row=r, column=store_col).value
        try: hooks = int(float(val)) if val else 0
        except: hooks = 0

        # "New" sheets: col=gender, col+1=kategori, col+2=hooks
        if "new" in sheet_name.lower():
            gender = str(ws.cell(row=r, column=store_col).value or "").strip()
            kategori = str(ws.cell(row=r, column=store_col + 1).value or "").strip()
            hooks_val = ws.cell(row=r, column=store_col + 2).value
            try: hooks = int(float(hooks_val)) if hooks_val else 0
            except: hooks = 0
        else:
            gender = ""
            kategori = ""

        if "back wall" in label or "backwall" in label:
            bw_idx += 1
            if hooks > 0:
                config["backwalls"].append({
                    "id": f"BW-{bw_idx}",
                    "hooks": hooks,
                    "gender": gender,
                    "kategori": kategori,
                })
        elif "gondola" in label:
            if hooks > 0:
                config["gondolas"].append({"hooks": hooks, "gender": gender, "kategori": kategori})
        elif "rak baby" in label or "rak baby kids" in label:
            config["rak_baby"] = hooks
        elif "basket" in label:
            config["basket"] = hooks
        elif "meja" in label or "table" in label:
            config["meja"] = hooks
        elif "gudang" in label or "storage" in label:
            config["gudang"] = hooks

    wb.close()
    print(f"Loaded config for {store_code}: {len(config['backwalls'])} BW, {len(config['gondolas'])} gondola")
    return config

def get_default_config(store_code):
    """Fallback config for unknown stores."""
    return {
        "backwalls": [
            {"id": "BW-1", "hooks": 49, "gender": "Men", "kategori": "jepit"},
            {"id": "BW-2", "hooks": 56, "gender": "Ladies", "kategori": "fashion"},
            {"id": "BW-3", "hooks": 56, "gender": "Men", "kategori": "fashion"},
            {"id": "BW-4", "hooks": 49, "gender": "Ladies/Kids", "kategori": "jepit"},
        ],
        "gondolas": [],
        "rak_baby": 2, "basket": 0, "meja": 1, "gudang": 0,
    }

# ============================================================
# PLANOGRAM ASSIGNMENT
# ============================================================
def assign_articles_to_display(articles, config):
    """Assign articles to backwalls based on gender/type matching and sales ranking."""
    # Sort by avg_sales desc (best sellers first)
    articles_sorted = sorted(articles, key=lambda a: -a["avg_sales"])

    assigned = []  # (article, backwall_id, zone)
    excluded = []

    for art in articles_sorted:
        placed = False
        for bw in config["backwalls"]:
            bw_gender = bw.get("gender", "").lower()
            bw_kat = bw.get("kategori", "").lower()
            art_gender = art["gender"].lower()
            art_tipe = art["tipe"].lower()

            # Gender match
            gender_match = (
                not bw_gender or
                bw_gender in art_gender.lower() or
                art_gender.lower() in bw_gender or
                "all" in bw_gender
            )
            # Type match
            type_match = (
                not bw_kat or
                bw_kat in art_tipe or
                art_tipe in bw_kat or
                "all" in bw_kat
            )

            if not (gender_match and type_match):
                continue

            # Check capacity
            hpa = HOOKS_PER_ARTICLE.get(art_tipe, 2)
            articles_per_col = NHOOKS // hpa
            bw_cols = bw["hooks"] // NHOOKS if bw["hooks"] >= NHOOKS else max(1, bw["hooks"] // hpa)
            max_boxes = bw_cols * articles_per_col

            # Count already assigned to this BW
            already = sum(a["reko_box"] for a, bid, _ in assigned if bid == bw["id"])
            if already + art["reko_box"] <= max_boxes:
                assigned.append((art, bw["id"], art["series"]))
                placed = True
                break

        if not placed:
            # Try baby rak / table display
            if art["gender"].lower() in ("boys", "girls", "baby") and config["rak_baby"] > 0:
                assigned.append((art, "RAK BABY", art["series"]))
            elif art["series"].upper() in ("AIRMOVE", "LUCA", "LUNA") and config["meja"] > 0:
                assigned.append((art, "TABLE DISPLAY", art["series"]))
            else:
                excluded.append((art, "No matching display space"))

    return assigned, excluded

# ============================================================
# VISUAL EXCEL (v4 format)
# ============================================================
def build_stacked_grid(sku_list, hooks_per_article, nhooks=7):
    """Build stacked grid: articles stack vertically within columns."""
    articles_per_col = nhooks // hooks_per_article
    boxes = []
    for art in sku_list:
        for i in range(art["reko_box"]):
            boxes.append((art, i))

    ncols = math.ceil(len(boxes) / articles_per_col) if boxes else 1
    grid = [[None for _ in range(ncols)] for _ in range(nhooks)]

    bi = 0
    for col in range(ncols):
        for slot in range(articles_per_col):
            if bi >= len(boxes):
                break
            art, box_idx = boxes[bi]
            start_hook = slot * hooks_per_article
            for h in range(hooks_per_article):
                if start_hook + h < nhooks:
                    label = f"{art['article_mix']}\n({art['series']})"
                    if box_idx == 0:
                        label += f"\nx{art['reko_box']}box"
                    grid[start_hook + h][col] = (art["article_mix"], art["series"], label)
            bi += 1
    return grid, ncols

def write_grid(ws, start_row, start_col, title, grid, ncols, nhooks):
    """Write stacked grid to worksheet."""
    ZONE_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    COL_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    HOOK_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    r = start_row
    cell = ws.cell(row=r, column=start_col, value=title)
    cell.font = Font(bold=True, size=10, color="FFFFFF")
    cell.fill = ZONE_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    for cc in range(start_col, start_col + ncols + 1):
        ws.cell(row=r, column=cc).border = THIN
    r += 1

    # Column headers
    ws.cell(row=r, column=start_col, value="").border = THIN
    for ci in range(ncols):
        cell = ws.cell(row=r, column=start_col + 1 + ci, value=f"Kol {ci+1}")
        cell.font = Font(bold=True, size=8)
        cell.fill = COL_FILL
        cell.alignment = CELL_ALIGN
        cell.border = THIN
    r += 1

    for ri in range(nhooks):
        cell = ws.cell(row=r + ri, column=start_col, value=f"Hook {ri+1}")
        cell.font = Font(size=8, italic=True)
        cell.alignment = CELL_ALIGN
        cell.border = THIN
        cell.fill = HOOK_FILL
        for ci in range(ncols):
            cell = ws.cell(row=r + ri, column=start_col + 1 + ci)
            cell.border = THIN
            cell.alignment = CELL_ALIGN
            if grid[ri][ci]:
                _, series, label = grid[ri][ci]
                cell.value = label
                cell.font = Font(size=7)
                hx = get_color(series)
                cell.fill = PatternFill(start_color=hx, end_color=hx, fill_type="solid")
            else:
                cell.value = "(KOSONG)"
                cell.font = Font(size=7, color="999999")
                cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    return r + nhooks

def generate_visual_excel(assigned, excluded, store_name, config, output_path):
    """Generate v4 format Excel with Layout Visual + Summary sheets."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Layout Visual"

    for ci in range(1, 50):
        ws.column_dimensions[get_column_letter(ci)].width = 18
    ws.column_dimensions['A'].width = 8

    ws.cell(row=1, column=1, value=f"LAYOUT PLANOGRAM — {store_name.upper()}").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"Generated by Plano-Agent — {datetime.now().strftime('%d %B %Y')}").font = Font(size=10, italic=True, color="666666")

    # Group assigned articles by backwall
    bw_groups = {}
    for art, bw_id, zone in assigned:
        bw_groups.setdefault(bw_id, []).append(art)

    cur_row = 4
    cur_col = 1
    max_row = cur_row

    for bw in config["backwalls"]:
        bw_id = bw["id"]
        if bw_id not in bw_groups:
            continue

        arts = bw_groups[bw_id]
        tipe = bw.get("kategori", "jepit").lower()
        if "fashion" in tipe:
            hpa = 3
        else:
            hpa = 2

        grid, ncols = build_stacked_grid(arts, hpa, NHOOKS)
        total_box = sum(a["reko_box"] for a in arts)
        title = f"{bw_id} — {bw.get('gender','')} {bw.get('kategori','')} ({ncols}x{NHOOKS}, {len(arts)} SKU, {total_box} box)"
        end_row = write_grid(ws, cur_row, cur_col, title, grid, ncols, NHOOKS)
        max_row = max(max_row, end_row)

        cur_col += ncols + 3
        if cur_col > 30:
            cur_col = 1
            cur_row = max_row + 2

    # Baby/Table
    for special in ["RAK BABY", "TABLE DISPLAY"]:
        if special in bw_groups:
            arts = bw_groups[special]
            grid, ncols = build_stacked_grid(arts, 1, 2)
            total_box = sum(a["reko_box"] for a in arts)
            title = f"{special} ({len(arts)} SKU, {total_box} box)"
            cur_row = max_row + 2
            max_row = write_grid(ws, cur_row, 1, title, grid, ncols, 2)

    # ── Summary Sheet ──
    ws2 = wb.create_sheet("Summary")
    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=9)

    ws2.column_dimensions['A'].width = 5
    ws2.column_dimensions['B'].width = 20
    ws2.column_dimensions['C'].width = 18
    ws2.column_dimensions['D'].width = 20
    ws2.column_dimensions['E'].width = 8
    ws2.column_dimensions['F'].width = 10

    ws2.cell(row=1, column=1, value=f"SUMMARY PLANOGRAM — {store_name.upper()}").font = Font(bold=True, size=14)
    ws2.cell(row=2, column=1, value=f"Generated by Plano-Agent — {datetime.now().strftime('%d %B %Y')}").font = Font(size=10, italic=True, color="666666")

    headers = ["No", "Zone", "Kode Artikel", "Series", "Box", "Pairs"]
    for ci, h in enumerate(headers):
        cell = ws2.cell(row=4, column=ci+1, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CELL_ALIGN
        cell.border = THIN

    r = 5
    no = 1
    grand_sku = grand_box = grand_pairs = 0

    # Group by zone
    zone_data = {}
    for art, bw_id, zone in assigned:
        key = f"{bw_id}: {zone}"
        zone_data.setdefault(key, []).append(art)

    for zone_name, arts in zone_data.items():
        zone_start = r
        zone_box = zone_pairs = 0
        for art in arts:
            pairs = art["reko_box"] * PAIRS_PER_BOX
            ws2.cell(row=r, column=1, value=no).alignment = CELL_ALIGN
            ws2.cell(row=r, column=1).border = THIN
            ws2.cell(row=r, column=2, value=zone_name).border = THIN
            ws2.cell(row=r, column=3, value=art["article_mix"]).border = THIN
            ws2.cell(row=r, column=3).font = Font(size=9)
            cell = ws2.cell(row=r, column=4, value=art["series"])
            cell.border = THIN
            hx = get_color(art["series"])
            cell.fill = PatternFill(start_color=hx, end_color=hx, fill_type="solid")
            ws2.cell(row=r, column=5, value=art["reko_box"]).alignment = CELL_ALIGN
            ws2.cell(row=r, column=5).border = THIN
            ws2.cell(row=r, column=6, value=pairs).alignment = CELL_ALIGN
            ws2.cell(row=r, column=6).border = THIN
            zone_box += art["reko_box"]
            zone_pairs += pairs
            no += 1
            r += 1

        if len(arts) > 1:
            ws2.cell(row=zone_start, column=2).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Subtotal
        sub_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
        for cc in range(1, 7):
            ws2.cell(row=r, column=cc).border = THIN
            ws2.cell(row=r, column=cc).fill = sub_fill
        ws2.cell(row=r, column=2, value=f"Subtotal {zone_name}").font = Font(bold=True, size=9)
        ws2.cell(row=r, column=5, value=zone_box).font = Font(bold=True, size=9)
        ws2.cell(row=r, column=5).alignment = CELL_ALIGN
        ws2.cell(row=r, column=6, value=zone_pairs).font = Font(bold=True, size=9)
        ws2.cell(row=r, column=6).alignment = CELL_ALIGN
        r += 1
        grand_sku += len(arts)
        grand_box += zone_box
        grand_pairs += zone_pairs

    # Grand total
    r += 1
    GT_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    GT_FONT = Font(bold=True, size=10, color="FFFFFF")
    for cc in range(1, 7):
        ws2.cell(row=r, column=cc).border = THIN
        ws2.cell(row=r, column=cc).fill = GT_FILL
    ws2.cell(row=r, column=1, value="GRAND TOTAL").font = GT_FONT
    ws2.cell(row=r, column=4, value=f"{grand_sku} SKU").font = GT_FONT
    ws2.cell(row=r, column=4).fill = GT_FILL
    ws2.cell(row=r, column=4).alignment = CELL_ALIGN
    ws2.cell(row=r, column=5, value=grand_box).font = GT_FONT
    ws2.cell(row=r, column=5).alignment = CELL_ALIGN
    ws2.cell(row=r, column=6, value=grand_pairs).font = GT_FONT
    ws2.cell(row=r, column=6).alignment = CELL_ALIGN

    # Excluded
    if excluded:
        r += 3
        ws2.cell(row=r, column=1, value="EXCLUDED").font = Font(bold=True, size=11, color="CC0000")
        r += 1
        EXC_FILL = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")
        for ci, h in enumerate(["No", "Kode Artikel", "Series", "Box", "Pairs", "Alasan"]):
            cell = ws2.cell(row=r, column=ci+1, value=h)
            cell.font = Font(bold=True, size=9, color="FFFFFF")
            cell.fill = EXC_FILL
            cell.alignment = CELL_ALIGN
            cell.border = THIN
        ws2.column_dimensions['F'].width = 30
        r += 1
        for i, (art, reason) in enumerate(excluded):
            ws2.cell(row=r, column=1, value=i+1).alignment = CELL_ALIGN
            ws2.cell(row=r, column=1).border = THIN
            ws2.cell(row=r, column=2, value=art["article_mix"]).border = THIN
            cell = ws2.cell(row=r, column=3, value=art["series"])
            cell.border = THIN
            hx = get_color(art["series"])
            cell.fill = PatternFill(start_color=hx, end_color=hx, fill_type="solid")
            ws2.cell(row=r, column=4, value=art["reko_box"]).alignment = CELL_ALIGN
            ws2.cell(row=r, column=4).border = THIN
            ws2.cell(row=r, column=5, value=art["reko_box"]*12).alignment = CELL_ALIGN
            ws2.cell(row=r, column=5).border = THIN
            ws2.cell(row=r, column=6, value=reason).border = THIN
            ws2.cell(row=r, column=6).font = Font(size=9, italic=True)
            r += 1

    # ── Sheet 3: Data Detail (raw planogram_paperclip data) ──
    ws3 = wb.create_sheet("Data Detail")
    ws3.cell(row=1, column=1, value=f"DATA DETAIL — {store_name.upper()}").font = Font(bold=True, size=14)
    ws3.cell(row=2, column=1, value="Back-calculation: semua artikel + alasan placement/exclusion. Data ini juga tersimpan di portal.planogram_paperclip.").font = Font(size=9, italic=True, color="666666")

    detail_headers = [
        "No", "Kode Artikel", "Gender", "Series", "Tipe", "Tier",
        "Avg Sales 3mo (pairs)", "Sales Mix %", "Reko Pairs", "Reko Box",
        "Status", "Zone / BW", "Alasan Exclude"
    ]
    for ci, h in enumerate(detail_headers):
        cell = ws3.cell(row=4, column=ci+1, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CELL_ALIGN
        cell.border = THIN

    # Column widths
    detail_widths = [5, 22, 10, 16, 10, 6, 18, 12, 12, 10, 10, 18, 30]
    for ci, w in enumerate(detail_widths):
        ws3.column_dimensions[get_column_letter(ci+1)].width = w

    r3 = 5
    no3 = 1

    # Assigned articles
    for art, bw_id, zone in sorted(assigned, key=lambda x: (-x[0]["avg_sales"], x[0]["article"])):
        ws3.cell(row=r3, column=1, value=no3).alignment = CELL_ALIGN
        ws3.cell(row=r3, column=1).border = THIN
        ws3.cell(row=r3, column=2, value=art["article_mix"]).border = THIN
        ws3.cell(row=r3, column=2).font = Font(size=9)
        ws3.cell(row=r3, column=3, value=art["gender"]).border = THIN
        cell = ws3.cell(row=r3, column=4, value=art["series"])
        cell.border = THIN
        hx = get_color(art["series"])
        cell.fill = PatternFill(start_color=hx, end_color=hx, fill_type="solid")
        ws3.cell(row=r3, column=5, value=art["tipe"]).border = THIN
        ws3.cell(row=r3, column=6, value=art["tier"]).alignment = CELL_ALIGN
        ws3.cell(row=r3, column=6).border = THIN
        ws3.cell(row=r3, column=7, value=round(art["avg_sales"], 2) if art["avg_sales"] else "").alignment = CELL_ALIGN
        ws3.cell(row=r3, column=7).border = THIN
        ws3.cell(row=r3, column=8, value=f"{art['sales_mix']*100:.2f}%" if art["sales_mix"] else "").alignment = CELL_ALIGN
        ws3.cell(row=r3, column=8).border = THIN
        ws3.cell(row=r3, column=9, value=round(art["reko_pairs"], 1) if art["reko_pairs"] else "").alignment = CELL_ALIGN
        ws3.cell(row=r3, column=9).border = THIN
        ws3.cell(row=r3, column=10, value=art["reko_box"]).alignment = CELL_ALIGN
        ws3.cell(row=r3, column=10).border = THIN
        cell_status = ws3.cell(row=r3, column=11, value="PLACED")
        cell_status.font = Font(size=9, bold=True, color="006100")
        cell_status.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        cell_status.alignment = CELL_ALIGN
        cell_status.border = THIN
        ws3.cell(row=r3, column=12, value=f"{bw_id}: {zone}").border = THIN
        ws3.cell(row=r3, column=13, value="").border = THIN
        no3 += 1
        r3 += 1

    # Excluded articles
    for art, reason in sorted(excluded, key=lambda x: (-x[0]["avg_sales"], x[0]["article"])):
        ws3.cell(row=r3, column=1, value=no3).alignment = CELL_ALIGN
        ws3.cell(row=r3, column=1).border = THIN
        ws3.cell(row=r3, column=2, value=art["article_mix"]).border = THIN
        ws3.cell(row=r3, column=2).font = Font(size=9)
        ws3.cell(row=r3, column=3, value=art["gender"]).border = THIN
        cell = ws3.cell(row=r3, column=4, value=art["series"])
        cell.border = THIN
        hx = get_color(art["series"])
        cell.fill = PatternFill(start_color=hx, end_color=hx, fill_type="solid")
        ws3.cell(row=r3, column=5, value=art["tipe"]).border = THIN
        ws3.cell(row=r3, column=6, value=art["tier"]).alignment = CELL_ALIGN
        ws3.cell(row=r3, column=6).border = THIN
        ws3.cell(row=r3, column=7, value=round(art["avg_sales"], 2) if art["avg_sales"] else "").alignment = CELL_ALIGN
        ws3.cell(row=r3, column=7).border = THIN
        ws3.cell(row=r3, column=8, value=f"{art['sales_mix']*100:.2f}%" if art["sales_mix"] else "").alignment = CELL_ALIGN
        ws3.cell(row=r3, column=8).border = THIN
        ws3.cell(row=r3, column=9, value=round(art["reko_pairs"], 1) if art["reko_pairs"] else "").alignment = CELL_ALIGN
        ws3.cell(row=r3, column=9).border = THIN
        ws3.cell(row=r3, column=10, value=art["reko_box"]).alignment = CELL_ALIGN
        ws3.cell(row=r3, column=10).border = THIN
        cell_status = ws3.cell(row=r3, column=11, value="EXCLUDED")
        cell_status.font = Font(size=9, bold=True, color="9C0006")
        cell_status.fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
        cell_status.alignment = CELL_ALIGN
        cell_status.border = THIN
        ws3.cell(row=r3, column=12, value="").border = THIN
        ws3.cell(row=r3, column=13, value=reason).border = THIN
        ws3.cell(row=r3, column=13).font = Font(size=9, italic=True)
        no3 += 1
        r3 += 1

    # Autofilter
    ws3.auto_filter.ref = f"A4:M{r3-1}"

    wb.save(output_path)
    print(f"Visual Excel saved: {output_path}")
    return grand_sku, grand_box, grand_pairs

# ============================================================
# DATABASE WRITE
# ============================================================
def write_to_db(assigned, excluded, store_db, area):
    """Write planogram to portal.planogram_paperclip (UPSERT per store)."""
    conn = psycopg2.connect(
        host=os.environ.get("PGHOST", "76.13.194.120"),
        port=os.environ.get("PGPORT", "5432"),
        dbname=os.environ.get("PGDATABASE", "openclaw_ops"),
        user=os.environ.get("PGUSER", "openclaw_app"),
        password=os.environ.get("PGPASSWORD", ""),
    )
    cur = conn.cursor()

    # Delete old data for this store
    cur.execute("DELETE FROM portal.planogram_paperclip WHERE store_name = %s", (store_db,))
    deleted = cur.rowcount
    print(f"Deleted {deleted} old rows for {store_db}")

    # Insert new
    inserted = 0
    for art, bw_id, zone in assigned:
        cur.execute("""
            INSERT INTO portal.planogram_paperclip
            (store_name, gender, series, article, tier, article_mix, kode_kecil,
             grand_total_pairs, box, area, tipe, avg_sales_3mo, sales_mix,
             rekomendasi_pairs, rekomendasi_box, updated_at, generated_by)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, now(), 'paperclip-plano-agent')
        """, (
            store_db, art["gender"], art["series"], art["article"], art["tier"],
            art["article_mix"], art["article_mix"],  # kode_kecil = article_mix
            str(art["reko_box"] * PAIRS_PER_BOX), str(art["reko_box"]),
            area, art["tipe"],
            str(art["avg_sales"]) if art["avg_sales"] else "",
            str(art["sales_mix"]) if art["sales_mix"] else "",
            str(art["reko_pairs"]) if art["reko_pairs"] else "",
            str(art["reko_box"]),
        ))
        inserted += 1

    conn.commit()
    cur.close()
    conn.close()
    print(f"Inserted {inserted} rows for {store_db}")

# ============================================================
# MAIN
# ============================================================
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--store-code", required=True)
    parser.add_argument("--store-name", required=True)
    parser.add_argument("--store-db", required=True)
    parser.add_argument("--area", default="Jatim")
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    print(f"=== Planogram Agent: {args.store_name} ({args.store_code}) ===")

    # 1. Load backwall config first (need total hooks for display capacity)
    config = load_backwall_config(args.store_code, args.area)
    total_hooks = sum(bw["hooks"] for bw in config.get("backwalls", []))
    total_hooks += sum(gd["hooks"] for gd in config.get("gondolas", []))
    total_display_pairs = total_hooks * 5  # ~5 pairs per hook average

    # 2. Calculate fresh planogram from sales history
    articles = load_store_planogram_data(args.store_name, args.store_db, args.area, total_display_pairs)
    if not articles:
        print(f"ERROR: No sales data found for {args.store_db}")
        sys.exit(1)

    # 3. Assign articles to display
    assigned, excluded = assign_articles_to_display(articles, config)
    print(f"Assigned: {len(assigned)} articles, Excluded: {len(excluded)}")

    # 4. Generate visual Excel
    sku, box, pairs = generate_visual_excel(assigned, excluded, args.store_name, config, args.output)
    print(f"Output: {sku} SKU, {box} box, {pairs} pairs")

    # 5. Write to database
    write_to_db(assigned, excluded, args.store_db, args.area)

    print(f"=== Done: {args.store_name} ===")

if __name__ == "__main__":
    main()
