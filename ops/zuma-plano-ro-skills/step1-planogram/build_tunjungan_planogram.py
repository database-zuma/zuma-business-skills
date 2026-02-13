"""
PLANOGRAM BUILDER -- Zuma Tunjungan Plaza
Based on SKILL_planogram_zuma_v3.md logic
Adapted from build_royal_planogram.py reference implementation
"""

import os
import psycopg2
import pandas as pd
import numpy as np
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
import math

# ============================================================
# CONFIG
# ============================================================
STORE_NAME = "Zuma Tunjungan Plaza"
DB_SALES_NAME = "zuma tunjungan plaza"
DB_STOCK_PATTERN = "%tunjungan plaza%"
STORAGE_CAPACITY = 75  # boxes

# Months for analysis (12 months: Feb 2025 - Jan 2026)
DATE_START = "2025-02-01"
DATE_END = "2026-01-31"

DB_CONFIG = {
    "host": "76.13.194.120",
    "port": 5432,
    "dbname": "openclaw_ops",
    "user": "openclaw_app",
    "password": "Zuma-0psCl4w-2026!",
}

OUTPUT_FILE = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "PLANOGRAM_Tunjungan_Plaza.xlsx"
)

# ============================================================
# DISPLAY UNIT DEFINITIONS — Tunjungan Plaza Layout
# ============================================================
# Backwalls with FIXED gender-type assignments
BACKWALLS = [
    {
        "id": "BW-1",
        "label": "Men Fashion",
        "orientation": "horizontal",
        "hooks": 56,
        "grid": "8x7",
        "hpa": 3,
        "max_articles": 18,
        "gender_type": "Men Fashion",
        "series_filter": None,  # All men fashion EXCEPT Airmove
        "series_exclude": ["AIRMOVE"],
    },
    {
        "id": "BW-2",
        "label": "Ladies Fashion",
        "orientation": "horizontal",
        "hooks": 56,
        "grid": "8x7",
        "hpa": 3,
        "max_articles": 18,
        "gender_type": "Ladies Fashion",
        "series_filter": None,  # All ladies fashion except Elsa (BW-4a-Elsa handles that)
        "series_exclude": ["ELSA"],
    },
    {
        "id": "BW-3",
        "label": "Men Jepit",
        "orientation": "vertical",
        "hooks": 63,
        "grid": "9x7",
        "hpa": 2,
        "max_articles": 31,
        "gender_type": "Men Jepit",
        "series_filter": ["ONYX", "CLASSIC", "STRIPE", "BLACKSERIES", "CLASSICEARTH"],
        "series_exclude": [],
    },
    {
        "id": "BW-4a-Elsa",
        "label": "Ladies Elsa",
        "orientation": "vertical",
        "hooks": 14,
        "grid": "part of BW-4a",
        "hpa": 3,
        "max_articles": 4,
        "gender_type": "Ladies Fashion",
        "series_filter": ["ELSA"],
        "series_exclude": [],
    },
    {
        "id": "BW-4a-Classic",
        "label": "Ladies Classic",
        "orientation": "vertical",
        "hooks": 28,
        "grid": "part of BW-4a",
        "hpa": 2,
        "max_articles": 14,
        "gender_type": "Ladies Jepit",
        "series_filter": ["CLASSIC", "CLASSICEARTH", "CLASSIC METALIC"],
        "series_exclude": [],
    },
    {
        "id": "BW-4b",
        "label": "Baby and Kids",
        "orientation": "vertical",
        "hooks": 49,
        "grid": "7x7",
        "hpa": 2,
        "max_articles": 24,
        "gender_type": "Baby & Kids",
        "series_filter": None,
        "series_exclude": [],
    },
]

# Special display units
SHELVING_AIRMOVE = {
    "id": "SHELVING-AIRMOVE",
    "slots": 3,
    "pairs_per_article": 10,
    "series": "AIRMOVE",
}
SHELVING_PUFFY = {
    "id": "SHELVING-PUFFY",
    "slots": 1,
    "pairs_per_article": 10,
    "series": "PUFFY",
}
TABLE_LUCALUNA = {
    "id": "TABLE-LUCALUNA",
    "slots": 5,
    "pairs_per_article": 11,
    "series": ["LUCA", "LUNA", "AIRMOVE"],
}
TABLE_BABY = {"id": "TABLE-BABY", "slots": 6, "pairs_per_article": 6, "gender": "BABY"}

# Color fills for tiers
TIER_COLORS = {
    "1": PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    ),  # Green - T1
    "8": PatternFill(
        start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"
    ),  # Blue - T8
    "2": PatternFill(
        start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"
    ),  # Yellow - T2
    "3": PatternFill(
        start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"
    ),  # Grey - T3
}
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14)
SUBTITLE_FONT = Font(bold=True, size=11)
AVAIL_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
SPECIAL_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
STORAGE_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


# ============================================================
# STEP 1: PULL DATA FROM DB
# ============================================================
def pull_data():
    conn = psycopg2.connect(
        host=DB_CONFIG["host"],
        port=DB_CONFIG["port"],
        dbname=DB_CONFIG["dbname"],
        user=DB_CONFIG["user"],
        password=DB_CONFIG["password"],
    )

    # 1a. Sales data -- USE core.sales_with_product (Rule 5: use core views, not raw)
    # Rule 7: exclude intercompany transactions
    # Use kode_mix for version-agnostic article aggregation (Rule 6)
    sales_query = f"""
        SELECT
            kode_mix,
            article,
            gender,
            series,
            tipe AS product_type,
            tier,
            TO_CHAR(transaction_date, 'YYYY-MM') as bulan,
            SUM(quantity) as total_qty,
            SUM(total_amount) as total_revenue
        FROM core.sales_with_product
        WHERE matched_store_name = TRIM(LOWER('{DB_SALES_NAME}'))
          AND transaction_date >= '{DATE_START}'
          AND transaction_date <= '{DATE_END}'
          AND quantity > 0
          AND (is_intercompany IS NULL OR is_intercompany = FALSE)
        GROUP BY kode_mix, article, gender, series, tipe, tier,
                 TO_CHAR(transaction_date, 'YYYY-MM')
        ORDER BY kode_mix, bulan
    """
    sales_df = pd.read_sql(sales_query, conn)

    # 1b. Master kodemix -- NO status filter (Rule 3: never filter by status)
    # Rule 1: use TRIM(LOWER()) for matching
    # Rule 2: deduplicate with DISTINCT ON
    master_query = """
        SELECT DISTINCT ON (TRIM(LOWER(kode_besar)))
            TRIM(LOWER(kode_besar)) AS kode_besar,
            kode, kode_mix, kode_mix_size, article, gender, series, tipe,
            tier_baru, status, version
        FROM portal.kodemix
        ORDER BY TRIM(LOWER(kode_besar)), no_urut
    """
    master_df = pd.read_sql(master_query, conn)

    # 1c. Stock data -- USE core.stock_with_product (Rule 5)
    stock_query = f"""
        SELECT
            kode_mix, article, gender, series, tipe AS product_type,
            tier, nama_gudang, SUM(quantity) as total_qty
        FROM core.stock_with_product
        WHERE TRIM(LOWER(nama_gudang)) LIKE '{DB_STOCK_PATTERN}'
        GROUP BY kode_mix, article, gender, series, tipe, tier, nama_gudang
        ORDER BY total_qty DESC
    """
    stock_df = pd.read_sql(stock_query, conn)

    conn.close()
    return sales_df, master_df, stock_df


# ============================================================
# STEP 2: PROCESS SALES (already enriched from core view)
# ============================================================
def process_sales(sales_df, master_df):
    """
    Since we use core.sales_with_product, data is already enriched with
    kode_mix, article, gender, series, tipe, tier. No manual joining needed.
    Just clean and rename columns.
    """
    # Rename tier column to match our convention
    sales_df = sales_df.rename(columns={"product_type": "tipe", "tier": "tier_baru"})

    # Drop rows with no kode_mix (unmatched products ~6%)
    unmatched = sales_df[sales_df["kode_mix"].isna()]
    if len(unmatched) > 0:
        print(f"  INFO: {len(unmatched)} rows have no kode_mix match (excluded)")
    sales_df = sales_df[sales_df["kode_mix"].notna()].copy()

    # Drop rows with no article name
    sales_df = sales_df[sales_df["article"].notna()].copy()

    # Drop non-product items
    non_product_patterns = [
        "SHOPPING BAG",
        "HANGER",
        "PAPER BAG",
        "THERMAL",
        "BOX LUCA",
    ]
    mask = (
        sales_df["article"]
        .str.upper()
        .apply(lambda x: not any(p in str(x) for p in non_product_patterns))
    )
    sales_df = sales_df[mask].copy()

    # Standardize tier values
    sales_df["tier_baru"] = sales_df["tier_baru"].fillna("3").astype(str)

    # Standardize gender
    sales_df["gender"] = sales_df["gender"].fillna("UNKNOWN").str.upper()
    sales_df = sales_df[sales_df["gender"] != "UNKNOWN"]

    # Standardize tipe
    sales_df["tipe"] = sales_df["tipe"].fillna("Jepit")

    print(f"  Clean sales rows: {len(sales_df)}")
    print(f"  Unique articles: {sales_df['article'].nunique()}")
    print(f"  Unique kode_mix: {sales_df['kode_mix'].nunique()}")

    return sales_df


# ============================================================
# STEP 3: COMPUTE ADJUSTED AVERAGE (per skill Section 1.4.2)
# ============================================================
def compute_adjusted_avg(merged_df):
    # Generate all months in range
    all_months = (
        pd.date_range(DATE_START, DATE_END, freq="MS").strftime("%Y-%m").tolist()
    )

    # Aggregate to article level
    article_monthly = (
        merged_df.groupby(["article", "gender", "series", "tipe", "tier_baru", "bulan"])
        .agg(total_qty=("total_qty", "sum"), total_revenue=("total_revenue", "sum"))
        .reset_index()
    )

    result_rows = []
    for (article, gender, series, tipe, tier), group in article_monthly.groupby(
        ["article", "gender", "series", "tipe", "tier_baru"]
    ):
        monthly = {}
        for _, row in group.iterrows():
            monthly[row["bulan"]] = float(row["total_qty"])

        # Fill missing months with 0
        month_values = [monthly.get(m, 0) for m in all_months]
        total_sales = sum(month_values)

        # Adjusted average per tier rules
        tier_str = str(tier)
        if tier_str == "1":
            # T1: exclude 0 months (likely OOS)
            nonzero = [v for v in month_values if v > 0]
            adj_avg = sum(nonzero) / len(nonzero) if nonzero else 0
        elif tier_str == "8":
            # T8: exclude leading 0s (not launched yet) and post-launch 0s (OOS)
            first_sale_idx = next(
                (i for i, v in enumerate(month_values) if v > 0), None
            )
            if first_sale_idx is not None:
                active = [v for v in month_values[first_sale_idx:] if v > 0]
                adj_avg = sum(active) / len(active) if active else 0
            else:
                adj_avg = 0
        elif tier_str in ("2", "3"):
            # T2/T3: contextual -- if 0 is surrounded by decent months, exclude
            overall_avg = total_sales / len(all_months) if len(all_months) > 0 else 0
            active_months = 0
            active_total = 0
            for i, v in enumerate(month_values):
                if v > 0:
                    active_months += 1
                    active_total += v
                else:
                    # Check surrounding months
                    surrounding = []
                    for j in range(max(0, i - 2), min(len(month_values), i + 3)):
                        if j != i and month_values[j] > 0:
                            surrounding.append(month_values[j])
                    if surrounding and np.mean(surrounding) > overall_avg * 0.5:
                        pass  # Exclude this 0 (likely OOS)
                    else:
                        active_months += 1  # Include 0 (genuine decline)
                        active_total += 0
            adj_avg = active_total / active_months if active_months > 0 else 0
        else:
            # T4/T5: full average
            adj_avg = total_sales / len(all_months) if len(all_months) > 0 else 0

        result_rows.append(
            {
                "article": article,
                "gender": gender,
                "series": series,
                "tipe": tipe,
                "tier": tier_str,
                "total_12mo": total_sales,
                "adj_avg": round(adj_avg, 1),
                "months_sold": sum(1 for v in month_values if v > 0),
                "month_values": month_values,
            }
        )

    result_df = pd.DataFrame(result_rows)
    return result_df, all_months


# ============================================================
# STEP 4: GENDER-TYPE MAPPING & SALES SHARE
# ============================================================
def map_gender_type(row):
    gender = str(row["gender"]).upper()
    tipe = str(row["tipe"])
    if gender == "MEN":
        return f"Men {tipe}"
    elif gender == "LADIES":
        return f"Ladies {tipe}"
    elif gender == "BABY":
        return "Baby & Kids"
    elif gender == "BOYS":
        return "Baby & Kids"
    elif gender == "GIRLS":
        return "Baby & Kids"
    elif gender == "JUNIOR":
        return "Baby & Kids"
    else:
        return f"{gender} {tipe}"


def compute_gender_shares(result_df):
    result_df["gender_type"] = result_df.apply(map_gender_type, axis=1)

    shares = (
        result_df.groupby("gender_type")
        .agg(total_adj_avg=("adj_avg", "sum"), article_count=("article", "count"))
        .reset_index()
    )
    total = shares["total_adj_avg"].sum()
    shares["pct_share"] = (shares["total_adj_avg"] / total * 100).round(1)
    shares = shares.sort_values("total_adj_avg", ascending=False).reset_index(drop=True)
    shares["rank"] = range(1, len(shares) + 1)

    return shares


# ============================================================
# STEP 5: PRE-ASSIGN SPECIAL DISPLAYS
# ============================================================
def pre_assign_airmove(result_df, shelving_config):
    """Assign top Airmove articles to SHELVING-AIRMOVE (3 slots)."""
    airmove = result_df[
        (result_df["series"].str.upper() == "AIRMOVE")
        & (result_df["tier"].isin(["1", "8", "2", "3"]))
    ].copy()
    tier_priority = {"1": 0, "8": 1, "2": 2, "3": 3}
    airmove["tier_sort"] = airmove["tier"].map(tier_priority).fillna(9)
    airmove = airmove.sort_values(["tier_sort", "adj_avg"], ascending=[True, False])
    assigned = airmove.head(shelving_config["slots"]).reset_index(drop=True)
    return assigned


def pre_assign_lucaluna(result_df, table_config, already_assigned_names):
    """Assign top Luca/Luna/Airmove articles to TABLE-LUCALUNA (5 slots).
    Airmove articles already on SHELVING-AIRMOVE are excluded."""
    mask = (
        result_df["series"].str.upper().isin(["LUCA", "LUNA", "AIRMOVE"])
        & result_df["tier"].isin(["1", "8", "2", "3"])
        & ~result_df["article"].isin(already_assigned_names)
    )
    candidates = result_df[mask].copy()
    tier_priority = {"1": 0, "8": 1, "2": 2, "3": 3}
    candidates["tier_sort"] = candidates["tier"].map(tier_priority).fillna(9)
    candidates = candidates.sort_values(
        ["tier_sort", "adj_avg"], ascending=[True, False]
    )
    assigned = candidates.head(table_config["slots"]).reset_index(drop=True)
    return assigned


def pre_assign_baby_table(result_df, table_config, already_assigned_names):
    """Assign top Baby articles to TABLE-BABY (6 slots)."""
    mask = (
        (result_df["gender"].str.upper().isin(["BABY", "BOYS", "GIRLS", "JUNIOR"]))
        & result_df["tier"].isin(["1", "8", "2", "3"])
        & ~result_df["article"].isin(already_assigned_names)
    )
    candidates = result_df[mask].copy()
    tier_priority = {"1": 0, "8": 1, "2": 2, "3": 3}
    candidates["tier_sort"] = candidates["tier"].map(tier_priority).fillna(9)
    candidates = candidates.sort_values(
        ["tier_sort", "adj_avg"], ascending=[True, False]
    )
    assigned = candidates.head(table_config["slots"]).reset_index(drop=True)
    return assigned


def pre_assign_puffy(result_df, shelving_config, already_assigned_names):
    """Assign top 1 Puffy article to SHELVING-PUFFY."""
    mask = (
        (result_df["series"].str.upper().str.contains("PUFFY", na=False))
        & result_df["tier"].isin(["1", "8", "2", "3"])
        & ~result_df["article"].isin(already_assigned_names)
    )
    candidates = result_df[mask].copy()
    tier_priority = {"1": 0, "8": 1, "2": 2, "3": 3}
    candidates["tier_sort"] = candidates["tier"].map(tier_priority).fillna(9)
    candidates = candidates.sort_values(
        ["tier_sort", "adj_avg"], ascending=[True, False]
    )
    assigned = candidates.head(shelving_config["slots"]).reset_index(drop=True)
    return assigned


# ============================================================
# STEP 6: ASSIGN ARTICLES TO BACKWALLS
# ============================================================
def assign_articles_to_backwalls(backwalls, result_df, pre_assigned_names):
    """
    Assign articles to backwalls with FIXED gender-type assignments.
    - Exclude pre-assigned articles (Airmove, Luca/Luna, Baby table, Puffy)
    - Exclude Luca/Luna/Airmove from hook backwalls entirely
    - Exclude T4/T5
    - Exclude articles already assigned to earlier backwalls (no duplication)
    - Priority: T1 -> T8 -> T2 -> T3, sorted by adj_avg DESC
    """
    planogram = {}
    # Track articles assigned across ALL backwalls to prevent duplication
    bw_assigned_names = set(pre_assigned_names)

    # Luca/Luna/Airmove exclusion patterns (table/shelving only, not backwall hooks)
    TABLE_VM_ONLY = ["LUCA", "LUNA", "AIRMOVE"]

    for bw in backwalls:
        bw_id = bw["id"]
        gt = bw["gender_type"]
        hpa = bw["hpa"]
        max_articles = bw["max_articles"]
        hooks = bw["hooks"]
        slots = hooks // hpa
        # Cap at max_articles if specified
        if max_articles and max_articles < slots:
            slots = max_articles

        series_filter = bw.get("series_filter")
        series_exclude = bw.get("series_exclude", [])

        # Filter articles for this gender-type
        articles = result_df[result_df["gender_type"] == gt].copy()

        # Apply series filter if specified (e.g., BW-3 only ONYX/CLASSIC/STRIPE/BLACKSERIES/CLASSICEARTH)
        if series_filter:
            articles = articles[
                articles["series"].str.upper().isin([s.upper() for s in series_filter])
            ]

        # Apply series exclusion (e.g., BW-1 excludes AIRMOVE)
        if series_exclude:
            for excl in series_exclude:
                articles = articles[
                    ~articles["series"].str.upper().str.contains(excl.upper(), na=False)
                ]

        # Exclude Luca/Luna/Airmove from ALL backwalls (Section 1.3.7)
        luca_mask = (
            articles["series"]
            .str.upper()
            .apply(lambda x: not any(p in str(x) for p in TABLE_VM_ONLY))
        )
        excluded_luca = articles[~luca_mask]
        if len(excluded_luca) > 0:
            for _, art in excluded_luca.iterrows():
                print(
                    f"  EXCLUDED from {bw_id}: {art['article']} (Luca/Luna/Airmove -> table/shelving only)"
                )
        articles = articles[luca_mask]

        # Exclude all already-assigned articles (pre-assigned + earlier backwalls)
        articles = articles[~articles["article"].isin(bw_assigned_names)]

        # Sort by: tier priority (T1 first, then T8, T2, T3), then adj_avg DESC
        tier_priority = {"1": 0, "8": 1, "2": 2, "3": 3, "4": 4, "5": 5}
        articles["tier_sort"] = articles["tier"].map(tier_priority).fillna(9)
        articles = articles.sort_values(
            ["tier_sort", "adj_avg"], ascending=[True, False]
        )

        # Filter: only T1, T8, T2, T3 (no T4/T5)
        articles = articles[articles["tier"].isin(["1", "8", "2", "3"])]

        # Check if we need compact mode
        t1_count = len(articles[articles["tier"] == "1"])
        t1_t8_count = len(articles[articles["tier"].isin(["1", "8"])])

        mode = "Full Box"
        actual_hpa = hpa

        # Decision: if T1+T8 > slots in full box mode, try compact
        if t1_t8_count > slots:
            # Try compact mode
            if "Jepit" in gt or gt == "Baby & Kids":
                compact_hpa = 1
            else:
                compact_hpa = 2
            compact_slots = hooks // compact_hpa
            if max_articles and max_articles < compact_slots:
                compact_slots = max_articles

            if t1_t8_count <= compact_slots:
                mode = "Compact"
                actual_hpa = compact_hpa
                slots = compact_slots
                print(
                    f"  {bw_id}: Switched to COMPACT mode ({t1_t8_count} T1+T8 > {hooks // hpa} full slots)"
                )
            else:
                # Even compact not enough, just fill what we can
                mode = "Compact"
                actual_hpa = compact_hpa
                slots = compact_slots
                print(
                    f"  WARNING: {bw_id}: Even compact mode insufficient ({t1_t8_count} T1+T8 > {compact_slots} compact slots)"
                )

        # Assign to slots
        assigned = articles.head(slots).reset_index(drop=True)

        # Check if all T1 fit
        if t1_count > slots:
            print(
                f"  WARNING: {bw_id} ({gt}): {t1_count} T1 articles but only {slots} slots!"
            )

        planogram[bw_id] = {
            "config": {**bw, "slots": slots, "mode": mode, "actual_hpa": actual_hpa},
            "articles": assigned,
            "available_slots": max(0, slots - len(assigned)),
        }

        # Track assigned articles to prevent duplication in subsequent backwalls
        for _, art in assigned.iterrows():
            bw_assigned_names.add(art["article"])

    return planogram


# ============================================================
# STEP 7: STORAGE ALLOCATION (75 box budget)
# ============================================================
def allocate_storage(
    planogram,
    airmove_articles,
    lucaluna_articles,
    baby_table_articles,
    puffy_articles,
    result_df,
    storage_capacity,
):
    """
    Allocate storage following priority order:
    1. Table/Shelving articles first (1 box each ~8 boxes)
    2. Table Baby (~3 boxes for 6 articles at ~6 pairs each)
    3. Compact overflow
    4. T1 backup
    5. T8 backup
    6. T2 remainder
    """
    storage_items = []
    used_boxes = 0

    # --- Priority 1: Table/Shelving articles (Airmove, Luca/Luna, Puffy) ---
    # Each gets 1 box in storage (display uses ~1-2 pairs, rest in storage)
    for _, art in airmove_articles.iterrows():
        storage_items.append(
            {
                "article": art["article"],
                "gender_type": art.get("gender_type", ""),
                "series": art["series"],
                "tier": art["tier"],
                "adj_avg": art["adj_avg"],
                "boxes": 1,
                "reason": "SHELVING-AIRMOVE backup (11 pairs in storage)",
                "priority": 1,
            }
        )
        used_boxes += 1

    for _, art in lucaluna_articles.iterrows():
        storage_items.append(
            {
                "article": art["article"],
                "gender_type": art.get("gender_type", ""),
                "series": art["series"],
                "tier": art["tier"],
                "adj_avg": art["adj_avg"],
                "boxes": 1,
                "reason": "TABLE-LUCALUNA backup (11 pairs in storage)",
                "priority": 1,
            }
        )
        used_boxes += 1

    for _, art in puffy_articles.iterrows():
        storage_items.append(
            {
                "article": art["article"],
                "gender_type": art.get("gender_type", ""),
                "series": art["series"],
                "tier": art["tier"],
                "adj_avg": art["adj_avg"],
                "boxes": 1,
                "reason": "SHELVING-PUFFY backup (11 pairs in storage)",
                "priority": 1,
            }
        )
        used_boxes += 1

    # --- Priority 2: Table Baby (6 articles, ~6 pairs each = ~3 boxes total) ---
    baby_pairs_total = len(baby_table_articles) * TABLE_BABY["pairs_per_article"]
    baby_boxes = math.ceil(baby_pairs_total / 12)
    if len(baby_table_articles) > 0:
        # Distribute baby boxes across articles
        for i, (_, art) in enumerate(baby_table_articles.iterrows()):
            box_alloc = 1 if i < baby_boxes else 0
            if box_alloc > 0:
                storage_items.append(
                    {
                        "article": art["article"],
                        "gender_type": art.get("gender_type", "Baby & Kids"),
                        "series": art["series"],
                        "tier": art["tier"],
                        "adj_avg": art["adj_avg"],
                        "boxes": box_alloc,
                        "reason": f"TABLE-BABY backup ({TABLE_BABY['pairs_per_article']} pairs on display)",
                        "priority": 2,
                    }
                )
                used_boxes += box_alloc

    # --- Priority 3: Compact mode overflow ---
    compact_overflow_pairs = 0
    for bw_id, data in planogram.items():
        config = data["config"]
        if config["mode"] == "Compact":
            for _, art in data["articles"].iterrows():
                tipe = str(art.get("tipe", ""))
                if "Jepit" in str(config.get("gender_type", "")) or tipe == "Jepit":
                    compact_overflow_pairs += 6  # half box
                else:
                    compact_overflow_pairs += 4  # 2/3 box fashion

    if compact_overflow_pairs > 0:
        compact_boxes = math.ceil(compact_overflow_pairs / 12)
        storage_items.append(
            {
                "article": "(Compact mode overflow)",
                "gender_type": "Mixed",
                "series": "-",
                "tier": "-",
                "adj_avg": 0,
                "boxes": compact_boxes,
                "reason": f"Compact mode overflow ({compact_overflow_pairs} pairs = {compact_boxes} box equiv)",
                "priority": 3,
            }
        )
        used_boxes += compact_boxes

    # --- Priority 4: T1 fast moving backup (proportional) ---
    remaining = storage_capacity - used_boxes
    if remaining > 0:
        # Collect all displayed T1 articles
        displayed_t1 = []
        all_displayed_names = set()

        # From backwalls
        for bw_id, data in planogram.items():
            for _, art in data["articles"].iterrows():
                all_displayed_names.add(art["article"])
                if str(art["tier"]) == "1":
                    displayed_t1.append(art)

        # From special displays
        for df in [
            airmove_articles,
            lucaluna_articles,
            puffy_articles,
            baby_table_articles,
        ]:
            for _, art in df.iterrows():
                all_displayed_names.add(art["article"])
                if str(art["tier"]) == "1":
                    displayed_t1.append(art)

        if displayed_t1:
            total_t1_avg = sum(a["adj_avg"] for a in displayed_t1)
            # Allocate up to 60% of remaining for T1
            t1_budget = min(remaining, int(remaining * 0.6))
            t1_allocated = 0

            for art in sorted(displayed_t1, key=lambda x: x["adj_avg"], reverse=True):
                if t1_allocated >= t1_budget:
                    break
                # Already has storage from special display?
                already_stored = any(
                    s["article"] == art["article"] for s in storage_items
                )
                if already_stored:
                    continue
                share = art["adj_avg"] / total_t1_avg if total_t1_avg > 0 else 0
                box_alloc = max(1, round(t1_budget * share))
                box_alloc = min(box_alloc, t1_budget - t1_allocated)
                if box_alloc > 0:
                    storage_items.append(
                        {
                            "article": art["article"],
                            "gender_type": art.get("gender_type", ""),
                            "series": art.get("series", ""),
                            "tier": "1",
                            "adj_avg": art["adj_avg"],
                            "boxes": box_alloc,
                            "reason": "T1 fast moving backup",
                            "priority": 4,
                        }
                    )
                    used_boxes += box_alloc
                    t1_allocated += box_alloc

    # --- Priority 5: T8 backup ---
    remaining = storage_capacity - used_boxes
    if remaining > 0:
        displayed_t8 = []
        for bw_id, data in planogram.items():
            for _, art in data["articles"].iterrows():
                if str(art["tier"]) == "8":
                    displayed_t8.append(art)
        for df in [
            airmove_articles,
            lucaluna_articles,
            puffy_articles,
            baby_table_articles,
        ]:
            for _, art in df.iterrows():
                if str(art["tier"]) == "8":
                    displayed_t8.append(art)

        t8_budget = min(remaining, int(remaining * 0.5))
        t8_allocated = 0
        for art in sorted(displayed_t8, key=lambda x: x["adj_avg"], reverse=True):
            if t8_allocated >= t8_budget:
                break
            already_stored = any(s["article"] == art["article"] for s in storage_items)
            if already_stored:
                continue
            storage_items.append(
                {
                    "article": art["article"],
                    "gender_type": art.get("gender_type", ""),
                    "series": art.get("series", ""),
                    "tier": "8",
                    "adj_avg": art["adj_avg"],
                    "boxes": 1,
                    "reason": "T8 new launch backup",
                    "priority": 5,
                }
            )
            used_boxes += 1
            t8_allocated += 1

    # --- Priority 6: T2 remainder ---
    remaining = storage_capacity - used_boxes
    if remaining > 0:
        displayed_t2 = []
        for bw_id, data in planogram.items():
            for _, art in data["articles"].iterrows():
                if str(art["tier"]) == "2":
                    displayed_t2.append(art)

        t2_allocated = 0
        for art in sorted(displayed_t2, key=lambda x: x["adj_avg"], reverse=True):
            if t2_allocated >= remaining:
                break
            already_stored = any(s["article"] == art["article"] for s in storage_items)
            if already_stored:
                continue
            storage_items.append(
                {
                    "article": art["article"],
                    "gender_type": art.get("gender_type", ""),
                    "series": art.get("series", ""),
                    "tier": "2",
                    "adj_avg": art["adj_avg"],
                    "boxes": 1,
                    "reason": "T2 secondary backup",
                    "priority": 6,
                }
            )
            used_boxes += 1
            t2_allocated += 1

    return storage_items, used_boxes


# ============================================================
# STEP 8: GENERATE XLSX OUTPUT
# ============================================================
def generate_xlsx(
    planogram,
    airmove_articles,
    lucaluna_articles,
    baby_table_articles,
    puffy_articles,
    storage_items,
    storage_used,
    result_df,
    shares,
    all_months,
):
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # -------------------------------------------------------
    # BACKWALL SHEETS (BW-1 through BW-4b)
    # -------------------------------------------------------
    for bw_id, data in planogram.items():
        config = data["config"]
        articles = data["articles"]
        available = data["available_slots"]
        gt = config["gender_type"]
        hooks = config["hooks"]
        hpa = config.get("actual_hpa", config["hpa"])
        slots = config["slots"]
        used = len(articles)
        mode = config["mode"]
        orientation = config.get("orientation", "horizontal")
        grid = config.get("grid", "")
        label = config.get("label", gt)

        sheet_name = f"{bw_id} {label}"
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]
        ws = wb.create_sheet(title=sheet_name)

        # Row 1: Title
        mode_label = "FULL BOX MODE" if mode == "Full Box" else "COMPACT MODE"
        title = f"BACKWALL {bw_id} - {label} - {hooks} Hooks ({grid}) - {mode_label}"
        max_col = min(slots * hpa + 1, 60)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        ws.cell(row=1, column=1, value=title).font = TITLE_FONT

        # Row 2: Mode description
        if mode == "Full Box":
            if hpa == 2:
                mode_desc = f"Mode: Full Box ({hpa} hook = 1 artikel = 12 pairs) | Capacity: {slots} artikel | Used: {used} | Available: {available} slot"
            else:
                mode_desc = f"Mode: Full Box ({hpa} hook = 1 artikel = 12 pairs) | Capacity: {slots} artikel | Used: {used} | Available: {available} slot"
        else:
            if "Jepit" in gt or gt == "Baby & Kids":
                mode_desc = f"Mode: Compact ({hpa} hook = 1 artikel = 6 pairs, sisa 6 ke storage) | Capacity: {slots} artikel | Used: {used} | Available: {available} slot"
            else:
                mode_desc = f"Mode: Compact ({hpa} hook = 1 artikel = 8 pairs, sisa 4 ke storage) | Capacity: {slots} artikel | Used: {used} | Available: {available} slot"
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
        ws.cell(row=2, column=1, value=mode_desc).font = SUBTITLE_FONT

        # Row 3: blank
        # Row 4: Hook labels
        ws.cell(row=4, column=1, value="Hook").font = Font(bold=True)
        col = 2
        for i in range(slots):
            start_hook = i * hpa + 1
            end_hook = start_hook + hpa - 1
            label_hook = (
                f"Hook {start_hook}-{end_hook}" if hpa > 1 else f"Hook {start_hook}"
            )
            if hpa > 1:
                ws.merge_cells(
                    start_row=4, start_column=col, end_row=4, end_column=col + hpa - 1
                )
            cell = ws.cell(row=4, column=col, value=label_hook)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER
            for mc in range(col, col + hpa):
                ws.cell(row=4, column=mc).border = THIN_BORDER
                ws.cell(row=4, column=mc).fill = HEADER_FILL
            col += hpa

        # Row 5: Series grouping
        ws.cell(row=5, column=1, value="Series").font = Font(bold=True)
        col = 2
        for i in range(slots):
            if i < len(articles):
                art_series = articles.iloc[i]["series"]
            else:
                art_series = "(Kosong)"
            if hpa > 1:
                ws.merge_cells(
                    start_row=5, start_column=col, end_row=5, end_column=col + hpa - 1
                )
            ws.cell(row=5, column=col, value=str(art_series)).font = Font(italic=True)
            ws.cell(row=5, column=col).alignment = Alignment(horizontal="center")
            col += hpa

        # Row 6: Article names
        ws.cell(row=6, column=1, value="Artikel").font = Font(bold=True)
        col = 2
        for i in range(slots):
            if i < len(articles):
                art = articles.iloc[i]
                name = art["article"]
                tier = str(art["tier"])
                fill = TIER_COLORS.get(tier, PatternFill())
            else:
                name = "AVAILABLE"
                fill = AVAIL_FILL
            if hpa > 1:
                ws.merge_cells(
                    start_row=6, start_column=col, end_row=6, end_column=col + hpa - 1
                )
            cell = ws.cell(row=6, column=col, value=name)
            cell.font = Font(bold=True, size=10)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.fill = fill
            cell.border = THIN_BORDER
            for mc in range(col, col + hpa):
                ws.cell(row=6, column=mc).border = THIN_BORDER
                ws.cell(row=6, column=mc).fill = fill
            col += hpa

        # Row 7: Tier + Avg
        ws.cell(row=7, column=1, value="Tier & Avg").font = Font(bold=True)
        col = 2
        for i in range(slots):
            if i < len(articles):
                art = articles.iloc[i]
                tier = str(art["tier"])
                avg = art["adj_avg"]
                tier_label = f"T{tier} | Avg: {avg}"
                fill = TIER_COLORS.get(tier, PatternFill())
            else:
                tier_label = ""
                fill = AVAIL_FILL
            if hpa > 1:
                ws.merge_cells(
                    start_row=7, start_column=col, end_row=7, end_column=col + hpa - 1
                )
            cell = ws.cell(row=7, column=col, value=tier_label)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = fill
            cell.border = THIN_BORDER
            for mc in range(col, col + hpa):
                ws.cell(row=7, column=mc).border = THIN_BORDER
                ws.cell(row=7, column=mc).fill = fill
            col += hpa

        # Row 8: Pairs
        ws.cell(row=8, column=1, value="Pairs").font = Font(bold=True)
        col = 2
        for i in range(slots):
            if i < len(articles):
                if mode == "Full Box":
                    pairs_label = "12 pairs (1 box)"
                else:
                    if hpa == 1:
                        pairs_label = "6 pairs (1/2 box)"
                    else:
                        pairs_label = "8 pairs (2/3 box)"
            else:
                pairs_label = ""
            if hpa > 1:
                ws.merge_cells(
                    start_row=8, start_column=col, end_row=8, end_column=col + hpa - 1
                )
            ws.cell(row=8, column=col, value=pairs_label).alignment = Alignment(
                horizontal="center"
            )
            col += hpa

        # Row 10: Legend
        ws.cell(row=10, column=1, value="LEGEND:").font = Font(bold=True)
        legend = [
            ("T1", "Fast Moving (MUST display)", "1"),
            ("T8", "New Launch (priority exposure)", "8"),
            ("T2", "Secondary (filler)", "2"),
            ("T3", "Tertiary (variety filler)", "3"),
        ]
        for j, (tier_label, desc, tier_key) in enumerate(legend):
            row_n = 11 + j
            cell = ws.cell(row=row_n, column=1, value=tier_label)
            cell.fill = TIER_COLORS.get(tier_key, PatternFill())
            cell.font = Font(bold=True)
            ws.cell(row=row_n, column=2, value=desc)

        # Row 16: Mode note
        if mode == "Full Box":
            note = f"FULL BOX MODE: Setiap artikel menempati {hpa} hook bersebelahan = 1 box penuh (12 pairs) di display. Tidak ada sisa ke storage."
        else:
            if hpa == 1:
                note = "COMPACT MODE: Setiap artikel menempati 1 hook = 6 pairs di display. Sisa 6 pairs per artikel ke storage."
            else:
                note = "COMPACT MODE: Setiap artikel menempati 2 hook = 8 pairs di display. Sisa 4 pairs per artikel ke storage."
        ws.merge_cells(start_row=16, start_column=1, end_row=16, end_column=max_col)
        ws.cell(row=16, column=1, value=note).font = Font(italic=True, size=9)

        # Column widths
        ws.column_dimensions["A"].width = 14
        for c in range(2, slots * hpa + 2):
            ws.column_dimensions[get_column_letter(c)].width = 12

    # -------------------------------------------------------
    # TABLE & SHELVING SHEET
    # -------------------------------------------------------
    ws = wb.create_sheet(title="Table & Shelving")
    ws.merge_cells("A1:J1")
    ws.cell(
        row=1, column=1, value=f"TABLE & SHELVING DISPLAYS - {STORE_NAME}"
    ).font = TITLE_FONT

    row = 3

    # --- SHELVING-AIRMOVE ---
    ws.cell(row=row, column=1, value="SHELVING-AIRMOVE").font = Font(bold=True, size=12)
    ws.cell(row=row, column=1).fill = SPECIAL_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    row += 1
    ws.cell(
        row=row,
        column=1,
        value=f"Capacity: {SHELVING_AIRMOVE['slots']} slots | ~{SHELVING_AIRMOVE['pairs_per_article']} pairs on display per article | 1 box backup in storage per article",
    ).font = SUBTITLE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    row += 1

    headers = [
        "Slot",
        "Article",
        "Series",
        "Tier",
        "Adj Avg",
        "Pairs on Display",
        "Storage Backup",
    ]
    for j, h in enumerate(headers):
        cell = ws.cell(row=row, column=j + 1, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
    row += 1

    for i in range(len(airmove_articles)):
        art = airmove_articles.iloc[i]
        ws.cell(row=row, column=1, value=f"Slot {i + 1}").border = THIN_BORDER
        cell = ws.cell(row=row, column=2, value=art["article"])
        cell.font = Font(bold=True)
        cell.fill = TIER_COLORS.get(str(art["tier"]), PatternFill())
        cell.border = THIN_BORDER
        ws.cell(row=row, column=3, value=art["series"]).border = THIN_BORDER
        ws.cell(row=row, column=4, value=f"T{art['tier']}").border = THIN_BORDER
        ws.cell(row=row, column=5, value=art["adj_avg"]).border = THIN_BORDER
        ws.cell(
            row=row, column=6, value=f"~{SHELVING_AIRMOVE['pairs_per_article']} pairs"
        ).border = THIN_BORDER
        ws.cell(row=row, column=7, value="1 box (11 pairs)").border = THIN_BORDER
        row += 1

    row += 1

    # --- SHELVING-PUFFY ---
    ws.cell(row=row, column=1, value="SHELVING-PUFFY").font = Font(bold=True, size=12)
    ws.cell(row=row, column=1).fill = SPECIAL_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    row += 1
    ws.cell(
        row=row,
        column=1,
        value=f"Capacity: {SHELVING_PUFFY['slots']} slot | Highest-selling PUFFY article | ~{SHELVING_PUFFY['pairs_per_article']} pairs on display",
    ).font = SUBTITLE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    row += 1

    for j, h in enumerate(headers):
        cell = ws.cell(row=row, column=j + 1, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
    row += 1

    for i in range(len(puffy_articles)):
        art = puffy_articles.iloc[i]
        ws.cell(row=row, column=1, value=f"Slot {i + 1}").border = THIN_BORDER
        cell = ws.cell(row=row, column=2, value=art["article"])
        cell.font = Font(bold=True)
        cell.fill = TIER_COLORS.get(str(art["tier"]), PatternFill())
        cell.border = THIN_BORDER
        ws.cell(row=row, column=3, value=art["series"]).border = THIN_BORDER
        ws.cell(row=row, column=4, value=f"T{art['tier']}").border = THIN_BORDER
        ws.cell(row=row, column=5, value=art["adj_avg"]).border = THIN_BORDER
        ws.cell(
            row=row, column=6, value=f"~{SHELVING_PUFFY['pairs_per_article']} pairs"
        ).border = THIN_BORDER
        ws.cell(row=row, column=7, value="1 box (11 pairs)").border = THIN_BORDER
        row += 1

    row += 1

    # --- TABLE-LUCALUNA ---
    ws.cell(row=row, column=1, value="TABLE-LUCALUNA").font = Font(bold=True, size=12)
    ws.cell(row=row, column=1).fill = SPECIAL_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    row += 1
    ws.cell(
        row=row,
        column=1,
        value=f"Capacity: {TABLE_LUCALUNA['slots']} slots | LUCA/LUNA/AIRMOVE series | ~{TABLE_LUCALUNA['pairs_per_article']} pairs on display per article | 1 box backup in storage",
    ).font = SUBTITLE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    row += 1

    for j, h in enumerate(headers):
        cell = ws.cell(row=row, column=j + 1, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
    row += 1

    for i in range(len(lucaluna_articles)):
        art = lucaluna_articles.iloc[i]
        ws.cell(row=row, column=1, value=f"Slot {i + 1}").border = THIN_BORDER
        cell = ws.cell(row=row, column=2, value=art["article"])
        cell.font = Font(bold=True)
        cell.fill = TIER_COLORS.get(str(art["tier"]), PatternFill())
        cell.border = THIN_BORDER
        ws.cell(row=row, column=3, value=art["series"]).border = THIN_BORDER
        ws.cell(row=row, column=4, value=f"T{art['tier']}").border = THIN_BORDER
        ws.cell(row=row, column=5, value=art["adj_avg"]).border = THIN_BORDER
        ws.cell(
            row=row, column=6, value=f"~{TABLE_LUCALUNA['pairs_per_article']} pairs"
        ).border = THIN_BORDER
        ws.cell(row=row, column=7, value="1 box (11 pairs)").border = THIN_BORDER
        row += 1

    row += 1

    # --- TABLE-BABY ---
    ws.cell(row=row, column=1, value="TABLE-BABY").font = Font(bold=True, size=12)
    ws.cell(row=row, column=1).fill = SPECIAL_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    row += 1
    ws.cell(
        row=row,
        column=1,
        value=f"Capacity: {TABLE_BABY['slots']} slots | BABY gender | ~{TABLE_BABY['pairs_per_article']} pairs on display per article",
    ).font = SUBTITLE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    row += 1

    baby_headers = [
        "Slot",
        "Article",
        "Series",
        "Tier",
        "Adj Avg",
        "Pairs on Display",
        "Storage Backup",
    ]
    for j, h in enumerate(baby_headers):
        cell = ws.cell(row=row, column=j + 1, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
    row += 1

    for i in range(len(baby_table_articles)):
        art = baby_table_articles.iloc[i]
        ws.cell(row=row, column=1, value=f"Slot {i + 1}").border = THIN_BORDER
        cell = ws.cell(row=row, column=2, value=art["article"])
        cell.font = Font(bold=True)
        cell.fill = TIER_COLORS.get(str(art["tier"]), PatternFill())
        cell.border = THIN_BORDER
        ws.cell(row=row, column=3, value=art["series"]).border = THIN_BORDER
        ws.cell(row=row, column=4, value=f"T{art['tier']}").border = THIN_BORDER
        ws.cell(row=row, column=5, value=art["adj_avg"]).border = THIN_BORDER
        ws.cell(
            row=row, column=6, value=f"~{TABLE_BABY['pairs_per_article']} pairs"
        ).border = THIN_BORDER
        ws.cell(row=row, column=7, value="Shared box backup").border = THIN_BORDER
        row += 1

    # Column widths for Table & Shelving sheet
    for c in range(1, 9):
        ws.column_dimensions[get_column_letter(c)].width = 20

    # -------------------------------------------------------
    # STORAGE ALLOCATION SHEET
    # -------------------------------------------------------
    ws = wb.create_sheet(title="Storage Allocation")
    ws.merge_cells("A1:I1")
    ws.cell(
        row=1,
        column=1,
        value=f"STORAGE ALLOCATION - {STORE_NAME} - Capacity: {STORAGE_CAPACITY} Box",
    ).font = TITLE_FONT

    # Summary section
    ws.merge_cells("A3:I3")
    ws.cell(
        row=3,
        column=1,
        value=f"Total Used: {storage_used} box | Remaining: {STORAGE_CAPACITY - storage_used} box | Utilization: {round(storage_used / STORAGE_CAPACITY * 100, 1) if STORAGE_CAPACITY > 0 else 0}%",
    ).font = Font(bold=True, size=11)

    # Priority breakdown
    row = 5
    ws.cell(row=row, column=1, value="STORAGE BREAKDOWN BY PRIORITY").font = Font(
        bold=True, size=12
    )
    row += 1

    priority_names = {
        1: "Table/Shelving backup (Airmove, Luca/Luna, Puffy)",
        2: "Table Baby backup",
        3: "Compact mode overflow",
        4: "T1 fast moving backup",
        5: "T8 new launch backup",
        6: "T2 secondary backup",
    }
    priority_totals = defaultdict(int)
    for item in storage_items:
        priority_totals[item["priority"]] += item["boxes"]

    for p in sorted(priority_totals.keys()):
        ws.cell(row=row, column=1, value=priority_names.get(p, f"Priority {p}"))
        ws.cell(row=row, column=5, value=f"{priority_totals[p]} box")
        row += 1

    ws.cell(row=row, column=1, value="TOTAL USED").font = Font(bold=True)
    ws.cell(row=row, column=5, value=f"{storage_used} box").font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value="REMAINING")
    ws.cell(row=row, column=5, value=f"{STORAGE_CAPACITY - storage_used} box")
    row += 2

    # Detailed allocation table
    ws.cell(row=row, column=1, value="DETAILED STORAGE ALLOCATION").font = Font(
        bold=True, size=12
    )
    row += 1

    headers = [
        "No",
        "Article",
        "Gender-Type",
        "Series",
        "Tier",
        "Adj Avg",
        "Boxes",
        "Reason",
    ]
    for j, h in enumerate(headers):
        cell = ws.cell(row=row, column=j + 1, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
    row += 1

    for i, item in enumerate(storage_items):
        ws.cell(row=row, column=1, value=i + 1).border = THIN_BORDER
        cell = ws.cell(row=row, column=2, value=item["article"])
        cell.border = THIN_BORDER
        cell.fill = TIER_COLORS.get(str(item["tier"]), PatternFill())
        ws.cell(row=row, column=3, value=item["gender_type"]).border = THIN_BORDER
        ws.cell(row=row, column=4, value=item["series"]).border = THIN_BORDER
        ws.cell(
            row=row, column=5, value=f"T{item['tier']}" if item["tier"] != "-" else "-"
        ).border = THIN_BORDER
        ws.cell(row=row, column=6, value=item["adj_avg"]).border = THIN_BORDER
        ws.cell(row=row, column=7, value=item["boxes"]).border = THIN_BORDER
        ws.cell(row=row, column=8, value=item["reason"]).border = THIN_BORDER
        row += 1

    # Total row
    ws.cell(row=row, column=1, value="").border = THIN_BORDER
    ws.cell(row=row, column=2, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=2).border = THIN_BORDER
    for c in range(3, 7):
        ws.cell(row=row, column=c, value="").border = THIN_BORDER
    ws.cell(row=row, column=7, value=storage_used).font = Font(bold=True)
    ws.cell(row=row, column=7).border = THIN_BORDER
    ws.cell(
        row=row,
        column=8,
        value=f"{round(storage_used / STORAGE_CAPACITY * 100, 1)}% utilization",
    ).border = THIN_BORDER

    for c in range(1, 9):
        ws.column_dimensions[get_column_letter(c)].width = 20

    # -------------------------------------------------------
    # SUMMARY REPORT SHEET
    # -------------------------------------------------------
    ws = wb.create_sheet(title="Summary Report")
    ws.merge_cells("A1:G1")
    ws.cell(
        row=1, column=1, value=f"PLANOGRAM SUMMARY - {STORE_NAME}"
    ).font = TITLE_FONT

    # Collect all displayed articles
    displayed_articles = set()
    all_displayed = []

    for bw_id, data in planogram.items():
        for _, art in data["articles"].iterrows():
            displayed_articles.add(art["article"])
            all_displayed.append(art)

    for df_label, df in [
        ("SHELVING-AIRMOVE", airmove_articles),
        ("TABLE-LUCALUNA", lucaluna_articles),
        ("TABLE-BABY", baby_table_articles),
        ("SHELVING-PUFFY", puffy_articles),
    ]:
        for _, art in df.iterrows():
            displayed_articles.add(art["article"])
            all_displayed.append(art)

    # Display utilization
    total_bw_slots = sum(p["config"]["slots"] for p in planogram.values())
    total_special_slots = (
        SHELVING_AIRMOVE["slots"]
        + SHELVING_PUFFY["slots"]
        + TABLE_LUCALUNA["slots"]
        + TABLE_BABY["slots"]
    )
    total_slots = total_bw_slots + total_special_slots
    total_bw_used = sum(len(p["articles"]) for p in planogram.values())
    total_special_used = (
        len(airmove_articles)
        + len(puffy_articles)
        + len(lucaluna_articles)
        + len(baby_table_articles)
    )
    total_used = total_bw_used + total_special_used
    total_avail = total_slots - total_used
    util_pct = round(total_used / total_slots * 100, 1) if total_slots > 0 else 0

    row = 3
    ws.cell(row=row, column=1, value="DISPLAY UTILIZATION").font = Font(
        bold=True, size=12
    )
    row += 1
    util_data = [
        ("Total slots available (BW + Special)", total_slots),
        ("  Backwall slots", total_bw_slots),
        ("  Special display slots (Shelving/Table)", total_special_slots),
        ("Total slots used", total_used),
        ("  Backwall used", total_bw_used),
        ("  Special display used", total_special_used),
        ("Utilization", f"{util_pct}%"),
        ("Empty slots", f"{total_avail} (available for expansion)"),
    ]
    for label, val in util_data:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=4, value=str(val))
        row += 1

    # Sales coverage
    row += 1
    ws.cell(row=row, column=1, value="SALES COVERAGE").font = Font(bold=True, size=12)
    row += 1
    total_adj = result_df["adj_avg"].sum()
    displayed_adj = result_df[result_df["article"].isin(displayed_articles)][
        "adj_avg"
    ].sum()
    coverage = round(displayed_adj / total_adj * 100, 1) if total_adj > 0 else 0

    cov_data = [
        ("Total adjusted avg (semua artikel)", f"{total_adj:.1f} pairs/bulan"),
        ("Adjusted avg dari artikel di display", f"{displayed_adj:.1f} pairs/bulan"),
        ("Sales coverage", f"{coverage}%"),
    ]
    for label, val in cov_data:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=4, value=str(val))
        row += 1

    # Tier distribution
    row += 1
    ws.cell(row=row, column=1, value="TIER DISTRIBUTION DI DISPLAY").font = Font(
        bold=True, size=12
    )
    row += 1
    headers = ["Tier", "Count", "% of Display", "Notes"]
    for j, h in enumerate(headers):
        cell = ws.cell(row=row, column=j + 1, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
    row += 1

    tier_counts = defaultdict(int)
    for art in all_displayed:
        tier_counts[str(art["tier"])] += 1

    total_displayed = len(all_displayed)
    tier_names = {
        "1": "Fast Moving",
        "8": "New Launch",
        "2": "Secondary",
        "3": "Tertiary",
    }

    total_t1 = len(result_df[result_df["tier"] == "1"])

    for t in ["1", "8", "2", "3"]:
        count = tier_counts.get(t, 0)
        pct = round(count / total_displayed * 100, 1) if total_displayed > 0 else 0
        note = ""
        if t == "1":
            note = f"{'All' if count >= total_t1 else f'{count} of {total_t1}'} T1 displayed"
        ws.cell(row=row, column=1, value=f"T{t} ({tier_names[t]})").border = THIN_BORDER
        ws.cell(row=row, column=1).fill = TIER_COLORS.get(t, PatternFill())
        ws.cell(row=row, column=2, value=count).border = THIN_BORDER
        ws.cell(row=row, column=3, value=f"{pct}%").border = THIN_BORDER
        ws.cell(row=row, column=4, value=note).border = THIN_BORDER
        row += 1

    # Storage utilization
    row += 1
    ws.cell(row=row, column=1, value="STORAGE UTILIZATION").font = Font(
        bold=True, size=12
    )
    row += 1
    ws.cell(row=row, column=1, value="Total capacity")
    ws.cell(row=row, column=4, value=f"{STORAGE_CAPACITY} box")
    row += 1
    ws.cell(row=row, column=1, value="Used")
    ws.cell(row=row, column=4, value=f"{storage_used} box")
    row += 1
    ws.cell(row=row, column=1, value="Remaining")
    ws.cell(row=row, column=4, value=f"{STORAGE_CAPACITY - storage_used} box")
    row += 1
    ws.cell(row=row, column=1, value="Utilization")
    ws.cell(
        row=row, column=4, value=f"{round(storage_used / STORAGE_CAPACITY * 100, 1)}%"
    )
    row += 1

    # Breakdown
    priority_names_short = {
        1: "Table/Shelving backup",
        2: "Table Baby backup",
        3: "Compact mode overflow",
        4: "T1 fast moving backup",
        5: "T8 new launch backup",
        6: "T2 secondary backup",
    }
    priority_totals_local = defaultdict(int)
    for item in storage_items:
        priority_totals_local[item["priority"]] += item["boxes"]

    ws.cell(row=row, column=1, value="Breakdown:").font = Font(italic=True)
    row += 1
    for p in sorted(priority_totals_local.keys()):
        ws.cell(row=row, column=1, value=f"  - {priority_names_short.get(p, f'P{p}')}")
        ws.cell(row=row, column=4, value=f"{priority_totals_local[p]} box")
        row += 1

    # Gender-type assignment
    row += 1
    ws.cell(row=row, column=1, value="DISPLAY UNIT ASSIGNMENT").font = Font(
        bold=True, size=12
    )
    row += 1
    headers = [
        "Display Unit",
        "Gender-Type / Series",
        "Hooks/Slots",
        "Capacity",
        "Used",
        "Available",
        "Sales Share",
    ]
    for j, h in enumerate(headers):
        cell = ws.cell(row=row, column=j + 1, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
    row += 1

    for bw_id, data in planogram.items():
        c = data["config"]
        gt = c["gender_type"]
        label = c.get("label", gt)
        share_row = shares[shares["gender_type"] == gt]
        share_pct = (
            f"{share_row['pct_share'].iloc[0]}%" if len(share_row) > 0 else "N/A"
        )
        ws.cell(row=row, column=1, value=bw_id).border = THIN_BORDER
        ws.cell(row=row, column=2, value=f"{label} ({gt})").border = THIN_BORDER
        ws.cell(row=row, column=3, value=f"{c['hooks']} hooks").border = THIN_BORDER
        ws.cell(row=row, column=4, value=c["slots"]).border = THIN_BORDER
        ws.cell(row=row, column=5, value=len(data["articles"])).border = THIN_BORDER
        ws.cell(row=row, column=6, value=data["available_slots"]).border = THIN_BORDER
        ws.cell(row=row, column=7, value=share_pct).border = THIN_BORDER
        row += 1

    # Special displays
    special_displays = [
        (
            "SHELVING-AIRMOVE",
            "AIRMOVE",
            f"{SHELVING_AIRMOVE['slots']} slots",
            SHELVING_AIRMOVE["slots"],
            len(airmove_articles),
        ),
        (
            "SHELVING-PUFFY",
            "PUFFY",
            f"{SHELVING_PUFFY['slots']} slot",
            SHELVING_PUFFY["slots"],
            len(puffy_articles),
        ),
        (
            "TABLE-LUCALUNA",
            "LUCA/LUNA/AIRMOVE",
            f"{TABLE_LUCALUNA['slots']} slots",
            TABLE_LUCALUNA["slots"],
            len(lucaluna_articles),
        ),
        (
            "TABLE-BABY",
            "BABY",
            f"{TABLE_BABY['slots']} slots",
            TABLE_BABY["slots"],
            len(baby_table_articles),
        ),
    ]
    for disp_id, disp_type, hook_label, capacity, used_count in special_displays:
        ws.cell(row=row, column=1, value=disp_id).border = THIN_BORDER
        ws.cell(row=row, column=2, value=disp_type).border = THIN_BORDER
        ws.cell(row=row, column=3, value=hook_label).border = THIN_BORDER
        ws.cell(row=row, column=4, value=capacity).border = THIN_BORDER
        ws.cell(row=row, column=5, value=used_count).border = THIN_BORDER
        ws.cell(
            row=row, column=6, value=max(0, capacity - used_count)
        ).border = THIN_BORDER
        ws.cell(row=row, column=7, value="-").border = THIN_BORDER
        row += 1

    # FLAGS & WARNINGS
    row += 1
    ws.cell(row=row, column=1, value="FLAGS & WARNINGS").font = Font(bold=True, size=12)
    row += 1

    flags = []

    # Check T1 not displayed
    all_displayed_names = displayed_articles
    t1_not_displayed = result_df[
        (result_df["tier"] == "1") & (~result_df["article"].isin(all_displayed_names))
    ]
    if len(t1_not_displayed) > 0:
        for _, art in t1_not_displayed.iterrows():
            flags.append(
                (
                    "CRITICAL",
                    f'T1 "{art["article"]}" ({art["gender_type"]}) NOT displayed -- Avg: {art["adj_avg"]}',
                )
            )

    # Check gender-types without display
    assigned_gts = set()
    for bw_id, data in planogram.items():
        assigned_gts.add(data["config"]["gender_type"])
    assigned_gts.add("Baby & Kids")  # table baby + BW-4b

    all_gts = set(result_df["gender_type"].unique())
    unassigned_gts = all_gts - assigned_gts
    for ugt in unassigned_gts:
        gt_sales = result_df[result_df["gender_type"] == ugt]["adj_avg"].sum()
        if gt_sales > 0:
            gt_share = shares[shares["gender_type"] == ugt]
            pct = f"{gt_share['pct_share'].iloc[0]}%" if len(gt_share) > 0 else "?"
            flags.append(
                (
                    "WARNING",
                    f"{ugt} has no display unit -- {pct} sales share, {gt_sales:.0f} adj_avg total",
                )
            )

    # Check T4/T5 in display
    t4t5_in_display = [a for a in all_displayed if str(a["tier"]) in ("4", "5")]
    if len(t4t5_in_display) > 0:
        for art in t4t5_in_display:
            flags.append(
                (
                    "WARNING",
                    f'T{art["tier"]} "{art["article"]}" is in display -- should be removed',
                )
            )

    # Storage warnings
    if storage_used > STORAGE_CAPACITY:
        flags.append(
            (
                "CRITICAL",
                f"Storage OVERFLOW: {storage_used} box used > {STORAGE_CAPACITY} box capacity",
            )
        )
    elif storage_used > STORAGE_CAPACITY * 0.9:
        flags.append(
            (
                "WARNING",
                f"Storage near full: {storage_used}/{STORAGE_CAPACITY} box ({round(storage_used / STORAGE_CAPACITY * 100, 1)}%)",
            )
        )

    # Luca/Luna/Airmove T1 not displayed
    luca_luna_all = result_df[
        result_df["series"].str.upper().isin(["LUCA", "LUNA", "AIRMOVE"])
    ]
    luca_t1_not_displayed = luca_luna_all[
        (luca_luna_all["tier"] == "1")
        & (~luca_luna_all["article"].isin(all_displayed_names))
    ]
    if len(luca_t1_not_displayed) > 0:
        for _, art in luca_t1_not_displayed.iterrows():
            flags.append(
                (
                    "WARNING",
                    f'T1 Luca/Luna/Airmove "{art["article"]}" not displayed -- table/shelving full',
                )
            )

    # Positive flags
    if len(t1_not_displayed) == 0:
        flags.append(("POSITIVE", "All T1 articles are displayed"))

    if len(t4t5_in_display) == 0:
        flags.append(("POSITIVE", "No T4/T5 in display"))

    if coverage >= 80:
        flags.append(("POSITIVE", f"Sales coverage {coverage}% (good)"))

    flag_colors = {
        "CRITICAL": Font(bold=True, color="FF0000"),
        "WARNING": Font(bold=True, color="FF8C00"),
        "POSITIVE": Font(bold=True, color="008000"),
    }

    for severity, msg in flags:
        icon = {"CRITICAL": "[CRITICAL]", "WARNING": "[WARNING]", "POSITIVE": "[OK]"}
        ws.cell(row=row, column=1, value=icon.get(severity, "")).font = flag_colors.get(
            severity, Font()
        )
        ws.cell(row=row, column=2, value=msg)
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 15

    # -------------------------------------------------------
    # FULL ARTICLE RANKING SHEET
    # -------------------------------------------------------
    ws = wb.create_sheet(title="Full Article Ranking")
    ws.cell(
        row=1,
        column=1,
        value=f"ALL ARTICLES RANKED -- {STORE_NAME} (Feb 2025 - Jan 2026)",
    ).font = TITLE_FONT
    ws.merge_cells("A1:M1")

    headers = [
        "Rank",
        "Article",
        "Gender",
        "Gender-Type",
        "Series",
        "Tipe",
        "Tier",
        "Adj Avg",
        "Total 12mo",
        "Months Sold",
        "In Display?",
        "Display Location",
        "Storage Boxes",
    ]
    for j, h in enumerate(headers):
        cell = ws.cell(row=3, column=j + 1, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    # Build display location map
    display_map = {}
    for bw_id, data in planogram.items():
        for idx, (_, art) in enumerate(data["articles"].iterrows()):
            hpa = data["config"].get("actual_hpa", data["config"]["hpa"])
            hook_start = idx * hpa + 1
            hook_end = hook_start + hpa - 1
            display_map[art["article"]] = f"{bw_id} Hook {hook_start}-{hook_end}"

    for i in range(len(airmove_articles)):
        art = airmove_articles.iloc[i]
        display_map[art["article"]] = f"SHELVING-AIRMOVE Slot {i + 1}"
    for i in range(len(lucaluna_articles)):
        art = lucaluna_articles.iloc[i]
        display_map[art["article"]] = f"TABLE-LUCALUNA Slot {i + 1}"
    for i in range(len(baby_table_articles)):
        art = baby_table_articles.iloc[i]
        display_map[art["article"]] = f"TABLE-BABY Slot {i + 1}"
    for i in range(len(puffy_articles)):
        art = puffy_articles.iloc[i]
        display_map[art["article"]] = f"SHELVING-PUFFY Slot {i + 1}"

    # Build storage map
    storage_map = {}
    for item in storage_items:
        if item["article"] not in storage_map:
            storage_map[item["article"]] = item["boxes"]
        else:
            storage_map[item["article"]] += item["boxes"]

    ranked = result_df.copy()
    tier_priority = {"1": 0, "8": 1, "2": 2, "3": 3, "4": 4, "5": 5}
    ranked["tier_sort"] = ranked["tier"].map(tier_priority).fillna(9)
    ranked = ranked.sort_values(
        ["tier_sort", "adj_avg"], ascending=[True, False]
    ).reset_index(drop=True)

    for i, (_, art) in enumerate(ranked.iterrows()):
        row_n = 4 + i
        in_display = "YES" if art["article"] in displayed_articles else "NO"
        loc = display_map.get(art["article"], "")
        stg = storage_map.get(art["article"], 0)

        ws.cell(row=row_n, column=1, value=i + 1).border = THIN_BORDER
        cell = ws.cell(row=row_n, column=2, value=art["article"])
        cell.border = THIN_BORDER
        cell.fill = TIER_COLORS.get(str(art["tier"]), PatternFill())
        ws.cell(row=row_n, column=3, value=art["gender"]).border = THIN_BORDER
        ws.cell(row=row_n, column=4, value=art["gender_type"]).border = THIN_BORDER
        ws.cell(row=row_n, column=5, value=art["series"]).border = THIN_BORDER
        ws.cell(row=row_n, column=6, value=art["tipe"]).border = THIN_BORDER
        ws.cell(row=row_n, column=7, value=f"T{art['tier']}").border = THIN_BORDER
        ws.cell(row=row_n, column=8, value=art["adj_avg"]).border = THIN_BORDER
        ws.cell(row=row_n, column=9, value=art["total_12mo"]).border = THIN_BORDER
        ws.cell(row=row_n, column=10, value=art["months_sold"]).border = THIN_BORDER
        ws.cell(row=row_n, column=11, value=in_display).border = THIN_BORDER
        ws.cell(row=row_n, column=12, value=loc).border = THIN_BORDER
        ws.cell(row=row_n, column=13, value=stg if stg > 0 else "").border = THIN_BORDER

    for c in range(1, 14):
        ws.column_dimensions[get_column_letter(c)].width = 18

    # -------------------------------------------------------
    # MONTHLY SALES DATA SHEET
    # -------------------------------------------------------
    ws = wb.create_sheet(title="Monthly Sales Data")
    ws.cell(
        row=1, column=1, value=f"MONTHLY SALES BY ARTICLE -- {STORE_NAME}"
    ).font = TITLE_FONT

    headers = (
        ["Article", "Gender-Type", "Tier", "Adj Avg"] + all_months + ["Total 12mo"]
    )
    for j, h in enumerate(headers):
        cell = ws.cell(row=3, column=j + 1, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    for i, (_, art) in enumerate(ranked.iterrows()):
        row_n = 4 + i
        ws.cell(row=row_n, column=1, value=art["article"]).border = THIN_BORDER
        ws.cell(row=row_n, column=2, value=art["gender_type"]).border = THIN_BORDER
        ws.cell(row=row_n, column=3, value=f"T{art['tier']}").border = THIN_BORDER
        ws.cell(row=row_n, column=4, value=art["adj_avg"]).border = THIN_BORDER
        for m_idx, val in enumerate(art["month_values"]):
            ws.cell(row=row_n, column=5 + m_idx, value=val).border = THIN_BORDER
        ws.cell(
            row=row_n, column=5 + len(all_months), value=art["total_12mo"]
        ).border = THIN_BORDER

    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 14

    # Save
    wb.save(OUTPUT_FILE)
    print(f"\nPLANOGRAM SAVED: {OUTPUT_FILE}")
    return wb


# ============================================================
# MAIN
# ============================================================
if __name__ == "__main__":
    print("=" * 60)
    print(f"BUILDING PLANOGRAM: {STORE_NAME}")
    print("=" * 60)

    # --- STEP 1: Pull data ---
    print("\n[1/8] Pulling data from VPS DB...")
    sales_df, master_df, stock_df = pull_data()
    print(f"  Sales: {len(sales_df)} rows")
    print(f"  Master: {len(master_df)} articles")
    print(f"  Stock: {len(stock_df)} items")

    # --- STEP 2: Process sales ---
    print("\n[2/8] Processing sales -> kodemix aggregation...")
    merged = process_sales(sales_df, master_df)
    print(f"  Merged: {len(merged)} rows")

    # --- STEP 3: Compute adjusted averages ---
    print("\n[3/8] Computing adjusted averages per tier...")
    result_df, all_months = compute_adjusted_avg(merged)
    print(f"  Articles: {len(result_df)}")
    print(f"  Months: {all_months}")

    # --- STEP 4: Gender-type mapping ---
    print("\n[4/8] Gender-type mapping & sales share...")
    shares = compute_gender_shares(result_df)
    print(shares.to_string(index=False))

    # --- STEP 5: Pre-assign special displays ---
    print("\n[5/8] Pre-assigning special displays...")
    pre_assigned_names = set()

    # 5a. Airmove -> SHELVING-AIRMOVE (top 3)
    airmove_articles = pre_assign_airmove(result_df, SHELVING_AIRMOVE)
    airmove_names = (
        set(airmove_articles["article"].tolist())
        if len(airmove_articles) > 0
        else set()
    )
    pre_assigned_names.update(airmove_names)
    print(f"\n  SHELVING-AIRMOVE ({SHELVING_AIRMOVE['slots']} slots):")
    for _, a in airmove_articles.iterrows():
        print(f"    T{a['tier']} | {a['article']:<35} | Avg: {a['adj_avg']}")

    # 5b. Luca/Luna/Airmove -> TABLE-LUCALUNA (top 5, excluding already assigned Airmove)
    lucaluna_articles = pre_assign_lucaluna(
        result_df, TABLE_LUCALUNA, pre_assigned_names
    )
    lucaluna_names = (
        set(lucaluna_articles["article"].tolist())
        if len(lucaluna_articles) > 0
        else set()
    )
    pre_assigned_names.update(lucaluna_names)
    print(f"\n  TABLE-LUCALUNA ({TABLE_LUCALUNA['slots']} slots):")
    for _, a in lucaluna_articles.iterrows():
        print(f"    T{a['tier']} | {a['article']:<35} | Avg: {a['adj_avg']}")

    # 5c. Baby -> TABLE-BABY (top 6)
    baby_table_articles = pre_assign_baby_table(
        result_df, TABLE_BABY, pre_assigned_names
    )
    baby_table_names = (
        set(baby_table_articles["article"].tolist())
        if len(baby_table_articles) > 0
        else set()
    )
    pre_assigned_names.update(baby_table_names)
    print(f"\n  TABLE-BABY ({TABLE_BABY['slots']} slots):")
    for _, a in baby_table_articles.iterrows():
        print(f"    T{a['tier']} | {a['article']:<35} | Avg: {a['adj_avg']}")

    # 5d. Puffy -> SHELVING-PUFFY (top 1)
    puffy_articles = pre_assign_puffy(result_df, SHELVING_PUFFY, pre_assigned_names)
    puffy_names = (
        set(puffy_articles["article"].tolist()) if len(puffy_articles) > 0 else set()
    )
    pre_assigned_names.update(puffy_names)
    print(f"\n  SHELVING-PUFFY ({SHELVING_PUFFY['slots']} slots):")
    for _, a in puffy_articles.iterrows():
        print(f"    T{a['tier']} | {a['article']:<35} | Avg: {a['adj_avg']}")

    # --- STEP 6: Assign articles to backwalls ---
    print("\n[6/8] Assigning articles to backwalls...")
    planogram = assign_articles_to_backwalls(BACKWALLS, result_df, pre_assigned_names)

    for bw_id, data in planogram.items():
        arts = data["articles"]
        avail = data["available_slots"]
        config = data["config"]
        print(
            f"\n  {bw_id} ({config['label']}, {config['gender_type']}, {config['mode']}):"
        )
        print(f"    Assigned: {len(arts)} articles, Available: {avail} slots")
        if len(arts) > 0:
            for _, a in arts.iterrows():
                print(f"      T{a['tier']} | {a['article']:<35} | Avg: {a['adj_avg']}")

    # --- STEP 7: Storage allocation ---
    print("\n[7/8] Allocating storage (capacity: {} boxes)...".format(STORAGE_CAPACITY))
    storage_items, storage_used = allocate_storage(
        planogram,
        airmove_articles,
        lucaluna_articles,
        baby_table_articles,
        puffy_articles,
        result_df,
        STORAGE_CAPACITY,
    )
    print(
        f"  Storage used: {storage_used}/{STORAGE_CAPACITY} boxes ({round(storage_used / STORAGE_CAPACITY * 100, 1)}%)"
    )
    print(f"  Storage items: {len(storage_items)}")

    # --- STEP 8: Generate XLSX ---
    print("\n[8/8] Generating XLSX output...")
    generate_xlsx(
        planogram,
        airmove_articles,
        lucaluna_articles,
        baby_table_articles,
        puffy_articles,
        storage_items,
        storage_used,
        result_df,
        shares,
        all_months,
    )

    # Summary
    total_bw_slots = sum(p["config"]["slots"] for p in planogram.values())
    total_special_slots = (
        SHELVING_AIRMOVE["slots"]
        + SHELVING_PUFFY["slots"]
        + TABLE_LUCALUNA["slots"]
        + TABLE_BABY["slots"]
    )
    total_slots = total_bw_slots + total_special_slots
    total_bw_used = sum(len(p["articles"]) for p in planogram.values())
    total_special_used = (
        len(airmove_articles)
        + len(puffy_articles)
        + len(lucaluna_articles)
        + len(baby_table_articles)
    )
    total_used = total_bw_used + total_special_used

    print(f"\n{'=' * 60}")
    print(
        f"DONE! Total: {total_used}/{total_slots} slots used ({round(total_used / total_slots * 100, 1)}%)"
    )
    print(f"  Backwall: {total_bw_used}/{total_bw_slots}")
    print(f"  Special:  {total_special_used}/{total_special_slots}")
    print(f"  Storage:  {storage_used}/{STORAGE_CAPACITY} boxes")
    print(f"Output: {OUTPUT_FILE}")
    print(f"{'=' * 60}")
