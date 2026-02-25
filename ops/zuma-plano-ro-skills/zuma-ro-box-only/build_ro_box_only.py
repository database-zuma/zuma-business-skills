#!/usr/bin/env python3
"""
Zuma RO Box-Only Generator
Logic: Max Stock (planogram + storage) vs On Hand → RO selisih jika kurang
Data Source: portal.temp_portal_plannogram (Jatim only)
"""

import math
import psycopg2
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# ============================================================================
# CONFIGURATION - Modify these for your store
# ============================================================================

STORE_NAME = "Zuma Royal Plaza"  # Display name
STORE_DB_PATTERN = "zuma royal plaza"  # For ILIKE match in stock query
OUTPUT_FILE = f"RO_BoxOnly_{STORE_NAME.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"

# Tiers excluded from RO (protol/clearance only)
EXCLUDED_TIERS = ["5", "6", "7"]

# Tier sort order: T1 → T8 → T2 → T3 → T4
TIER_ORDER = {"1": 0, "8": 1, "2": 2, "3": 3, "4": 4}

# Database connection
DB_CONFIG = {
    "host": "76.13.194.120",
    "port": 5432,
    "database": "openclaw_ops",
    "user": "openclaw_app",
    "password": "Zuma-0psCl4w-2026!",
}

# ============================================================================
# DATABASE QUERIES
# ============================================================================


def get_connection():
    return psycopg2.connect(**DB_CONFIG)


def get_planogram(store_name):
    """Fetch planogram including max stock columns"""
    query = """
    SELECT
        store_name,
        gender,
        series,
        article,
        tier,
        article_mix,
        total_planogram::int   AS planogram_qty,
        storage_new::int       AS storage_qty,
        (total_planogram::int + storage_new::int) AS max_stock
    FROM portal.temp_portal_plannogram
    WHERE store_name ILIKE %s
    ORDER BY tier, article
    """
    conn = get_connection()
    df = pd.read_sql(query, conn, params=(store_name,))
    conn.close()
    return df


def get_store_stock(store_pattern):
    """Get current on-hand stock at the store (per article, all sizes summed)"""
    query = """
    SELECT
        TRIM(LOWER(kode_mix)) AS article_mix,
        SUM(quantity)         AS on_hand
    FROM core.stock_with_product
    WHERE LOWER(nama_gudang) ILIKE %s
    GROUP BY TRIM(LOWER(kode_mix))
    """
    conn = get_connection()
    df = pd.read_sql(query, conn, params=(f"%{store_pattern}%",))
    conn.close()
    return df


def get_wh_stock(article_mixes):
    """Get WH Pusat stock split by DDD and LJBB (exact match, no Protol/Reject)"""
    if not article_mixes:
        return pd.DataFrame(columns=["article_mix", "wh_qty_ddd", "wh_qty_ljbb"])

    query = """
    SELECT
        TRIM(LOWER(kode_mix)) AS article_mix,
        SUM(CASE WHEN warehouse_code = 'DDD'  THEN quantity ELSE 0 END) AS wh_qty_ddd,
        SUM(CASE WHEN warehouse_code = 'LJBB' THEN quantity ELSE 0 END) AS wh_qty_ljbb
    FROM core.stock_with_product
    WHERE nama_gudang = 'Warehouse Pusat'
      AND warehouse_code IN ('DDD', 'LJBB')
      AND TRIM(LOWER(kode_mix)) = ANY(%s)
    GROUP BY TRIM(LOWER(kode_mix))
    """
    conn = get_connection()
    article_mixes_lower = [str(a).lower().strip() for a in article_mixes]
    df = pd.read_sql(query, conn, params=(article_mixes_lower,))
    conn.close()
    return df


# ============================================================================
# EXCEL GENERATION
# ============================================================================

ZUMA_GREEN = "00E273"
ZUMA_GREEN_DARK = "2E7D32"


def header_style():
    return {
        "font": Font(bold=True, size=10, color="FFFFFF"),
        "fill": PatternFill(
            start_color=ZUMA_GREEN_DARK, end_color=ZUMA_GREEN_DARK, fill_type="solid"
        ),
        "alignment": Alignment(horizontal="center", vertical="center"),
        "border": Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        ),
    }


def apply_style(cell, s):
    for attr in ("font", "fill", "alignment", "border"):
        if attr in s:
            setattr(cell, attr, s[attr])


def generate_excel(ro_df, summary_stats, output_path):
    wb = Workbook()

    # =========================================================================
    # Sheet 1: Cover
    # =========================================================================
    ws = wb.active
    ws.title = "RO Request"

    ws["B2"] = "WEEKLY RO REQUEST — BOX ONLY"
    ws["B2"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["B2"].fill = PatternFill(
        start_color=ZUMA_GREEN, end_color=ZUMA_GREEN, fill_type="solid"
    )
    ws.merge_cells("B2:H2")
    ws["B2"].alignment = Alignment(horizontal="center", vertical="center")

    ws["B3"] = STORE_NAME
    ws["B3"].font = Font(bold=True, size=14)
    ws.merge_cells("B3:H3")

    ws["B5"] = "Generated:"
    ws["C5"] = datetime.now().strftime("%d %B %Y, %H:%M")

    ws["B7"] = "SUMMARY"
    ws["B7"].font = Font(bold=True, size=12, color="FFFFFF")
    ws["B7"].fill = PatternFill(
        start_color=ZUMA_GREEN, end_color=ZUMA_GREEN, fill_type="solid"
    )
    ws.merge_cells("B7:H7")
    ws["B7"].alignment = Alignment(horizontal="center")

    summary_rows = [
        ("Total Articles (RO Box):", summary_stats["total_ro_articles"]),
        ("Total Boxes to Order:", f"{summary_stats['total_boxes']} boxes"),
        ("Total Pairs to Order:", f"{summary_stats['total_pairs']} pairs"),
        ("WH Source:", "Warehouse Pusat — DDD (priority) + LJBB"),
        ("Logic:", "Max Stock (Planogram + Storage) vs On Hand → RO selisih"),
        ("Tier Priority:", "T1 → T8 → T2 → T3 → T4"),
    ]

    r = 8
    for label, val in summary_rows:
        ws[f"B{r}"] = label
        ws[f"B{r}"].font = Font(bold=True)
        ws[f"C{r}"] = val
        r += 1

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 45

    # =========================================================================
    # Sheet 2: Daftar RO Box
    # =========================================================================
    ws_ro = wb.create_sheet("Daftar RO Box")

    headers = [
        "No",
        "Article",
        "Kode Mix",
        "Gender",
        "Series",
        "Tier",
        "Max Stock (Pairs)",
        "On Hand (Pairs)",
        "Selisih (Pairs)",
        "Box Qty",
        "WH Qty DDD (Pairs)",
        "WH Qty LJBB (Pairs)",
    ]

    for col, h in enumerate(headers, 1):
        apply_style(ws_ro.cell(row=1, column=col, value=h), header_style())

    row_num = 2
    for idx, (_, d) in enumerate(ro_df.iterrows(), 1):
        ws_ro.cell(row=row_num, column=1, value=idx)
        ws_ro.cell(row=row_num, column=2, value=d["article"])
        ws_ro.cell(row=row_num, column=3, value=d["article_mix"])
        ws_ro.cell(row=row_num, column=4, value=d["gender"])
        ws_ro.cell(row=row_num, column=5, value=d["series"])
        ws_ro.cell(row=row_num, column=6, value=d["tier"])
        ws_ro.cell(row=row_num, column=7, value=int(d["max_stock"]))
        ws_ro.cell(row=row_num, column=8, value=int(d["on_hand"]))
        ws_ro.cell(row=row_num, column=9, value=int(d["selisih"]))
        ws_ro.cell(row=row_num, column=10, value=int(d["box_qty"]))
        ws_ro.cell(row=row_num, column=11, value=int(d["wh_qty_ddd"]))
        ws_ro.cell(row=row_num, column=12, value=int(d["wh_qty_ljbb"]))
        row_num += 1

    # Total row
    total_row = row_num + 1
    ws_ro.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws_ro.cell(row=total_row, column=9, value=summary_stats["total_pairs"]).font = Font(
        bold=True
    )
    ws_ro.cell(
        row=total_row, column=10, value=summary_stats["total_boxes"]
    ).font = Font(bold=True)

    col_widths = [5, 38, 15, 10, 16, 6, 18, 16, 16, 10, 20, 20]
    for i, w in enumerate(col_widths, 1):
        ws_ro.column_dimensions[chr(64 + i)].width = w

    # =========================================================================
    # Sheet 3: Reference
    # =========================================================================
    ws_ref = wb.create_sheet("Reference")
    ref_data = [
        ("Field", "Value"),
        ("Store", STORE_NAME),
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Logic", "Max Stock vs On Hand → RO selisih"),
        ("Max Stock", "total_planogram + storage_new (from temp_portal_plannogram)"),
        ("WH Source", "Warehouse Pusat — DDD + LJBB (no Protol/Reject)"),
        ("Tiers Excluded", ", ".join(EXCLUDED_TIERS)),
        ("Tier Priority", "T1 → T8 → T2 → T3 → T4"),
        ("Total Articles Scanned", summary_stats["total_scanned"]),
        ("Surplus (skipped)", summary_stats["surplus_count"]),
        ("No WH Stock (skipped)", summary_stats["no_wh_count"]),
        ("Final RO Articles", summary_stats["total_ro_articles"]),
        ("Total Boxes", summary_stats["total_boxes"]),
        ("Total Pairs", summary_stats["total_pairs"]),
    ]
    for i, (f, v) in enumerate(ref_data, 1):
        ws_ref.cell(row=i, column=1, value=f)
        ws_ref.cell(row=i, column=2, value=v)
        if i == 1:
            ws_ref.cell(row=i, column=1).font = Font(bold=True)
            ws_ref.cell(row=i, column=2).font = Font(bold=True)

    ws_ref.column_dimensions["A"].width = 30
    ws_ref.column_dimensions["B"].width = 55

    wb.save(output_path)
    print(f"✅ RO Box-Only Excel generated: {output_path}")


# ============================================================================
# MAIN EXECUTION
# ============================================================================


def main():
    print(f"🔍 Generating RO Box-Only for: {STORE_NAME}")
    print("=" * 60)

    # Step 1: Planogram
    print("📊 Fetching planogram...")
    plano_df = get_planogram(STORE_NAME)
    print(f"   Found {len(plano_df)} articles")

    # Exclude protol tiers
    before = len(plano_df)
    plano_df = plano_df[~plano_df["tier"].isin(EXCLUDED_TIERS)].copy()
    print(
        f"   Excluded tiers {EXCLUDED_TIERS}: {before - len(plano_df)} articles → {len(plano_df)} remain"
    )

    if plano_df.empty:
        print("❌ No planogram data. Check store name.")
        return

    # Step 2: Store on-hand stock
    print("📦 Fetching store stock...")
    stock_df = get_store_stock(STORE_DB_PATTERN)
    stock_df = stock_df.rename(columns={"article_mix": "article_mix_norm"})
    print(f"   Found {len(stock_df)} articles with stock")

    # Step 3: Merge planogram + on-hand → compute selisih
    print("🧮 Computing Max Stock vs On Hand...")
    plano_df["article_mix_norm"] = plano_df["article_mix"].str.lower().str.strip()
    merged = plano_df.merge(stock_df, on="article_mix_norm", how="left")
    merged["on_hand"] = merged["on_hand"].fillna(0).astype(int)
    merged["selisih"] = merged["max_stock"] - merged["on_hand"]

    surplus = merged[merged["selisih"] <= 0]
    candidates = merged[merged["selisih"] > 0].copy()
    print(f"   Surplus (skip): {len(surplus)} articles")
    print(f"   Need restock:   {len(candidates)} articles")

    if candidates.empty:
        print("✅ No restock needed — all articles at or above max stock.")
        return

    # Step 4: RO box qty = ceil(selisih / 12)
    candidates["box_qty"] = candidates["selisih"].apply(lambda x: math.ceil(x / 12))

    # Step 5: WH availability
    print("🏭 Checking WH Pusat stock...")
    wh_df = get_wh_stock(candidates["article_mix"].tolist())
    wh_df["article_mix_norm"] = wh_df["article_mix"].str.lower().str.strip()

    candidates = candidates.merge(
        wh_df[["article_mix_norm", "wh_qty_ddd", "wh_qty_ljbb"]],
        on="article_mix_norm",
        how="left",
    )
    candidates["wh_qty_ddd"] = candidates["wh_qty_ddd"].fillna(0).astype(int)
    candidates["wh_qty_ljbb"] = candidates["wh_qty_ljbb"].fillna(0).astype(int)
    candidates["wh_total"] = candidates["wh_qty_ddd"] + candidates["wh_qty_ljbb"]

    no_wh = candidates[candidates["wh_total"] < 12]
    final_ro = candidates[candidates["wh_total"] >= 12].copy()
    print(f"   WH available:   {len(final_ro)} articles")
    print(f"   WH insufficient (excluded): {len(no_wh)} articles")

    if final_ro.empty:
        print("⚠️  No articles have sufficient WH stock (≥12 pairs).")
        return

    # Step 6: Sort — tier priority, then DDD qty desc within tier
    final_ro["tier_sort"] = final_ro["tier"].map(TIER_ORDER).fillna(99)
    final_ro = (
        final_ro.sort_values(["tier_sort", "wh_qty_ddd"], ascending=[True, False])
        .drop(columns="tier_sort")
        .reset_index(drop=True)
    )

    total_boxes = int(final_ro["box_qty"].sum())
    total_pairs = int(final_ro["selisih"].sum())

    summary_stats = {
        "total_scanned": len(plano_df),
        "surplus_count": len(surplus),
        "no_wh_count": len(no_wh),
        "total_ro_articles": len(final_ro),
        "total_boxes": total_boxes,
        "total_pairs": total_pairs,
    }

    print("\n📋 SUMMARY:")
    print(f"   Total RO Articles : {len(final_ro)}")
    print(f"   Total Boxes       : {total_boxes}")
    print(f"   Total Pairs       : {total_pairs}")

    print("\n📄 Generating Excel...")
    generate_excel(final_ro, summary_stats, OUTPUT_FILE)
    print("\n✨ Done!")


if __name__ == "__main__":
    main()
