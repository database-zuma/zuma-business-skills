#!/usr/bin/env python3
"""
=============================================================================
  RO REQUEST GENERATOR — Universal (All Stores)

  Generates Restock (RO Box / RO Protol) + Surplus Pull List
  based on planogram targets vs actual store stock.

  Data sources:
    - Planogram: portal.planogram_existing_q1_2026 (DB)
    - Stock:     openclaw_ops DB (core.stock_with_product)
    - Sales:     openclaw_ops DB (core.sales_with_product)
    - Warehouse: Warehouse Pusat Box (DDD+LJBB), Warehouse Pusat Protol (DDD)

  Business Rules (from SKILL.md — zuma-distribution-flow):
    - RO Protol: when <50% sizes empty → send individual pairs from Gudang Protol
    - RO Box:    when >=50% sizes empty → send full box (12 pairs) from Gudang Box
    - Surplus:   when tier actual% > tier ideal% → pull lowest-selling articles
    - STORAGE = 0 → every RO Box creates surplus risk (pre-plan redistribution)

  CLI Usage:
    python build_ro_request.py --store "Icon Mall Gresik" --storage 0
    python build_ro_request.py --store "Royal Plaza" --storage 75
    python build_ro_request.py --store "Galaxy Mall" --storage 0
    python build_ro_request.py --store "Royal Plaza" --storage 75 --threshold 0.60
    python build_ro_request.py --store "Icon Mall Gresik" --output /tmp/my_ro.xlsx

  Args:
    --store       (required) Store display name. Used as ILIKE pattern on
                  nama_gudang (stock) and store_name (planogram).
    --storage     (optional, default 0) Storage capacity in boxes.
    --output      (optional) Output .xlsx path. Default: Desktop/DN PO ENTITAS/
                  RO_Request_{StoreName}_{date}.xlsx
    --threshold   (optional, default 0.50) % size kosong threshold for RO Box.

=============================================================================
"""

import argparse
import os
import sys
import io
import re
from datetime import datetime, date
from collections import defaultdict

# Fix Windows console encoding for Unicode symbols
if sys.stdout.encoding != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

try:
    import psycopg2
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing dependencies...")
    os.system(f"{sys.executable} -m pip install psycopg2-binary openpyxl")
    import psycopg2
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# =============================================================================
# DATABASE CONFIG
# =============================================================================

DB_CONFIG = {
    "host": "76.13.194.120",
    "port": 5432,
    "dbname": "openclaw_ops",
    "user": "openclaw_app",
    "password": "Zuma-0psCl4w-2026!",
}

# =============================================================================
# PLANOGRAM — size columns in portal.planogram_existing_q1_2026
# =============================================================================

PLANOGRAM_TABLE = "portal.planogram_existing_q1_2026"

# 42 size columns: 25 individual + 17 paired
PLANOGRAM_SIZE_COLS = [
    # Individual sizes (25)
    "size_21", "size_22", "size_23", "size_24", "size_25",
    "size_27", "size_28", "size_29", "size_30", "size_31",
    "size_32", "size_33", "size_34", "size_35", "size_36",
    "size_37", "size_38", "size_39", "size_40", "size_41",
    "size_42", "size_43", "size_44", "size_45", "size_46",
    # Paired sizes (17)
    "size_18_19", "size_20_21", "size_21_22", "size_22_23", "size_23_24",
    "size_24_25", "size_25_26", "size_27_28", "size_29_30", "size_31_32",
    "size_33_34", "size_35_36", "size_37_38", "size_39_40", "size_41_42",
    "size_43_44", "size_45_46",
]

# Surplus tiers to check (T1, T2, T3 only — T4/T5 excluded, T8 protected)
SURPLUS_CHECK_TIERS = [1, 2, 3]

# =============================================================================
# EXCEL STYLING
# =============================================================================

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL_DARK   = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
HEADER_FILL_GREEN  = PatternFill(start_color="0D6B3E", end_color="0D6B3E", fill_type="solid")
HEADER_FILL_RED    = PatternFill(start_color="8B1A1A", end_color="8B1A1A", fill_type="solid")
HEADER_FILL_ORANGE = PatternFill(start_color="B35900", end_color="B35900", fill_type="solid")
HEADER_FILL_PURPLE = PatternFill(start_color="4A1B6D", end_color="4A1B6D", fill_type="solid")
HEADER_FILL_TEAL   = PatternFill(start_color="0E4D5C", end_color="0E4D5C", fill_type="solid")
FILL_LIGHT_GREEN   = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
FILL_LIGHT_RED     = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
FILL_LIGHT_YELLOW  = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
FILL_LIGHT_GRAY    = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
FILL_ZEBRA         = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


# =============================================================================
# HELPERS
# =============================================================================

def col_to_size_label(col_name: str) -> str:
    """
    Convert planogram DB column name to stock size label.
    Examples:
      size_35      → 35
      size_35_36   → 35/36
      size_18_19   → 18/19
    """
    label = col_name[5:]       # strip 'size_'
    label = label.replace("_", "/")
    return label


def build_plano_pattern(store_arg: str) -> str:
    """
    Build a flexible ILIKE pattern for planogram store_name.
    Removes common bridging words ('Mall', 'Icon' by itself) that sometimes
    differ between the planogram table and the stock table.

    Examples:
      "Icon Mall Gresik" → "%Icon%Gresik%"  (matches "Zuma Icon Gresik")
      "Royal Plaza"      → "%Royal%Plaza%"  (matches "Zuma Royal Plaza")
      "Galaxy Mall"      → "%Galaxy%"        (matches "Zuma Galaxy Mall")
    """
    skip_words = {"mall"}
    words = [w for w in store_arg.split() if w.lower() not in skip_words]
    if not words:
        words = store_arg.split()
    return "%" + "%".join(words) + "%"


def to_float(val, default=0.0) -> float:
    """Safely convert a DB value (possibly text) to float."""
    if val is None or val == "":
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def style_header_row(ws, row_num, max_col, fill=HEADER_FILL_DARK):
    """Apply dark header styling to a given row."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def auto_width(ws, min_width=10, max_width=35):
    """Auto-adjust column widths based on content."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 2, max_width))


# =============================================================================
# 1. READ PLANOGRAM FROM DATABASE
# =============================================================================

def read_planogram(conn, store_arg: str) -> dict:
    """
    Read planogram targets from portal.planogram_existing_q1_2026.
    Uses ILIKE '%{pattern}%' on store_name (flexible — handles naming differences).

    Returns:
      { kode_mix: {gender, series, article, tier, kode_mix, sizes, total_planogram, box} }
    """
    plano_pattern = build_plano_pattern(store_arg)
    print(f"[1/8] Reading planogram from DB (store_name ILIKE '{plano_pattern}')...")

    cur = conn.cursor()
    size_cols_sql = ", ".join(PLANOGRAM_SIZE_COLS)
    cur.execute(f"""
        SELECT article_mix, kode_kecil, article, gender, series, tier,
               {size_cols_sql},
               box
        FROM {PLANOGRAM_TABLE}
        WHERE LOWER(store_name) ILIKE LOWER(%s)
        ORDER BY article_mix
    """, (plano_pattern,))
    rows = cur.fetchall()
    cur.close()

    if not rows:
        # Fallback: try exact ILIKE with original store_arg
        print(f"       No results with pattern '{plano_pattern}', trying '%{store_arg}%'...")
        cur = conn.cursor()
        cur.execute(f"""
            SELECT article_mix, kode_kecil, article, gender, series, tier,
                   {size_cols_sql},
                   box
            FROM {PLANOGRAM_TABLE}
            WHERE LOWER(store_name) ILIKE LOWER(%s)
            ORDER BY article_mix
        """, (f"%{store_arg}%",))
        rows = cur.fetchall()
        cur.close()

    articles = {}
    n_size_cols = len(PLANOGRAM_SIZE_COLS)

    for row in rows:
        article_mix = row[0]
        kode_kecil = row[1] if len(row) > 1 else ""
        article     = row[2]
        gender      = row[3]
        series      = row[3]
        tier_raw    = row[4]
        size_values = row[5:5 + n_size_cols]
        box_val     = row[5 + n_size_cols]

        if not article_mix:
            continue

        kode_mix = str(article_mix).strip()

        # Build sizes dict (only columns with non-zero values)
        sizes = {}
        total_planogram = 0.0
        for col_name, val in zip(PLANOGRAM_SIZE_COLS, size_values):
            fval = to_float(val)
            if fval > 0:
                size_label = col_to_size_label(col_name)
                sizes[size_label] = fval
                total_planogram += fval

        if not sizes:
            continue  # Article has no sizes in planogram, skip

        try:
            tier = int(tier_raw) if tier_raw else 0
        except (ValueError, TypeError):
            tier = 0

        articles[kode_mix] = {
            "gender":            gender or "",
            "series":            series or "",
            "article":           article or kode_mix,
            "tier":              tier,
            "kode_mix":          kode_mix,
            "kode_kecil":        kode_kecil or "",
            "sizes":             sizes,
            "total_planogram":   total_planogram,
            "box":               to_float(box_val),
            "avg_sales_3m_pairs": 0.0,
            "avg_sales_3m_box":   0.0,
        }

    print(f"       → {len(rows)} planogram rows from DB, {len(articles)} with sizes")

    # Tier breakdown
    tier_counts = defaultdict(int)
    for a in articles.values():
        tier_counts[a["tier"]] += 1
    for t in sorted(tier_counts):
        print(f"         Tier {t}: {tier_counts[t]} articles")

    return articles


# =============================================================================
# 2. QUERY DATABASE — STOCK, SALES, WAREHOUSE
# =============================================================================

def query_db(conn, store_arg: str) -> dict:
    """
    Pull all required data from openclaw_ops for the given store.
    Uses ILIKE '%{store_arg}%' on nama_gudang / matched_store_name.

    Returns dict with keys:
      store_stock, wh_box_stock, wh_protol_stock, sales_3m,
      all_store_articles, snapshot_date
    """
    store_pattern = f"%{store_arg}%"
    print(f"[2/8] Querying database (stock/sales ILIKE '{store_pattern}')...")

    cur = conn.cursor()

    # --- 2a. Store actual stock per kode_mix per ukuran ---
    print("       → Store actual stock...")
    cur.execute("""
        SELECT kode_mix, ukuran, SUM(quantity) as qty
        FROM core.stock_with_product
        WHERE LOWER(nama_gudang) ILIKE LOWER(%s)
          AND kode_mix IS NOT NULL
        GROUP BY kode_mix, ukuran
        ORDER BY kode_mix, ukuran
    """, (store_pattern,))

    store_stock = defaultdict(lambda: defaultdict(float))
    for kode_mix, ukuran, qty in cur.fetchall():
        store_stock[str(kode_mix)][str(ukuran) if ukuran else ""] = to_float(qty)

    print(f"         {len(store_stock)} articles with stock data")

    # --- 2b. Warehouse Pusat Box stock (DDD + LJBB) per kode_mix per ukuran ---
    print("       → Warehouse Pusat BOX stock (DDD + LJBB)...")
    cur.execute("""
        SELECT kode_mix, ukuran, SUM(quantity) as qty
        FROM core.stock_with_product
        WHERE LOWER(nama_gudang) = 'warehouse pusat'
          AND source_entity IN ('DDD', 'LJBB')
          AND kode_mix IS NOT NULL
        GROUP BY kode_mix, ukuran
        ORDER BY kode_mix, ukuran
    """)

    wh_box_stock = defaultdict(lambda: defaultdict(float))
    for kode_mix, ukuran, qty in cur.fetchall():
        wh_box_stock[str(kode_mix)][str(ukuran) if ukuran else ""] = to_float(qty)

    print(f"         {len(wh_box_stock)} articles in WH Pusat Box")

    # --- 2c. Warehouse Pusat Protol stock (DDD only) per kode_mix per ukuran ---
    print("       → Warehouse Pusat PROTOL stock (DDD only)...")
    cur.execute("""
        SELECT kode_mix, ukuran, SUM(quantity) as qty
        FROM core.stock_with_product
        WHERE LOWER(nama_gudang) = 'warehouse pusat protol'
          AND source_entity = 'DDD'
          AND kode_mix IS NOT NULL
        GROUP BY kode_mix, ukuran
        ORDER BY kode_mix, ukuran
    """)

    wh_protol_stock = defaultdict(lambda: defaultdict(float))
    for kode_mix, ukuran, qty in cur.fetchall():
        wh_protol_stock[str(kode_mix)][str(ukuran) if ukuran else ""] = to_float(qty)

    print(f"         {len(wh_protol_stock)} articles in WH Pusat Protol")

    # --- 2d. Sales last 3 months at store ---
    print("       → Sales last 3 months at store...")
    cur.execute("""
        SELECT kode_mix, SUM(quantity) as total_pairs
        FROM core.sales_with_product
        WHERE LOWER(matched_store_name) ILIKE LOWER(%s)
          AND transaction_date >= CURRENT_DATE - INTERVAL '3 months'
          AND kode_mix IS NOT NULL
        GROUP BY kode_mix
        ORDER BY kode_mix
    """, (store_pattern,))

    sales_3m = {}
    for kode_mix, total_pairs in cur.fetchall():
        sales_3m[str(kode_mix)] = to_float(total_pairs)

    print(f"         {len(sales_3m)} articles with sales data")

    # --- 2e. All articles at store (for off-planogram detection) ---
    print("       → Full article list at store (for off-planogram check)...")
    cur.execute("""
        SELECT kode_mix, article, gender, series, tier, ukuran, SUM(quantity) as qty
        FROM core.stock_with_product
        WHERE LOWER(nama_gudang) ILIKE LOWER(%s)
          AND kode_mix IS NOT NULL
          AND quantity > 0
        GROUP BY kode_mix, article, gender, series, tier, ukuran
        ORDER BY kode_mix, ukuran
    """, (store_pattern,))

    all_store_articles = defaultdict(
        lambda: {"article": "", "gender": "", "series": "", "tier": 0, "sizes": {}}
    )
    for kode_mix, article, gender, series, tier, ukuran, qty in cur.fetchall():
        rec = all_store_articles[str(kode_mix)]
        rec["article"] = article or str(kode_mix)
        rec["gender"]  = gender or ""
        rec["series"]  = series or ""
        rec["tier"]    = int(tier) if tier else 0
        rec["sizes"][str(ukuran) if ukuran else ""] = to_float(qty)

    print(f"         {len(all_store_articles)} articles with stock > 0 at store")

    # --- 2f. Snapshot date ---
    cur.execute("""
        SELECT MAX(snapshot_date) FROM core.stock_with_product
        WHERE LOWER(nama_gudang) ILIKE LOWER(%s)
    """, (store_pattern,))
    result = cur.fetchone()
    snapshot_date = result[0] if result and result[0] else date.today()
    print(f"         Stock snapshot date: {snapshot_date}")

    cur.close()

    return {
        "store_stock":        dict(store_stock),
        "wh_box_stock":       dict(wh_box_stock),
        "wh_protol_stock":    dict(wh_protol_stock),
        "sales_3m":           sales_3m,
        "all_store_articles": dict(all_store_articles),
        "snapshot_date":      snapshot_date,
    }


# =============================================================================
# 3. GAP ANALYSIS
# =============================================================================

def calculate_gaps(planogram: dict, db_data: dict, ro_box_threshold: float) -> list:
    """
    For each displayed article, compare planogram target vs actual stock per size.
    Decides RO type: RO_BOX (>=threshold% sizes empty) or RO_PROTOL (<threshold%).
    """
    print(f"[3/8] Calculating gaps (threshold={ro_box_threshold:.0%})...")

    store_stock = db_data["store_stock"]
    results = []

    for kode_mix, plano in planogram.items():
        target_sizes = plano["sizes"]
        actual_sizes = store_stock.get(kode_mix, {})

        total_assortment = len(target_sizes)
        if total_assortment == 0:
            continue

        sizes_detail = []
        sizes_zero = 0
        total_gap = 0
        total_actual = 0
        total_target = 0

        for size_label, target in sorted(target_sizes.items(), key=lambda x: x[0]):
            actual = actual_sizes.get(size_label, 0)
            gap = max(0, target - actual)
            is_zero = actual <= 0

            sizes_detail.append({
                "size":    size_label,
                "target":  target,
                "actual":  actual,
                "gap":     gap,
                "is_zero": is_zero,
            })

            if is_zero:
                sizes_zero += 1
            total_gap    += gap
            total_actual += actual
            total_target += target

        pct_size_kosong = sizes_zero / total_assortment if total_assortment > 0 else 0

        if total_gap <= 0:
            ro_type = "NO_RESTOCK"
        elif pct_size_kosong >= ro_box_threshold:
            ro_type = "RO_BOX"
        else:
            ro_type = "RO_PROTOL"

        results.append({
            **plano,
            "sizes_detail":     sizes_detail,
            "total_assortment": total_assortment,
            "sizes_zero":       sizes_zero,
            "pct_size_kosong":  pct_size_kosong,
            "total_gap":        total_gap,
            "total_actual":     total_actual,
            "total_target":     total_target,
            "ro_type":          ro_type,
        })

    ro_box    = sum(1 for r in results if r["ro_type"] == "RO_BOX")
    ro_protol = sum(1 for r in results if r["ro_type"] == "RO_PROTOL")
    no_rest   = sum(1 for r in results if r["ro_type"] == "NO_RESTOCK")
    print(f"       → RO Box: {ro_box}, RO Protol: {ro_protol}, No restock: {no_rest}")

    return results


# =============================================================================
# 4. RO DECISIONS + WAREHOUSE AVAILABILITY
# =============================================================================

def generate_ro_decisions(gap_results: list, db_data: dict, storage_capacity: int):
    """
    For each article needing restock:
    - RO_PROTOL: check WH Pusat Protol availability per size
    - RO_BOX:    check WH Pusat Box availability (full box)
    Adds surplus pre-plan warnings for STORAGE=0 stores.
    """
    print(f"[4/8] Checking warehouse availability...")

    wh_box    = db_data["wh_box_stock"]
    wh_protol = db_data["wh_protol_stock"]

    ro_protol_list  = []
    ro_box_list     = []
    no_restock_list = []

    for art in gap_results:
        kode_mix = art["kode_mix"]
        ro_type  = art["ro_type"]

        if ro_type == "NO_RESTOCK":
            no_restock_list.append(art)
            continue

        if ro_type == "RO_PROTOL":
            protol_available  = wh_protol.get(kode_mix, {})
            protol_can_fill   = []
            protol_cannot_fill = []

            for sd in art["sizes_detail"]:
                if sd["gap"] <= 0:
                    continue
                avail     = protol_available.get(sd["size"], 0)
                can_fill  = min(sd["gap"], avail)
                remaining = sd["gap"] - can_fill

                sd["protol_available"] = avail
                sd["protol_fill"]      = can_fill
                sd["protol_remaining"] = remaining

                if can_fill > 0:
                    protol_can_fill.append(sd)
                if remaining > 0:
                    protol_cannot_fill.append(sd)

            art["protol_can_fill"]    = protol_can_fill
            art["protol_cannot_fill"] = protol_cannot_fill

            if protol_cannot_fill:
                box_avail         = wh_box.get(kode_mix, {})
                total_box_pairs   = sum(box_avail.values())
                art["fallback_box_available"]    = total_box_pairs > 0
                art["fallback_box_total_pairs"]  = total_box_pairs

                if total_box_pairs > 0 and storage_capacity == 0:
                    sizes_with_stock = [sd for sd in art["sizes_detail"]
                                        if sd["actual"] > 0 and sd["gap"] <= 0]
                    art["surplus_preplan_sizes"]   = sizes_with_stock
                    art["surplus_preplan_warning"] = (
                        f"⚠ STORAGE=0: Box will add surplus in "
                        f"{len(sizes_with_stock)} sizes already stocked. "
                        f"Pre-plan redistribution BEFORE approving."
                    )
            else:
                art["fallback_box_available"]   = False
                art["fallback_box_total_pairs"]  = 0

            ro_protol_list.append(art)

        elif ro_type == "RO_BOX":
            box_avail       = wh_box.get(kode_mix, {})
            total_box_pairs = sum(box_avail.values())

            art["box_available"]    = total_box_pairs > 0
            art["box_total_pairs"]  = total_box_pairs
            art["box_sizes"]        = dict(box_avail)

            if total_box_pairs > 0 and storage_capacity == 0:
                sizes_with_stock = [sd for sd in art["sizes_detail"] if sd["actual"] > 0]
                art["surplus_preplan_sizes"] = sizes_with_stock
                art["surplus_preplan_warning"] = (
                    f"⚠ STORAGE=0: Box will add surplus in "
                    f"{len(sizes_with_stock)} sizes still stocked. Pre-plan redistribution."
                ) if sizes_with_stock else ""
            else:
                art["surplus_preplan_warning"] = ""

            ro_box_list.append(art)

    print(f"       → RO Protol requests: {len(ro_protol_list)}")
    print(f"       → RO Box requests:    {len(ro_box_list)}")
    print(f"       → No restock needed:  {len(no_restock_list)}")

    return ro_protol_list, ro_box_list, no_restock_list


# =============================================================================
# 5. TO / COVERAGE CALCULATION
# =============================================================================

def calculate_to(gap_results: list, db_data: dict):
    """
    Calculate stock_coverage_months and turnover_rate for each article.
    For surplus: sorts by avg_monthly_sales ASC (slowest sellers first).
    """
    print(f"[5/8] Calculating TO / stock coverage...")

    sales_3m = db_data["sales_3m"]

    for art in gap_results:
        kode_mix         = art["kode_mix"]
        total_pairs_3m   = sales_3m.get(kode_mix, 0)
        avg_monthly_sales = total_pairs_3m / 3.0 if total_pairs_3m > 0 else 0

        current_stock = art["total_actual"]

        if avg_monthly_sales > 0:
            stock_coverage_months = current_stock / avg_monthly_sales
        else:
            stock_coverage_months = float("inf") if current_stock > 0 else 0

        if current_stock > 0:
            turnover_rate = avg_monthly_sales / current_stock
        else:
            turnover_rate = float("inf") if avg_monthly_sales > 0 else 0

        art["total_pairs_3m"]       = total_pairs_3m
        art["avg_monthly_sales"]    = avg_monthly_sales
        art["stock_coverage_months"] = stock_coverage_months
        art["turnover_rate"]        = turnover_rate

    print(f"       → TO calculated for {len(gap_results)} articles")


# =============================================================================
# 6. SURPLUS CALCULATION (Tier-Based)
# =============================================================================

def calculate_surplus(planogram: dict, gap_results: list, db_data: dict) -> dict:
    """
    Surplus = when a tier is OVER-CAPACITY at the store.
    Only T1, T2, T3 are checked. T4/T5 excluded (clearance). T8 excluded (protected).

    Steps:
    1. Derive ideal tier distribution from planogram (article count per tier)
    2. Calculate actual tier distribution from current store stock
    3. For over-capacity tiers: rank articles by avg_monthly_sales ASC, pull slowest N
    4. Also detect off-planogram stock
    """
    print(f"[6/8] Calculating surplus...")

    all_store_articles = db_data["all_store_articles"]
    store_stock        = db_data["store_stock"]

    # --- Ideal tier distribution (from planogram) ---
    ideal_tier_counts = defaultdict(int)
    for art in planogram.values():
        ideal_tier_counts[art["tier"]] += 1

    total_planogram_articles = sum(ideal_tier_counts.values())
    ideal_tier_pct = {t: c / total_planogram_articles if total_planogram_articles > 0 else 0
                      for t, c in ideal_tier_counts.items()}

    print(f"       Ideal tier distribution (planogram, {total_planogram_articles} articles):")
    for t in sorted(ideal_tier_pct):
        print(f"         Tier {t}: {ideal_tier_counts[t]} articles ({ideal_tier_pct[t] * 100:.1f}%)")

    # --- Actual tier distribution (articles with stock at store) ---
    actual_tier_articles = defaultdict(list)

    for art in gap_results:
        if art["total_actual"] > 0:
            size_stock = store_stock.get(art["kode_mix"], {})
            actual_tier_articles[art["tier"]].append({
                "kode_mix":             art["kode_mix"],
                "article":              art["article"],
                "gender":               art["gender"],
                "series":               art["series"],
                "tier":                 art["tier"],
                "total_stock":          art["total_actual"],
                "sizes":                {s: q for s, q in size_stock.items() if q > 0},
                "avg_monthly_sales":    art.get("avg_monthly_sales", 0),
                "stock_coverage_months": art.get("stock_coverage_months", 0),
                "on_planogram":         True,
            })

    off_planogram = []
    for kode_mix, info in all_store_articles.items():
        if kode_mix not in planogram:
            total_stock = sum(info["sizes"].values())
            if total_stock > 0:
                sales_val  = db_data["sales_3m"].get(kode_mix, 0)
                avg_monthly = sales_val / 3.0 if sales_val > 0 else 0
                coverage    = (total_stock / avg_monthly if avg_monthly > 0
                               else (float("inf") if total_stock > 0 else 0))
                art_info = {
                    "kode_mix":             kode_mix,
                    "article":              info["article"],
                    "gender":               info["gender"],
                    "series":               info["series"],
                    "tier":                 info["tier"],
                    "total_stock":          total_stock,
                    "sizes":                info["sizes"],
                    "avg_monthly_sales":    avg_monthly,
                    "stock_coverage_months": coverage,
                    "on_planogram":         False,
                }
                off_planogram.append(art_info)
                actual_tier_articles[info["tier"]].append(art_info)

    actual_tier_counts = {t: len(arts) for t, arts in actual_tier_articles.items()}
    total_actual_articles = sum(actual_tier_counts.values())
    actual_tier_pct = {t: c / total_actual_articles if total_actual_articles > 0 else 0
                       for t, c in actual_tier_counts.items()}

    print(f"\n       Actual tier distribution ({total_actual_articles} articles with stock):")
    for t in sorted(actual_tier_counts):
        ideal  = ideal_tier_counts.get(t, 0)
        actual = actual_tier_counts[t]
        diff   = actual - ideal
        status = "OVER" if diff > 0 else ("UNDER" if diff < 0 else "OK")
        check  = "✓ check" if t in SURPLUS_CHECK_TIERS else "✗ skip"
        print(f"         Tier {t}: {actual} actual vs {ideal} ideal "
              f"({'+' if diff > 0 else ''}{diff}) [{status}] {check}")

    # --- Surplus candidates (only SURPLUS_CHECK_TIERS) ---
    surplus_pull_list = []

    for tier in SURPLUS_CHECK_TIERS:
        ideal_count  = ideal_tier_counts.get(tier, 0)
        actual_count = actual_tier_counts.get(tier, 0)

        if actual_count <= ideal_count:
            continue

        over_by        = actual_count - ideal_count
        tier_articles  = actual_tier_articles[tier]
        tier_sorted    = sorted(tier_articles, key=lambda x: x["avg_monthly_sales"])

        for art in tier_sorted[:over_by]:
            surplus_pull_list.append({
                **art,
                "surplus_reason": f"Tier {tier} over-capacity by {over_by} articles",
                "pull_priority":  "HIGH" if art["avg_monthly_sales"] == 0 else "MEDIUM",
            })

    print(f"\n       → Surplus candidates:          {len(surplus_pull_list)}")
    print(f"       → Off-planogram with stock:    {len(off_planogram)}")

    return {
        "surplus_pull_list":        surplus_pull_list,
        "off_planogram":            off_planogram,
        "ideal_tier_counts":        dict(ideal_tier_counts),
        "ideal_tier_pct":           dict(ideal_tier_pct),
        "actual_tier_counts":       dict(actual_tier_counts),
        "actual_tier_pct":          dict(actual_tier_pct),
        "total_planogram_articles": total_planogram_articles,
        "total_actual_articles":    total_actual_articles,
    }


# =============================================================================
# 7. EXCEL OUTPUT
# =============================================================================

def write_excel(
    store_name: str,
    storage_capacity: int,
    output_file: str,
    ro_protol_list: list,
    ro_box_list: list,
    no_restock_list: list,
    surplus_data: dict,
    gap_results: list,
    planogram: dict,
    db_data: dict,
):
    """
    Write 5-sheet RO Request Excel workbook.

    Sheets:
      1. RO Request       — Cover page: summary, instructions, signatures
      2. Daftar RO Protol — Picking list for WH Protol (one row per article)
      3. Daftar RO Box    — Picking list for WH Box (one row per article)
      4. Daftar Surplus   — Pull list size-level for store staff
      5. Reference        — Tier analysis + full article status (internal)
    """
    print(f"\n[7/8] Writing Excel output: {output_file}")

    os.makedirs(os.path.dirname(output_file), exist_ok=True)

    wb       = openpyxl.Workbook()
    snapshot = db_data.get("snapshot_date", date.today())
    today    = datetime.now()
    has_ro_box = len(ro_box_list) > 0

    # Extra fonts for form layout
    TITLE_FONT         = Font(bold=True, size=20, color="1B2A4A")
    STORE_FONT         = Font(bold=True, size=15, color="0D6B3E")
    SECTION_TITLE_FONT = Font(bold=True, size=13, color="1B2A4A")
    LABEL_FONT         = Font(bold=True, size=11)
    VALUE_FONT         = Font(size=11)
    SMALL_FONT         = Font(size=10, color="666666")
    SIGN_FONT          = Font(size=10)
    TOTAL_FONT         = Font(bold=True, size=11, color="1B2A4A")

    def write_form_header(ws, row, title, subtitle, total_cols):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
        ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=16, color="1B2A4A")
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
        ws.cell(row=row, column=1, value=store_name).font = Font(bold=True, size=12, color="0D6B3E")
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
        ws.cell(row=row, column=1, value=f"Week of {today.strftime('%d %B %Y')}  |  Stock Snapshot: {snapshot}").font = SMALL_FONT
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
        ws.cell(row=row, column=1, value=subtitle).font = Font(size=11, color="333333")
        row += 2
        return row

    # Compute totals
    total_protol_pairs = 0
    for art in ro_protol_list:
        for sd in art["sizes_detail"]:
            if sd["gap"] > 0:
                total_protol_pairs += sd.get("protol_fill", sd.get("gap", 0))

    total_surplus_pairs = sum(a["total_stock"] for a in surplus_data["surplus_pull_list"])

    # =========================================================================
    # SHEET 1: RO REQUEST — Cover Page
    # =========================================================================
    ws = wb.active
    ws.title = "RO Request"
    ws.sheet_properties.tabColor = "1B2A4A"

    for col, width in zip("ABCDE", [24, 22, 22, 28, 22]):
        ws.column_dimensions[col].width = width

    row = 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1, value="WEEKLY RO REQUEST").font = TITLE_FONT
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1, value=store_name).font = STORE_FONT
    row += 2

    for label, val in [
        ("Week of:",         today.strftime("%d %B %Y")),
        ("Stock Snapshot:",  str(snapshot)),
        ("Storage Capacity:", f"{storage_capacity} boxes"),
    ]:
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        ws.cell(row=row, column=2, value=val).font = VALUE_FONT
        row += 1

    row += 1
    for role, name in [("From:", "Area Supervisor"), ("To:", "Warehouse Supervisor")]:
        ws.cell(row=row, column=1, value=role).font = LABEL_FONT
        ws.cell(row=row, column=2, value=name).font = VALUE_FONT
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
        ws.cell(row=row, column=3, value="___________________________").font = SIGN_FONT
        row += 1

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1, value="REQUEST SUMMARY").font = SECTION_TITLE_FONT
    row += 1

    for ci, h in enumerate(["Type", "Articles", "Total", "Source / Destination", "See Sheet"], 1):
        ws.cell(row=row, column=ci, value=h)
    style_header_row(ws, row, 5, HEADER_FILL_DARK)
    row += 1

    for label, count, total, source, sheet in [
        ("RO PROTOL",    len(ro_protol_list),
         f"{int(total_protol_pairs)} pairs",  "FROM: WH Pusat Protol", "Daftar RO Protol"),
        ("RO BOX",       len(ro_box_list) if has_ro_box else 0,
         f"{len(ro_box_list)} boxes" if has_ro_box else "NONE this week",
         "FROM: WH Pusat Box" if has_ro_box else "-",
         "Daftar RO Box" if has_ro_box else "-"),
        ("SURPLUS PULL", len(surplus_data["surplus_pull_list"]),
         f"{int(total_surplus_pairs)} pairs", "TO: WH Pusat Protol", "Daftar Surplus"),
    ]:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=count)
        ws.cell(row=row, column=3, value=total)
        ws.cell(row=row, column=4, value=source)
        ws.cell(row=row, column=5, value=sheet)
        for ci in range(1, 6):
            ws.cell(row=row, column=ci).border = THIN_BORDER
        row += 1

    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1, value="INSTRUCTIONS").font = SECTION_TITLE_FONT
    row += 1
    for inst in [
        "1. RO Protol: Pick individual pairs from Warehouse Pusat Protol per article list.",
        "2. RO Box: Pick full boxes (1 box = 12 pairs) from Warehouse Pusat Box per article list.",
        "3. Surplus: Pull listed items from store display, return to Warehouse Pusat Protol.",
        "4. Restock + Surplus happen SAME DAY: deliver RO items IN, pull surplus items OUT.",
        "5. Priority: RO Protol first. RO Box only when >=50% sizes are empty.",
    ]:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.cell(row=row, column=1, value=inst).font = Font(size=10)
        row += 1

    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1, value="SIGNATURES").font = SECTION_TITLE_FONT
    row += 1
    for ci, h in enumerate(["", "Prepared by", "Approved by", "Received by"], 1):
        ws.cell(row=row, column=ci, value=h).font = LABEL_FONT
        ws.cell(row=row, column=ci).alignment = Alignment(horizontal="center")
    row += 1
    for field in ["Name:", "Date:", "Signature:"]:
        ws.cell(row=row, column=1, value=field).font = SIGN_FONT
        for ci in range(2, 5):
            ws.cell(row=row, column=ci, value="___________________").font = SIGN_FONT
            ws.cell(row=row, column=ci).alignment = Alignment(horizontal="center")
        row += 1

    # =========================================================================
    # SHEET 2: DAFTAR RO PROTOL
    # =========================================================================
    ws2 = wb.create_sheet("Daftar RO Protol")
    ws2.sheet_properties.tabColor = "0D6B3E"

    headers2    = ["No", "Article (Kode Mix)", "Tier", "Gender", "Series",
                   "Sizes Needed (size:qty)", "Total Pairs"]
    total_cols2 = len(headers2)

    row = 1
    row = write_form_header(
        ws2, row, "DAFTAR RO PROTOL",
        f"Source: Warehouse Pusat Protol (DDD)  |  "
        f"Total: {len(ro_protol_list)} articles, {int(total_protol_pairs)} pairs",
        total_cols2,
    )
    for ci, h in enumerate(headers2, 1):
        ws2.cell(row=row, column=ci, value=h)
    style_header_row(ws2, row, total_cols2, HEADER_FILL_GREEN)
    row += 1

    num = 1
    for art in sorted(ro_protol_list, key=lambda x: (x["tier"], x["article"])):
        sizes_needed = []
        total_fill   = 0
        for sd in art["sizes_detail"]:
            if sd["gap"] > 0:
                fill = sd.get("protol_fill", sd.get("gap", 0))
                sizes_needed.append(f"{sd['size']}:{int(sd['gap'])}")
                total_fill += fill
        if not sizes_needed:
            continue

        ws2.cell(row=row, column=1, value=num)
        ws2.cell(row=row, column=2, value=art["kode_mix"])
        ws2.cell(row=row, column=3, value=art["tier"])
        ws2.cell(row=row, column=4, value=art["gender"])
        ws2.cell(row=row, column=5, value=art["series"])
        ws2.cell(row=row, column=6, value=", ".join(sizes_needed))
        ws2.cell(row=row, column=7, value=int(total_fill))

        if num % 2 == 0:
            for ci in range(1, total_cols2 + 1):
                ws2.cell(row=row, column=ci).fill = FILL_ZEBRA
        for ci in range(1, total_cols2 + 1):
            ws2.cell(row=row, column=ci).border = THIN_BORDER
        num += 1
        row += 1

    # Total row
    for ci in range(1, total_cols2 + 1):
        ws2.cell(row=row, column=ci).border = THIN_BORDER
        ws2.cell(row=row, column=ci).fill   = FILL_LIGHT_GRAY
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws2.cell(row=row, column=1, value="TOTAL PAIRS").font = TOTAL_FONT
    ws2.cell(row=row, column=7, value=int(total_protol_pairs)).font = TOTAL_FONT

    auto_width(ws2)

    # =========================================================================
    # SHEET 3: DAFTAR RO BOX
    # =========================================================================
    ws3 = wb.create_sheet("Daftar RO Box")
    ws3.sheet_properties.tabColor = "8B1A1A"

    headers3    = ["No", "Article (Kode Mix)", "Kode Kecil", "Tier", "Gender", "Series",
                   "Box Qty", "WH Available"]
    total_cols3 = len(headers3)

    row = 1
    row = write_form_header(
        ws3, row, "DAFTAR RO BOX",
        (f"Source: Warehouse Pusat Box (DDD + LJBB)  |  Total: {len(ro_box_list)} boxes"
         if has_ro_box else "DAFTAR RO BOX  |  Tidak ada RO Box minggu ini"),
        total_cols3,
    )
    for ci, h in enumerate(headers3, 1):
        ws3.cell(row=row, column=ci, value=h)
    style_header_row(ws3, row, total_cols3, HEADER_FILL_RED)
    row += 1

    if has_ro_box:
        num = 1
        for art in sorted(ro_box_list, key=lambda x: (x["tier"], x["article"])):
            ws3.cell(row=row, column=1, value=num)
            ws3.cell(row=row, column=2, value=art["kode_mix"])
            ws3.cell(row=row, column=3, value=art.get("kode_kecil", ""))
            ws3.cell(row=row, column=4, value=art["tier"])
            ws3.cell(row=row, column=5, value=art["gender"])
            ws3.cell(row=row, column=6, value=art["series"])
            ws3.cell(row=row, column=7, value=1)
            ws3.cell(row=row, column=8, value="YES" if art.get("box_available") else "NO")

            if not art.get("box_available"):
                for ci in range(1, total_cols3 + 1):
                    ws3.cell(row=row, column=ci).fill = FILL_LIGHT_RED
            elif num % 2 == 0:
                for ci in range(1, total_cols3 + 1):
                    ws3.cell(row=row, column=ci).fill = FILL_ZEBRA
            for ci in range(1, total_cols3 + 1):
                ws3.cell(row=row, column=ci).border = THIN_BORDER
            num += 1
            row += 1

        for ci in range(1, total_cols3 + 1):
            ws3.cell(row=row, column=ci).border = THIN_BORDER
            ws3.cell(row=row, column=ci).fill   = FILL_LIGHT_GRAY
        ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws3.cell(row=row, column=1, value="TOTAL BOXES").font = TOTAL_FONT
        ws3.cell(row=row, column=6, value=len(ro_box_list)).font = TOTAL_FONT
    else:
        ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols3)
        ws3.cell(row=row, column=1, value="Tidak ada RO Box minggu ini").font = Font(italic=True, color="888888")

    auto_width(ws3)

    # =========================================================================
    # SHEET 4: DAFTAR SURPLUS — Size-Level Pull List
    # =========================================================================
    ws4 = wb.create_sheet("Daftar Surplus")
    ws4.sheet_properties.tabColor = "B35900"

    headers4    = ["No", "Article (Kode Mix)", "Tier", "Size", "Pairs to Pull"]
    total_cols4 = len(headers4)

    row = 1
    row = write_form_header(
        ws4, row, "DAFTAR SURPLUS PULL",
        f"Destination: Warehouse Pusat Protol  |  "
        f"Total: {len(surplus_data['surplus_pull_list'])} articles, {int(total_surplus_pairs)} pairs",
        total_cols4,
    )
    for ci, h in enumerate(headers4, 1):
        ws4.cell(row=row, column=ci, value=h)
    style_header_row(ws4, row, total_cols4, HEADER_FILL_ORANGE)
    row += 1

    num = 1
    running_total = 0
    for art in surplus_data["surplus_pull_list"]:
        sizes = art.get("sizes", {})
        if not sizes:
            ws4.cell(row=row, column=1, value=num)
            ws4.cell(row=row, column=2, value=art["kode_mix"])
            ws4.cell(row=row, column=3, value=art["tier"])
            ws4.cell(row=row, column=4, value="(all)")
            ws4.cell(row=row, column=5, value=int(art["total_stock"]))
            running_total += int(art["total_stock"])
            if num % 2 == 0:
                for ci in range(1, total_cols4 + 1):
                    ws4.cell(row=row, column=ci).fill = FILL_ZEBRA
            for ci in range(1, total_cols4 + 1):
                ws4.cell(row=row, column=ci).border = THIN_BORDER
            num += 1
            row += 1
        else:
            for size_label in sorted(sizes.keys()):
                qty = sizes[size_label]
                if qty <= 0:
                    continue
                ws4.cell(row=row, column=1, value=num)
                ws4.cell(row=row, column=2, value=art["kode_mix"])
                ws4.cell(row=row, column=3, value=art["tier"])
                ws4.cell(row=row, column=4, value=size_label)
                ws4.cell(row=row, column=5, value=int(qty))
                running_total += int(qty)
                if num % 2 == 0:
                    for ci in range(1, total_cols4 + 1):
                        ws4.cell(row=row, column=ci).fill = FILL_ZEBRA
                for ci in range(1, total_cols4 + 1):
                    ws4.cell(row=row, column=ci).border = THIN_BORDER
                num += 1
                row += 1

    for ci in range(1, total_cols4 + 1):
        ws4.cell(row=row, column=ci).border = THIN_BORDER
        ws4.cell(row=row, column=ci).fill   = FILL_LIGHT_GRAY
    ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws4.cell(row=row, column=1, value="TOTAL PAIRS TO PULL").font = TOTAL_FONT
    ws4.cell(row=row, column=5, value=int(running_total)).font = TOTAL_FONT

    auto_width(ws4)

    # =========================================================================
    # SHEET 5: REFERENCE — Tier Analysis + Full Status + Off-Planogram
    # =========================================================================
    ws5 = wb.create_sheet("Reference")
    ws5.sheet_properties.tabColor = "4A1B6D"

    row = 1
    ws5.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
    ws5.cell(row=row, column=1, value="REFERENCE DATA (Internal Use)").font = Font(bold=True, size=14, color="4A1B6D")
    row += 2

    # --- Tier Capacity Analysis ---
    ws5.cell(row=row, column=1, value="TIER CAPACITY ANALYSIS").font = Font(bold=True, size=12)
    row += 1
    tier_headers = ["Tier", "Ideal (Planogram)", "Ideal %", "Actual (Stock)", "Actual %", "Diff", "Status"]
    for ci, h in enumerate(tier_headers, 1):
        ws5.cell(row=row, column=ci, value=h)
    style_header_row(ws5, row, len(tier_headers), HEADER_FILL_DARK)
    row += 1

    all_tiers = sorted(set(list(surplus_data["ideal_tier_counts"].keys()) +
                            list(surplus_data["actual_tier_counts"].keys())))
    for tier in all_tiers:
        ideal   = surplus_data["ideal_tier_counts"].get(tier, 0)
        actual  = surplus_data["actual_tier_counts"].get(tier, 0)
        ideal_p = surplus_data["ideal_tier_pct"].get(tier, 0)
        actual_p = surplus_data["actual_tier_pct"].get(tier, 0)
        diff    = actual - ideal
        status  = "OVER" if diff > 0 else ("UNDER" if diff < 0 else "OK")

        ws5.cell(row=row, column=1, value=f"Tier {tier}")
        ws5.cell(row=row, column=2, value=ideal)
        ws5.cell(row=row, column=3, value=f"{ideal_p * 100:.1f}%")
        ws5.cell(row=row, column=4, value=actual)
        ws5.cell(row=row, column=5, value=f"{actual_p * 100:.1f}%")
        ws5.cell(row=row, column=6, value=f"{'+' if diff > 0 else ''}{diff}")
        ws5.cell(row=row, column=7, value=status)

        fill = None
        if diff > 0 and tier in SURPLUS_CHECK_TIERS:
            fill = FILL_LIGHT_RED
        elif diff < 0:
            fill = FILL_LIGHT_YELLOW
        elif diff == 0:
            fill = FILL_LIGHT_GREEN
        if fill:
            for ci in range(1, 8):
                ws5.cell(row=row, column=ci).fill = fill
        for ci in range(1, 8):
            ws5.cell(row=row, column=ci).border = THIN_BORDER
        row += 1

    row += 2

    # --- Full Article Status ---
    ws5.cell(row=row, column=1, value="FULL ARTICLE STATUS").font = Font(bold=True, size=12)
    row += 1
    status_headers = [
        "Kode Mix", "Article", "Gender", "Series", "Tier",
        "Target Pairs", "Actual Pairs", "Gap",
        "% Kosong", "RO Type", "Avg Monthly Sales", "Stock Coverage (months)",
    ]
    for ci, h in enumerate(status_headers, 1):
        ws5.cell(row=row, column=ci, value=h)
    style_header_row(ws5, row, len(status_headers), HEADER_FILL_PURPLE)
    row += 1

    for art in sorted(gap_results, key=lambda x: (x["tier"], x["kode_mix"])):
        ws5.cell(row=row, column=1,  value=art["kode_mix"])
        ws5.cell(row=row, column=2,  value=art["article"])
        ws5.cell(row=row, column=3,  value=art["gender"])
        ws5.cell(row=row, column=4,  value=art["series"])
        ws5.cell(row=row, column=5,  value=art["tier"])
        ws5.cell(row=row, column=6,  value=art["total_target"])
        ws5.cell(row=row, column=7,  value=art["total_actual"])
        ws5.cell(row=row, column=8,  value=art["total_gap"])
        ws5.cell(row=row, column=9,  value=f"{art['pct_size_kosong'] * 100:.0f}%")
        ws5.cell(row=row, column=10, value=art["ro_type"])
        ws5.cell(row=row, column=11, value=round(art.get("avg_monthly_sales", 0), 1))
        cov = art.get("stock_coverage_months", 0)
        ws5.cell(row=row, column=12, value="dead stock" if cov == float("inf") else round(cov, 1))

        fill = (FILL_LIGHT_RED    if art["ro_type"] == "RO_BOX"
                else FILL_LIGHT_YELLOW if art["ro_type"] == "RO_PROTOL"
                else FILL_LIGHT_GREEN)
        for ci in range(1, len(status_headers) + 1):
            ws5.cell(row=row, column=ci).fill   = fill
            ws5.cell(row=row, column=ci).border = THIN_BORDER
        row += 1

    row += 2

    # --- Off-Planogram Stock ---
    ws5.cell(row=row, column=1, value="OFF-PLANOGRAM STOCK").font = Font(bold=True, size=12)
    row += 1
    off_headers = ["Kode Mix", "Article", "Gender", "Series", "Tier",
                   "Stock (pairs)", "Sizes", "Recommendation"]
    for ci, h in enumerate(off_headers, 1):
        ws5.cell(row=row, column=ci, value=h)
    style_header_row(ws5, row, len(off_headers), HEADER_FILL_TEAL)
    row += 1

    for art in sorted(surplus_data["off_planogram"], key=lambda x: (-x["total_stock"], x["kode_mix"])):
        sizes_str = ", ".join(
            f"{s}({int(q)})" for s, q in sorted(art["sizes"].items()) if q > 0
        )
        tier  = art["tier"]
        sales = art["avg_monthly_sales"]
        reco  = ("CLEARANCE"    if tier in (4, 5)
                 else "T8 PROTECTION" if tier == 8
                 else "DEAD STOCK - pull" if sales == 0
                 else "SURPLUS - pull")

        ws5.cell(row=row, column=1, value=art["kode_mix"])
        ws5.cell(row=row, column=2, value=art["article"])
        ws5.cell(row=row, column=3, value=art["gender"])
        ws5.cell(row=row, column=4, value=art["series"])
        ws5.cell(row=row, column=5, value=tier)
        ws5.cell(row=row, column=6, value=art["total_stock"])
        ws5.cell(row=row, column=7, value=sizes_str)
        ws5.cell(row=row, column=8, value=reco)
        for ci in range(1, len(off_headers) + 1):
            ws5.cell(row=row, column=ci).border = THIN_BORDER
        row += 1

    auto_width(ws5)

    # Save workbook
    wb.save(output_file)
    print(f"\n       Saved: {output_file}")
    print(f"       Sheets: {[s.title for s in wb.worksheets]}")


# =============================================================================
# 8. MAIN
# =============================================================================

def parse_args():
    parser = argparse.ArgumentParser(
        description="RO Request Generator — Universal (All Stores)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python build_ro_request.py --store "Icon Mall Gresik" --storage 0
  python build_ro_request.py --store "Royal Plaza" --storage 75
  python build_ro_request.py --store "Galaxy Mall" --storage 0
  python build_ro_request.py --store "Royal Plaza" --storage 75 --threshold 0.60
  python build_ro_request.py --store "Icon Mall Gresik" --output /tmp/test.xlsx
        """,
    )
    parser.add_argument(
        "--store", required=True,
        help="Store display name (e.g. 'Royal Plaza', 'Icon Mall Gresik', 'Galaxy Mall'). "
             "Used as ILIKE pattern on nama_gudang and planogram store_name.",
    )
    parser.add_argument(
        "--storage", type=int, default=0,
        help="Storage capacity in boxes (default: 0).",
    )
    parser.add_argument(
        "--output", default=None,
        help="Output .xlsx path. Default: ~/Desktop/DN PO ENTITAS/RO_Request_{store}_{date}.xlsx",
    )
    parser.add_argument(
        "--threshold", type=float, default=0.50,
        help="Fraction of sizes empty to trigger RO Box (default: 0.50 = 50%%)",
    )
    return parser.parse_args()


def main():
    args = parse_args()

    # Derive display store name (capitalise words nicely for the doc)
    store_display = args.store.strip()

    # Build output path
    if args.output:
        output_file = args.output
    else:
        today_str  = datetime.now().strftime("%Y-%m-%d")
        safe_name  = re.sub(r"[^\w\-]", "_", store_display)
        output_dir = os.path.expanduser("~/Desktop/DN PO ENTITAS")
        output_file = os.path.join(output_dir, f"RO_Request_{safe_name}_{today_str}.xlsx")

    print("=" * 70)
    print(f"  RO REQUEST GENERATOR — {store_display}")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"  Storage: {args.storage} boxes  |  Threshold: {args.threshold:.0%}")
    print(f"  Output:  {output_file}")
    print("=" * 70)
    print()

    # Connect to DB
    print("Connecting to database...")
    conn = psycopg2.connect(**DB_CONFIG)
    print("Connected.\n")

    # Step 1: Planogram from DB
    planogram = read_planogram(conn, store_display)
    if not planogram:
        print(f"\nERROR: No planogram data found for store '{store_display}'.")
        print("       Check store name spelling or run:")
        print(f"       SELECT DISTINCT store_name FROM {PLANOGRAM_TABLE};")
        conn.close()
        sys.exit(1)

    # Step 2: Stock, sales, warehouse from DB
    db_data = query_db(conn, store_display)
    conn.close()

    # Step 3: Gap analysis
    gap_results = calculate_gaps(planogram, db_data, args.threshold)

    # Step 4: RO decisions + warehouse availability
    ro_protol_list, ro_box_list, no_restock_list = generate_ro_decisions(
        gap_results, db_data, args.storage
    )

    # Step 5: TO / stock coverage
    calculate_to(gap_results, db_data)

    # Step 6: Surplus
    surplus_data = calculate_surplus(planogram, gap_results, db_data)

    # Step 7: Excel output
    write_excel(
        store_name=store_display,
        storage_capacity=args.storage,
        output_file=output_file,
        ro_protol_list=ro_protol_list,
        ro_box_list=ro_box_list,
        no_restock_list=no_restock_list,
        surplus_data=surplus_data,
        gap_results=gap_results,
        planogram=planogram,
        db_data=db_data,
    )

    # Summary
    total_protol_pairs = 0
    for art in ro_protol_list:
        for sd in art["sizes_detail"]:
            if sd["gap"] > 0:
                total_protol_pairs += sd.get("protol_fill", sd.get("gap", 0))

    total_surplus_pairs = sum(a["total_stock"] for a in surplus_data["surplus_pull_list"])

    print()
    print("=" * 70)
    print("  SUMMARY")
    print("=" * 70)
    print(f"  Store:                 {store_display}")
    print(f"  Storage:               {args.storage} boxes")
    print(f"  Planogram articles:    {len(planogram)}")
    print(f"  RO Protol requests:    {len(ro_protol_list)} articles, {int(total_protol_pairs)} pairs")
    print(f"  RO Box requests:       {len(ro_box_list)} articles, {len(ro_box_list)} boxes")
    print(f"  No restock needed:     {len(no_restock_list)}")
    print(f"  Surplus pull list:     {len(surplus_data['surplus_pull_list'])} articles, {int(total_surplus_pairs)} pairs")
    print(f"  Off-planogram stock:   {len(surplus_data['off_planogram'])}")
    print(f"  Output:                {output_file}")
    print("=" * 70)


if __name__ == "__main__":
    main()
