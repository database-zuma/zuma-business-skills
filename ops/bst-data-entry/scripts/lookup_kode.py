#!/usr/bin/env python3
"""
Fuzzy lookup kode artikel BST.

Usage:
  python lookup_kode.py <kode_tulisan_tangan> [warna_hint]
  python lookup_kode.py M1SPV2162 "Cocoa White Tan"
  python lookup_kode.py LIEAV2102 "Silver Navy"

Outputs matching kode + nama artikel, ranked by similarity.
"""
import json, sys, os, re
from difflib import SequenceMatcher

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
REF_PATH = os.path.join(SCRIPT_DIR, '..', 'references', 'kode_nama.json')

def load_kode_db(excel_path=None):
    """Load kode database. If excel_path given, also extract from file."""
    db = {}
    if os.path.exists(REF_PATH):
        with open(REF_PATH) as f:
            db = json.load(f)

    if excel_path:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(excel_path)
            ws = wb['DATA ENTRY']
            for row in range(3, ws.max_row + 1):
                kode = str(ws.cell(row=row, column=5).value or "").strip()
                fname = str(ws.cell(row=row, column=6).value or "")
                if 'COMPUTED_VALUE' in fname:
                    m = re.search(r',\"([^\"]+)\"\)', fname)
                    if m: fname = m.group(1)
                if kode and fname and kode not in db:
                    db[kode] = fname
        except Exception as e:
            print(f"Warning: Could not read Excel: {e}", file=sys.stderr)
    return db

def normalize_kode(raw):
    """Strip trailing extra digit and fix common OCR/handwriting swaps."""
    s = raw.upper().strip().replace(' ', '').replace('.', '')
    # Try removing trailing digit (common BST typo)
    variants = [s]
    if len(s) > 4:
        variants.append(s[:-1])
    # Common character swaps
    swap_map = {'I': '1', '1': 'I', 'O': '0', '0': 'O', 'Y': 'V', 'V': 'Y',
                'U': 'V', 'S': '5', '5': 'S', 'N': 'M', 'M': 'N', 'A': 'P', 'P': 'A'}
    return variants

def fuzzy_match(raw_kode, color_hint, db):
    """Find best matches for a handwritten kode."""
    raw = raw_kode.upper().strip().replace(' ', '').replace('.', '')
    results = []

    for kode, nama in db.items():
        # Exact match
        if raw == kode:
            results.append((kode, nama, 1.0, 'exact'))
            continue

        # Match without trailing digit
        if len(raw) > 4 and raw[:-1] == kode:
            results.append((kode, nama, 0.95, 'trailing_digit'))
            continue

        # Sequence similarity
        ratio = SequenceMatcher(None, raw, kode).ratio()
        if ratio > 0.6:
            results.append((kode, nama, ratio, 'fuzzy'))

        # Also try without trailing digit
        if len(raw) > 4:
            ratio2 = SequenceMatcher(None, raw[:-1], kode).ratio()
            if ratio2 > ratio and ratio2 > 0.6:
                results.append((kode, nama, ratio2, 'fuzzy_trimmed'))

    # Boost score if color hint matches
    if color_hint:
        hint = color_hint.upper()
        boosted = []
        for kode, nama, score, method in results:
            if hint in nama.upper():
                boosted.append((kode, nama, min(score + 0.2, 1.0), method + '+color'))
            else:
                boosted.append((kode, nama, score, method))
        results = boosted

    # Sort by score desc, deduplicate
    seen = set()
    unique = []
    for item in sorted(results, key=lambda x: -x[2]):
        if item[0] not in seen:
            seen.add(item[0])
            unique.append(item)
    return unique[:5]

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python lookup_kode.py <kode> [warna_hint] [excel_path]")
        sys.exit(1)

    raw_kode = sys.argv[1]
    color_hint = sys.argv[2] if len(sys.argv) > 2 else ""
    excel_path = sys.argv[3] if len(sys.argv) > 3 else None

    db = load_kode_db(excel_path)
    matches = fuzzy_match(raw_kode, color_hint, db)

    if matches:
        print(f"Matches for '{raw_kode}' (hint: '{color_hint}'):")
        for kode, nama, score, method in matches:
            print(f"  {score:.0%} [{method}] {kode} → {nama}")
    else:
        print(f"No matches found for '{raw_kode}'")
