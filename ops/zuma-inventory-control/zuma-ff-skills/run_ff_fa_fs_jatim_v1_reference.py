#!/usr/bin/env python3
"""
FF/FA/FS calculation for Jatim — 2026-02-11
Uses planogram from "Copy of FF FA FS ST Jatim Q1 2026.xlsx" (Planogram sheet)
Stock from VPS fact_stock_unified + portal.kodemix (fast path, ~2s)
"""

import pandas as pd
import psycopg2
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import os

# ==============================================================================
# CONFIG
# ==============================================================================

PLANOGRAM_FILE = r"D:\WAYAN\Work\0.KANTOR\ZUMA INDONESIA\0. DATA\0. N8N\00. CLAUDE CODE - 2\zuma-frontier-project\0. multi-moltbolt-project\0. zuma inventory control skills\0. zuma FF FA FS skills\Copy of FF FA FS ST Jatim Q1 2026.xlsx"
OUTPUT_FILE = os.path.join(os.path.dirname(__file__), "FF_FA_FS_Jatim_2026-02-11.xlsx")
TARGET_DATE = "2026-02-11"
REGION = "Jatim"

# SIZE_COLUMN_MAP: planogram column key → list of DB size values to sum
# Planogram has BOTH paired ("18/19") and individual (34, 35, ...) columns.
# DB also has both paired ("21/22") and individual ("34") size values.
# Strategy: for paired planogram cols, match the paired DB size AND individual parts.
# For individual planogram cols, match the individual DB size only.
# NOTE: planogram column keys can be str ("18/19") or int (34).
# We normalize them to str for consistent lookup.
SIZE_COLUMN_MAP = {
    # Paired planogram columns (str keys)
    "18/19": ["18/19", "18", "19"],
    "20/21": ["20/21", "20", "21"],
    "21/22": ["21/22", "21", "22"],
    "22/23": ["22/23", "22", "23"],
    "23/24": ["23/24", "23", "24"],
    "24/25": ["24/25", "24", "25"],
    "25/26": ["25/26", "25", "26"],
    "27/28": ["27/28", "27", "28"],
    "29/30": ["29/30", "29", "30"],
    "31/32": ["31/32", "31", "32"],
    "33/34": ["33/34", "33", "34"],
    "35/36": ["35/36", "35", "36"],
    "37/38": ["37/38", "37", "38"],
    "39/40": ["39/40", "39", "40"],
    "41/42": ["41/42", "41", "42"],
    "43/44": ["43/44", "43", "44"],
    "45/46": ["45/46", "45", "46"],
    # Individual planogram columns (originally int keys, normalized to str)
    "34": ["34"],
    "35": ["35"],
    "36": ["36"],
    "37": ["37"],
    "38": ["38"],
    "39": ["39"],
    "40": ["40"],
    "41": ["41"],
    "42": ["42"],
    "43": ["43"],
    "44": ["44"],
}
SIZE_COLS = list(SIZE_COLUMN_MAP.keys())

# STORE_NAME_MAP: planogram "Store Name" → DB fact_stock_unified "nama_gudang"
# These are the ACTUAL nama_gudang values from fact_stock_unified for Jatim area.
STORE_NAME_MAP = {
    "Zuma Matos": "Zuma MATOS",
    "Zuma Galaxy Mall": "ZUMA GALAXY MALL",
    "Zuma Tunjungan Plaza": "Zuma Tunjungan Plaza 3",
    "ZUMA PTC": "Zuma PTC",
    "Zuma Icon Gresik": "Zuma Icon Mall Gresik",
    "Zuma Lippo Sidoarjo": "Zuma Lippo Sidoarjo",
    "Zuma Lippo Batu": "Zuma Lippo Batu",
    "Zuma Royal Plaza": "Zuma Royal Plaza",
    "Zuma City Of Tomorrow Mall": "Zuma City Of Tomorrow Mall",
    "Zuma Sunrise Mall": "Zuma Sunrise Mall Mojokerto",
    "Zuma Mall Olympic Garden": "Zuma Mall Olympic Garden",
}

# Reverse map: DB nama_gudang (lowered) → planogram Store Name
DB_TO_PLANO = {}
for plano_name, db_name in STORE_NAME_MAP.items():
    DB_TO_PLANO[db_name.lower().strip()] = plano_name

DB_CONFIG = {
    "host": "76.13.194.120",
    "port": "5432",
    "dbname": "openclaw_ops",
    "user": "openclaw_app",
    "password": os.getenv("OPENCLAW_DB_PASS", ""),
}

# ==============================================================================
# LOAD PLANOGRAM
# ==============================================================================

print("=" * 60)
print("  FF / FA / FS — Jatim Report Generator")
print("=" * 60)

print(f"\nLoading planogram from: {os.path.basename(PLANOGRAM_FILE)}")
planogram_df = pd.read_excel(PLANOGRAM_FILE, sheet_name="Planogram", engine="openpyxl")

# Find article_mix column
for col in planogram_df.columns:
    if str(col).lower().strip() in [
        "article mix",
        "article_mix",
        "kode_mix",
        "kodemix",
    ]:
        planogram_df = planogram_df.rename(columns={col: "article_mix"})
        break

# Normalize article_mix: strip whitespace, lowercase for matching
planogram_df["article_mix"] = (
    planogram_df["article_mix"].astype(str).str.strip().str.lower()
)

# Normalize planogram column names: int columns (34, 35, ...) → str ("34", "35", ...)
planogram_df.columns = [str(c) for c in planogram_df.columns]

# Ensure size columns are numeric (plan quantities can be floats like 2.4)
# We round to nearest int for plan qty (planogram is pairs-based capacity)
for sc in SIZE_COLS:
    if sc in planogram_df.columns:
        planogram_df[sc] = (
            pd.to_numeric(planogram_df[sc], errors="coerce")
            .fillna(0)
            .round()
            .astype(int)
        )

# Ensure size columns are int
for sc in SIZE_COLS:
    if sc in planogram_df.columns:
        planogram_df[sc] = (
            pd.to_numeric(planogram_df[sc], errors="coerce").fillna(0).astype(int)
        )

stores = sorted(planogram_df["Store Name"].unique())
print(f"Planogram: {len(planogram_df):,} rows, {len(stores)} stores")
for s in stores:
    cnt = len(planogram_df[planogram_df["Store Name"] == s])
    print(f"  {s}: {cnt} articles")

# ==============================================================================
# FETCH STOCK FROM VPS — FAST PATH (fact_stock_unified + portal.kodemix)
# ==============================================================================

print(f"\nConnecting to VPS PostgreSQL...")
conn = psycopg2.connect(**DB_CONFIG)

# Build the list of DB store names we care about (exclude warehouses)
store_db_names = [STORE_NAME_MAP.get(s, s) for s in stores]
placeholders = ",".join(["%s"] * len(store_db_names))

# Fast query: fact_stock_unified joined with portal.kodemix for kode_mix + size
# This avoids the slow core.stock_with_product view (which has heavy multi-table JOINs)
query = f"""
    SELECT
        s.nama_gudang,
        km.kode_mix,
        km.size,
        s.quantity
    FROM fact_stock_unified s
    LEFT JOIN (
        SELECT DISTINCT ON (trim(lower(kode_besar)))
            trim(lower(kode_besar)) AS kode_besar,
            kode_mix,
            size
        FROM portal.kodemix
        ORDER BY trim(lower(kode_besar)), no_urut
    ) km ON s.matched_kode_besar = km.kode_besar
    LEFT JOIN portal.stock_capacity sc
        ON trim(lower(s.nama_gudang)) = trim(lower(sc.stock_location))
    WHERE sc.area = %s
      AND s.snapshot_date = %s
      AND s.quantity IS NOT NULL
      AND s.nama_gudang IN ({placeholders})
"""
params = tuple([REGION, TARGET_DATE] + store_db_names)

print(f"Fetching stock (fast path: fact_stock_unified)...")
cur = conn.cursor()
cur.execute(query, params)
rows = cur.fetchall()
stock_df = pd.DataFrame(rows, columns=["nama_gudang", "kode_mix", "size", "quantity"])
cur.close()
conn.close()

# Normalize kode_mix for matching
stock_df["kode_mix"] = stock_df["kode_mix"].astype(str).str.strip().str.lower()
stock_df["size"] = stock_df["size"].astype(str).str.strip()
stock_df["quantity"] = pd.to_numeric(stock_df["quantity"], errors="coerce").fillna(0)

# Map DB nama_gudang back to planogram Store Name
stock_df["store_label"] = (
    stock_df["nama_gudang"].str.lower().str.strip().map(DB_TO_PLANO)
)

print(f"Stock fetched: {len(stock_df):,} rows for {TARGET_DATE}")
for db_name in sorted(stock_df["nama_gudang"].unique()):
    cnt = len(stock_df[stock_df["nama_gudang"] == db_name])
    label = DB_TO_PLANO.get(db_name.lower().strip(), "???")
    print(f"  {db_name} -> {label}: {cnt} SKUs")

# Check for unmapped stores
unmapped = stock_df[stock_df["store_label"].isna()]["nama_gudang"].unique()
if len(unmapped) > 0:
    print(f"\nWARNING: {len(unmapped)} DB stores not mapped to planogram:")
    for u in unmapped:
        print(f"    {u}")

# ==============================================================================
# BUILD HASIL GABUNGAN
# ==============================================================================

print(f"\nBuilding Hasil Gabungan...")

# Pre-index stock by (store_label, kode_mix, size) for fast lookup
stock_grouped = (
    stock_df.groupby(["store_label", "kode_mix", "size"])["quantity"]
    .sum()
    .reset_index()
)
stock_lookup = {}
for _, srow in stock_grouped.iterrows():
    key = (srow["store_label"], srow["kode_mix"], srow["size"])
    stock_lookup[key] = srow["quantity"]

results = []

for _, plan_row in planogram_df.iterrows():
    store_label = plan_row["Store Name"]
    article_mix = plan_row["article_mix"]  # already lowered + stripped

    row = {
        "Store_Name": store_label,
        "Article_Mix": article_mix,
        "Gender": plan_row.get("Gender", ""),
        "Series": plan_row.get("Series", ""),
        "Article": plan_row.get("Article", ""),
        "Tier": plan_row.get("Tier", ""),
    }

    total_plan = 0
    total_stock = 0
    count_plan_pos = 0
    count_stock_pos = 0

    for size_col in SIZE_COLS:
        plan_qty = int(plan_row.get(size_col, 0) or 0)

        # Sum stock across all possible DB size representations for this planogram column
        stock_qty = 0
        for db_size in SIZE_COLUMN_MAP[size_col]:
            key = (store_label, article_mix, db_size)
            qty = stock_lookup.get(key, 0)
            if qty > 0:
                stock_qty += qty

        row[f"{size_col}_Plan"] = plan_qty
        row[f"{size_col}_Stock"] = stock_qty

        total_plan += plan_qty
        total_stock += stock_qty

        if plan_qty > 0:
            count_plan_pos += 1
        if plan_qty > 0 and stock_qty > 0:
            count_stock_pos += 1

    row["Total_Plan"] = total_plan
    row["Total_Stock"] = total_stock
    row["Count_Plan_Positive"] = count_plan_pos
    row["Count_Stock_Positive"] = count_stock_pos
    row["Count_Article_Plan"] = 1 if total_plan > 0 else 0
    # FA: only count article as "stocked" if it's also in the planogram
    row["Count_Article_Stock"] = 1 if (total_plan > 0 and total_stock > 0) else 0

    results.append(row)

hasil_df = pd.DataFrame(results)
print(f"Hasil Gabungan: {len(hasil_df):,} rows")

# Quick debug: how many articles matched at least 1 size?
matched = hasil_df[hasil_df["Total_Stock"] > 0]
print(
    f"Articles with stock match: {len(matched):,} / {len(hasil_df):,} ({len(matched) / len(hasil_df) * 100:.1f}%)"
)

# ==============================================================================
# CALCULATE METRICS
# ==============================================================================

print(f"\n{'=' * 60}")
print(f"  RESULTS — {TARGET_DATE}")
print(f"{'=' * 60}")
print(f"{'Store':<35s} {'FF':>7s} {'FA':>7s} {'FS':>7s}")
print(f"{'-' * 35} {'-' * 7} {'-' * 7} {'-' * 7}")

report_rows = []
for store in stores:
    s = hasil_df[hasil_df["Store_Name"] == store]

    ff_num = s["Count_Stock_Positive"].sum()
    ff_den = s["Count_Plan_Positive"].sum()
    ff = round(ff_num / ff_den * 100, 2) if ff_den > 0 else 0.0

    fa_num = s["Count_Article_Stock"].sum()
    fa_den = s["Count_Article_Plan"].sum()
    fa = round(fa_num / fa_den * 100, 2) if fa_den > 0 else 0.0

    fs_num = s["Total_Stock"].sum()
    fs_den = s["Total_Plan"].sum()
    fs = round(fs_num / fs_den * 100, 2) if fs_den > 0 else 0.0

    print(f"{store:<35s} {ff:>6.1f}% {fa:>6.1f}% {fs:>6.1f}%")
    report_rows.append({"store": store, "FF": ff, "FA": fa, "FS": fs})

# Region totals
all_data = hasil_df
ff_total_num = all_data["Count_Stock_Positive"].sum()
ff_total_den = all_data["Count_Plan_Positive"].sum()
ff_total = round(ff_total_num / ff_total_den * 100, 2) if ff_total_den > 0 else 0.0

fa_total_num = all_data["Count_Article_Stock"].sum()
fa_total_den = all_data["Count_Article_Plan"].sum()
fa_total = round(fa_total_num / fa_total_den * 100, 2) if fa_total_den > 0 else 0.0

fs_total_num = all_data["Total_Stock"].sum()
fs_total_den = all_data["Total_Plan"].sum()
fs_total = round(fs_total_num / fs_total_den * 100, 2) if fs_total_den > 0 else 0.0

print(f"{'-' * 35} {'-' * 7} {'-' * 7} {'-' * 7}")
print(f"{'TOTAL ' + REGION:<35s} {ff_total:>6.1f}% {fa_total:>6.1f}% {fs_total:>6.1f}%")

# ==============================================================================
# GENERATE EXCEL
# ==============================================================================

print(f"\nGenerating Excel report...")

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Report"

header_font = Font(bold=True, size=11)
metric_font = Font(bold=True, size=12, color="FFFFFF")
value_align = Alignment(horizontal="center")

METRIC_COLORS = {
    "FF": "2E7D32",
    "FA": "1565C0",
    "FS": "E65100",
}


def value_fill(val, metric):
    if metric == "FS":
        if val >= 100:
            return PatternFill(start_color="C8E6C9", fill_type="solid")
        elif val >= 80:
            return PatternFill(start_color="FFF9C4", fill_type="solid")
        else:
            return PatternFill(start_color="FFCDD2", fill_type="solid")
    else:
        if val >= 90:
            return PatternFill(start_color="C8E6C9", fill_type="solid")
        elif val >= 70:
            return PatternFill(start_color="FFF9C4", fill_type="solid")
        else:
            return PatternFill(start_color="FFCDD2", fill_type="solid")


current_row = 1
dates = [TARGET_DATE]

for metric in ["FF", "FA", "FS"]:
    color = METRIC_COLORS[metric]
    banner = PatternFill(start_color=color, end_color=color, fill_type="solid")

    # Date header
    ws.cell(row=current_row, column=1, value="Store").font = header_font
    for ci, d in enumerate(dates, start=2):
        c = ws.cell(row=current_row, column=ci, value=d)
        c.font = header_font
        c.alignment = value_align
    current_row += 1

    # Metric banner
    c = ws.cell(row=current_row, column=1, value=metric)
    c.font = metric_font
    c.fill = banner
    for ci in range(2, len(dates) + 2):
        ws.cell(row=current_row, column=ci).fill = banner
    current_row += 1

    # Store rows
    for rr in report_rows:
        ws.cell(row=current_row, column=1, value=rr["store"]).font = Font(size=10)
        val = rr[metric]
        c = ws.cell(row=current_row, column=2, value=round(val, 1))
        c.number_format = "0.0"
        c.alignment = value_align
        if val > 0:
            c.fill = value_fill(val, metric)
        current_row += 1

    # Region total
    total_val = {"FF": ff_total, "FA": fa_total, "FS": fs_total}[metric]
    ws.cell(row=current_row, column=1, value=f"TOTAL {REGION}").font = Font(
        bold=True, size=10
    )
    c = ws.cell(row=current_row, column=2, value=round(total_val, 1))
    c.number_format = "0.0"
    c.alignment = value_align
    c.font = Font(bold=True, size=10)
    if total_val > 0:
        c.fill = value_fill(total_val, metric)
    current_row += 2

ws.column_dimensions["A"].width = 35
ws.column_dimensions["B"].width = 14

# Add Hasil Gabungan debug sheet
ws_hg = wb.create_sheet(title="Hasil Gabungan")
for ci, col_name in enumerate(hasil_df.columns, start=1):
    ws_hg.cell(row=1, column=ci, value=col_name).font = Font(bold=True)
for ri, (_, hg_row) in enumerate(hasil_df.iterrows(), start=2):
    for ci, val in enumerate(hg_row, start=1):
        ws_hg.cell(row=ri, column=ci, value=val)

wb.save(OUTPUT_FILE)
print(f"\nReport saved: {OUTPUT_FILE}")
print("Done!")
