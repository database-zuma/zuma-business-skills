"""
Step 2: Planogram Visualization — Zuma Tunjungan Plaza
Reads PLANOGRAM_Tunjungan_Plaza.xlsx (Step 1 output) and generates:
  Option A: Excel floor plan (VISUAL_PLANOGRAM_Tunjungan_Plaza.xlsx)
  Option B: ASCII floor plan  (VISUAL_PLANOGRAM_Tunjungan_Plaza.txt)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import re
from datetime import datetime

# ============================================================
# CONFIG
# ============================================================
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_FILE = os.path.join(SCRIPT_DIR, "..", "PLANOGRAM_Tunjungan_Plaza.xlsx")
EXCEL_OUTPUT = os.path.join(SCRIPT_DIR, "VISUAL_PLANOGRAM_Tunjungan_Plaza.xlsx")
ASCII_OUTPUT = os.path.join(SCRIPT_DIR, "VISUAL_PLANOGRAM_Tunjungan_Plaza.txt")

STORE_NAME = "Zuma Tunjungan Plaza"
DATE_STR = datetime.now().strftime("%d %b %Y")

# Series -> hex color (without #, for openpyxl)
SERIES_COLORS = {
    "ZORRO": "EAD1DC",
    "DALLAS": "E6B8AF",
    "SLIDE": "FFF2CC",
    "SLIDE BASIC": "FFF2CC",
    "SLIDE MAX": "FFF2CC",
    "SLIDE PUFFY": "FFF2CC",
    "FLO": "F9CB9C",
    "PUFFY": "A4C2F4",
    "ELSA": "CCCCCC",
    "CLASSIC": "FFF2CC",
    "CLASSIC METALIC": "FFD966",
    "BLACKSERIES": "D9D9D9",
    "BLACK SERIES": "D9D9D9",
    "STRIPE": "D9D2E9",
    "ONYX": "FCE5CD",
    "ONYX Z": "FCE5CD",
    "MERCI": "D9D2E9",
    "STITCH": "B4E7CE",
    "WEDGES": "D7CCC8",
    "LUNA": "D5A6BD",
    "LUCA": "D9EAD3",
    "AIRMOVE": "CFE2F3",
    "WBB": "B7E1CD",
    "POOH": "FFFFB3",
    "VELCRO": "E06666",
    "VELCRO BRICKS": "E06666",
    "VELCRO STITCH": "B4E7CE",
    "LOTSO": "F48FB1",
    "BOYS COLLAB": "CE93D8",
    "GIRLS COLLAB": "F8BBD0",
    "DISNEY": "C9DAF8",
    "DISNEY MINNIE": "C9DAF8",
    "DISNEY MICKEY": "C9DAF8",
    "BATMAN": "90A4AE",
    "MICKEY": "F4CCCC",
    "MICKEY & FRIENDS": "F4CCCC",
    "SPIDER-MAN": "FF7043",
    "MILTON": "B39DDB",
    "OXFORD": "80CBC4",
    "DOVER": "A1887F",
    "BABY CLASSIC": "80DEEA",
    "BABY KARAKTER": "B7E1CD",
    "BABY VELCRO": "E06666",
    "KIDS CLASSIC": "B6D7A8",
    "KIDS KARAKTER": "B6D7A8",
    "TOY STORY": "81D4FA",
    "(KOSONG)": "F5F5F5",
}

# Series abbreviation for ASCII (max 3 chars)
SERIES_ABBREV = {
    "CLASSIC": "CLS",
    "CLASSIC METALIC": "CLM",
    "BLACKSERIES": "BLK",
    "BLACK SERIES": "BLK",
    "STRIPE": "STP",
    "ZORRO": "ZRO",
    "DALLAS": "DAL",
    "SLIDE": "SLD",
    "SLIDE BASIC": "SLB",
    "SLIDE MAX": "SLM",
    "SLIDE PUFFY": "SLP",
    "FLO": "FLO",
    "PUFFY": "PFY",
    "ELSA": "ELS",
    "MERCI": "MRC",
    "STITCH": "STH",
    "WEDGES": "WDG",
    "ONYX": "ONX",
    "ONYX Z": "ONX",
    "WBB": "WBB",
    "VELCRO": "VLC",
    "VELCRO BRICKS": "VLB",
    "VELCRO STITCH": "VLS",
    "POOH": "POO",
    "LOTSO": "LOT",
    "BOYS COLLAB": "BCL",
    "GIRLS COLLAB": "GCL",
    "DISNEY": "DSN",
    "DISNEY MINNIE": "DMN",
    "DISNEY MICKEY": "DMK",
    "BATMAN": "BAT",
    "MICKEY": "MCK",
    "MICKEY & FRIENDS": "M&F",
    "LUNA": "LNA",
    "LUCA": "LCA",
    "AIRMOVE": "AIR",
    "SPIDER-MAN": "SPD",
    "MILTON": "MLT",
    "OXFORD": "OXF",
    "DOVER": "DVR",
    "TOY STORY": "TOY",
    "(KOSONG)": "---",
    "AVAILABLE": "---",
}


def get_series_color_hex(series_name):
    """Get color hex (no #) for a series. Fallback to auto-generated pastel."""
    s = series_name.upper().strip()
    if s in SERIES_COLORS:
        return SERIES_COLORS[s]
    # Try partial match
    for key, color in SERIES_COLORS.items():
        if key in s or s in key:
            return color
    # Auto-generate pastel from hash
    import hashlib

    h = int(hashlib.md5(s.encode()).hexdigest()[:6], 16)
    r = 180 + (h % 75)
    g = 180 + ((h >> 8) % 75)
    b = 180 + ((h >> 16) % 75)
    return f"{r:02x}{g:02x}{b:02x}"


def get_series_abbrev(series_name):
    """Get 3-char abbreviation for ASCII view."""
    s = series_name.upper().strip()
    if s in SERIES_ABBREV:
        return SERIES_ABBREV[s]
    # Try partial match
    for key, abbr in SERIES_ABBREV.items():
        if key in s or s in key:
            return abbr
    # Fallback: first 3 chars
    return s[:3] if len(s) >= 3 else s


# ============================================================
# PARSE PLANOGRAM XLSX
# ============================================================
def parse_planogram_xlsx(filepath):
    """
    Parse PLANOGRAM_Tunjungan_Plaza.xlsx -> dict with backwall data + table/shelving.
    Returns:
      {
        "backwalls": {
          "BW-1": {"gender_type": ..., "hooks": ..., "hpa": ..., "slots": ...,
                    "articles": [{"name": ..., "series": ..., "tier": ..., "avg": ...}, ...]},
          ...
        },
        "tables": {
          "SHELVING-AIRMOVE": {"articles": [...]},
          "SHELVING-PUFFY": {"articles": [...]},
          "TABLE-LUCALUNA": {"articles": [...]},
          "TABLE-BABY": {"articles": [...]},
        },
        "summary": {"total_slots": ..., "total_used": ..., "util_pct": ...}
      }
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    data = {"backwalls": {}, "tables": {}, "summary": {}}

    # BW sheet name mapping to our IDs
    bw_sheet_map = {
        "BW-1 Men Fashion": "BW-1",
        "BW-2 Ladies Fashion": "BW-2",
        "BW-3 Men Jepit": "BW-3",
        "BW-4a-Elsa Ladies Elsa": "BW-4a-Elsa",
        "BW-4a-Classic Ladies Classic": "BW-4a-Classic",
        "BW-4b Baby and Kids": "BW-4b",
    }

    for sheet_name, bw_id in bw_sheet_map.items():
        if sheet_name in wb.sheetnames:
            bw_data = parse_backwall_sheet(wb[sheet_name], bw_id, sheet_name)
            data["backwalls"][bw_id] = bw_data

    # Table & Shelving
    if "Table & Shelving" in wb.sheetnames:
        data["tables"] = parse_table_shelving_sheet(wb["Table & Shelving"])

    # Summary from the summary sheet
    ws_sum = wb["Summary Report"] if "Summary Report" in wb.sheetnames else None
    if ws_sum:
        total_slots = 0
        total_used = 0
        util_pct = 0
        for r in range(3, ws_sum.max_row + 1):
            label = str(ws_sum.cell(row=r, column=1).value or "")
            val = str(ws_sum.cell(row=r, column=4).value or "").strip()
            if "Total slots available" in label:
                total_slots = int(val) if val.isdigit() else 0
            elif "Total slots used" in label:
                total_used = int(val) if val.isdigit() else 0
            elif label.strip() == "Utilization":
                util_match = re.search(r"([\d.]+)%", val)
                if util_match:
                    util_pct = float(util_match.group(1))
        if total_slots > 0 and util_pct == 0:
            util_pct = round(total_used / total_slots * 100, 1)
        data["summary"] = {
            "total_slots": total_slots,
            "total_used": total_used,
            "util_pct": util_pct,
        }
    else:
        # Compute from parsed data
        total_slots = sum(bw["slots"] for bw in data["backwalls"].values())
        total_slots += sum(len(t["articles"]) for t in data["tables"].values())
        total_used = sum(
            len([a for a in bw["articles"] if a["name"] != "AVAILABLE"])
            for bw in data["backwalls"].values()
        ) + sum(len(t["articles"]) for t in data["tables"].values())
        data["summary"] = {
            "total_slots": total_slots,
            "total_used": total_used,
            "util_pct": round(total_used / total_slots * 100, 1)
            if total_slots > 0
            else 0,
        }

    wb.close()
    return data


def parse_backwall_sheet(ws, bw_id, sheet_name):
    """Parse a single BW-N sheet."""
    # Extract gender_type from sheet name
    parts = sheet_name.split(" ", 1)
    gender_type = parts[1] if len(parts) > 1 else sheet_name

    # Row 1: title — extract hooks count and dimension
    title = str(ws.cell(row=1, column=1).value or "")
    hooks_match = re.search(r"(\d+)\s*Hooks", title)
    hooks = int(hooks_match.group(1)) if hooks_match else 0
    dim_match = re.search(r"\((\d+x\d+)\)", title)
    dimension = dim_match.group(1) if dim_match else ""

    # Row 2: mode description — extract hooks_per_article
    mode_desc = str(ws.cell(row=2, column=1).value or "")
    hpa_match = re.search(r"(\d+)\s*hook\s*=\s*1\s*artikel", mode_desc)
    hpa = int(hpa_match.group(1)) if hpa_match else 2

    # Extract capacity from mode desc
    cap_match = re.search(r"Capacity:\s*(\d+)\s*artikel", mode_desc)
    slots = int(cap_match.group(1)) if cap_match else (hooks // hpa if hpa > 0 else 0)

    # Parse articles by stepping through columns with stride = hpa
    articles = []
    col = 2  # articles start at column 2
    slot_idx = 0
    while slot_idx < slots:
        series = str(ws.cell(row=5, column=col).value or "").strip()
        article = str(ws.cell(row=6, column=col).value or "").strip()
        tier_str = str(ws.cell(row=7, column=col).value or "").strip()

        # Parse tier: "T1 | Avg: 6.3" -> tier=1, avg=6.3
        tier = ""
        avg = 0.0
        tier_match = re.match(r"T(\d+)", tier_str)
        if tier_match:
            tier = tier_match.group(1)
        avg_match = re.search(r"Avg:\s*([\d.]+)", tier_str)
        if avg_match:
            avg = float(avg_match.group(1))

        if article and article != "None" and article.upper() != "AVAILABLE":
            articles.append(
                {
                    "name": article,
                    "series": series if series and series != "(Kosong)" else "UNKNOWN",
                    "tier": tier,
                    "avg": avg,
                }
            )
        elif article.upper() == "AVAILABLE":
            articles.append(
                {
                    "name": "AVAILABLE",
                    "series": "(KOSONG)",
                    "tier": "",
                    "avg": 0,
                }
            )

        col += hpa
        slot_idx += 1

    return {
        "gender_type": gender_type,
        "hooks": hooks,
        "hpa": hpa,
        "slots": slots,
        "dimension": dimension,
        "articles": articles,
    }


def parse_table_shelving_sheet(ws):
    """Parse Table & Shelving sheet into sections."""
    tables = {}
    current_section = None
    current_articles = []

    for r in range(1, ws.max_row + 1):
        c1 = str(ws.cell(row=r, column=1).value or "").strip()

        # Detect section headers
        if c1.startswith("SHELVING-") or c1.startswith("TABLE-"):
            # Save previous section
            if current_section:
                tables[current_section] = {"articles": current_articles}
            current_section = c1
            current_articles = []
            continue

        # Detect data rows (Slot N — but skip header row "Slot")
        if re.match(r"Slot\s*\d+", c1) and current_section:
            article = str(ws.cell(row=r, column=2).value or "").strip()
            series = str(ws.cell(row=r, column=3).value or "").strip()
            tier_raw = str(ws.cell(row=r, column=4).value or "").strip()
            avg_val = ws.cell(row=r, column=5).value

            tier = ""
            tier_match = re.match(r"T(\d+)", tier_raw)
            if tier_match:
                tier = tier_match.group(1)

            # Safely convert avg_val to float
            avg = 0.0
            if avg_val is not None:
                try:
                    avg = float(avg_val)
                except (ValueError, TypeError):
                    avg = 0.0

            if article:
                current_articles.append(
                    {
                        "name": article,
                        "series": series if series else "UNKNOWN",
                        "tier": tier,
                        "avg": avg,
                    }
                )

    # Save last section
    if current_section:
        tables[current_section] = {"articles": current_articles}

    return tables


# ============================================================
# TUNJUNGAN PLAZA LAYOUT CONFIG
# ============================================================
# Store layout (bird's eye, entrance at bottom):
#
# TOP WALL:
# BW-1 Men Fashion (8x7, 18 art, horizontal) | BW-2 Ladies Fashion (8x7, 18 art, horizontal)
#   + AIRMOVE SHELVING below BW-1              |   + PUFFY SHELVING below BW-2
#
# LEFT:                CENTER:                  RIGHT:
# KASIR               TABLE LUCA/LUNA (5 art)   BW-4a-Elsa (4 art, vertical)
# BW-3 Men Jepit      KURSI                     BW-4a-Classic (14 art, vertical)
#  (9x7, 26+5 art)    TABLE BABY (6 art)        BW-4b Baby&Kids (24 art, vertical)
#  vertical            [Depan]
#
#                    ENTRANCE

TUNJUNGAN_LAYOUT = {
    "store_name": "Zuma Tunjungan Plaza",
    "grid_width": 100,
    "grid_height": 80,
    "title_row": 1,
    "display_units": [
        # ── TOP WALL ──
        # BW-1 Men Fashion — top-left horizontal (18 articles x 2 cols = 36 cols)
        {
            "bw_id": "BW-1",
            "label": "BW-1 MEN FASHION",
            "dimension": "8x7",
            "orientation": "horizontal",
            "grid_col": 3,
            "grid_row": 7,
            "label_row": 4,
            "dim_row": 5,
        },
        # BW-2 Ladies Fashion — top-right horizontal (18 articles x 2 cols = 36 cols)
        {
            "bw_id": "BW-2",
            "label": "BW-2 LADIES FASHION",
            "dimension": "8x7",
            "orientation": "horizontal",
            "grid_col": 55,
            "grid_row": 7,
            "label_row": 4,
            "dim_row": 5,
        },
        # SHELVING-AIRMOVE — below BW-1
        {
            "bw_id": "SHELVING-AIRMOVE",
            "label": "SHELVING AIRMOVE",
            "type": "shelving",
            "grid_col": 3,
            "grid_row": 12,
            "label_row": 11,
        },
        # SHELVING-PUFFY — below BW-2
        {
            "bw_id": "SHELVING-PUFFY",
            "label": "SHELVING PUFFY",
            "type": "shelving",
            "grid_col": 55,
            "grid_row": 12,
            "label_row": 11,
        },
        # ── LEFT SIDE ──
        # KASIR
        {
            "bw_id": "KASIR",
            "label": "KASIR",
            "type": "landmark",
            "grid_col": 3,
            "grid_row": 17,
            "width": 8,
            "height": 3,
        },
        # BW-3 Men Jepit — left vertical (26 used + 5 empty = 31 slots)
        {
            "bw_id": "BW-3",
            "label": "BW-3 MEN JEPIT",
            "dimension": "9x7",
            "orientation": "vertical",
            "grid_col": 3,
            "grid_row": 25,
            "label_row": 21,
            "dim_row": 22,
        },
        # ── CENTER ──
        # TABLE LUCA/LUNA
        {
            "bw_id": "TABLE-LUCALUNA",
            "label": "TABLE LUCA/LUNA",
            "type": "table",
            "grid_col": 20,
            "grid_row": 25,
            "label_row": 24,
        },
        # KURSI (seating area)
        {
            "bw_id": "KURSI",
            "label": "KURSI",
            "type": "landmark",
            "grid_col": 42,
            "grid_row": 25,
            "width": 8,
            "height": 3,
        },
        # TABLE BABY
        {
            "bw_id": "TABLE-BABY",
            "label": "TABLE BABY",
            "type": "table",
            "grid_col": 42,
            "grid_row": 34,
            "label_row": 33,
        },
        # [Depan] label
        {
            "bw_id": "DEPAN",
            "label": "[Depan]",
            "type": "landmark",
            "grid_col": 25,
            "grid_row": 44,
            "width": 10,
            "height": 2,
        },
        # ── RIGHT SIDE ──
        # BW-4a-Elsa — right vertical (4 articles)
        {
            "bw_id": "BW-4a-Elsa",
            "label": "BW-4a LADIES ELSA",
            "dimension": "part of BW-4a",
            "orientation": "vertical",
            "grid_col": 88,
            "grid_row": 18,
            "label_row": 15,
            "dim_row": 16,
        },
        # BW-4a-Classic — right vertical (14 articles)
        {
            "bw_id": "BW-4a-Classic",
            "label": "BW-4a LADIES CLASSIC",
            "dimension": "part of BW-4a",
            "orientation": "vertical",
            "grid_col": 88,
            "grid_row": 26,
            "label_row": 23,
            "dim_row": 24,
        },
        # BW-4b Baby & Kids — right vertical (24 articles)
        {
            "bw_id": "BW-4b",
            "label": "BW-4b BABY & KIDS",
            "dimension": "7x7",
            "orientation": "vertical",
            "grid_col": 88,
            "grid_row": 44,
            "label_row": 41,
            "dim_row": 42,
        },
    ],
    "entrance": {
        "grid_col": 35,
        "grid_row": 72,
        "label": "ENTRANCE",
    },
}


# ============================================================
# OPTION A: EXCEL VISUAL
# ============================================================
def generate_excel_visual(planogram_data, layout, output_path):
    """Generate an Excel floor plan with colored grid cells showing article names."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Layout Visual"

    # Narrow columns for grid effect
    for c in range(1, layout["grid_width"] + 5):
        ws.column_dimensions[get_column_letter(c)].width = 8
    for r in range(1, layout["grid_height"] + 10):
        ws.row_dimensions[r].height = 20

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    thick_border = Border(
        left=Side(style="medium"),
        right=Side(style="medium"),
        top=Side(style="medium"),
        bottom=Side(style="medium"),
    )

    # Title
    title_col = 20
    ws.merge_cells(
        start_row=1, start_column=title_col, end_row=1, end_column=title_col + 25
    )
    title_cell = ws.cell(
        row=1, column=title_col, value=f"LAYOUT {layout['store_name'].upper()}"
    )
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center")

    # Date subtitle
    ws.merge_cells(
        start_row=2, start_column=title_col, end_row=2, end_column=title_col + 25
    )
    date_cell = ws.cell(
        row=2,
        column=title_col,
        value=f"Planogram Period: Feb 2025 - Jan 2026 | Generated: {DATE_STR}",
    )
    date_cell.font = Font(size=9, italic=True)
    date_cell.alignment = Alignment(horizontal="center")

    # Draw each display unit
    for unit in layout["display_units"]:
        unit_type = unit.get("type", "backwall")

        if unit_type == "landmark":
            _excel_draw_landmark(ws, unit, thin_border, thick_border)
        elif unit_type in ("shelving", "table"):
            table_id = unit["bw_id"]
            table_data = planogram_data["tables"].get(table_id)
            if table_data:
                _excel_draw_table(
                    ws, unit, table_data["articles"], thin_border, thick_border
                )
            else:
                _excel_draw_landmark(ws, unit, thin_border, thick_border)
        else:
            bw_id = unit["bw_id"]
            bw_data = planogram_data["backwalls"].get(bw_id)
            if bw_data:
                _excel_draw_backwall(ws, unit, bw_data, thin_border, thick_border)

    # Entrance
    ent = layout["entrance"]
    ws.merge_cells(
        start_row=ent["grid_row"],
        start_column=ent["grid_col"],
        end_row=ent["grid_row"] + 1,
        end_column=ent["grid_col"] + 18,
    )
    ent_cell = ws.cell(
        row=ent["grid_row"], column=ent["grid_col"], value="=== ENTRANCE ==="
    )
    ent_cell.font = Font(bold=True, size=14)
    ent_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Legend
    _excel_draw_legend(ws, planogram_data, layout)

    # Summary stats at bottom
    summary = planogram_data["summary"]
    sum_row = layout["grid_height"] + 4
    ws.cell(
        row=sum_row,
        column=2,
        value=f"Total: {summary['total_used']}/{summary['total_slots']} slots used ({summary['util_pct']}%)",
    )
    ws.cell(row=sum_row, column=2).font = Font(bold=True, size=10)

    wb.save(output_path)
    print(f"  Excel visual saved: {output_path}")


def _apply_fill_border(ws, row, col, fill, border):
    """Apply fill and border to a cell."""
    c = ws.cell(row=row, column=col)
    c.fill = fill
    c.border = border


def _excel_draw_backwall(ws, unit, bw_data, thin_border, thick_border):
    """Draw a single backwall on the Excel grid."""
    articles = bw_data["articles"]
    orientation = unit.get("orientation", "horizontal")
    start_col = unit["grid_col"]
    start_row = unit["grid_row"]
    label_row = unit.get("label_row", start_row - 2)
    dim_row = unit.get("dim_row", start_row - 1)
    dimension = unit.get("dimension", bw_data.get("dimension", ""))

    n = len(articles)
    if n == 0:
        return

    COLS_PER_ART = 2  # each article gets 2 merged columns

    if orientation == "horizontal":
        total_cols = n * COLS_PER_ART

        # Label row (backwall name)
        ws.merge_cells(
            start_row=label_row,
            start_column=start_col,
            end_row=label_row,
            end_column=start_col + total_cols - 1,
        )
        label_cell = ws.cell(row=label_row, column=start_col, value=unit["label"])
        label_cell.font = Font(bold=True, size=10)
        label_cell.alignment = Alignment(horizontal="center")

        # Dimension row
        ws.merge_cells(
            start_row=dim_row,
            start_column=start_col,
            end_row=dim_row,
            end_column=start_col + total_cols - 1,
        )
        dim_cell = ws.cell(
            row=dim_row,
            column=start_col,
            value=f"[ {dimension} ]  --  {n} articles  --  {bw_data['hooks']} hooks  --  {bw_data['hpa']} hooks/artikel",
        )
        dim_cell.font = Font(bold=True, size=9, color="333333")
        dim_cell.alignment = Alignment(horizontal="center")

        # 3 data rows per article: Series, Article Name, Tier & Avg
        for i, art in enumerate(articles):
            col = start_col + i * COLS_PER_ART
            series = art["series"]
            color_hex = get_series_color_hex(series)
            fill = PatternFill(
                start_color=color_hex, end_color=color_hex, fill_type="solid"
            )

            # Row 1: Series name
            ws.merge_cells(
                start_row=start_row,
                start_column=col,
                end_row=start_row,
                end_column=col + COLS_PER_ART - 1,
            )
            cell_s = ws.cell(row=start_row, column=col, value=series)
            cell_s.fill = fill
            cell_s.border = thin_border
            cell_s.alignment = Alignment(horizontal="center", vertical="center")
            cell_s.font = Font(size=8, italic=True)
            for mc in range(col, col + COLS_PER_ART):
                _apply_fill_border(ws, start_row, mc, fill, thin_border)

            # Row 2: Article name (FULL)
            ws.merge_cells(
                start_row=start_row + 1,
                start_column=col,
                end_row=start_row + 1,
                end_column=col + COLS_PER_ART - 1,
            )
            cell_a = ws.cell(row=start_row + 1, column=col, value=art["name"])
            cell_a.fill = fill
            cell_a.border = thin_border
            cell_a.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell_a.font = Font(size=8, bold=True)
            for mc in range(col, col + COLS_PER_ART):
                _apply_fill_border(ws, start_row + 1, mc, fill, thin_border)

            # Row 3: Tier & Avg
            tier_label = f"T{art['tier']} | Avg: {art['avg']}" if art["tier"] else ""
            ws.merge_cells(
                start_row=start_row + 2,
                start_column=col,
                end_row=start_row + 2,
                end_column=col + COLS_PER_ART - 1,
            )
            cell_t = ws.cell(row=start_row + 2, column=col, value=tier_label)
            cell_t.fill = fill
            cell_t.border = thin_border
            cell_t.alignment = Alignment(horizontal="center", vertical="center")
            cell_t.font = Font(size=7)
            for mc in range(col, col + COLS_PER_ART):
                _apply_fill_border(ws, start_row + 2, mc, fill, thin_border)

    elif orientation == "vertical":
        VCOLS = 4  # series, article(x2 merged), tier

        # Label row
        ws.merge_cells(
            start_row=label_row,
            start_column=start_col,
            end_row=label_row,
            end_column=start_col + VCOLS - 1,
        )
        label_cell = ws.cell(row=label_row, column=start_col, value=unit["label"])
        label_cell.font = Font(bold=True, size=9)
        label_cell.alignment = Alignment(horizontal="center")

        # Dimension row
        ws.merge_cells(
            start_row=dim_row,
            start_column=start_col,
            end_row=dim_row,
            end_column=start_col + VCOLS - 1,
        )
        used = len([a for a in articles if a["name"] != "AVAILABLE"])
        dim_cell = ws.cell(
            row=dim_row,
            column=start_col,
            value=f"[ {dimension} ] {used} articles",
        )
        dim_cell.font = Font(bold=True, size=8, color="333333")
        dim_cell.alignment = Alignment(horizontal="center")

        # Header row
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True, size=7)
        hr = start_row - 1
        for hi, hdr in enumerate(["Series", "Article", "", "Tier"]):
            hc = ws.cell(row=hr, column=start_col + hi, value=hdr)
            hc.fill = header_fill
            hc.font = header_font
            hc.border = thin_border
            hc.alignment = Alignment(horizontal="center")
        ws.merge_cells(
            start_row=hr,
            start_column=start_col + 1,
            end_row=hr,
            end_column=start_col + 2,
        )

        # Articles top-to-bottom
        for i, art in enumerate(articles):
            row = start_row + i
            series = art["series"]
            color_hex = get_series_color_hex(series)
            fill = PatternFill(
                start_color=color_hex, end_color=color_hex, fill_type="solid"
            )

            # Col 1: Series
            cell = ws.cell(row=row, column=start_col, value=series)
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(size=7)

            # Col 2-3 merged: Article name
            ws.merge_cells(
                start_row=row,
                start_column=start_col + 1,
                end_row=row,
                end_column=start_col + 2,
            )
            cell2 = ws.cell(row=row, column=start_col + 1, value=art["name"])
            cell2.fill = fill
            cell2.border = thin_border
            cell2.alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True
            )
            cell2.font = Font(size=8, bold=True)
            _apply_fill_border(ws, row, start_col + 2, fill, thin_border)

            # Col 4: Tier & Avg
            tier_label = f"T{art['tier']}|{art['avg']}" if art["tier"] else ""
            cell3 = ws.cell(row=row, column=start_col + 3, value=tier_label)
            cell3.fill = fill
            cell3.border = thin_border
            cell3.alignment = Alignment(horizontal="center", vertical="center")
            cell3.font = Font(size=7)


def _excel_draw_table(ws, unit, articles, thin_border, thick_border):
    """Draw a table/shelving section on Excel grid."""
    start_col = unit["grid_col"]
    start_row = unit["grid_row"]
    label_row = unit.get("label_row", start_row - 1)

    width = 8  # merged cols for readable text

    # Label
    ws.merge_cells(
        start_row=label_row,
        start_column=start_col,
        end_row=label_row,
        end_column=start_col + width - 1,
    )
    label_cell = ws.cell(row=label_row, column=start_col, value=unit["label"])
    label_cell.font = Font(bold=True, size=10)
    label_cell.alignment = Alignment(horizontal="center")

    for i, art in enumerate(articles):
        row = start_row + i
        series = art["series"]
        color_hex = get_series_color_hex(series)
        fill = PatternFill(
            start_color=color_hex, end_color=color_hex, fill_type="solid"
        )

        ws.merge_cells(
            start_row=row,
            start_column=start_col,
            end_row=row,
            end_column=start_col + width - 1,
        )
        tier_str = f"T{art['tier']}" if art["tier"] else ""
        cell = ws.cell(
            row=row,
            column=start_col,
            value=f"{art['name']} ({series}) {tier_str}|Avg:{art['avg']}",
        )
        cell.fill = fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(size=9, bold=True)
        for mc in range(start_col, start_col + width):
            _apply_fill_border(ws, row, mc, fill, thin_border)


def _excel_draw_landmark(ws, unit, thin_border, thick_border):
    """Draw a landmark (KASIR, KURSI, etc.) on Excel grid."""
    start_col = unit["grid_col"]
    start_row = unit["grid_row"]
    width = unit.get("width", 8)
    height = unit.get("height", 3)

    ws.merge_cells(
        start_row=start_row,
        start_column=start_col,
        end_row=start_row + height - 1,
        end_column=start_col + width - 1,
    )
    cell = ws.cell(row=start_row, column=start_col, value=unit["label"])
    cell.font = Font(bold=True, size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thick_border

    # Apply border to all cells in the merged area
    for r in range(start_row, start_row + height):
        for c in range(start_col, start_col + width):
            ws.cell(row=r, column=c).border = thick_border


def _excel_draw_legend(ws, planogram_data, layout):
    """Draw a legend section below the floor plan."""
    legend_row = layout["grid_height"] - 3
    legend_col = 2

    ws.cell(row=legend_row, column=legend_col, value="LEGEND (Series Colors):")
    ws.cell(row=legend_row, column=legend_col).font = Font(bold=True, size=10)

    # Collect unique series from all backwalls + tables
    all_series = set()
    for bw_data in planogram_data["backwalls"].values():
        for art in bw_data["articles"]:
            if art["series"] and art["series"] != "(KOSONG)":
                all_series.add(art["series"])
    for table_data in planogram_data["tables"].values():
        for art in table_data["articles"]:
            if art["series"]:
                all_series.add(art["series"])

    sorted_series = sorted(all_series)
    row_offset = 1
    items_per_row = 6

    for i, series in enumerate(sorted_series):
        r = legend_row + row_offset + (i // items_per_row)
        c = legend_col + (i % items_per_row) * 4

        color_hex = get_series_color_hex(series)
        fill = PatternFill(
            start_color=color_hex, end_color=color_hex, fill_type="solid"
        )

        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c + 1)
        cell = ws.cell(row=r, column=c, value=series)
        cell.fill = fill
        cell.font = Font(size=8, bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        _apply_fill_border(
            ws,
            r,
            c + 1,
            fill,
            Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            ),
        )


# ============================================================
# OPTION B: ASCII VISUAL
# ============================================================
def generate_ascii_visual(planogram_data, layout, output_path):
    """Generate an ASCII art floor plan."""
    lines = []
    W = 120  # total width

    # Header
    lines.append("")
    lines.append(center_text(f"LAYOUT {layout['store_name'].upper()}", W))
    lines.append(center_text("=" * 50, W))
    lines.append(
        center_text(f"Planogram Feb 2025 - Jan 2026 | Generated: {DATE_STR}", W)
    )
    lines.append("")

    # --- TOP WALL: BW-1 Men Fashion + BW-2 Ladies Fashion ---
    bw1 = planogram_data["backwalls"].get("BW-1")
    bw2 = planogram_data["backwalls"].get("BW-2")

    bw1_block = (
        _ascii_horizontal_block("BW-1 MEN FASHION (8x7)", bw1, 55) if bw1 else []
    )
    bw2_block = (
        _ascii_horizontal_block("BW-2 LADIES FASHION (8x7)", bw2, 55) if bw2 else []
    )

    # Side by side with gap
    max_lines_top = max(len(bw1_block), len(bw2_block))
    for i in range(max_lines_top):
        left = bw1_block[i] if i < len(bw1_block) else " " * 55
        right = bw2_block[i] if i < len(bw2_block) else " " * 55
        lines.append(f"  {left}  {right}")

    # Shelving below top wall
    shelv_airmove = planogram_data["tables"].get("SHELVING-AIRMOVE")
    shelv_puffy = planogram_data["tables"].get("SHELVING-PUFFY")

    if shelv_airmove or shelv_puffy:
        lines.append("")
        left_shelv = _ascii_shelving_line("AIRMOVE SHELVING", shelv_airmove, 55)
        right_shelv = _ascii_shelving_line("PUFFY SHELVING", shelv_puffy, 55)
        lines.append(f"  {left_shelv}  {right_shelv}")

    lines.append("")

    # --- MIDDLE: BW-3 (left) | Center (Tables, Kursi) | Right (BW-4a, BW-4b) ---
    bw3 = planogram_data["backwalls"].get("BW-3")
    bw4a_elsa = planogram_data["backwalls"].get("BW-4a-Elsa")
    bw4a_classic = planogram_data["backwalls"].get("BW-4a-Classic")
    bw4b = planogram_data["backwalls"].get("BW-4b")

    # Left column: KASIR + BW-3
    left_block = _ascii_left_column(bw3, 22)

    # Center column: TABLE LUCA/LUNA, KURSI, TABLE BABY
    table_lucaluna = planogram_data["tables"].get("TABLE-LUCALUNA")
    table_baby = planogram_data["tables"].get("TABLE-BABY")
    center_block = _ascii_center_block(table_lucaluna, table_baby, 50)

    # Right column: BW-4a-Elsa + BW-4a-Classic + BW-4b
    right_block = _ascii_right_column(bw4a_elsa, bw4a_classic, bw4b, 40)

    # Combine: left | center | right
    max_lines_mid = max(len(left_block), len(center_block), len(right_block))
    for i in range(max_lines_mid):
        left = left_block[i] if i < len(left_block) else " " * 22
        mid = center_block[i] if i < len(center_block) else " " * 50
        right = right_block[i] if i < len(right_block) else " " * 40
        lines.append(f"{left} {mid} {right}")

    lines.append("")

    # --- ENTRANCE ---
    lines.append(center_text("=== ENTRANCE ===", W))
    lines.append("")

    # --- LEGEND ---
    lines.append("  " + "-" * 100)
    lines.append("  LEGEND (Series Abbreviations):")
    lines.append("")

    # Collect all series used
    all_series = set()
    for bw_data in planogram_data["backwalls"].values():
        for art in bw_data["articles"]:
            if art["series"] and art["series"] != "(KOSONG)":
                all_series.add(art["series"])
    for table_data in planogram_data["tables"].values():
        for art in table_data["articles"]:
            if art["series"]:
                all_series.add(art["series"])

    sorted_series = sorted(all_series)
    legend_line = "  "
    for i, series in enumerate(sorted_series):
        abbr = get_series_abbrev(series)
        entry = f"  {abbr} = {series}"
        if len(legend_line) + len(entry) > W - 5:
            lines.append(legend_line)
            legend_line = "  "
        legend_line += entry
    if legend_line.strip():
        lines.append(legend_line)

    lines.append("")

    # --- SUMMARY ---
    lines.append("  " + "-" * 100)
    summary = planogram_data["summary"]
    lines.append(
        f"  SUMMARY: {summary['total_used']}/{summary['total_slots']} slots used "
        f"({summary['util_pct']}% utilization)"
    )
    lines.append("")

    bw_summaries = []
    for bw_id in ["BW-1", "BW-2", "BW-3", "BW-4a-Elsa", "BW-4a-Classic", "BW-4b"]:
        bw = planogram_data["backwalls"].get(bw_id)
        if bw:
            used = len([a for a in bw["articles"] if a["name"] != "AVAILABLE"])
            bw_summaries.append(f"{bw_id} {bw['gender_type']}: {used}/{bw['slots']}")
    lines.append(f"  Backwalls: {' | '.join(bw_summaries)}")

    table_summaries = []
    for tid, tdata in planogram_data["tables"].items():
        table_summaries.append(f"{tid}: {len(tdata['articles'])}")
    lines.append(f"  Tables/Shelving: {' | '.join(table_summaries)}")
    lines.append("  " + "-" * 100)
    lines.append("")

    output = "\n".join(lines)

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(output)

    print(f"  ASCII visual saved: {output_path}")
    return output


def center_text(text, width):
    """Center text within given width."""
    padding = max(0, (width - len(text)) // 2)
    return " " * padding + text


def _ascii_horizontal_block(title, bw_data, width):
    """
    Draw a horizontal backwall as ASCII block.
    Articles displayed left-to-right in rows of cells.
    """
    articles = bw_data["articles"]
    n = len(articles)
    if n == 0:
        return [f"  {title}", "  (empty)"]

    cell_w = 6  # chars per cell including border
    cols_per_row = min(n, (width - 2) // cell_w)
    if cols_per_row < 1:
        cols_per_row = 1

    lines = []
    lines.append(f"  {title}")

    # Process articles in chunks of cols_per_row
    for chunk_start in range(0, n, cols_per_row):
        chunk = articles[chunk_start : chunk_start + cols_per_row]
        actual_cols = len(chunk)

        # Top border
        top = "  +" + "+".join(["-----"] * actual_cols) + "+"
        lines.append(top)

        # Series row
        series_row = "  |"
        for art in chunk:
            abbr = get_series_abbrev(art["series"])
            series_row += f"{abbr:^5}|"
        lines.append(series_row)

        # Tier row
        tier_row = "  |"
        for art in chunk:
            t = f"T{art['tier']}" if art["tier"] else "   "
            tier_row += f"{t:^5}|"
        lines.append(tier_row)

        # Bottom border
        bot = "  +" + "+".join(["-----"] * actual_cols) + "+"
        lines.append(bot)

    return lines


def _ascii_shelving_line(title, shelving_data, width):
    """Draw a single-line shelving summary."""
    if not shelving_data:
        return " " * width
    arts = shelving_data["articles"]
    art_names = ", ".join(
        f"{get_series_abbrev(a['series'])}({a['name'].split()[-1]})" for a in arts
    )
    line = f"  [{title}: {art_names}]"
    return line.ljust(width)


def _ascii_left_column(bw3, width):
    """Build left column: KASIR + BW-3 Men Jepit (vertical)."""
    lines = []

    # KASIR
    lines.append(f"+{'-' * (width - 2)}+")
    lines.append(f"|{'KASIR':^{width - 2}}|")
    lines.append(f"+{'-' * (width - 2)}+")
    lines.append("")

    # BW-3 Men Jepit
    if bw3:
        articles = bw3["articles"]
        used = len([a for a in articles if a["name"] != "AVAILABLE"])
        title = f"BW-3 MEN JEPIT (9x7)"
        lines.append(f"+{'-' * (width - 2)}+")
        lines.append(f"|{title:^{width - 2}}|")
        lines.append(f"|{f'{used} articles':^{width - 2}}|")
        lines.append(f"+{'-' * (width - 2)}+")

        for art in articles:
            abbr = get_series_abbrev(art["series"])
            tier = f"T{art['tier']}" if art["tier"] else "  "
            if art["name"] == "AVAILABLE":
                label = "  (empty)   "
            else:
                label = f" {abbr} {tier}"
            lines.append(f"|{label:<{width - 2}}|")

        lines.append(f"+{'-' * (width - 2)}+")

    # Pad to ensure enough height
    while len(lines) < 45:
        lines.append(" " * width)

    return lines


def _ascii_center_block(table_lucaluna, table_baby, width):
    """Draw center area: TABLE LUCA/LUNA, KURSI, TABLE BABY."""
    lines = []

    # Pad top
    for _ in range(2):
        lines.append(" " * width)

    # TABLE LUCA/LUNA
    tbl_w = 38
    tbl_pad = 4
    if table_lucaluna:
        arts = table_lucaluna["articles"]
        lines.append((" " * tbl_pad + "+" + "-" * (tbl_w - 2) + "+").ljust(width))
        lines.append(
            (" " * tbl_pad + "|" + f"{'TABLE LUCA/LUNA':^{tbl_w - 2}}" + "|").ljust(
                width
            )
        )
        lines.append((" " * tbl_pad + "+" + "-" * (tbl_w - 2) + "+").ljust(width))
        for art in arts:
            abbr = get_series_abbrev(art["series"])
            tier = f"T{art['tier']}" if art["tier"] else ""
            label = f"{art['name'][:22]} ({abbr}) {tier}"
            lines.append(
                (" " * tbl_pad + "|" + f"{label:^{tbl_w - 2}}" + "|").ljust(width)
            )
        lines.append((" " * tbl_pad + "+" + "-" * (tbl_w - 2) + "+").ljust(width))

    lines.append(" " * width)

    # KURSI
    kursi_w = 20
    kursi_pad = 12
    lines.append((" " * kursi_pad + "+" + "-" * (kursi_w - 2) + "+").ljust(width))
    lines.append(
        (" " * kursi_pad + "|" + f"{'KURSI':^{kursi_w - 2}}" + "|").ljust(width)
    )
    lines.append((" " * kursi_pad + "+" + "-" * (kursi_w - 2) + "+").ljust(width))

    lines.append(" " * width)

    # TABLE BABY
    if table_baby:
        arts = table_baby["articles"]
        lines.append((" " * tbl_pad + "+" + "-" * (tbl_w - 2) + "+").ljust(width))
        lines.append(
            (" " * tbl_pad + "|" + f"{'TABLE BABY':^{tbl_w - 2}}" + "|").ljust(width)
        )
        lines.append((" " * tbl_pad + "+" + "-" * (tbl_w - 2) + "+").ljust(width))
        for art in arts:
            abbr = get_series_abbrev(art["series"])
            tier = f"T{art['tier']}" if art["tier"] else ""
            label = f"{art['name'][:22]} ({abbr}) {tier}"
            lines.append(
                (" " * tbl_pad + "|" + f"{label:^{tbl_w - 2}}" + "|").ljust(width)
            )
        lines.append((" " * tbl_pad + "+" + "-" * (tbl_w - 2) + "+").ljust(width))

    lines.append(" " * width)

    # [Depan] label
    lines.append(center_text("[Depan]", width).ljust(width))

    # Pad bottom
    while len(lines) < 45:
        lines.append(" " * width)

    return lines


def _ascii_right_column(bw4a_elsa, bw4a_classic, bw4b, width):
    """Build right column: BW-4a-Elsa + BW-4a-Classic + BW-4b (all vertical)."""
    lines = []
    inner_w = width - 4

    def _draw_vertical_bw(bw_data, title, lines_out):
        if not bw_data:
            return
        articles = bw_data["articles"]
        used = len([a for a in articles if a["name"] != "AVAILABLE"])
        lines_out.append(f"+{'-' * (width - 2)}+")
        lines_out.append(f"|{title:^{width - 2}}|")
        lines_out.append(f"|{f'{used} articles':^{width - 2}}|")
        lines_out.append(f"+{'-' * (width - 2)}+")

        for art in articles:
            abbr = get_series_abbrev(art["series"])
            tier = f"T{art['tier']}" if art["tier"] else "  "
            if art["name"] == "AVAILABLE":
                label = "  (empty)   "
            else:
                name_short = art["name"][: inner_w - 10]
                label = f" {abbr} {name_short:<{inner_w - 8}} {tier}"
            lines_out.append(f"|{label:<{width - 2}}|")

        lines_out.append(f"+{'-' * (width - 2)}+")

    # BW-4a-Elsa
    _draw_vertical_bw(bw4a_elsa, "BW-4a LADIES ELSA", lines)
    lines.append("")

    # BW-4a-Classic
    _draw_vertical_bw(bw4a_classic, "BW-4a LADIES CLASSIC", lines)
    lines.append("")

    # BW-4b Baby & Kids
    _draw_vertical_bw(bw4b, "BW-4b BABY & KIDS", lines)

    # Pad bottom
    while len(lines) < 45:
        lines.append(" " * width)

    return lines


# ============================================================
# MAIN
# ============================================================
if __name__ == "__main__":
    print("=" * 60)
    print(f"PLANOGRAM VISUALIZATION: {STORE_NAME}")
    print("=" * 60)

    print(f"\n[1/3] Parsing planogram XLSX: {INPUT_FILE}")
    if not os.path.exists(INPUT_FILE):
        print(f"ERROR: Input file not found: {INPUT_FILE}")
        exit(1)

    data = parse_planogram_xlsx(INPUT_FILE)

    # Print parsed summary
    for bw_id, bw in data["backwalls"].items():
        used = len([a for a in bw["articles"] if a["name"] != "AVAILABLE"])
        print(f"  {bw_id} ({bw['gender_type']}): {used}/{bw['slots']} articles")
    for tid, tdata in data["tables"].items():
        print(f"  {tid}: {len(tdata['articles'])} articles")
    print(
        f"  Summary: {data['summary']['total_used']}/{data['summary']['total_slots']} "
        f"({data['summary']['util_pct']}%)"
    )

    print(f"\n[2/3] Generating Excel visual...")
    generate_excel_visual(data, TUNJUNGAN_LAYOUT, EXCEL_OUTPUT)

    print(f"\n[3/3] Generating ASCII visual...")
    ascii_output = generate_ascii_visual(data, TUNJUNGAN_LAYOUT, ASCII_OUTPUT)

    print(f"\n{'=' * 60}")
    print("DONE!")
    print(f"  Excel: {EXCEL_OUTPUT}")
    print(f"  ASCII: {ASCII_OUTPUT}")
    print(f"{'=' * 60}")

    # Print the ASCII output for verification
    print("\n--- ASCII OUTPUT PREVIEW ---")
    print(ascii_output)
