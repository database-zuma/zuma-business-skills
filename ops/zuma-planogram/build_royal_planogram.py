"""
PLANOGRAM BUILDER — Zuma Royal Plaza
Based on SKILL_planogram_zuma_v2.md logic
"""

import psycopg2
import pandas as pd
import numpy as np
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
import math

# ============================================================
# CONFIG
# ============================================================
STORE_NAME = "Zuma Royal Plaza"
DB_STORE_NAME = "Zuma Royal Plaza"
STORAGE_CAPACITY = 0  # Royal has NO storage

# Royal Plaza layout from Jatim sheet
BACKWALLS = [
    {"id": "BW-1", "hooks": 48},
    {"id": "BW-2", "hooks": 56},
    {"id": "BW-3", "hooks": 56},
    {"id": "BW-4", "hooks": 52},
]
RAK_BABY = {"layers": 2, "pairs_per_layer": 6}  # 12 total pairs, 6 articles capacity
# No gondola, no table, no VM, no basket

TOTAL_HOOKS = sum(bw["hooks"] for bw in BACKWALLS)

# Months for analysis (12 months: Feb 2025 - Jan 2026)
DATE_START = "2025-02-01"
DATE_END = "2026-01-31"

OUTPUT_FILE = r"D:\WAYAN\Work\0.KANTOR\ZUMA INDONESIA\0. DATA\0. N8N\00. CLAUDE CODE - 2\zuma-frontier-project\0. multi-moltbolt-project\0. zuma plannogram skills\PLANOGRAM_Royal_Plaza.xlsx"

# Color fills for tiers
TIER_COLORS = {
    "1": PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    ),  # Green - T1
    "8": PatternFill(
        start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"
    ),  # Blue - T8
    "2": PatternFill(
        start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"
    ),  # Yellow - T2
    "3": PatternFill(
        start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"
    ),  # Grey - T3
}
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14)
SUBTITLE_FONT = Font(bold=True, size=11)
AVAIL_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


# ============================================================
# STEP 1: PULL DATA FROM DB
# ============================================================
def pull_data():
    conn = psycopg2.connect(
        host="76.13.194.120",
        dbname="openclaw_ops",
        user="openclaw_app",
        password="Zuma-0psCl4w-2026!",
    )

    # 1a. Sales data — USE core.sales_with_product (Rule 5: use core views, not raw)
    # Rule 7: exclude intercompany transactions
    # Use kode_mix for version-agnostic article aggregation (Rule 6)
    sales_query = f"""
        SELECT
            kode_mix,
            article,
            gender,
            series,
            tipe AS product_type,
            tier,
            TO_CHAR(transaction_date, 'YYYY-MM') as bulan,
            SUM(quantity) as total_qty,
            SUM(total_amount) as total_revenue
        FROM core.sales_with_product
        WHERE matched_store_name = TRIM(LOWER('{DB_STORE_NAME}'))
          AND transaction_date >= '{DATE_START}'
          AND transaction_date <= '{DATE_END}'
          AND quantity > 0
          AND (is_intercompany IS NULL OR is_intercompany = FALSE)
        GROUP BY kode_mix, article, gender, series, tipe, tier,
                 TO_CHAR(transaction_date, 'YYYY-MM')
        ORDER BY kode_mix, bulan
    """
    sales_df = pd.read_sql(sales_query, conn)

    # 1b. Master kodemix — NO status filter (Rule 3: never filter by status)
    # Rule 1: use TRIM(LOWER()) for matching
    # Rule 2: deduplicate with DISTINCT ON
    master_query = """
        SELECT DISTINCT ON (TRIM(LOWER(kode_besar)))
            TRIM(LOWER(kode_besar)) AS kode_besar,
            kode, kode_mix, kode_mix_size, article, gender, series, tipe,
            tier_baru, status, version
        FROM portal.kodemix
        ORDER BY TRIM(LOWER(kode_besar)), no_urut
    """
    master_df = pd.read_sql(master_query, conn)

    # 1c. Stock data — USE core.stock_with_product (Rule 5)
    stock_query = f"""
        SELECT
            kode_mix, article, gender, series, tipe AS product_type,
            tier, nama_gudang, SUM(quantity) as total_qty
        FROM core.stock_with_product
        WHERE TRIM(LOWER(nama_gudang)) = TRIM(LOWER('{DB_STORE_NAME}'))
        GROUP BY kode_mix, article, gender, series, tipe, tier, nama_gudang
        ORDER BY total_qty DESC
    """
    stock_df = pd.read_sql(stock_query, conn)

    conn.close()
    return sales_df, master_df, stock_df


# ============================================================
# STEP 2: PROCESS SALES (already enriched from core view)
# ============================================================
def process_sales(sales_df, master_df):
    """
    Since we use core.sales_with_product, data is already enriched with
    kode_mix, article, gender, series, tipe, tier. No manual joining needed.
    Just clean and rename columns.
    """
    # Rename tier column to match our convention
    sales_df = sales_df.rename(columns={"product_type": "tipe", "tier": "tier_baru"})

    # Drop rows with no kode_mix (unmatched products ~6%)
    unmatched = sales_df[sales_df["kode_mix"].isna()]
    if len(unmatched) > 0:
        print(f"  INFO: {len(unmatched)} rows have no kode_mix match (excluded)")
    sales_df = sales_df[sales_df["kode_mix"].notna()].copy()

    # Drop rows with no article name
    sales_df = sales_df[sales_df["article"].notna()].copy()

    # Drop non-product items
    non_product_patterns = [
        "SHOPPING BAG",
        "HANGER",
        "PAPER BAG",
        "THERMAL",
        "BOX LUCA",
    ]
    mask = (
        sales_df["article"]
        .str.upper()
        .apply(lambda x: not any(p in str(x) for p in non_product_patterns))
    )
    sales_df = sales_df[mask].copy()

    # Standardize tier values
    sales_df["tier_baru"] = sales_df["tier_baru"].fillna("3").astype(str)

    # Standardize gender
    sales_df["gender"] = sales_df["gender"].fillna("UNKNOWN").str.upper()
    sales_df = sales_df[sales_df["gender"] != "UNKNOWN"]

    # Standardize tipe
    sales_df["tipe"] = sales_df["tipe"].fillna("Jepit")

    print(f"  Clean sales rows: {len(sales_df)}")
    print(f"  Unique articles: {sales_df['article'].nunique()}")
    print(f"  Unique kode_mix: {sales_df['kode_mix'].nunique()}")

    return sales_df


# ============================================================
# STEP 3: COMPUTE ADJUSTED AVERAGE (per skill Section 1.4.2)
# ============================================================
def compute_adjusted_avg(merged_df):
    # Generate all months in range
    all_months = (
        pd.date_range(DATE_START, DATE_END, freq="MS").strftime("%Y-%m").tolist()
    )

    # Pivot: article x month
    # First aggregate to article level (combine all kode variants of same article)
    article_monthly = (
        merged_df.groupby(["article", "gender", "series", "tipe", "tier_baru", "bulan"])
        .agg(total_qty=("total_qty", "sum"), total_revenue=("total_revenue", "sum"))
        .reset_index()
    )

    # Pivot
    articles = (
        article_monthly.groupby(["article", "gender", "series", "tipe", "tier_baru"])
        .apply(
            lambda g: g.set_index("bulan")["total_qty"].reindex(
                all_months, fill_value=0
            )
        )
        .reset_index()
    )

    # Reshape - the apply returns a multi-index df
    # Let's do this more carefully
    result_rows = []
    for (article, gender, series, tipe, tier), group in article_monthly.groupby(
        ["article", "gender", "series", "tipe", "tier_baru"]
    ):
        monthly = {}
        for _, row in group.iterrows():
            monthly[row["bulan"]] = float(row["total_qty"])

        # Fill missing months with 0
        month_values = [monthly.get(m, 0) for m in all_months]
        total_sales = sum(month_values)

        # Adjusted average per tier rules
        tier_str = str(tier)
        if tier_str == "1":
            # T1: exclude 0 months (likely OOS)
            nonzero = [v for v in month_values if v > 0]
            adj_avg = sum(nonzero) / len(nonzero) if nonzero else 0
        elif tier_str == "8":
            # T8: exclude leading 0s (not launched yet) and post-launch 0s (OOS)
            first_sale_idx = next(
                (i for i, v in enumerate(month_values) if v > 0), None
            )
            if first_sale_idx is not None:
                active = [v for v in month_values[first_sale_idx:] if v > 0]
                adj_avg = sum(active) / len(active) if active else 0
            else:
                adj_avg = 0
        elif tier_str in ("2", "3"):
            # T2/T3: contextual — if 0 is surrounded by decent months, exclude
            overall_avg = total_sales / len(all_months) if len(all_months) > 0 else 0
            active_months = 0
            active_total = 0
            for i, v in enumerate(month_values):
                if v > 0:
                    active_months += 1
                    active_total += v
                else:
                    # Check surrounding months
                    surrounding = []
                    for j in range(max(0, i - 2), min(len(month_values), i + 3)):
                        if j != i and month_values[j] > 0:
                            surrounding.append(month_values[j])
                    if surrounding and np.mean(surrounding) > overall_avg * 0.5:
                        pass  # Exclude this 0 (likely OOS)
                    else:
                        active_months += 1  # Include 0 (genuine decline)
                        active_total += 0
            adj_avg = active_total / active_months if active_months > 0 else 0
        else:
            # T4/T5: full average
            adj_avg = total_sales / len(all_months) if len(all_months) > 0 else 0

        result_rows.append(
            {
                "article": article,
                "gender": gender,
                "series": series,
                "tipe": tipe,
                "tier": tier_str,
                "total_12mo": total_sales,
                "adj_avg": round(adj_avg, 1),
                "months_sold": sum(1 for v in month_values if v > 0),
                "month_values": month_values,
            }
        )

    result_df = pd.DataFrame(result_rows)
    return result_df, all_months


# ============================================================
# STEP 4: GENDER-TYPE MAPPING & SALES SHARE
# ============================================================
def map_gender_type(row):
    gender = str(row["gender"]).upper()
    tipe = str(row["tipe"])
    if gender == "MEN":
        return f"Men {tipe}"
    elif gender == "LADIES":
        return f"Ladies {tipe}"
    elif gender == "BABY":
        return "Baby & Kids"
    elif gender == "BOYS":
        return "Boys"
    elif gender == "GIRLS":
        return "Girls"
    elif gender == "JUNIOR":
        return "Junior"
    else:
        return f"{gender} {tipe}"


def compute_gender_shares(result_df):
    result_df["gender_type"] = result_df.apply(map_gender_type, axis=1)

    shares = (
        result_df.groupby("gender_type")
        .agg(total_adj_avg=("adj_avg", "sum"), article_count=("article", "count"))
        .reset_index()
    )
    total = shares["total_adj_avg"].sum()
    shares["pct_share"] = (shares["total_adj_avg"] / total * 100).round(1)
    shares = shares.sort_values("total_adj_avg", ascending=False).reset_index(drop=True)
    shares["rank"] = range(1, len(shares) + 1)

    return shares


# ============================================================
# STEP 5: ASSIGN GENDER-TYPE TO BACKWALLS
# ============================================================
def assign_gender_to_backwalls(shares, backwalls):
    # Sort backwalls by hooks descending
    bw_sorted = sorted(backwalls, key=lambda x: x["hooks"], reverse=True)

    # Get top gender-types (we have 4 backwalls)
    top_genders = shares.head(len(bw_sorted))["gender_type"].tolist()

    assignments = []
    for i, bw in enumerate(bw_sorted):
        if i < len(top_genders):
            gt = top_genders[i]
            # Determine tipe from gender_type
            if "Fashion" in gt:
                hooks_per_article = 3
            else:
                hooks_per_article = 2  # Jepit or Baby (default jepit-like)
            # Special: Baby/Boys/Girls could be either
            if gt in ("Baby & Kids", "Boys", "Girls", "Junior"):
                hooks_per_article = 2  # Default jepit for kids
            slots = bw["hooks"] // hooks_per_article
            assignments.append(
                {
                    **bw,
                    "gender_type": gt,
                    "hooks_per_article": hooks_per_article,
                    "slots": slots,
                    "mode": "Full Box",
                }
            )
        else:
            assignments.append(
                {
                    **bw,
                    "gender_type": "UNASSIGNED",
                    "hooks_per_article": 2,
                    "slots": bw["hooks"] // 2,
                    "mode": "Full Box",
                }
            )

    return assignments


# ============================================================
# STEP 6: ASSIGN ARTICLES TO BACKWALLS
# ============================================================
def assign_articles(bw_assignments, result_df, rak_article_names=None):
    """
    Assign articles to backwalls.
    - Exclude Luca/Luna/Airmove (table/VM only per skill Section 1.5)
    - Exclude articles already assigned to Rak Baby (avoid duplication)
    """
    planogram = {}
    if rak_article_names is None:
        rak_article_names = set()

    # Luca/Luna/Airmove exclusion patterns (table/VM only, not backwall)
    TABLE_VM_ONLY = ["LUCA", "LUNA", "AIRMOVE"]

    for bw in bw_assignments:
        gt = bw["gender_type"]
        slots = bw["slots"]

        # Filter articles for this gender-type
        articles = result_df[result_df["gender_type"] == gt].copy()

        # Exclude Luca/Luna/Airmove — these are table/VM only (Section 1.5)
        luca_mask = (
            articles["article"]
            .str.upper()
            .apply(lambda x: not any(p in str(x) for p in TABLE_VM_ONLY))
        )
        excluded_luca = articles[~luca_mask]
        if len(excluded_luca) > 0:
            for _, art in excluded_luca.iterrows():
                print(
                    f"  EXCLUDED from {bw['id']}: {art['article']} (Luca/Luna/Airmove -> table/VM only)"
                )
        articles = articles[luca_mask]

        # Exclude articles already on Rak Baby (avoid duplication)
        if rak_article_names:
            rak_mask = ~articles["article"].isin(rak_article_names)
            articles = articles[rak_mask]

        # Sort by: tier priority (T1 first, then T8, T2, T3), then adj_avg DESC
        tier_priority = {"1": 0, "8": 1, "2": 2, "3": 3, "4": 4, "5": 5}
        articles["tier_sort"] = articles["tier"].map(tier_priority).fillna(9)
        articles = articles.sort_values(
            ["tier_sort", "adj_avg"], ascending=[True, False]
        )

        # Filter: only T1, T8, T2, T3 (no T4/T5)
        articles = articles[articles["tier"].isin(["1", "8", "2", "3"])]

        # Assign to slots
        assigned = articles.head(slots).reset_index(drop=True)

        # Check if all T1 fit
        t1_count = len(articles[articles["tier"] == "1"])
        if t1_count > slots:
            print(
                f"WARNING: {bw['id']} ({gt}): {t1_count} T1 articles but only {slots} slots!"
            )

        planogram[bw["id"]] = {
            "config": bw,
            "articles": assigned,
            "available_slots": max(0, slots - len(assigned)),
        }

    return planogram


# ============================================================
# STEP 7: ASSIGN RAK BABY
# ============================================================
def assign_rak_baby(result_df, rak_config):
    # Filter baby articles (Baby & Kids gender)
    baby = result_df[result_df["gender_type"] == "Baby & Kids"].copy()
    tier_priority = {"1": 0, "8": 1, "2": 2, "3": 3}
    baby["tier_sort"] = baby["tier"].map(tier_priority).fillna(9)
    baby = baby[baby["tier"].isin(["1", "8", "2", "3"])]
    baby = baby.sort_values(["tier_sort", "adj_avg"], ascending=[True, False])

    # Full mode: 1 article per layer, 6 pairs per layer
    layers = rak_config["layers"]
    assigned = baby.head(layers).reset_index(drop=True)

    return assigned


# ============================================================
# STEP 8: GENERATE XLSX OUTPUT (matching example format)
# ============================================================
def generate_xlsx(planogram, rak_articles, result_df, shares, all_months):
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # --- BACKWALL SHEETS ---
    for bw_id, data in planogram.items():
        config = data["config"]
        articles = data["articles"]
        available = data["available_slots"]
        gt = config["gender_type"]
        hooks = config["hooks"]
        hpa = config["hooks_per_article"]
        slots = config["slots"]
        used = len(articles)

        sheet_name = f"{bw_id} {gt}"
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]
        ws = wb.create_sheet(title=sheet_name)

        # Row 1: Title
        mode_label = "FULL BOX MODE"
        title = f"BACKWALL {bw_id[-1]} - {gt} - {hooks} Hooks - {mode_label}"
        ws.merge_cells(
            start_row=1, start_column=1, end_row=1, end_column=min(slots * hpa, 30)
        )
        ws.cell(row=1, column=1, value=title).font = TITLE_FONT

        # Row 2: Mode description
        if "Jepit" in gt or gt in ("Baby & Kids", "Boys", "Girls", "Junior"):
            mode_desc = f"Mode: Full Box (2 hook = 1 artikel = 12 pairs) | Capacity: {slots} artikel | Used: {used} | Available: {available} slot"
        else:
            mode_desc = f"Mode: Full Box (3 hook = 1 artikel = 12 pairs) | Capacity: {slots} artikel | Used: {used} | Available: {available} slot"
        ws.merge_cells(
            start_row=2, start_column=1, end_row=2, end_column=min(slots * hpa, 30)
        )
        ws.cell(row=2, column=1, value=mode_desc).font = SUBTITLE_FONT

        # Row 3: blank
        # Row 4: Hook labels
        ws.cell(row=4, column=1, value="Hook").font = Font(bold=True)
        col = 2
        for i in range(slots):
            start_hook = i * hpa + 1
            end_hook = start_hook + hpa - 1
            label = f"Hook {start_hook}-{end_hook}"
            if hpa > 1:
                ws.merge_cells(
                    start_row=4, start_column=col, end_row=4, end_column=col + hpa - 1
                )
            cell = ws.cell(row=4, column=col, value=label)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER
            # Apply border to merged cells too
            for mc in range(col, col + hpa):
                ws.cell(row=4, column=mc).border = THIN_BORDER
                ws.cell(row=4, column=mc).fill = HEADER_FILL
            col += hpa

        # Row 5: Series grouping
        ws.cell(row=5, column=1, value="Series").font = Font(bold=True)
        col = 2
        current_series = None
        series_start = col
        for i in range(slots):
            if i < len(articles):
                art_series = articles.iloc[i]["series"]
            else:
                art_series = "(Kosong)"
            if art_series != current_series:
                if current_series is not None and col > series_start:
                    # Could merge previous series group - skip for simplicity
                    pass
                current_series = art_series
                series_start = col
            if hpa > 1:
                ws.merge_cells(
                    start_row=5, start_column=col, end_row=5, end_column=col + hpa - 1
                )
            ws.cell(row=5, column=col, value=str(art_series)).font = Font(italic=True)
            ws.cell(row=5, column=col).alignment = Alignment(horizontal="center")
            col += hpa

        # Row 6: Article names
        ws.cell(row=6, column=1, value="Artikel").font = Font(bold=True)
        col = 2
        for i in range(slots):
            if i < len(articles):
                art = articles.iloc[i]
                name = art["article"]
                tier = str(art["tier"])
                fill = TIER_COLORS.get(tier, PatternFill())
            else:
                name = "AVAILABLE"
                fill = AVAIL_FILL
            if hpa > 1:
                ws.merge_cells(
                    start_row=6, start_column=col, end_row=6, end_column=col + hpa - 1
                )
            cell = ws.cell(row=6, column=col, value=name)
            cell.font = Font(bold=True, size=10)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.fill = fill
            cell.border = THIN_BORDER
            for mc in range(col, col + hpa):
                ws.cell(row=6, column=mc).border = THIN_BORDER
                ws.cell(row=6, column=mc).fill = fill
            col += hpa

        # Row 7: Tier + Avg
        ws.cell(row=7, column=1, value="Tier & Avg").font = Font(bold=True)
        col = 2
        for i in range(slots):
            if i < len(articles):
                art = articles.iloc[i]
                tier = str(art["tier"])
                avg = art["adj_avg"]
                label = f"T{tier} | Avg: {avg}"
                fill = TIER_COLORS.get(tier, PatternFill())
            else:
                label = ""
                fill = AVAIL_FILL
            if hpa > 1:
                ws.merge_cells(
                    start_row=7, start_column=col, end_row=7, end_column=col + hpa - 1
                )
            cell = ws.cell(row=7, column=col, value=label)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = fill
            cell.border = THIN_BORDER
            for mc in range(col, col + hpa):
                ws.cell(row=7, column=mc).border = THIN_BORDER
                ws.cell(row=7, column=mc).fill = fill
            col += hpa

        # Row 8: Pairs
        ws.cell(row=8, column=1, value="Pairs").font = Font(bold=True)
        col = 2
        for i in range(slots):
            if i < len(articles):
                label = "12 pairs (1 box)"
            else:
                label = ""
            if hpa > 1:
                ws.merge_cells(
                    start_row=8, start_column=col, end_row=8, end_column=col + hpa - 1
                )
            ws.cell(row=8, column=col, value=label).alignment = Alignment(
                horizontal="center"
            )
            col += hpa

        # Row 10: Legend
        ws.cell(row=10, column=1, value="LEGEND:").font = Font(bold=True)
        legend = [
            ("T1", "Fast Moving (MUST display)", "1"),
            ("T8", "New Launch (priority exposure)", "8"),
            ("T2", "Secondary (filler)", "2"),
            ("T3", "Tertiary (variety filler)", "3"),
        ]
        for j, (tier_label, desc, tier_key) in enumerate(legend):
            row_n = 11 + j
            cell = ws.cell(row=row_n, column=1, value=tier_label)
            cell.fill = TIER_COLORS.get(tier_key, PatternFill())
            cell.font = Font(bold=True)
            ws.cell(row=row_n, column=2, value=desc)

        # Row 16: Mode note
        if "Jepit" in gt or gt in ("Baby & Kids", "Boys", "Girls", "Junior"):
            note = "FULL BOX MODE: Setiap artikel menempati 2 hook bersebelahan = 1 box penuh (12 pairs) di display. Tidak ada sisa ke storage."
        else:
            note = "FULL BOX MODE: Setiap artikel menempati 3 hook bersebelahan = 1 box penuh (12 pairs) di display. Tidak ada sisa ke storage."
        ws.merge_cells(
            start_row=16, start_column=1, end_row=16, end_column=min(slots * hpa, 30)
        )
        ws.cell(row=16, column=1, value=note).font = Font(italic=True, size=9)

        # Column widths
        ws.column_dimensions["A"].width = 14
        for c in range(2, slots * hpa + 2):
            ws.column_dimensions[get_column_letter(c)].width = 12

    # --- RAK BABY SHEET ---
    ws = wb.create_sheet(title="Rak Baby")
    ws.merge_cells("A1:G1")
    ws.cell(
        row=1, column=1, value=f"RAK BABY - {RAK_BABY['layers']} Layer - FULL MODE"
    ).font = TITLE_FONT
    ws.merge_cells("A2:G2")
    ws.cell(
        row=2,
        column=1,
        value=f"Mode: Full (1 artikel per layer = 6 pairs) | Capacity: {RAK_BABY['layers']} artikel | Storage impact: 6 pairs/artikel",
    ).font = SUBTITLE_FONT

    # Headers
    headers = ["Layer", "Artikel", "Tier", "Adj Avg", "Pairs on Display"]
    for j, h in enumerate(headers):
        cell = ws.cell(row=4, column=j + 1, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    for i in range(len(rak_articles)):
        art = rak_articles.iloc[i]
        row_n = 5 + i
        ws.cell(row=row_n, column=1, value=f"Layer {i + 1}").border = THIN_BORDER
        cell = ws.cell(row=row_n, column=2, value=art["article"])
        cell.font = Font(bold=True)
        cell.fill = TIER_COLORS.get(str(art["tier"]), PatternFill())
        cell.border = THIN_BORDER
        ws.cell(row=row_n, column=3, value=f"T{art['tier']}").border = THIN_BORDER
        ws.cell(row=row_n, column=4, value=art["adj_avg"]).border = THIN_BORDER
        ws.cell(row=row_n, column=5, value="6 pairs (half box)").border = THIN_BORDER

    note_row = 5 + len(rak_articles) + 1
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=5)
    ws.cell(
        row=note_row,
        column=1,
        value="Full mode: 6 pairs per artikel per layer. Sisa 6 pairs/artikel ke storage (TAPI storage = 0, jadi display only).",
    ).font = Font(italic=True, size=9)

    for c in range(1, 6):
        ws.column_dimensions[get_column_letter(c)].width = 18

    # --- STORAGE ALLOCATION SHEET ---
    ws = wb.create_sheet(title="Storage Allocation")
    ws.merge_cells("A1:I1")
    ws.cell(
        row=1,
        column=1,
        value=f"STORAGE ALLOCATION - {STORE_NAME} - Capacity: {STORAGE_CAPACITY} Box",
    ).font = TITLE_FONT

    ws.merge_cells("A3:I3")
    ws.cell(
        row=3,
        column=1,
        value="Royal Plaza TIDAK memiliki storage/gudang. Semua artikel di-display dalam Full Box Mode (12 pairs = 1 box penuh).",
    ).font = Font(bold=True, color="FF0000")

    ws.merge_cells("A4:I4")
    ws.cell(
        row=4,
        column=1,
        value="Implikasi: Tidak bisa display Luca/Luna/Airmove (butuh storage). Tidak bisa pakai Compact Mode. Replenish harus dari gudang pusat.",
    ).font = Font(italic=True)

    ws.merge_cells("A6:I6")
    ws.cell(row=6, column=1, value="STORAGE SUMMARY").font = Font(bold=True, size=12)

    summary_data = [
        ("Table display backup (Luca/Luna/AM)", "0 box", "Tidak ada table/VM display"),
        ("Compact mode overflow", "0 box", "Full box mode only (no storage)"),
        ("T1 fast moving backup", "0 box", "No storage available"),
        ("", "", ""),
        ("TOTAL TERPAKAI", "0 box", ""),
        ("TOTAL CAPACITY", f"{STORAGE_CAPACITY} box", ""),
        ("SISA", f"{STORAGE_CAPACITY} box", ""),
    ]
    for i, (label, val, note) in enumerate(summary_data):
        row_n = 7 + i
        ws.cell(row=row_n, column=1, value=label).font = (
            Font(bold=True) if "TOTAL" in label else Font()
        )
        ws.cell(row=row_n, column=5, value=val)
        ws.cell(row=row_n, column=7, value=note)

    # --- SUMMARY REPORT SHEET ---
    ws = wb.create_sheet(title="Summary Report")
    ws.merge_cells("A1:F1")
    ws.cell(
        row=1, column=1, value=f"PLANOGRAM SUMMARY - {STORE_NAME}"
    ).font = TITLE_FONT

    # Display utilization
    total_slots = (
        sum(p["config"]["slots"] for p in planogram.values()) + RAK_BABY["layers"]
    )
    total_used = sum(len(p["articles"]) for p in planogram.values()) + len(rak_articles)
    total_avail = total_slots - total_used
    util_pct = round(total_used / total_slots * 100, 1) if total_slots > 0 else 0

    row = 3
    ws.cell(row=row, column=1, value="DISPLAY UTILIZATION").font = Font(
        bold=True, size=12
    )
    row += 1
    util_data = [
        ("Total hooks/slots available", total_slots),
        ("Total hooks/slots used", total_used),
        ("Utilization", f"{util_pct}%"),
        ("Empty slots", f"{total_avail} (available for expansion)"),
    ]
    for label, val in util_data:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=3, value=str(val))
        row += 1

    # Sales coverage
    row += 1
    ws.cell(row=row, column=1, value="SALES COVERAGE").font = Font(bold=True, size=12)
    row += 1
    total_adj = result_df["adj_avg"].sum()
    # Collect all displayed articles
    displayed_articles = set()
    for p in planogram.values():
        for _, art in p["articles"].iterrows():
            displayed_articles.add(art["article"])
    for _, art in rak_articles.iterrows():
        displayed_articles.add(art["article"])

    displayed_adj = result_df[result_df["article"].isin(displayed_articles)][
        "adj_avg"
    ].sum()
    coverage = round(displayed_adj / total_adj * 100, 1) if total_adj > 0 else 0

    cov_data = [
        ("Total adjusted avg (semua artikel)", f"{total_adj:.1f} pairs/bulan"),
        ("Adjusted avg dari artikel di display", f"{displayed_adj:.1f} pairs/bulan"),
        ("Sales coverage", f"{coverage}%"),
    ]
    for label, val in cov_data:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=3, value=str(val))
        row += 1

    # Tier distribution
    row += 1
    ws.cell(row=row, column=1, value="TIER DISTRIBUTION DI DISPLAY").font = Font(
        bold=True, size=12
    )
    row += 1
    headers = ["Tier", "Count", "% of Display", "Notes"]
    for j, h in enumerate(headers):
        cell = ws.cell(row=row, column=j + 1, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
    row += 1

    all_displayed = []
    for p in planogram.values():
        for _, art in p["articles"].iterrows():
            all_displayed.append(art)
    for _, art in rak_articles.iterrows():
        all_displayed.append(art)

    tier_counts = defaultdict(int)
    for art in all_displayed:
        tier_counts[str(art["tier"])] += 1

    total_displayed = len(all_displayed)
    tier_names = {
        "1": "Fast Moving",
        "8": "New Launch",
        "2": "Secondary",
        "3": "Tertiary",
    }

    # Count total T1 in store (not just displayed)
    total_t1 = len(result_df[result_df["tier"] == "1"])

    for t in ["1", "8", "2", "3"]:
        count = tier_counts.get(t, 0)
        pct = round(count / total_displayed * 100, 1) if total_displayed > 0 else 0
        note = ""
        if t == "1":
            note = f"{'All' if count >= total_t1 else f'{count} of {total_t1}'} T1 displayed"
        ws.cell(row=row, column=1, value=f"T{t} ({tier_names[t]})").border = THIN_BORDER
        ws.cell(row=row, column=1).fill = TIER_COLORS.get(t, PatternFill())
        ws.cell(row=row, column=2, value=count).border = THIN_BORDER
        ws.cell(row=row, column=3, value=f"{pct}%").border = THIN_BORDER
        ws.cell(row=row, column=4, value=note).border = THIN_BORDER
        row += 1

    # Storage utilization
    row += 1
    ws.cell(row=row, column=1, value="STORAGE UTILIZATION").font = Font(
        bold=True, size=12
    )
    row += 1
    ws.cell(row=row, column=1, value="Total capacity")
    ws.cell(row=row, column=3, value=f"{STORAGE_CAPACITY} box (NO STORAGE)")
    row += 1
    ws.cell(row=row, column=1, value="Used")
    ws.cell(row=row, column=3, value="0 box")
    row += 1

    # Gender-type assignment
    row += 1
    ws.cell(row=row, column=1, value="GENDER-TYPE ASSIGNMENT").font = Font(
        bold=True, size=12
    )
    row += 1
    headers = [
        "Display Unit",
        "Gender-Type",
        "Hooks",
        "Slots",
        "Used",
        "Available",
        "Sales Share",
    ]
    for j, h in enumerate(headers):
        cell = ws.cell(row=row, column=j + 1, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
    row += 1

    for bw_id, data in planogram.items():
        c = data["config"]
        gt = c["gender_type"]
        share_row = shares[shares["gender_type"] == gt]
        share_pct = (
            f"{share_row['pct_share'].iloc[0]}%" if len(share_row) > 0 else "N/A"
        )
        ws.cell(row=row, column=1, value=bw_id).border = THIN_BORDER
        ws.cell(row=row, column=2, value=gt).border = THIN_BORDER
        ws.cell(row=row, column=3, value=c["hooks"]).border = THIN_BORDER
        ws.cell(row=row, column=4, value=c["slots"]).border = THIN_BORDER
        ws.cell(row=row, column=5, value=len(data["articles"])).border = THIN_BORDER
        ws.cell(row=row, column=6, value=data["available_slots"]).border = THIN_BORDER
        ws.cell(row=row, column=7, value=share_pct).border = THIN_BORDER
        row += 1

    # Rak Baby row
    ws.cell(row=row, column=1, value="Rak Baby").border = THIN_BORDER
    ws.cell(row=row, column=2, value="Baby & Kids").border = THIN_BORDER
    ws.cell(
        row=row, column=3, value=f"{RAK_BABY['layers']} layers"
    ).border = THIN_BORDER
    ws.cell(row=row, column=4, value=RAK_BABY["layers"]).border = THIN_BORDER
    ws.cell(row=row, column=5, value=len(rak_articles)).border = THIN_BORDER
    ws.cell(
        row=row, column=6, value=max(0, RAK_BABY["layers"] - len(rak_articles))
    ).border = THIN_BORDER
    baby_share = shares[shares["gender_type"] == "Baby & Kids"]
    ws.cell(
        row=row,
        column=7,
        value=f"{baby_share['pct_share'].iloc[0]}%" if len(baby_share) > 0 else "N/A",
    ).border = THIN_BORDER
    row += 1

    # FLAGS & WARNINGS
    row += 1
    ws.cell(row=row, column=1, value="FLAGS & WARNINGS").font = Font(bold=True, size=12)
    row += 1

    flags = []

    # Check T1 not displayed
    all_displayed_names = displayed_articles
    t1_not_displayed = result_df[
        (result_df["tier"] == "1") & (~result_df["article"].isin(all_displayed_names))
    ]
    if len(t1_not_displayed) > 0:
        for _, art in t1_not_displayed.iterrows():
            flags.append(
                (
                    "CRITICAL",
                    f'T1 "{art["article"]}" ({art["gender_type"]}) NOT displayed — Avg: {art["adj_avg"]}',
                )
            )

    # Check gender-types without display
    assigned_gts = set(p["config"]["gender_type"] for p in planogram.values())
    assigned_gts.add("Baby & Kids")  # rak baby
    all_gts = set(result_df["gender_type"].unique())
    unassigned_gts = all_gts - assigned_gts
    for ugt in unassigned_gts:
        gt_sales = result_df[result_df["gender_type"] == ugt]["adj_avg"].sum()
        if gt_sales > 0:
            gt_share = shares[shares["gender_type"] == ugt]
            pct = f"{gt_share['pct_share'].iloc[0]}%" if len(gt_share) > 0 else "?"
            flags.append(
                (
                    "WARNING",
                    f"{ugt} has no display unit — {pct} sales share, {gt_sales:.0f} adj_avg total",
                )
            )

    # No storage warning
    flags.append(
        (
            "WARNING",
            "No storage (gudang) — cannot use Compact Mode, cannot display Luca/Luna/Airmove",
        )
    )

    # No table/VM
    luca_luna = result_df[
        result_df["article"].str.contains("LUCA|LUNA|AIRMOVE", case=False, na=False)
    ]
    luca_t1 = luca_luna[luca_luna["tier"] == "1"]
    if len(luca_t1) > 0:
        for _, art in luca_t1.iterrows():
            flags.append(
                (
                    "WARNING",
                    f'T1 Luca/Luna/Airmove "{art["article"]}" cannot display — no table/VM in store',
                )
            )

    # Positive flags
    if len(t1_not_displayed) == 0:
        # Check only for assigned gender-types
        t1_in_assigned = result_df[
            (result_df["tier"] == "1") & (result_df["gender_type"].isin(assigned_gts))
        ]
        t1_displayed_count = len(
            t1_in_assigned[t1_in_assigned["article"].isin(all_displayed_names)]
        )
        if t1_displayed_count == len(t1_in_assigned):
            flags.append(("POSITIVE", "All T1 for assigned gender-types are displayed"))

    t4t5_in_display = [a for a in all_displayed if str(a["tier"]) in ("4", "5")]
    if len(t4t5_in_display) == 0:
        flags.append(("POSITIVE", "No T4/T5 in display"))

    flag_colors = {
        "CRITICAL": Font(bold=True, color="FF0000"),
        "WARNING": Font(bold=True, color="FF8C00"),
        "POSITIVE": Font(bold=True, color="008000"),
    }

    for severity, msg in flags:
        icon = {"CRITICAL": "[CRITICAL]", "WARNING": "[WARNING]", "POSITIVE": "[OK]"}
        ws.cell(row=row, column=1, value=icon.get(severity, "")).font = flag_colors.get(
            severity, Font()
        )
        ws.cell(row=row, column=2, value=msg)
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 15

    # --- RANKED ARTICLES SHEET (bonus: full data) ---
    ws = wb.create_sheet(title="Full Article Ranking")
    ws.cell(
        row=1,
        column=1,
        value=f"ALL ARTICLES RANKED — {STORE_NAME} (Feb 2025 - Jan 2026)",
    ).font = TITLE_FONT
    ws.merge_cells("A1:L1")

    headers = [
        "Rank",
        "Article",
        "Gender",
        "Gender-Type",
        "Series",
        "Tipe",
        "Tier",
        "Adj Avg",
        "Total 12mo",
        "Months Sold",
        "In Display?",
        "Display Location",
    ]
    for j, h in enumerate(headers):
        cell = ws.cell(row=3, column=j + 1, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    # Build display location map
    display_map = {}
    for bw_id, data in planogram.items():
        for idx, (_, art) in enumerate(data["articles"].iterrows()):
            hpa = data["config"]["hooks_per_article"]
            hook_start = idx * hpa + 1
            hook_end = hook_start + hpa - 1
            display_map[art["article"]] = f"{bw_id} Hook {hook_start}-{hook_end}"
    for i in range(len(rak_articles)):
        art = rak_articles.iloc[i]
        display_map[art["article"]] = f"Rak Baby Layer {i + 1}"

    ranked = result_df.copy()
    tier_priority = {"1": 0, "8": 1, "2": 2, "3": 3, "4": 4, "5": 5}
    ranked["tier_sort"] = ranked["tier"].map(tier_priority).fillna(9)
    ranked = ranked.sort_values(
        ["tier_sort", "adj_avg"], ascending=[True, False]
    ).reset_index(drop=True)

    for i, (_, art) in enumerate(ranked.iterrows()):
        row_n = 4 + i
        in_display = "YES" if art["article"] in displayed_articles else "NO"
        loc = display_map.get(art["article"], "")

        ws.cell(row=row_n, column=1, value=i + 1).border = THIN_BORDER
        cell = ws.cell(row=row_n, column=2, value=art["article"])
        cell.border = THIN_BORDER
        cell.fill = TIER_COLORS.get(str(art["tier"]), PatternFill())
        ws.cell(row=row_n, column=3, value=art["gender"]).border = THIN_BORDER
        ws.cell(row=row_n, column=4, value=art["gender_type"]).border = THIN_BORDER
        ws.cell(row=row_n, column=5, value=art["series"]).border = THIN_BORDER
        ws.cell(row=row_n, column=6, value=art["tipe"]).border = THIN_BORDER
        ws.cell(row=row_n, column=7, value=f"T{art['tier']}").border = THIN_BORDER
        ws.cell(row=row_n, column=8, value=art["adj_avg"]).border = THIN_BORDER
        ws.cell(row=row_n, column=9, value=art["total_12mo"]).border = THIN_BORDER
        ws.cell(row=row_n, column=10, value=art["months_sold"]).border = THIN_BORDER
        ws.cell(row=row_n, column=11, value=in_display).border = THIN_BORDER
        ws.cell(row=row_n, column=12, value=loc).border = THIN_BORDER

    for c in range(1, 13):
        ws.column_dimensions[get_column_letter(c)].width = 18

    # --- MONTHLY SALES SHEET (bonus: raw monthly data) ---
    ws = wb.create_sheet(title="Monthly Sales Data")
    ws.cell(
        row=1, column=1, value=f"MONTHLY SALES BY ARTICLE — {STORE_NAME}"
    ).font = TITLE_FONT

    headers = (
        ["Article", "Gender-Type", "Tier", "Adj Avg"] + all_months + ["Total 12mo"]
    )
    for j, h in enumerate(headers):
        cell = ws.cell(row=3, column=j + 1, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    for i, (_, art) in enumerate(ranked.iterrows()):
        row_n = 4 + i
        ws.cell(row=row_n, column=1, value=art["article"]).border = THIN_BORDER
        ws.cell(row=row_n, column=2, value=art["gender_type"]).border = THIN_BORDER
        ws.cell(row=row_n, column=3, value=f"T{art['tier']}").border = THIN_BORDER
        ws.cell(row=row_n, column=4, value=art["adj_avg"]).border = THIN_BORDER
        for m_idx, val in enumerate(art["month_values"]):
            ws.cell(row=row_n, column=5 + m_idx, value=val).border = THIN_BORDER
        ws.cell(
            row=row_n, column=5 + len(all_months), value=art["total_12mo"]
        ).border = THIN_BORDER

    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 14

    # Save
    wb.save(OUTPUT_FILE)
    print(f"\nPLANOGRAM SAVED: {OUTPUT_FILE}")
    return wb


# ============================================================
# MAIN
# ============================================================
if __name__ == "__main__":
    print("=" * 60)
    print(f"BUILDING PLANOGRAM: {STORE_NAME}")
    print("=" * 60)

    print("\n[1/7] Pulling data from VPS DB...")
    sales_df, master_df, stock_df = pull_data()
    print(f"  Sales: {len(sales_df)} rows")
    print(f"  Master: {len(master_df)} articles")
    print(f"  Stock: {len(stock_df)} items")

    print("\n[2/7] Processing sales -> kodemix aggregation...")
    merged = process_sales(sales_df, master_df)
    print(f"  Merged: {len(merged)} rows")

    print("\n[3/7] Computing adjusted averages per tier...")
    result_df, all_months = compute_adjusted_avg(merged)
    print(f"  Articles: {len(result_df)}")
    print(f"  Months: {all_months}")

    print("\n[4/7] Gender-type mapping & sales share...")
    shares = compute_gender_shares(result_df)
    print(shares.to_string(index=False))

    print("\n[5/7] Assigning gender-types to backwalls...")
    bw_assignments = assign_gender_to_backwalls(shares, BACKWALLS)
    for bw in bw_assignments:
        print(
            f"  {bw['id']}: {bw['gender_type']} ({bw['hooks']} hooks -> {bw['slots']} slots, {bw['mode']})"
        )

    print("\n[6/7] Assigning articles to display...")
    # Assign Rak Baby FIRST so we can exclude those articles from backwalls
    rak_articles = assign_rak_baby(result_df, RAK_BABY)
    rak_article_names = (
        set(rak_articles["article"].tolist()) if len(rak_articles) > 0 else set()
    )
    planogram = assign_articles(
        bw_assignments, result_df, rak_article_names=rak_article_names
    )

    for bw_id, data in planogram.items():
        arts = data["articles"]
        avail = data["available_slots"]
        print(f"\n  {bw_id} ({data['config']['gender_type']}):")
        print(f"    Assigned: {len(arts)} articles, Available: {avail} slots")
        if len(arts) > 0:
            for _, a in arts.iterrows():
                print(f"      T{a['tier']} | {a['article']:<35} | Avg: {a['adj_avg']}")

    print(f"\n  Rak Baby:")
    for _, a in rak_articles.iterrows():
        print(f"    T{a['tier']} | {a['article']:<35} | Avg: {a['adj_avg']}")

    print("\n[7/7] Generating XLSX output...")
    generate_xlsx(planogram, rak_articles, result_df, shares, all_months)

    # Summary
    total_slots = (
        sum(p["config"]["slots"] for p in planogram.values()) + RAK_BABY["layers"]
    )
    total_used = sum(len(p["articles"]) for p in planogram.values()) + len(rak_articles)
    print(f"\n{'=' * 60}")
    print(
        f"DONE! Total: {total_used}/{total_slots} slots used ({round(total_used / total_slots * 100, 1)}%)"
    )
    print(f"Output: {OUTPUT_FILE}")
    print(f"{'=' * 60}")
